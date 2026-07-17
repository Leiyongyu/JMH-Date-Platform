"""从旧版 Excel 文件导入出口明细和进货明细到数据库。"""
import sys, os
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, '.')

from datetime import datetime
from openpyxl import load_workbook
from infrastructure.database import get_conn

EXPORT_FILE = r"D:\JMH\出口业务收汇情况表\外汇退税\数据源\2025-9~12出口明细.xlsx"
PURCHASE_FILE = r"D:\JMH\出口业务收汇情况表\外汇退税\数据源\2025-9~12进货明细.xlsx"

# ── 导入出口明细 ──
print("=" * 60)
print("导入出口明细...")
wb = load_workbook(EXPORT_FILE)
ws = wb.active

conn = get_conn()
cursor = conn.cursor()
cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
export_count = 0
for row in range(2, ws.max_row + 1):
    申报年月 = str(ws.cell(row, 1).value or '').strip()
    申报批次 = str(ws.cell(row, 2).value or '').strip()
    序号 = str(ws.cell(row, 3).value or '').strip()
    关联号 = str(ws.cell(row, 4).value or '').strip()
    出口发票号 = str(ws.cell(row, 5).value or '').strip()
    报关单号_21 = str(ws.cell(row, 6).value or '').strip()
    代理证明号 = str(ws.cell(row, 7).value or '').strip() or None
    出口日期_raw = ws.cell(row, 8).value
    商品代码 = str(ws.cell(row, 9).value or '').strip()
    商品名称 = str(ws.cell(row, 10).value or '').strip()
    单位 = str(ws.cell(row, 11).value or '').strip()
    数量 = ws.cell(row, 12).value
    FOB = ws.cell(row, 13).value
    申报商品代码 = str(ws.cell(row, 14).value or '').strip() or None
    业务类型 = str(ws.cell(row, 15).value or '').strip() or None
    备注 = str(ws.cell(row, 16).value or '').strip() or None

    if not 报关单号_21 or len(报关单号_21) < 18:
        continue

    # 拆分 21 位报关单号
    报关单号 = 报关单号_21[:18]
    项号 = 报关单号_21[18:] if len(报关单号_21) >= 21 else '001'

    # 日期处理
    if isinstance(出口日期_raw, datetime):
        出口日期 = 出口日期_raw.date()
    elif 出口日期_raw:
        try:
            出口日期 = datetime.strptime(str(出口日期_raw)[:10], '%Y-%m-%d').date()
        except ValueError:
            出口日期 = None
    else:
        出口日期 = None

    try:
        cursor.execute("""
            INSERT INTO export_detail (
                customs_declaration_no, customs_item_no, export_date,
                contract_no, export_invoice_no, agency_certificate_no,
                export_product_code, export_product_name,
                unit, export_quantity, fob_amount, currency_code,
                declaration_month, declaration_batch, sequence_no, relation_no,
                declared_product_code, tax_business_type, remark,
                customs_match_status,
                source_file_name, source_file_hash, import_batch_id, created_by,
                parse_status
            ) VALUES (
                %s, %s, %s,
                NULL, %s, %s,
                %s, %s,
                %s, %s, %s, 'USD',
                %s, %s, %s, %s,
                %s, %s, %s,
                'MATCHED',
                'old_import.xlsx', 'old_import', 0, 'IMPORT_SCRIPT',
                'CONFIRMED'
            )
            ON DUPLICATE KEY UPDATE
                export_date = VALUES(export_date),
                export_invoice_no = VALUES(export_invoice_no),
                export_product_code = VALUES(export_product_code),
                export_product_name = VALUES(export_product_name),
                unit = VALUES(unit),
                export_quantity = VALUES(export_quantity),
                fob_amount = VALUES(fob_amount),
                declaration_month = VALUES(declaration_month),
                declaration_batch = VALUES(declaration_batch),
                sequence_no = VALUES(sequence_no),
                relation_no = VALUES(relation_no),
                customs_match_status = 'MATCHED'
        """, (
            报关单号, 项号, 出口日期,
            出口发票号 or None, 代理证明号,
            商品代码, 商品名称,
            单位, 数量, FOB,
            申报年月 or None, 申报批次 or None, 序号.zfill(8) if 序号 else None, 关联号 or None,
            申报商品代码, 业务类型, 备注,
        ))
        export_count += 1
    except Exception as e:
        print(f"  SKIP row {row}: {e}")

conn.commit()
cursor.close()
conn.close()
print(f"  导入 {export_count} 条出口明细")

# ── 导入进货明细 ──
print("=" * 60)
print("导入进货明细...")
wb = load_workbook(PURCHASE_FILE)
ws = wb.active

conn = get_conn()
cursor = conn.cursor()
cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
purchase_count = 0
for row in range(2, ws.max_row + 1):
    申报年月 = str(ws.cell(row, 1).value or '').strip()
    申报批次 = str(ws.cell(row, 2).value or '').strip()
    序号 = str(ws.cell(row, 3).value or '').strip()
    关联号 = str(ws.cell(row, 4).value or '').strip()
    税种 = str(ws.cell(row, 5).value or '').strip()
    可退税额 = ws.cell(row, 6).value
    纳税号 = str(ws.cell(row, 7).value or '').strip()
    发票号 = str(ws.cell(row, 8).value or '').strip()
    开票日期_raw = ws.cell(row, 9).value
    商品代码 = str(ws.cell(row, 10).value or '').strip()
    商品名称 = str(ws.cell(row, 11).value or '').strip()
    单位 = str(ws.cell(row, 12).value or '').strip()
    数量 = ws.cell(row, 13).value
    计税金额 = ws.cell(row, 14).value
    征税率_pct = ws.cell(row, 15).value
    退税率_pct = ws.cell(row, 16).value
    备注 = str(ws.cell(row, 17).value or '').strip() or None

    if not 发票号:
        continue

    # 日期处理
    if isinstance(开票日期_raw, datetime):
        开票日期 = 开票日期_raw.date()
    elif 开票日期_raw:
        try:
            开票日期 = datetime.strptime(str(开票日期_raw)[:10], '%Y-%m-%d').date()
        except ValueError:
            开票日期 = None
    else:
        开票日期 = None

    # 税率转换（百分比 → 小数）
    征税率 = float(征税率_pct) / 100 if 征税率_pct else 0.13
    退税率 = float(退税率_pct) / 100 if 退税率_pct else 0.13
    可退税额_val = float(可退税额) if 可退税额 else None
    数量_val = float(数量) if 数量 else 0

    try:
        # 税额反推: 计税金额 * 征税率
        tax_amount = float(计税金额 or 0) * 征税率

        cursor.execute("""
            INSERT INTO purchase_inventory (
                invoice_no, invoice_date, invoice_item_no,
                supplier_tax_no, tax_type,
                product_name, unit,
                purchased_quantity, allocated_quantity, remaining_quantity,
                taxable_amount, tax_rate, refund_rate, tax_amount,
                refundable_tax_amount, inventory_status, remark,
                declaration_month, declaration_batch, sequence_no, relation_no,
                source_file_name, source_file_hash, import_batch_id, created_by,
                parse_status
            ) VALUES (
                %s, %s, %s,
                %s, %s,
                %s, %s,
                %s, 0, %s,
                %s, %s, %s, %s,
                %s, 'AVAILABLE', %s,
                %s, %s, %s, %s,
                'old_import.xlsx', 'old_import', 0, 'IMPORT_SCRIPT',
                'CONFIRMED'
            )
            ON DUPLICATE KEY UPDATE
                invoice_date = VALUES(invoice_date),
                supplier_tax_no = VALUES(supplier_tax_no),
                tax_type = VALUES(tax_type),
                product_name = VALUES(product_name),
                unit = VALUES(unit),
                purchased_quantity = VALUES(purchased_quantity),
                remaining_quantity = VALUES(purchased_quantity) - allocated_quantity,
                taxable_amount = VALUES(taxable_amount),
                tax_rate = VALUES(tax_rate),
                refund_rate = VALUES(refund_rate),
                tax_amount = VALUES(tax_amount),
                refundable_tax_amount = VALUES(refundable_tax_amount),
                declaration_month = VALUES(declaration_month),
                declaration_batch = VALUES(declaration_batch),
                sequence_no = VALUES(sequence_no),
                relation_no = VALUES(relation_no)
        """, (
            发票号, 开票日期, 1,  # invoice_item_no 固定为 1（旧数据无项号）
            纳税号, 税种 or 'V|增值税',
            商品名称, 单位,
            数量_val, 数量_val,
            计税金额, 征税率, 退税率, tax_amount,
            可退税额_val, 备注,
            申报年月 or None, 申报批次 or None, 序号.zfill(8) if 序号 else None, 关联号 or None,
        ))
        purchase_count += 1
    except Exception as e:
        print(f"  SKIP row {row}: {e}")

conn.commit()
cursor.close()
conn.close()
print(f"  导入 {purchase_count} 条进货明细")
print("=" * 60)
print("完成!")
