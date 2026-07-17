import unittest
from datetime import date

from openpyxl import load_workbook

from modules.tax_refund.parsers.export_report import EXPORT_HEADERS, build_export_detail_workbook


class ExportReportTests(unittest.TestCase):
    def test_builds_expected_16_column_export_format(self):
        output = build_export_detail_workbook([{
            'declaration_month': '202601', 'declaration_batch': '001',
            'sequence_no': '00000001', 'relation_no': '20260100100000001',
            'export_invoice_no': 'INV15L7CCK57',
            'customs_declaration_no': '531620260000035324', 'customs_item_no': '1',
            'export_date': date(2026, 1, 6), 'export_product_code': '8302100000',
            'export_product_name': '车门铰链', 'unit': '个', 'export_quantity': 20,
            'fob_amount': 391.8,
        }])
        workbook = load_workbook(output, data_only=True)
        sheet = workbook['出口明细']
        self.assertEqual(sheet['D2'].number_format, '@')
        self.assertTrue(sheet['D2'].quotePrefix)
        self.assertEqual(list(EXPORT_HEADERS), [cell.value for cell in sheet[1]])
        self.assertEqual('531620260000035324001', sheet['F2'].value)
        self.assertEqual('83021000', sheet['I2'].value)
        self.assertEqual(20, sheet['L2'].value)
        self.assertEqual(391.8, sheet['M2'].value)
        workbook.close()

    def test_optional_batch_and_sequence_stay_blank(self):
        output = build_export_detail_workbook([{
            'declaration_month': '202601',
            'declaration_batch': None,
            'sequence_no': None,
            'relation_no': None,
            'customs_declaration_no': '531620260000035324',
            'customs_item_no': '1',
        }])
        workbook = load_workbook(output, data_only=True)
        sheet = workbook['出口明细']
        self.assertIsNone(sheet['B2'].value)
        self.assertIsNone(sheet['C2'].value)
        self.assertIsNone(sheet['D2'].value)
        workbook.close()


if __name__ == '__main__':
    unittest.main()
