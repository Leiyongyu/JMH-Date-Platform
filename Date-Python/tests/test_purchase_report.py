import unittest
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from modules.tax_refund.parsers.purchase_report import PURCHASE_HEADERS, build_purchase_detail_workbook


class PurchaseReportTests(unittest.TestCase):
    def test_builds_expected_17_column_purchase_format(self):
        output = build_purchase_detail_workbook([{
            'invoice_no': '26442000001113119491',
            'invoice_date': date(2026, 1, 29),
            'invoice_item_no': 1,
            'supplier_tax_no': '91440101MA59EYHH03',
            'tax_type': 'V|增值税',
            'product_name': '空气弹簧',
            'unit': '个',
            'purchased_quantity': Decimal('10'),
            'taxable_amount': Decimal('2123.89'),
            'tax_rate': Decimal('0.13'),
            'refund_rate': Decimal('0.13'),
            'tax_amount': Decimal('276.11'),
            'refundable_tax_amount': Decimal('276.11'),
            'remark': '',
        }])
        workbook = load_workbook(output, data_only=True)
        sheet = workbook['进货明细']
        self.assertEqual(list(PURCHASE_HEADERS), [cell.value for cell in sheet[1]])
        self.assertEqual([None, None, '00000001', None], [sheet.cell(2, i).value for i in range(1, 5)])
        self.assertEqual('V|增值税', sheet['E2'].value)
        self.assertEqual('26442000001113119491', sheet['H2'].value)
        self.assertEqual('空气弹簧', sheet['K2'].value)
        self.assertEqual(Decimal('276.11'), Decimal(str(sheet['N2'].value)))
        self.assertEqual(13, sheet['O2'].value)
        self.assertEqual(13, sheet['P2'].value)
        self.assertIsNone(sheet['Q2'].value)

    def test_sequence_restarts_from_00000001_for_each_invoice(self):
        common = {
            'invoice_date': date(2026, 1, 29),
            'supplier_tax_no': '91440101MA59EYHH03',
            'product_name': '空气弹簧', 'unit': '个',
            'purchased_quantity': 1, 'tax_amount': 13,
            'taxable_amount': 100, 'tax_rate': Decimal('0.13'),
            'refund_rate': Decimal('0.13'), 'refundable_tax_amount': 13,
        }
        output = build_purchase_detail_workbook([
            {**common, 'invoice_no': 'A', 'invoice_item_no': 1},
            {**common, 'invoice_no': 'A', 'invoice_item_no': 2},
            {**common, 'invoice_no': 'B', 'invoice_item_no': 1},
        ])
        sheet = load_workbook(output, data_only=True)['进货明细']
        self.assertEqual(
            ['00000001', '00000002', '00000001'],
            [sheet[f'C{row}'].value for row in range(2, 5)],
        )


if __name__ == '__main__':
    unittest.main()
