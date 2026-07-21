import unittest
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from modules.tax_refund.parsers.forex_report import (
    FOREX_HEADERS,
    build_forex_summary_workbook,
)


class ForexReportTests(unittest.TestCase):
    def test_builds_expected_forex_summary_format(self):
        output = build_forex_summary_workbook([{
            'contract_no': 'RVG10645-260529-0004',
            'customs_declaration_no': '531620260000035324',
            'export_date': date(2026, 5, 29),
            'customs_port': '成都海关',
            'customs_contract_usd': Decimal('12822.00'),
            'export_amount_usd': Decimal('12822.00'),
            'monthly_exchange_rate': Decimal('7.160600'),
            'received_amount_usd': Decimal('10000.00'),
            'receipt_total_usd': Decimal('12000.00'),
            'core_transaction_no': 'TX202605290001',
            'actual_exchange_rate': Decimal('7.170000'),
            'settlement_receipt_rmb': Decimal('86040.00'),
            'receipt_date': date(2026, 6, 2),
            'difference_usd': Decimal('2822.00'),
            'business_entity': '玖马赫',
        }])

        workbook = load_workbook(output, data_only=True)
        sheet = workbook['回款汇总']
        self.assertEqual(list(FOREX_HEADERS), [cell.value for cell in sheet[1]])
        self.assertEqual('531620260000035324', sheet['C2'].value)
        self.assertEqual(Decimal('10000'), Decimal(str(sheet['I2'].value)))
        self.assertEqual(Decimal('12000'), Decimal(str(sheet['J2'].value)))
        self.assertEqual('yyyy-mm-dd', sheet['D2'].number_format)
        self.assertEqual('yyyy-mm-dd', sheet['N2'].number_format)
        self.assertEqual('A2', sheet.freeze_panes)
        self.assertEqual('A1:P2', sheet.auto_filter.ref)
        workbook.close()

    def test_empty_export_still_has_headers(self):
        workbook = load_workbook(build_forex_summary_workbook([]), data_only=True)
        sheet = workbook['回款汇总']
        self.assertEqual(list(FOREX_HEADERS), [cell.value for cell in sheet[1]])
        self.assertEqual(1, sheet.max_row)
        workbook.close()


if __name__ == '__main__':
    unittest.main()
