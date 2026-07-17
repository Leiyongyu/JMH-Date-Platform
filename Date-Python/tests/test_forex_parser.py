import hashlib
import os
import tempfile
import unittest
from datetime import date, datetime

from openpyxl import Workbook

from modules.tax_refund.parsers.forex_normalizer import normalize_date
from modules.tax_refund.parsers.forex_excel import parse_forex_workbook


HEADERS = [
    '序号', '合同编号', '报关单号', '出口日期', '出口口岸',
    '报关合同金额（USD)', '出口金额（USD)', '月度汇率',
    '回款金额（USD)', '收汇金额（USD）', '核心流水号', '实际汇率',
    '回单结汇金额（RMB）', '收汇时间', '差额（USD)', '结汇时间',
]


class ForexParserTests(unittest.TestCase):
    def _workbook_path(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = os.path.join(temp_dir.name, 'forex.xlsx')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'
        sheet.append(HEADERS)
        sheet.append([1, 'C1', '123456789012345678', '2025-01-01', '海关', 40, None,
                      710, 40, 100, 'TX1', 716.06, 71606, datetime(2025, 1, 2), None,
                      '一段跨多组的结汇说明'])
        sheet.append([2, 'C2', '223456789012345678', '2025-01-01', '海关', 60, None,
                      710, 60, 100, 'TX1', 716.06, 71606, datetime(2025, 1, 2), None, None])
        sheet.append([3, 'C3', '323456789012345678', '2025-01-01', '海关', 30, None,
                      710, 30, 30, 'TX2', 715, 21450, datetime(2025, 1, 3), 0.25, None])
        sheet.merge_cells('P2:P4')

        ignored = workbook.create_sheet('玖马赫')
        ignored.append(HEADERS)
        ignored.append([1, 'IGNORED', '999999999999999999'])
        workbook.save(path)
        workbook.close()
        return path

    def test_only_sheet1_is_parsed_and_core_transaction_defines_group(self):
        path = self._workbook_path()
        with open(path, 'rb') as source:
            file_hash = hashlib.sha256(source.read()).hexdigest()
        aux_map, results = parse_forex_workbook(path, file_hash)

        self.assertEqual({}, aux_map)
        self.assertEqual(['Sheet1'], [result['sheet_name'] for result in results])
        rows = results[0]['rows']
        self.assertEqual(3, len(rows))
        self.assertEqual(rows[0]['receipt_group_key'], rows[1]['receipt_group_key'])
        self.assertNotEqual(rows[0]['receipt_group_key'], rows[2]['receipt_group_key'])
        self.assertEqual(100, rows[0]['receipt_total_usd'])
        self.assertEqual(40, rows[0]['allocation_amount_usd'])
        self.assertEqual(60, rows[1]['allocation_amount_usd'])
        self.assertEqual(0.25, rows[2]['difference_usd'])
        self.assertNotIn('settlement_time_raw', rows[0])

    def test_datetime_normalizes_to_date(self):
        value = normalize_date(datetime(2025, 1, 2, 13, 30))
        self.assertEqual(date(2025, 1, 2), value)
        self.assertNotIsInstance(value, datetime)


if __name__ == '__main__':
    unittest.main()
