"""eBay 财务数据：酷长利润 Excel 清洗、增量覆盖与查询。"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import hashlib
import json
import os
import re
from tempfile import NamedTemporaryFile
from typing import Any

from fastapi import APIRouter, File, Header, Query, UploadFile
from openpyxl import load_workbook
from starlette.concurrency import run_in_threadpool

from infrastructure.database import get_conn
from core.errors import NotFoundError, ValidationError as AppValidationError


router = APIRouter(prefix="/ebay-finance", tags=["eBay finance"])
FILE_PATTERN = re.compile(r"^(.+?)-(.+?)-(\d{8})-(\d{8})\.xlsx$", re.IGNORECASE)
REQUIRED_HEADERS = ("SKU", "订单总额", "利润", "利润率")

FIELD_HEADERS = {
    "image_url": "图片", "multi_attribute": "是否多属性", "order_total": "订单总额",
    "order_amount": "订单金额", "units_sold": "售出数", "order_count": "订单数",
    "tax_amount": "税费", "profit": "利润", "profit_margin": "利润率",
    "product_sales_amount": "商品销售额", "shipping_revenue": "应收运费",
    "platform_fee": "平台费用", "payment_fee": "收款手续费", "purchase_cost": "采购成本",
    "first_leg_freight": "头程运费", "tail_freight": "尾程运费", "refund_amount": "退款金额",
    "advertising_fee": "广告费", "platform_other_fee": "平台其他费",
}
DECIMAL_FIELDS = tuple(k for k in FIELD_HEADERS if k not in ("image_url", "multi_attribute", "units_sold", "order_count"))
EDIT_FIELDS = (
    "order_total", "order_amount", "units_sold", "order_count", "tax_amount", "profit",
    "profit_margin", "product_sales_amount", "platform_fee", "purchase_cost",
    "advertising_fee", "refund_amount",
)


def _serialize(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _success(data: Any, *, meta: dict | None = None):
    body = {"success": True, "data": data}
    if meta is not None:
        body["meta"] = meta
    return body


def _parse_filename(filename: str) -> tuple[str, str, date, date]:
    clean_name = os.path.basename((filename or "").replace("\\", "/"))
    match = FILE_PATTERN.fullmatch(clean_name)
    if not match:
        raise ValueError("文件名必须符合：平台-站点-开始日期-结束日期.xlsx")
    try:
        start = datetime.strptime(match.group(3), "%Y%m%d").date()
        end = datetime.strptime(match.group(4), "%Y%m%d").date()
    except ValueError as exc:
        raise ValueError("文件名中的日期无效，应为 yyyyMMdd") from exc
    if end < start:
        raise ValueError("文件名中的结束日期不能早于开始日期")
    return match.group(1).strip().lower(), match.group(2).strip(), start, end


def _decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() in ("", "--"):
        return None
    try:
        return Decimal(str(value).replace(",", "").replace("$", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, bool)):
        return value.strip() if isinstance(value, str) else value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (float, Decimal)):
        return float(value)
    return str(value)


def _read_workbook(path: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        original_headers = next(iterator, None)
        if not original_headers:
            raise ValueError("Excel 文件缺少表头")
        headers, first_index, occurrences = [], {}, {}
        for index, value in enumerate(original_headers):
            name = str(value).strip() if value is not None else f"未命名列{index + 1}"
            first_index.setdefault(name, index)
            occurrences[name] = occurrences.get(name, 0) + 1
            headers.append(name if occurrences[name] == 1 else f"{name}__{occurrences[name]}")
        for required in REQUIRED_HEADERS:
            if required not in first_index:
                raise ValueError(f"Excel 缺少必要字段：{required}")

        result, seen = [], set()
        for excel_row, values in enumerate(iterator, start=2):
            values = tuple(values)
            sku_value = values[first_index["SKU"]] if first_index["SKU"] < len(values) else None
            sku = str(sku_value).strip() if sku_value is not None else ""
            if not sku:
                continue
            if sku in seen:
                raise ValueError(f"文件中 SKU 重复：{sku}（第{excel_row}行）")
            seen.add(sku)
            raw = {header: _json_value(values[i] if i < len(values) else None) for i, header in enumerate(headers)}
            item = {"sku": sku, "raw_data_json": json.dumps(raw, ensure_ascii=False, separators=(",", ":"))}
            for field, header in FIELD_HEADERS.items():
                value = values[first_index[header]] if header in first_index and first_index[header] < len(values) else None
                if field in ("image_url", "multi_attribute"):
                    item[field] = str(value).strip() if value is not None else None
                elif field in ("units_sold", "order_count"):
                    number = _decimal(value)
                    item[field] = int(number) if number is not None else None
                else:
                    item[field] = _decimal(value)
            result.append(item)
        if not result:
            raise ValueError("文件中没有可导入的 SKU 数据")
        return result
    finally:
        workbook.close()


def import_workbook(path: str, filename: str, operator: str) -> dict[str, Any]:
    platform, site, period_start, period_end = _parse_filename(filename)
    rows = _read_workbook(path)
    with open(path, "rb") as stream:
        file_hash = hashlib.sha256(stream.read()).hexdigest()
    operator = (operator or "ERP").strip()[:64]
    conn = get_conn()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT sku FROM ebay_finance_profit WHERE platform=%s AND site=%s AND period_start=%s AND period_end=%s",
            (platform, site, period_start, period_end),
        )
        existing = {row["sku"] for row in cursor.fetchall()}
        inserted = sum(1 for row in rows if row["sku"] not in existing)
        updated = len(rows) - inserted
        cursor.execute(
            """INSERT INTO ebay_finance_import_batch
               (platform,site,period_start,period_end,file_name,file_hash,total_rows,inserted_rows,updated_rows,operator,status,create_time,update_time)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'SUCCESS',NOW(),NOW())
               ON DUPLICATE KEY UPDATE id=LAST_INSERT_ID(id),file_name=VALUES(file_name),file_hash=VALUES(file_hash),
               total_rows=VALUES(total_rows),inserted_rows=VALUES(inserted_rows),updated_rows=VALUES(updated_rows),
               operator=VALUES(operator),status='SUCCESS',error_message=NULL,update_time=NOW()""",
            (platform, site, period_start, period_end, filename, file_hash, len(rows), inserted, updated, operator),
        )
        batch_id = cursor.lastrowid
        columns = ["image_url", "multi_attribute", "order_total", "order_amount", "units_sold", "order_count",
                   "tax_amount", "profit", "profit_margin", "product_sales_amount", "shipping_revenue", "platform_fee",
                   "payment_fee", "purchase_cost", "first_leg_freight", "tail_freight", "refund_amount",
                   "advertising_fee", "platform_other_fee"]
        sql = f"""INSERT INTO ebay_finance_profit
            (batch_id,platform,site,period_start,period_end,sku,{','.join(columns)},raw_data_json,source_file_name,created_by,updated_by,create_time,update_time)
            VALUES ({','.join(['%s'] * (6 + len(columns) + 6))})
            ON DUPLICATE KEY UPDATE batch_id=VALUES(batch_id),{','.join(f'{c}=VALUES({c})' for c in columns)},
            raw_data_json=VALUES(raw_data_json),source_file_name=VALUES(source_file_name),updated_by=VALUES(updated_by),update_time=NOW()"""
        params = []
        for row in rows:
            params.append((batch_id, platform, site, period_start, period_end, row["sku"],
                           *(row.get(column) for column in columns), row["raw_data_json"], filename,
                           operator, operator, datetime.now(), datetime.now()))
        cursor.executemany(sql, params)
        skus = [row["sku"] for row in rows]
        placeholders = ",".join(["%s"] * len(skus))
        cursor.execute(
            f"DELETE FROM ebay_finance_profit WHERE platform=%s AND site=%s AND period_start=%s AND period_end=%s AND sku NOT IN ({placeholders})",
            (platform, site, period_start, period_end, *skus),
        )
        deleted = cursor.rowcount
        conn.commit()
        return {"batchId": batch_id, "fileName": filename, "platform": platform, "site": site,
                "periodStart": period_start.isoformat(), "periodEnd": period_end.isoformat(),
                "totalRows": len(rows), "insertedRows": inserted, "updatedRows": updated, "deletedRows": deleted}
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()


def _row_to_camel(row: dict[str, Any]) -> dict[str, Any]:
    result = {}
    for key, value in row.items():
        parts = key.split("_")
        camel = parts[0] + "".join(part.capitalize() for part in parts[1:])
        result[camel] = _serialize(value)
    return result


@router.post("/import", summary="导入酷长利润 Excel")
async def import_chief_profit(file: UploadFile = File(...), x_erp_user: str | None = Header(None)):
    suffix = os.path.splitext(file.filename or "")[1].lower()
    if suffix != ".xlsx":
        raise AppValidationError("仅支持 .xlsx 文件")
    temp_path = ""
    try:
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as temp:
            temp_path = temp.name
            while chunk := await file.read(1024 * 1024):
                temp.write(chunk)
        try:
            result = await run_in_threadpool(
                import_workbook, temp_path, file.filename or "", x_erp_user or "ERP"
            )
        except ValueError as exc:
            raise AppValidationError(str(exc)) from exc
        return _success(result)
    finally:
        await file.close()
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


@router.get("", summary="分页查询 eBay 财务明细")
def list_profit(page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=500),
                platform: str | None = None, site: str | None = None, sku: str | None = None,
                period_start: date | None = None, period_end: date | None = None):
    where, params = [], []
    for column, value in (("platform", platform), ("site", site)):
        if value:
            where.append(f"{column}=%s"); params.append(value.strip())
    if sku:
        where.append("sku LIKE %s"); params.append(f"%{sku.strip()}%")
    if period_start:
        where.append("period_start>=%s"); params.append(period_start)
    if period_end:
        where.append("period_end<=%s"); params.append(period_end)
    clause = " WHERE " + " AND ".join(where) if where else ""
    conn = get_conn(); cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT COUNT(*) total FROM ebay_finance_profit" + clause, params)
        total = cursor.fetchone()["total"]
        columns = "id,batch_id,platform,site,period_start,period_end,sku,image_url,multi_attribute,order_total,order_amount,units_sold,order_count,tax_amount,profit,profit_margin,product_sales_amount,shipping_revenue,platform_fee,payment_fee,purchase_cost,first_leg_freight,tail_freight,refund_amount,advertising_fee,platform_other_fee,source_file_name,created_by,updated_by,create_time,update_time"
        cursor.execute(f"SELECT {columns} FROM ebay_finance_profit{clause} ORDER BY period_end DESC,profit DESC,id DESC LIMIT %s OFFSET %s", (*params, page_size, (page - 1) * page_size))
        return _success([_row_to_camel(row) for row in cursor.fetchall()], meta={"page": page, "page_size": page_size, "total": total})
    finally:
        cursor.close(); conn.close()


@router.get("/imports", summary="分页查询 eBay 导入记录")
def list_imports(page: int = Query(1, ge=1), page_size: int = Query(10, ge=1, le=100),
                 platform: str | None = None, site: str | None = None):
    where, params = [], []
    for column, value in (("platform", platform), ("site", site)):
        if value:
            where.append(f"{column}=%s"); params.append(value.strip())
    clause = " WHERE " + " AND ".join(where) if where else ""
    conn = get_conn(); cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT COUNT(*) total FROM ebay_finance_import_batch" + clause, params)
        total = cursor.fetchone()["total"]
        cursor.execute("SELECT * FROM ebay_finance_import_batch" + clause + " ORDER BY period_end DESC,update_time DESC,id DESC LIMIT %s OFFSET %s", (*params, page_size, (page - 1) * page_size))
        return _success([_row_to_camel(row) for row in cursor.fetchall()], meta={"page": page, "page_size": page_size, "total": total})
    finally:
        cursor.close(); conn.close()


@router.get("/{row_id}", summary="查询 eBay 财务明细")
def get_profit(row_id: int):
    conn = get_conn(); cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM ebay_finance_profit WHERE id=%s", (row_id,))
        row = cursor.fetchone()
        if row is None:
            raise NotFoundError("财务明细不存在")
        return _success(_row_to_camel(row))
    finally:
        cursor.close(); conn.close()


@router.put("/{row_id}", summary="编辑 eBay 财务明细")
def update_profit(row_id: int, payload: dict[str, Any], x_erp_user: str | None = Header(None)):
    values = []
    for field in EDIT_FIELDS:
        camel = field.split("_")[0] + "".join(part.capitalize() for part in field.split("_")[1:])
        value = payload.get(camel)
        if field in ("units_sold", "order_count"):
            number = _decimal(value); values.append(int(number) if number is not None else None)
        else:
            values.append(_decimal(value))
    conn = get_conn(); cursor = conn.cursor()
    try:
        assignments = ",".join(f"{field}=%s" for field in EDIT_FIELDS)
        cursor.execute(f"UPDATE ebay_finance_profit SET {assignments},updated_by=%s,update_time=NOW() WHERE id=%s",
                       (*values, (x_erp_user or "ERP").strip()[:64], row_id))
        if cursor.rowcount == 0:
            cursor.execute("SELECT 1 FROM ebay_finance_profit WHERE id=%s", (row_id,))
            if cursor.fetchone() is None:
                raise NotFoundError("财务明细不存在或已被删除")
        conn.commit()
        return _success({"updated": 1})
    except Exception:
        conn.rollback(); raise
    finally:
        cursor.close(); conn.close()
