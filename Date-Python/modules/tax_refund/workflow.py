"""数据库驱动的外汇退税汇总文件生成工作流。"""

from __future__ import annotations

import json
import re
import shutil
import uuid
from collections import defaultdict
from copy import copy
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from modules.tax_refund.repository import (
    get_refund_exports,
    get_refund_forex_rows,
    get_refund_purchases,
)
from modules.tax_refund.inventory_service import (
    confirm_generation,
    get_generation_by_task,
    mark_generation_published,
    release_generation,
    reserve_plan,
)
from modules.tax_refund.customs_numbers import customs_item_number


TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"
PURCHASE_TEMPLATE = TEMPLATE_DIR / "purchase_template.xlsx"
EXPORT_TEMPLATE = TEMPLATE_DIR / "export_template.xlsx"
RECEIPT_TEMPLATE = TEMPLATE_DIR / "receipt_template.xlsx"
ENTERPRISE_RECEIPT_TEMPLATE = TEMPLATE_DIR / "enterprise_receipt_template.xlsx"


@dataclass(slots=True)
class WorkflowOptions:
    output_parent_dir: str
    declaration_month: str = "202512"
    overwrite: bool = False
    payer_name: str = "Hong Kong Cammy Yeson Limited"
    export_ids: list[int] | None = None
    task_id: int | None = None
    idempotency_key: str | None = None
    operator_id: str = "ERP"
    operator_name: str = "ERP"


@dataclass(slots=True)
class WorkflowResult:
    success: bool = False
    output_dir: str | None = None
    generation_id: int | None = None
    supplier_count: int = 0
    generated_files: int = 0
    purchase_rows: int = 0
    export_rows: int = 0
    receipt_rows: int = 0
    source_purchase_rows: int = 0
    source_export_rows: int = 0
    source_forex_rows: int = 0
    unmatched_export_rows: int = 0
    missing_forex_records: int = 0
    multiple_forex_records: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class RefundWorkflowError(RuntimeError):
    """可直接展示给测试页面或 API 调用方的业务错误。"""


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value: Any) -> Decimal:
    if value is None or _text(value) == "":
        return Decimal("0")
    try:
        return Decimal(_text(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _money(value: Decimal | float | int | None) -> float | None:
    if value is None:
        return None
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _rate_percent(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(_decimal(value) * Decimal("100"))


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s*（）()]", "", _text(value)).replace("％", "%")


def _header_map(ws, scan_rows: int = 30) -> tuple[int, dict[str, int]]:
    best_row = 0
    best: dict[str, int] = {}
    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        current = {}
        for col_idx in range(1, ws.max_column + 1):
            key = _normalize_header(ws.cell(row_idx, col_idx).value)
            if key:
                current[key] = col_idx
        if len(current) > len(best):
            best_row, best = row_idx, current
    return best_row, best


def _set_cell(ws, row: int, col: int, value: Any, fmt: str = "") -> None:
    """写入单元格，可选日期格式。"""
    cell = ws.cell(row, col)
    cell.value = value
    if fmt == "date" and value is not None:
        cell.number_format = "yyyy-mm-dd"


def _format_date(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return _text(value) or None


def _capture_row_style(ws, row_idx: int) -> list[dict[str, Any]]:
    return [
        {
            "style": copy(ws.cell(row_idx, col)._style),
            "number_format": ws.cell(row_idx, col).number_format,
            "alignment": copy(ws.cell(row_idx, col).alignment),
            "protection": copy(ws.cell(row_idx, col).protection),
        }
        for col in range(1, ws.max_column + 1)
    ]


def _copy_row_style(ws, row_idx: int, styles: list[dict[str, Any]]) -> None:
    for col_idx, style in enumerate(styles, start=1):
        cell = ws.cell(row_idx, col_idx)
        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])


def _prepare_template(template_path: Path):
    wb = load_workbook(template_path)
    ws = wb.active
    header_row, columns = _header_map(ws)
    if not header_row:
        wb.close()
        raise RefundWorkflowError(f"模板未识别到表头: {template_path.name}")
    sample_row = min(header_row + 1, ws.max_row)
    styles = _capture_row_style(ws, sample_row)
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)
    return wb, ws, header_row, columns, styles


def _write_mapped_rows(
    template_path: Path,
    output_path: Path,
    rows: list[dict[str, Any]],
    mapping: dict[str, str],
) -> int:
    wb, ws, header_row, columns, styles = _prepare_template(template_path)
    try:
        for offset, source in enumerate(rows, start=1):
            row_idx = header_row + offset
            _copy_row_style(ws, row_idx, styles)
            for template_name, source_name in mapping.items():
                col = columns.get(_normalize_header(template_name))
                if col:
                    ws.cell(row_idx, col).value = source.get(source_name)
            for date_name in ("出口日期", "开票日期"):
                col = columns.get(_normalize_header(date_name))
                if col:
                    ws.cell(row_idx, col).value = _format_date(ws.cell(row_idx, col).value)
                    ws.cell(row_idx, col).number_format = "yyyy-mm-dd"
            voucher_col = columns.get(_normalize_header("凭证种类"))
            if voucher_col:
                ws.cell(row_idx, voucher_col).value = "02|增值税专用发票"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    finally:
        wb.close()
    return len(rows)


def _customs_item_number(row: dict[str, Any]) -> str:
    return customs_item_number(
        row.get("customs_declaration_no"), row.get("customs_item_no"))


def _product_code(value: Any) -> str:
    digits = "".join(character for character in _text(value) if character.isdigit())
    return digits[:8] if len(digits) >= 8 else digits


def _exchange_rate(value: Any) -> Decimal | None:
    rate = _decimal(value)
    if not rate:
        return None
    return rate / Decimal("100") if rate > Decimal("20") else rate


class RefundWorkflow:
    """从当前退税数据库生成按供货方分组的三套申报文件。"""

    purchase_mapping = {
        "申报年月": "申报年月",
        "申报批次": "申报批次",
        "关联号": "关联号",
        "税种": "税种",
        "进货凭证号": "进货凭证号",
        "开票日期": "开票日期",
        "供货方纳税人识别号": "供货方纳税号",
        "出口商品代码": "出口商品代码",
        "商品名称": "出口商品名称",
        "计量单位": "计量单位",
        "数量": "数量",
        "计税金额": "计税金额",
        "征税率%": "征税率",
        "退税率%": "退税率",
        "可退税额": "可退税额",
        "备注": "备注",
    }
    export_mapping = {
        "申报年月": "申报年月",
        "申报批次": "申报批次",
        "关联号": "关联号",
        "出口货物报关单号": "出口货物报关单号",
        "代理出口货物证明号": "代理出口货物证明号",
        "出口发票号码": "出口发票号码",
        "出口日期": "出口日期",
        "出口商品代码": "出口商品代码",
        "出口商品名称": "出口商品名称",
        "计量单位": "计量单位",
        "出口数量": "出口数量",
        "美元离岸价": "美元离岸价",
        "申报商品代码": "申报商品代码",
        "退免税业务类型": "退免税业务类型",
        "备注": "备注",
    }

    def validate(self, options: WorkflowOptions) -> dict[str, Any]:
        if not re.fullmatch(r"\d{6}", options.declaration_month):
            raise RefundWorkflowError("申报年月必须为6位数字，例如202512")
        if not _text(options.output_parent_dir):
            raise RefundWorkflowError("请填写生成文件保存路径")
        missing_templates = [
            path.name for path in (PURCHASE_TEMPLATE, EXPORT_TEMPLATE, RECEIPT_TEMPLATE)
            if not path.is_file()
        ]
        if missing_templates:
            raise RefundWorkflowError("缺少内置模板: " + "、".join(missing_templates))

        output_parent = Path(options.output_parent_dir).expanduser().resolve()
        target = (output_parent / "汇总").resolve()
        if target == output_parent or output_parent not in target.parents:
            raise RefundWorkflowError("输出目录校验失败")

        preview_result = WorkflowResult()
        plan = self._build_plan(options, preview_result)
        return {
            "output_parent_dir": str(output_parent),
            "target_dir": str(target),
            "target_exists": target.exists(),
            "purchase_template": str(PURCHASE_TEMPLATE),
            "export_template": str(EXPORT_TEMPLATE),
            "receipt_template": str(RECEIPT_TEMPLATE),
            "source_purchase_rows": preview_result.source_purchase_rows,
            "source_export_rows": preview_result.source_export_rows,
            "source_forex_rows": preview_result.source_forex_rows,
            "supplier_count": len(plan),
            "matched_export_rows": sum(len(item["exports"]) for item in plan),
            "purchase_allocation_rows": sum(len(item["purchases"]) for item in plan),
            "unmatched_export_rows": preview_result.unmatched_export_rows,
            "missing_forex_records": preview_result.missing_forex_records,
            "warnings": preview_result.warnings,
        }

    def run(self, options: WorkflowOptions) -> WorkflowResult:
        result = WorkflowResult()
        generation_id = None
        inventory_committed = False
        staging = None
        try:
            if options.task_id is None:
                raise RefundWorkflowError("正式生成退税文件必须关联API任务ID")
            existing = get_generation_by_task(options.task_id)
            if existing and existing["generation_status"] == "COMMITTED":
                return self._result_from_generation(existing)
            if existing and existing["generation_status"] == "FILE_PENDING":
                staging = Path(existing["staging_directory"])
                target = Path(existing["output_directory"])
                if not target.exists():
                    if not staging.exists():
                        raise RefundWorkflowError("库存已扣减，但待发布临时目录不存在，请人工检查")
                    staging.replace(target)
                mark_generation_published(existing["id"], str(target))
                existing["generation_status"] = "COMMITTED"
                return self._result_from_generation(existing)
            if existing and existing["generation_status"] in ("PREPARING", "RESERVED"):
                release_generation(existing["id"], "任务恢复时释放上次未完成的库存预占")

            info = self.validate(options)
            output_parent = Path(info["output_parent_dir"])
            target = Path(info["target_dir"])
            # 始终加时间戳，避免覆盖
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            target = target.parent / f"汇总_{ts}_{options.task_id}"
            output_parent.mkdir(parents=True, exist_ok=True)

            plan = self._build_plan(options, result)
            if not plan:
                raise RefundWorkflowError("没有出口商品能够被当前进货库存完整匹配")

            staging = output_parent / f".汇总.tmp-{uuid.uuid4().hex[:8]}"
            staging.mkdir(parents=True)
            try:
                generation_id = reserve_plan(
                    task_id=options.task_id,
                    idempotency_key=options.idempotency_key,
                    declaration_month=options.declaration_month,
                    operator_id=options.operator_id,
                    operator_name=options.operator_name,
                    target_dir=str(target),
                    staging_dir=str(staging),
                    plan=plan,
                )
                result.generation_id = generation_id
                self._generate(staging, plan, options, result)
                result.success = True
                result.output_dir = str(target)
                (staging / "生成结果.json").write_text(
                    json.dumps(result.to_dict(), ensure_ascii=False, indent=2, default=str),
                    encoding="utf-8",
                )
                confirm_generation(generation_id, result.to_dict(), options.operator_id)
                inventory_committed = True
                staging.replace(target)
                mark_generation_published(generation_id, str(target))
            except Exception:
                if generation_id and not inventory_committed:
                    release_generation(generation_id, "退税文件生成失败，已释放库存预占")
                if staging.exists() and not inventory_committed:
                    shutil.rmtree(staging, ignore_errors=True)
                raise
        except Exception as exc:
            result.success = False
            result.output_dir = None
            result.errors.append(str(exc))
        return result

    @staticmethod
    def _result_from_generation(generation: dict[str, Any]) -> WorkflowResult:
        payload = generation.get("result_payload") or {}
        fields = WorkflowResult.__dataclass_fields__
        values = {key: value for key, value in payload.items() if key in fields}
        result = WorkflowResult(**values)
        result.success = True
        result.generation_id = int(generation["id"])
        result.output_dir = generation.get("output_directory")
        return result

    def _build_plan(
        self,
        options: WorkflowOptions,
        result: WorkflowResult,
    ) -> list[dict[str, Any]]:
        purchases = get_refund_purchases()
        exports = get_refund_exports()
        # 如果指定了 export_ids，只处理选中的出口记录
        if options.export_ids:
            id_set = set(options.export_ids)
            exports = [e for e in exports if e["id"] in id_set]
        forex_rows = get_refund_forex_rows()
        result.source_purchase_rows = len(purchases)
        result.source_export_rows = len(exports)
        result.source_forex_rows = len(forex_rows)

        inventory: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(
            lambda: defaultdict(list)
        )
        for purchase in purchases:
            sku = _text(purchase.get("sku_normalized"))
            supplier = _text(purchase.get("supplier_tax_no"))
            if not sku or not supplier:
                continue
            lot = dict(purchase)
            lot["available"] = _decimal(purchase.get("remaining_quantity"))
            inventory[sku][supplier].append(lot)

        grouped: dict[str, dict[str, Any]] = defaultdict(
            lambda: {"exports": [], "purchases": [], "refund_total": Decimal("0")}
        )
        # 关联号兜底匹配（旧数据无 SKU 时使用）
        relation_purchase_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for purchase in purchases:
            rel = _text(purchase.get("relation_no"))
            if rel:
                lot = dict(purchase)
                lot["available"] = _decimal(purchase.get("remaining_quantity"))
                relation_purchase_map[rel].append(lot)

        for export in exports:
            sku = _text(export.get("sku_normalized"))
            quantity = _decimal(export.get("export_quantity"))
            relation = _text(export.get("relation_no"))

            # 旧数据兜底：无 SKU 时走关联号匹配
            use_relation_match = (not sku) and relation and relation in relation_purchase_map

            if not use_relation_match and (not sku or quantity <= 0):
                result.unmatched_export_rows += 1
                result.warnings.append(
                    f"出口 {export['customs_declaration_no']}-{export['customs_item_no']}: "
                    "缺少可用于进货匹配的SKU或数量"
                )
                continue

            if use_relation_match:
                # 关联号匹配：取同一关联号下的所有进货记录
                matched_purchases = relation_purchase_map[relation]
                supplier = _text(matched_purchases[0].get("supplier_tax_no"))
                # 确认该关联号下有可用库存
                total_avail = sum((p["available"] for p in matched_purchases), Decimal("0"))
                if total_avail < quantity:
                    result.unmatched_export_rows += 1
                    result.warnings.append(
                        f"出口 {export['customs_declaration_no']}-{export['customs_item_no']} "
                        f"关联号 {relation}: 库存不足 (需要 {quantity}, 可用 {total_avail})"
                    )
                    continue
                remaining = quantity
                allocations = []
                for lot in matched_purchases:
                    if remaining <= 0:
                        break
                    allocated = min(remaining, lot["available"])
                    if allocated <= 0:
                        continue
                    lot["available"] -= allocated
                    remaining -= allocated
                    # 关联号匹配：每条进货已预分配，不拆分金额，用原始值
                    allocation = {
                        "lot_id": lot["id"],
                        "export_id": export["id"],
                        "allocated_quantity": allocated,
                        "match_mode": "RELATION",
                        "申报年月": options.declaration_month,
                        "申报批次": None,
                        "关联号": None,
                        "税种": _text(lot.get("tax_type")) or "V|增值税",
                        "可退税额": _money(lot.get("refundable_tax_amount")),
                        "供货方纳税号": supplier,
                        "进货凭证号": _text(lot.get("invoice_no")),
                        "开票日期": lot.get("invoice_date"),
                        "出口商品代码": _product_code(export.get("export_product_code")),
                        "出口商品名称": _text(export.get("export_product_name")) or _text(lot.get("product_name")),
                        "计量单位": _text(lot.get("unit")) or _text(export.get("unit")),
                        "数量": float(allocated),
                        "计税金额": _money(lot.get("taxable_amount")),
                        "征税率": _rate_percent(lot.get("tax_rate")),
                        "退税率": _rate_percent(lot.get("refund_rate")),
                        "备注": _text(lot.get("remark")),
                    }
                    allocations.append(allocation)
                    grouped[supplier]["refund_total"] += _decimal(lot.get("refundable_tax_amount"))

                export_row = dict(export)
                export_row["出口货物报关单号"] = _customs_item_number(export)
                grouped[supplier]["exports"].append(export_row)
                grouped[supplier]["purchases"].extend(allocations)
                continue  # 跳过 SKU 匹配流程

            supplier_lots = inventory.get(sku, {})

            supplier_lots = inventory.get(sku, {})
            eligible = []
            for supplier, lots in supplier_lots.items():
                total_available = sum((lot["available"] for lot in lots), Decimal("0"))
                if total_available >= quantity:
                    first_lot = next((lot for lot in lots if lot["available"] > 0), None)
                    if first_lot:
                        eligible.append((first_lot["invoice_date"], first_lot["id"], supplier, lots))
            if not eligible:
                result.unmatched_export_rows += 1
                result.warnings.append(
                    f"出口 {export['customs_declaration_no']}-{export['customs_item_no']} "
                    f"SKU {sku}: 没有单一供货方库存能够完整覆盖数量 {quantity}"
                )
                continue

            _, _, supplier, lots = min(eligible, key=lambda item: (item[0], item[1], item[2]))
            remaining = quantity
            allocations = []
            for lot in lots:
                if remaining <= 0:
                    break
                allocated = min(remaining, lot["available"])
                if allocated <= 0:
                    continue
                lot["available"] -= allocated
                remaining -= allocated
                purchased_quantity = _decimal(lot.get("purchased_quantity"))
                ratio = allocated / purchased_quantity if purchased_quantity else Decimal("0")
                tax_amount = _decimal(lot.get("tax_amount")) * ratio
                refundable = _decimal(lot.get("refundable_tax_amount")) * ratio
                allocation = {
                    "lot_id": lot["id"],
                    "export_id": export["id"],
                    "allocated_quantity": allocated,
                    "match_mode": "SKU",
                    "申报年月": options.declaration_month,
                    "申报批次": None,
                    "关联号": None,
                    "税种": _text(lot.get("tax_type")) or "V|增值税",
                    "可退税额": _money(refundable),
                    "供货方纳税号": supplier,
                    "进货凭证号": _text(lot.get("invoice_no")),
                    "开票日期": lot.get("invoice_date"),
                    "出口商品代码": _product_code(export.get("export_product_code")),
                    "出口商品名称": _text(export.get("export_product_name")) or _text(lot.get("product_name")),
                    "计量单位": _text(lot.get("unit")) or _text(export.get("unit")),
                    "数量": float(allocated),
                    # 按已确认业务规则：计税金额取发票税额，并按分配数量拆分。
                    "计税金额": _money(tax_amount),
                    "征税率": _rate_percent(lot.get("tax_rate")),
                    "退税率": _rate_percent(lot.get("refund_rate")),
                    "备注": _text(lot.get("remark")),
                }
                allocations.append(allocation)
                grouped[supplier]["refund_total"] += refundable

            export_row = dict(export)
            export_row["出口货物报关单号"] = _customs_item_number(export)
            grouped[supplier]["exports"].append(export_row)
            grouped[supplier]["purchases"].extend(allocations)

        ordered_suppliers = sorted(
            grouped.items(),
            key=lambda item: (-item[1]["refund_total"], item[0]),
        )
        used_sequences = set()
        next_sequence = 1
        plan = []
        for code_number, (supplier, item) in enumerate(ordered_suppliers, start=1):
            code = f"{code_number:03d}"
            relation_by_export = {}
            for export in item["exports"]:
                raw_sequence = _text(export.get("sequence_no"))
                if re.fullmatch(r"\d{1,8}", raw_sequence) and int(raw_sequence) not in used_sequences:
                    sequence_number = int(raw_sequence)
                else:
                    while next_sequence in used_sequences:
                        next_sequence += 1
                    sequence_number = next_sequence
                    next_sequence += 1
                used_sequences.add(sequence_number)
                sequence = f"{sequence_number:08d}"
                relation = f"{options.declaration_month}{code}{sequence}"
                relation_by_export[export["id"]] = relation
                export.update({
                    "申报年月": options.declaration_month,
                    "申报批次": code,
                    "关联号": relation,
                    "出口发票号码": _text(export.get("export_invoice_no")),
                    "代理出口货物证明号": export.get("agency_certificate_no"),
                    "出口日期": export.get("export_date"),
                    "出口商品代码": _product_code(export.get("export_product_code")),
                    "出口商品名称": _text(export.get("export_product_name")),
                    "计量单位": _text(export.get("unit")),
                    "出口数量": export.get("export_quantity"),
                    "美元离岸价": export.get("fob_amount"),
                    "申报商品代码": export.get("declared_product_code"),
                    "退免税业务类型": export.get("tax_business_type"),
                    "备注": _text(export.get("remark")),
                })
            for purchase in item["purchases"]:
                purchase["申报批次"] = code
                purchase["关联号"] = relation_by_export[purchase["export_id"]]
            item.update({"code": code, "tax_id": supplier})
            plan.append(item)

        forex_grouped = defaultdict(list)
        for row in forex_rows:
            key = _text(row.get("customs_no_match_key"))
            if key:
                forex_grouped[key].append(row)
        for item in plan:
            item["forex"] = forex_grouped
            for export in item["exports"]:
                customs_key = _text(export.get("customs_declaration_no"))[:18]
                matched_receipts = [
                    row for row in forex_grouped.get(customs_key, [])
                    if row.get("core_transaction_no")
                ]
                if not matched_receipts:
                    result.missing_forex_records += 1
                elif len(matched_receipts) > 1:
                    result.multiple_forex_records += 1
                    result.warnings.append(
                        f"报关单 {customs_key} 存在 {len(matched_receipts)} 笔回款，"
                        "收汇情况表暂取最早一笔"
                    )
        return plan

    def _generate(
        self,
        staging: Path,
        plan: list[dict[str, Any]],
        options: WorkflowOptions,
        result: WorkflowResult,
    ) -> None:
        for item in plan:
            code = item["code"]
            tax_id = item["tax_id"]
            folder = staging / code
            purchase_name = f"《外贸企业出口退税进货明细申报表》导入模板-{tax_id}-{code}.xlsx"
            export_name = f"《外贸企业出口退税出口明细申报表》导入模板-{tax_id}-{code}.xlsx"
            receipt_name = f"出口业务收汇情况表_{tax_id}_{code}.xlsx"
            result.purchase_rows += _write_mapped_rows(
                PURCHASE_TEMPLATE, folder / purchase_name, item["purchases"], self.purchase_mapping)
            result.export_rows += _write_mapped_rows(
                EXPORT_TEMPLATE, folder / export_name, item["exports"], self.export_mapping)
            result.receipt_rows += self._write_receipt(
                folder / receipt_name, tax_id, code, item["exports"], item["forex"],
                options, result)
            result.generated_files += 3
            result.supplier_count += 1

        # 生成汇总版收汇情况表（所有供货方合并到一个文件）
        result.receipt_rows += self._write_receipt_consolidated(
            staging / "出口业务收汇情况表_汇总.xlsx", plan, options, result)
        result.generated_files += 1

        # 生成企业收汇情况表（汇总所有供货方）
        enterprise_rows = self._write_enterprise_receipt(
            staging / "企业收汇情况表.xlsx", plan, options, result)
        result.generated_files += 1
        if enterprise_rows:
            result.receipt_rows += enterprise_rows

    def _write_receipt(
        self,
        output: Path,
        tax_id: str,
        code: str,
        exports: list[dict[str, Any]],
        forex: dict[str, list[dict[str, Any]]],
        options: WorkflowOptions,
        result: WorkflowResult,
    ) -> int:
        wb, ws, header_row, columns, styles = _prepare_template(RECEIPT_TEMPLATE)
        count = 0
        try:
            for export in exports:
                customs_key = _text(export.get("customs_declaration_no"))[:18]
                matches = forex.get(customs_key, [])
                matched_receipts = [row for row in matches if row.get("core_transaction_no")]
                if not matched_receipts:
                    forex_row = matches[0] if matches else {}
                else:
                    forex_row = matched_receipts[0]

                rate = _exchange_rate(forex_row.get("monthly_exchange_rate"))
                usd = _decimal(export.get("fob_amount"))
                rmb = _money(usd * rate) if rate else None
                count += 1
                row_idx = header_row + count
                _copy_row_style(ws, row_idx, styles)
                values = {
                    "供货方纳税号": tax_id,
                    "申报年月": options.declaration_month,
                    "申报批次": code,
                    "出口货物报关单号": export["出口货物报关单号"],
                    "代理出口货物证明号": export.get("agency_certificate_no"),
                    "出口发票号码": _text(export.get("export_invoice_no")),
                    "出口货物销售额币种": "USD|美元",
                    "出口货物销售额": float(usd),
                    "出口货物销售额币种汇率": float(rate) if rate else None,
                    "出口货物销售额折合人民币金额": rmb,
                    "收汇日期": forex_row.get("receipt_date"),
                    "收汇凭证号": forex_row.get("core_transaction_no"),
                    "出口货物收汇币种": "USD|美元",
                    "凭证总金额": forex_row.get("settlement_receipt_rmb"),
                    "出口货物收汇金额": float(_decimal(forex_row.get("allocated_amount_usd"))) if forex_row.get("allocated_amount_usd") else None,
                                        "出口货物收汇币种汇率": float(_exchange_rate(forex_row.get("actual_exchange_rate"))) if forex_row.get("actual_exchange_rate") else None,
                    "出口货物收汇折合人民币金额": rmb,
                    "付汇人": options.payer_name,
                    "出口合同号": forex_row.get("contract_no") or export.get("contract_no"),
                }
                for name, value in values.items():
                    col = columns.get(_normalize_header(name))
                    if col:
                        ws.cell(row_idx, col).value = _format_date(value) if "日期" in name else value
                        if "日期" in name:
                            ws.cell(row_idx, col).number_format = "yyyy-mm-dd"
            output.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output)
        finally:
            wb.close()
        return count

    def _write_receipt_consolidated(
        self, output: Path, plan: list[dict[str, Any]],
        options: WorkflowOptions, result: WorkflowResult,
    ) -> int:
        """所有供货方的收汇情况表合并到一个文件。"""
        wb, ws, header_row, columns, styles = _prepare_template(RECEIPT_TEMPLATE)
        count = 0
        try:
            for item in plan:
                tax_id = item["tax_id"]
                code = item["code"]
                forex = item["forex"]
                for export in item["exports"]:
                    customs_key = _text(export.get("customs_declaration_no"))[:18]
                    matches = forex.get(customs_key, [])
                    matched_receipts = [row for row in matches if row.get("core_transaction_no")]
                    if not matched_receipts:
                        forex_row = matches[0] if matches else {}
                    else:
                        forex_row = matched_receipts[0]

                    rate = _exchange_rate(forex_row.get("monthly_exchange_rate"))
                    usd = _decimal(export.get("fob_amount"))
                    rmb = _money(usd * rate) if rate else None
                    count += 1
                    row_idx = header_row + count
                    _copy_row_style(ws, row_idx, styles)
                    values = {
                        "供货方纳税号": tax_id,
                        "申报年月": options.declaration_month,
                        "申报批次": code,
                        "出口货物报关单号": export["出口货物报关单号"],
                        "代理出口货物证明号": export.get("agency_certificate_no"),
                        "出口发票号码": _text(export.get("export_invoice_no")),
                        "出口货物销售额币种": "USD|美元",
                        "出口货物销售额": float(usd),
                        "出口货物销售额币种汇率": float(rate) if rate else None,
                        "出口货物销售额折合人民币金额": rmb,
                        "收汇日期": forex_row.get("receipt_date"),
                        "收汇凭证号": forex_row.get("core_transaction_no"),
                        "出口货物收汇币种": "USD|美元",
                        "凭证总金额": forex_row.get("settlement_receipt_rmb"),
                        "出口货物收汇金额": float(_decimal(forex_row.get("allocated_amount_usd"))) if forex_row.get("allocated_amount_usd") else None,
                        "出口货物收汇币种汇率": float(_exchange_rate(forex_row.get("actual_exchange_rate"))) if forex_row.get("actual_exchange_rate") else None,
                        "出口货物收汇折合人民币金额": rmb,
                        "付汇人": options.payer_name,
                        "出口合同号": forex_row.get("contract_no") or export.get("contract_no"),
                    }
                    for name, value in values.items():
                        col = columns.get(_normalize_header(name))
                        if col:
                            ws.cell(row_idx, col).value = _format_date(value) if "日期" in name else value
                            if "日期" in name:
                                ws.cell(row_idx, col).number_format = "yyyy-mm-dd"
            output.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output)
        finally:
            wb.close()
        return count

    def _write_enterprise_receipt(
        self,
        output: Path,
        plan: list[dict[str, Any]],
        options: WorkflowOptions,
        result: WorkflowResult,
    ) -> int:
        """生成企业收汇情况表（汇总所有供货方）。使用固定列号写入。"""
        if not ENTERPRISE_RECEIPT_TEMPLATE.is_file():
            return 0

        wb, ws, header_row, columns, styles = _prepare_template(ENTERPRISE_RECEIPT_TEMPLATE)
        count = 0
        try:
            for item in plan:
                tax_id = item["tax_id"]
                code = item["code"]
                forex = item["forex"]
                for export in item["exports"]:
                    customs_21 = _text(export.get("出口货物报关单号") or export.get("customs_declaration_no", ""))
                    customs_18 = customs_21[:18] if len(customs_21) >= 18 else customs_21
                    fob = _decimal(export.get("fob_amount"))

                    # 匹配外汇数据
                    matches = forex.get(customs_18, [])
                    matched_receipts = [row for row in matches if row.get("core_transaction_no")]
                    if not matched_receipts:
                        forex_row = matches[0] if matches else {}
                    else:
                        forex_row = matched_receipts[0]

                    rate = _exchange_rate(forex_row.get("monthly_exchange_rate"))
                    rmb = _money(fob * rate) if rate else None

                    count += 1
                    row_idx = header_row + count
                    _copy_row_style(ws, row_idx, styles)

                    export_date = export.get("出口日期") or export.get("export_date")
                    receipt_date = forex_row.get("receipt_date")
                    contract_no = forex_row.get("contract_no") or export.get("contract_no")
                    settlement_rmb = forex_row.get("settlement_receipt_rmb")
                    allocated_usd = forex_row.get("allocated_amount_usd")
                    usd_float = float(fob) if fob else None
                    rate_float = float(rate) if rate else None
                    rmb_float = float(rmb) if rmb else None

                    # 固定列号映射（匹配企业收汇情况表模版）
                    _set_cell(ws, row_idx, 1, count)                                   # A: 序号
                    _set_cell(ws, row_idx, 2, options.declaration_month)                # B: 退税申报属期
                    _set_cell(ws, row_idx, 3, code)                                     # C: 退税申报批次
                    _set_cell(ws, row_idx, 4, customs_21)                               # D: 出口报关单号（21位）
                    _set_cell(ws, row_idx, 5, _format_date(export_date), "date")        # E: 报关单出口日期
                    _set_cell(ws, row_idx, 6, rmb)                                      # F: 人民币离岸价
                    _set_cell(ws, row_idx, 7, usd_float)                                # G: 美元离岸价
                    _set_cell(ws, row_idx, 8, rmb)                                      # H: 申报退税销售额
                    _set_cell(ws, row_idx, 10, float(_decimal(allocated_usd)) if allocated_usd else None)  # J: 已收汇金额(分配金额)
                    _set_cell(ws, row_idx, 11, _format_date(receipt_date), "date")      # K: 收汇日期
                    _set_cell(ws, row_idx, 12, forex_row.get("core_transaction_no"))    # L: 收汇凭证号
                    _set_cell(ws, row_idx, 13, settlement_rmb)                          # M: 收汇凭证总金额(RMB)
                    _set_cell(ws, row_idx, 14, None)                                    # N: 未收汇金额
                    _set_cell(ws, row_idx, 17, _text(contract_no) or None)              # Q: 对应合同编号
                    _set_cell(ws, row_idx, 18, rmb)                                     # R: 合同总金额（人民币离岸价）

            output.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output)
        finally:
            wb.close()
        return count
