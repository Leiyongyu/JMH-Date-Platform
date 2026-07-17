"""外汇回款 Excel 解析器。

当前业务只导入 ``Sheet1``。其余工作表无论名称和内容均不参与解析、
预览或入库。
"""
import hashlib
import re

from openpyxl import load_workbook


TARGET_SHEET = 'Sheet1'
BUSINESS_ENTITY = '玖马赫'
ERROR_VALUES = {'-', '#N/A', '#REF!', '#VALUE!', '#NAME?', '#DIV/0!'}


def classify_sheet(name):
    """保留给页面层使用；只有 Sheet1 属于本次导入范围。"""
    clean = str(name or '').strip()
    if clean == TARGET_SHEET:
        return 'MAIN', BUSINESS_ENTITY
    return 'SKIP', None


def safe_read(ws, row, col, merged_values=None):
    if not col:
        return None
    return ws.cell(row, col).value


def _text(value):
    if value is None:
        return ''
    text = str(value).replace('\u00a0', ' ').strip()
    return '' if text in ERROR_VALUES else text


def _header_key(value):
    return re.sub(r'\s+', '', _text(value)).replace('（', '(').replace('）', ')').lower()


def _find_header(ws, merged_values, max_scan=20):
    for row in range(1, min(ws.max_row, max_scan) + 1):
        headers = {}
        for col in range(1, ws.max_column + 1):
            key = _header_key(safe_read(ws, row, col, merged_values))
            if key:
                headers[key] = col
        if '报关单号' in headers:
            return row, headers
    raise ValueError('Sheet1 未找到包含“报关单号”的表头')


def _column(headers, *aliases):
    for alias in aliases:
        col = headers.get(_header_key(alias))
        if col:
            return col
    return None


def _split_customs_ids(raw):
    """拆分单元格中的报关单号并保留18位匹配键。"""
    if raw is None:
        return []
    result = []
    for part in re.split(r'[\r\n,;，；]+', str(raw)):
        text = part.strip()
        if not text or text in ERROR_VALUES:
            continue
        match = re.search(r'\d{18,21}', text)
        customs_no = (match.group(0) if match else text)[:18]
        if len(customs_no) == 18:
            result.append(customs_no)
    return result


def _receipt_group(file_hash, ws, row, columns, merged_values=None):
    """新版Sheet1按核心流水号分组，不处理任何合并区域。"""
    evidence_columns = (
        columns['receipt_total'], columns['core_tx'], columns['actual_rate'],
        columns['settlement_rmb'], columns['receipt_date'],
    )
    if not any(_text(safe_read(ws, row, col, merged_values)) for col in evidence_columns if col):
        return None

    core_transaction_no = _text(safe_read(ws, row, columns['core_tx'], merged_values))
    if core_transaction_no:
        raw = f'{file_hash}|{TARGET_SHEET}|core|{core_transaction_no}'
    else:
        raw = f'{file_hash}|{TARGET_SHEET}|row|{row}'
    return {
        'group_key': hashlib.sha256(raw.encode('utf-8')).hexdigest()[:16],
    }


def _row_has_business_data(ws, row, merged_values):
    return any(_text(safe_read(ws, row, col, merged_values))
               for col in range(1, ws.max_column + 1))


def parse_forex_workbook(file_path, file_hash):
    """解析 Sheet1，返回兼容原调用的 ``(aux_map, sheet_results)``。"""
    workbook = load_workbook(file_path, data_only=True)
    try:
        if TARGET_SHEET not in workbook.sheetnames:
            raise ValueError('工作簿中不存在 Sheet1，无法导入外汇回款数据')

        ws = workbook[TARGET_SHEET]
        merged_values = None
        header_row, headers = _find_header(ws, merged_values)
        columns = {
            'serial_no': _column(headers, '序号'),
            'contract_no': _column(headers, '合同编号'),
            'customs_no': _column(headers, '报关单号'),
            'export_date': _column(headers, '出口日期'),
            'customs_port': _column(headers, '出口口岸'),
            'customs_contract_usd': _column(headers, '报关合同金额(USD)'),
            'export_amount_usd': _column(headers, '出口金额(USD)'),
            'monthly_rate': _column(headers, '月度汇率'),
            'allocation_amount': _column(headers, '回款金额(USD)'),
            'receipt_total': _column(headers, '收汇金额(USD)'),
            'core_tx': _column(headers, '核心流水号'),
            'actual_rate': _column(headers, '实际汇率'),
            'settlement_rmb': _column(headers, '回单结汇金额(RMB)'),
            'receipt_date': _column(headers, '收汇时间'),
            'difference_usd': _column(headers, '差额(USD)'),
        }
        required = ('contract_no', 'customs_no', 'allocation_amount', 'receipt_total', 'core_tx')
        missing = [name for name in required if not columns[name]]
        if missing:
            raise ValueError(f'Sheet1 缺少必要列: {", ".join(missing)}')

        rows = []
        errors = []
        seen_customs = {}

        for row_no in range(header_row + 1, ws.max_row + 1):
            if not _row_has_business_data(ws, row_no, merged_values):
                continue

            customs_raw = safe_read(ws, row_no, columns['customs_no'], merged_values)
            customs_list = _split_customs_ids(customs_raw)
            if not customs_list:
                errors.append({'sheet': TARGET_SHEET, 'row': row_no, 'message': '报关单号为空或格式无效'})
                continue

            group = _receipt_group(file_hash, ws, row_no, columns, merged_values)
            for customs_no in customs_list:
                contract_no = _text(safe_read(ws, row_no, columns['contract_no'], merged_values)) or None
                skip_import = False
                previous = seen_customs.get(customs_no)
                if previous:
                    skip_import = True
                    errors.append({
                        'sheet': TARGET_SHEET,
                        'row': row_no,
                        'message': (
                            f'报关单号 {customs_no} 与第{previous["row"]}行重复；'
                            f'合同分别为 {previous["contract"] or "空"} / {contract_no or "空"}，本行不入库'
                        ),
                    })
                else:
                    seen_customs[customs_no] = {'row': row_no, 'contract': contract_no}

                receipt_total = safe_read(ws, row_no, columns['receipt_total'], merged_values)
                rows.append({
                    'sheet_name': TARGET_SHEET,
                    'business_entity': BUSINESS_ENTITY,
                    'source_row_no': row_no,
                    'serial_no': _text(safe_read(ws, row_no, columns['serial_no'], merged_values)) or None,
                    'customs_declaration_no': customs_no,
                    'customs_no_match_key': customs_no,
                    'contract_no': contract_no,
                    'export_date': safe_read(ws, row_no, columns['export_date'], merged_values),
                    'customs_port': _text(safe_read(ws, row_no, columns['customs_port'], merged_values)) or None,
                    'customs_contract_usd': safe_read(ws, row_no, columns['customs_contract_usd'], merged_values),
                    'export_amount_usd': safe_read(ws, row_no, columns['export_amount_usd'], merged_values),
                    'monthly_exchange_rate_raw': safe_read(ws, row_no, columns['monthly_rate'], merged_values),
                    'allocation_amount_usd': safe_read(ws, row_no, columns['allocation_amount'], merged_values),
                    'receipt_amount_usd': safe_read(ws, row_no, columns['allocation_amount'], merged_values),
                    'receipt_total_usd': receipt_total,
                    'received_amount_usd': receipt_total,
                    'core_transaction_no': _text(safe_read(ws, row_no, columns['core_tx'], merged_values)) or None,
                    'actual_exchange_rate_raw': safe_read(ws, row_no, columns['actual_rate'], merged_values),
                    'settlement_receipt_rmb': safe_read(ws, row_no, columns['settlement_rmb'], merged_values),
                    'receipt_date': safe_read(ws, row_no, columns['receipt_date'], merged_values),
                    'difference_usd': safe_read(ws, row_no, columns['difference_usd'], merged_values),
                    'receipt_group_key': group['group_key'] if group else None,
                    'skip_import': skip_import,
                })

        return {}, [{
            'sheet_name': TARGET_SHEET,
            'category': 'MAIN',
            'business_entity': BUSINESS_ENTITY,
            'rows': rows,
            'errors': errors,
        }]
    finally:
        workbook.close()
