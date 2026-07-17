"""按退税出口明细模板生成可下载工作簿。"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


EXPORT_HEADERS = (
    '申报年月', '申报批次', '序号', '关联号', '出口发票号码', '出口货物报关单号',
    '代理出口货物证明号', '出口日期', '出口商品代码', '出口商品名称', '计量单位',
    '出口数量', '美元离岸价', '申报商品代码', '退（免）税业务类型', '备注',
)


def _customs_item_number(row):
    customs_no = str(row.get('customs_declaration_no') or '').strip()
    item_no = str(row.get('customs_item_no') or '').strip().lstrip('0') or '0'
    if len(customs_no) >= 21 and customs_no[-3:] == item_no.zfill(3):
        return customs_no
    return f'{customs_no}{item_no.zfill(3)}'


def _tax_product_code(value):
    digits = ''.join(character for character in str(value or '') if character.isdigit())
    return digits[:8] if len(digits) >= 8 else digits


def _optional_padded(value, width):
    text = str(value or '').strip()
    return text.zfill(width) if text else ''


def build_export_detail_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '出口明细'
    sheet.append(EXPORT_HEADERS)

    for row in rows:
        sheet.append((
            str(row.get('declaration_month') or ''),
            _optional_padded(row.get('declaration_batch'), 3),
            _optional_padded(row.get('sequence_no'), 8),
            str(row.get('relation_no') or ''),
            row.get('export_invoice_no') or '',
            _customs_item_number(row),
            row.get('agency_certificate_no') or '',
            row.get('export_date'),
            _tax_product_code(row.get('export_product_code')),
            row.get('export_product_name') or '',
            row.get('unit') or '',
            row.get('export_quantity'),
            row.get('fob_amount'),
            row.get('declared_product_code') or '',
            row.get('tax_business_type') or '',
            row.get('remark') or '',
        ))

    header_fill = PatternFill('solid', fgColor='BFBFBF')
    thin = Side(style='thin', color='404040')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        for column in (1, 2, 3, 4, 5, 6, 9, 14, 15):
            row[column - 1].number_format = '@'
            row[column - 1].quotePrefix = True
        row[7].number_format = 'yyyy-mm-dd'
        row[11].number_format = '0.######'
        row[12].number_format = '0.00####'

    widths = (11, 10, 12, 21, 20, 25, 20, 13, 15, 28, 11, 12, 14, 16, 20, 20)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:P{max(1, sheet.max_row)}'
    sheet.row_dimensions[1].height = 22

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
