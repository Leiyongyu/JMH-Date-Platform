"""生成外汇回款汇总 Excel。"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


FOREX_HEADERS = (
    '序号', '合同编号', '报关单号', '出口日期', '出口口岸',
    '报关合同金额(USD)', '出口金额(USD)', '月度汇率',
    '报关单分配回款金额(USD)', '银行回款总额(USD)', '核心流水号',
    '实际汇率', '回单结汇金额(RMB)', '收汇日期', '差额(USD)', '业务主体',
)


def build_forex_summary_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '回款汇总'
    sheet.append(FOREX_HEADERS)

    for index, row in enumerate(rows, start=1):
        sheet.append((
            index,
            row.get('contract_no') or '',
            row.get('customs_declaration_no') or '',
            row.get('export_date'),
            row.get('customs_port') or '',
            row.get('customs_contract_usd'),
            row.get('export_amount_usd'),
            row.get('monthly_exchange_rate'),
            row.get('received_amount_usd'),
            row.get('receipt_total_usd'),
            row.get('core_transaction_no') or '',
            row.get('actual_exchange_rate'),
            row.get('settlement_receipt_rmb'),
            row.get('receipt_date'),
            row.get('difference_usd'),
            row.get('business_entity') or '',
        ))

    header = sheet[1]
    for cell in header:
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Microsoft YaHei', size=10)
            cell.alignment = Alignment(vertical='center')
        for column in (2, 3, 5, 11, 16):
            row[column - 1].number_format = '@'
            row[column - 1].quotePrefix = True
        row[3].number_format = 'yyyy-mm-dd'
        row[13].number_format = 'yyyy-mm-dd'
        for column in (6, 7, 9, 10, 13, 15):
            row[column - 1].number_format = '#,##0.00'
        for column in (8, 12):
            row[column - 1].number_format = '0.000000'

    widths = (9, 18, 24, 13, 16, 20, 18, 14, 24, 22, 22, 14, 20, 13, 15, 16)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:P{max(1, sheet.max_row)}'
    sheet.row_dimensions[1].height = 30
    sheet.sheet_view.showGridLines = False

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
