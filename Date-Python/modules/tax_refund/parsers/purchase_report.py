"""按外贸企业出口退税进货明细格式生成可下载工作簿。"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PURCHASE_HEADERS = (
    '申报年月', '申报批次', '序号', '关联号', '税种', '可退税额',
    '供货方纳税号', '进货凭证号', '开票日期', '出口商品代码', '出口商品名称',
    '计量单位', '数量', '计税金额', '征税率(%)', '退税率(%)', '备注',
)


def _rate_percent(value):
    if value in (None, ''):
        return None
    return float(value) * 100


def build_purchase_detail_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '进货明细'
    sheet.append(PURCHASE_HEADERS)

    for item in rows:
        sheet.append((
            '',  # 申报年月：后续形成申报任务时填写
            '',  # 申报批次：后续形成申报任务时填写
            str(item.get('invoice_item_no') or '').zfill(8),  # 每张发票内部从00000001开始
            '',  # 关联号：分配到出口明细后使用出口年月+批次+序号
            item.get('tax_type') or 'V|增值税',
            item.get('refundable_tax_amount'),
            item.get('supplier_tax_no') or '',
            item.get('invoice_no') or '',
            item.get('invoice_date'),
            item.get('export_product_code') or '',
            item.get('export_product_name') or item.get('product_name') or '',
            item.get('unit') or '',
            item.get('purchased_quantity'),
            item.get('tax_amount'),
            _rate_percent(item.get('tax_rate')),
            _rate_percent(item.get('refund_rate')),
            item.get('remark') or '',
        ))

    header_fill = PatternFill('solid', fgColor='BFBFBF')
    thin = Side(style='thin', color='404040')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        for column in (1, 2, 3, 4, 5, 7, 8, 10, 17):
            row[column - 1].number_format = '@'
            row[column - 1].quotePrefix = True
        row[5].number_format = '0.00'
        row[8].number_format = 'yyyy-mm-dd'
        row[12].number_format = '0.######'
        row[13].number_format = '0.00'
        row[14].number_format = '0.######'
        row[15].number_format = '0.######'

    widths = (11, 10, 12, 21, 13, 13, 23, 24, 13, 15, 28, 11, 12, 14, 13, 13, 20)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:Q{max(1, sheet.max_row)}'
    sheet.row_dimensions[1].height = 22

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
