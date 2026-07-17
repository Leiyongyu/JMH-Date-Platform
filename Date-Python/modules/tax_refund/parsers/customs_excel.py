"""
报关资料 Excel 解析模块
========================
兼容两种格式的报关单Sheet商品解析：
- V1_COMBINED: 商品名称及规格型号 合一列
- V2_SKU_SPEC: 独立的 商品名称、SKU、规格型号 列
"""
import re
import os
import hashlib
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from modules.tax_refund.parsers.sku_normalizer import normalize_sku


# 停止行关键词
STOP_KEYWORDS = ['报关人员', '申报单位', '兹申明', '海关批注及签章']


def parse_customs_excel(file_path):
    """
    解析报关资料 Excel 的"报关单"Sheet。
    返回: (contract_no, format_version, items, error_msg)
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return None, None, [], f'无法打开文件: {e}'

    # 1. 找到"报关单"Sheet
    target_sheet = None
    for name in wb.sheetnames:
        if name.strip().replace(' ', '') == '报关单':
            target_sheet = name
            break
    if not target_sheet:
        wb.close()
        return None, None, [], '未找到"报关单"工作表'

    ws = wb[target_sheet]

    # 2. 提取合同协议号
    contract_no = _find_contract_no(ws)

    # 3. 查找商品表头行
    header_row, header_cols = _find_header_row(ws)
    if not header_row:
        wb.close()
        return None, None, [], '未找到商品表头行（需含项号、商品编码、数量及单位、单价/总价/币制）'

    # 4. 识别格式版本
    format_ver = _detect_format(header_cols)

    # 5. 从商业发票/装箱单抬头提取出口发票号，并解析商品行
    export_invoice_no = _extract_export_invoice_no(wb)
    items = _parse_item_rows(
        ws, header_row, header_cols, format_ver, contract_no, export_invoice_no, file_path)

    wb.close()
    return contract_no, format_ver, items, None


def _find_contract_no(ws):
    """在报关单Sheet中查找合同协议号"""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        for i, cell in enumerate(row):
            if cell and '合同协议号' in str(cell):
                # 右侧单元格
                if i + 1 < len(row) and row[i + 1]:
                    val = str(row[i + 1]).strip()
                    # 去除换行和空格
                    val = re.sub(r'\s+', '', val)
                    return val.upper()
    return ''


def _find_header_row(ws):
    """
    在前30行查找含 项号、商品编码、数量及单位、单价/总价/币制 的行。
    同时检测 商品名称、商品名称及规格型号、SKU、规格型号 列。
    返回: (row_num, column_map_dict)
    """
    required = ['项号', '商品编码', '数量及单位']
    price_keys = ['单价/总价/币制', '单价', '总价']
    optional = ['商品名称及规格型号', '商品名称', 'SKU', '规格型号']

    for row_idx in range(1, min(ws.max_row + 1, 31)):
        row_vals = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                row_vals.append((col_idx, str(val).strip().replace('\n', ' ')))
            else:
                row_vals.append((col_idx, ''))

        found = {}
        for col_idx, val in row_vals:
            clean = val.replace(' ', '').replace('*', '')
            # 检查必需列
            for req in required:
                if req.replace(' ', '') in clean and req not in found:
                    found[req] = col_idx
            # 检查价格列
            for pk in price_keys:
                if pk.replace(' ', '').replace('/', '') in clean and '单价/总价/币制' not in found:
                    found['单价/总价/币制'] = col_idx
            # 检查可选列
            for opt in optional:
                opt_clean = opt.replace(' ', '').replace('*', '')
                if opt_clean in clean and opt not in found:
                    found[opt] = col_idx

        # 必须有：项号、商品编码、数量及单位、单价（任意形式）
        has_price = '单价/总价/币制' in found
        if all(r in found for r in required) and has_price:
            return row_idx, found

    return None, None


def _detect_format(header_cols):
    """识别格式版本：根据表头列名判断"""
    # Collect all header keys
    all_keys = ' '.join(str(k) for k in header_cols.keys())
    has_sku_col = 'SKU' in all_keys.upper() or 'sku' in all_keys
    has_spec_col = '规格型号' in all_keys
    has_name_col = '商品名称' in all_keys
    has_combined = '商品名称及规格型号' in all_keys

    # V2: 有独立的 商品名称 + SKU + 规格型号 列
    if has_sku_col and has_spec_col and has_name_col and not has_combined:
        return 'V2_SKU_SPEC'
    return 'V1_COMBINED'


def _parse_item_rows(ws, header_row, header_cols, format_ver, contract_no,
                     export_invoice_no, file_path):
    """解析商品行数据"""
    items = []
    source_file = os.path.basename(file_path)
    # Compute hash once
    sha = hashlib.sha256()
    with open(file_path, 'rb') as f:
        sha.update(f.read())
    file_hash = sha.hexdigest()

    # 列映射
    seq_col = header_cols.get('项号', 1)
    code_col = header_cols.get('商品编码', 2)
    qty_col = header_cols.get('数量及单位')
    price_col = header_cols.get('单价/总价/币制')

    for row_idx in range(header_row + 1, ws.max_row + 1):
        seq_val = _cell_str(ws, row_idx, seq_col)

        # 检查停止条件
        if not seq_val or seq_val in STOP_KEYWORDS:
            # 也检查整行是否有停止关键字
            row_text = ''
            for c in range(1, ws.max_column + 1):
                row_text += _cell_str(ws, row_idx, c)
            if any(kw in row_text for kw in STOP_KEYWORDS):
                break
            if not seq_val:
                continue

        # 项号标准化
        seq_norm = _normalize_seq(seq_val)
        if not seq_norm:
            continue

        # 商品编码
        code_raw = _cell_str(ws, row_idx, code_col)
        commodity_code = _extract_commodity_code(code_raw)

        # 根据格式版本解析
        if format_ver == 'V2_SKU_SPEC':
            name_col = header_cols.get('商品名称', 3)
            sku_col = header_cols.get('SKU')
            spec_col = header_cols.get('规格型号')
            product_name = _cell_str(ws, row_idx, name_col) if name_col else ''
            sku_raw = _cell_str(ws, row_idx, sku_col) if sku_col else ''
            spec_raw = _cell_str(ws, row_idx, spec_col) if spec_col else ''
            sku_normalized = normalize_sku(sku_raw) if sku_raw else None
            desc_raw = f'{product_name} | {sku_raw} | {spec_raw}'
        else:
            # V1_COMBINED: 商品名称及规格型号列（在商品编码右边）
            combined_col = header_cols.get('商品名称及规格型号', code_col + 1)
            combined_val = _cell_str(ws, row_idx, combined_col)
            # 下一个列通常是 SKU / 型号
            sku_spec_col = combined_col + 1
            sku_spec_val = _cell_str(ws, row_idx, sku_spec_col)

            # 商品名称取第一个换行前
            product_name = combined_val.split('\n')[0].strip() if combined_val else ''
            # SKU从下一个列取
            sku_normalized = normalize_sku(sku_spec_val) if sku_spec_val else None
            spec_raw = sku_spec_val or ''
            desc_raw = f'{combined_val} | {sku_spec_val}'

        # 数量及单位拆分
        qty_raw = _cell_str(ws, row_idx, qty_col) if qty_col else ''
        t_qty, t_unit, s_qty, s_unit = _split_quantity_unit(qty_raw)

        # 单价/总价/币制拆分
        price_raw = _cell_str(ws, row_idx, price_col) if price_col else ''
        unit_price, total_price, currency = _split_price(price_raw)

        # 金额校验
        parse_msg = ''
        if t_qty and unit_price and total_price:
            try:
                expected = Decimal(str(t_qty)) * Decimal(str(unit_price))
                diff = abs(expected - Decimal(str(total_price)))
                if diff > Decimal('0.03'):
                    parse_msg = f'金额校验差异: {float(diff):.2f}'
            except:
                pass

        # V2格式SKU必填校验
        if format_ver == 'V2_SKU_SPEC' and (not sku_normalized or sku_normalized == ''):
            parse_msg = (parse_msg + '; ' if parse_msg else '') + 'V2格式SKU不能为空'

        item = {
            'contract_agreement_no': contract_no,
            'export_invoice_no': export_invoice_no,
            'product_sequence_no': seq_val,
            'product_sequence_normalized': seq_norm,
            'commodity_code': commodity_code,
            'product_name': product_name,
            'sku': sku_normalized or (sku_raw if format_ver == 'V2_SKU_SPEC' else None),
            'specification_model': spec_raw if spec_raw and spec_raw != '无型号' else '',
            'transaction_quantity': t_qty,
            'transaction_unit': t_unit,
            'statutory_quantity': s_qty,
            'statutory_unit': s_unit,
            'unit_price': unit_price,
            'total_price': total_price,
            'currency_code': currency or 'USD',
            'format_version': format_ver,
            'source_file_name': source_file,
            'source_file_hash': file_hash,
            'source_row_no': row_idx,
            'parse_status': 'ERROR' if parse_msg else 'PENDING',
            'parse_message': parse_msg or None,
        }
        items.append(item)

    return items


def _extract_export_invoice_no(workbook):
    """从商业发票或装箱单抬头中提取出口发票号码。"""
    pattern = re.compile(r'\bINV[A-Z0-9][A-Z0-9-]{4,}\b', re.IGNORECASE)
    preferred = sorted(
        workbook.worksheets,
        key=lambda sheet: 0 if ('INVOICE' in sheet.title.upper() or '发票' in sheet.title) else 1,
    )
    for sheet in preferred:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20),
                                   min_col=1, max_col=min(sheet.max_column, 15)):
            for cell in row:
                if cell.value is None:
                    continue
                match = pattern.search(str(cell.value).strip())
                if match:
                    return match.group(0).upper()
    return None


def _cell_str(ws, row, col):
    """获取单元格字符串值"""
    if col is None:
        return ''
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ''
    return str(val).strip()


def _normalize_seq(seq_val):
    """标准化商品项号：去前导零，去空格"""
    if not seq_val:
        return None
    seq = str(seq_val).strip().upper()
    if seq.isdigit():
        return str(int(seq))
    return seq


def _extract_commodity_code(raw):
    """提取开头8-10位商品编码"""
    if not raw:
        return None
    m = re.match(r'^\s*(\d{8,10})', str(raw))
    return m.group(1) if m else None


def _split_quantity_unit(raw):
    """拆分数量及单位: '20个/21千克' → (20, '个', 21, '千克')"""
    if not raw:
        return None, None, None, None

    # 统一分隔符
    text = str(raw).strip().replace('\n', '/')
    # 全角斜杠
    text = text.replace('／', '/')

    parts = [p.strip() for p in text.split('/') if p.strip()]

    def _parse_one(part):
        m = re.match(r'^(\d+\.?\d*)\s*(\D+)', part)
        if m:
            return float(m.group(1)), m.group(2).strip()
        return None, None

    t_qty, t_unit = _parse_one(parts[0]) if parts else (None, None)

    s_qty, s_unit = None, None
    if len(parts) >= 2:
        s_qty, s_unit = _parse_one(parts[1])

    return t_qty, t_unit, s_qty, s_unit


def _split_price(raw):
    """拆分单价/总价/币制: '19.59/391.8/USD' → (19.59, 391.8, 'USD')"""
    if not raw:
        return None, None, None

    text = str(raw).strip().replace('\n', '/').replace('／', '/')
    parts = [p.strip() for p in text.split('/') if p.strip()]

    unit_price = None
    total_price = None
    currency = None

    if len(parts) >= 1:
        try:
            unit_price = float(parts[0])
        except ValueError:
            pass
    if len(parts) >= 2:
        try:
            total_price = float(parts[1])
        except ValueError:
            pass
    if len(parts) >= 3:
        currency = parts[2].upper().strip()

    return unit_price, total_price, currency


if __name__ == '__main__':
    import json
    # Test File 1 (V1_COMBINED)
    path1 = r'D:\JMH\出口业务收汇情况表\外汇退税\测试数据\FBA15L7CCK57\FBA15L7CCK57\FBA15L7CCK57报关资料.xlsx'
    contract, fmt, items, err = parse_customs_excel(path1)
    print(f'File 1: contract={contract}, format={fmt}, items={len(items)}, err={err}')
    for item in items[:3]:
        print(f'  #{item["product_sequence_normalized"]}: {item["product_name"][:20]} | '
              f'{item["transaction_quantity"]}{item["transaction_unit"]} | '
              f'{item["total_price"]} {item["currency_code"]} | SKU={item["sku"]}')

    print()
    # Test File 2 (V2_SKU_SPEC) - find via walk
    import os
    path2 = None
    for root, dirs, files in os.walk(r'D:\ran冉'):
        for f in files:
            if 'RVG10645-260529-0004' in f and f.endswith('.xlsx') and not f.startswith('~$'):
                path2 = os.path.join(root, f)
                break
    if path2:
        contract2, fmt2, items2, err2 = parse_customs_excel(path2)
        print(f'File 2: contract={contract2}, format={fmt2}, items={len(items2)}, err={err2}')
        for item in items2[:3]:
            print(f'  #{item["product_sequence_normalized"]}: {item["product_name"][:20]} | '
                  f'{item["transaction_quantity"]}{item["transaction_unit"]} | '
                  f'{item["total_price"]} {item["currency_code"]} | SKU={item["sku"]}')
