from __future__ import annotations

import re
import shutil
import zipfile
from collections import defaultdict
from copy import copy
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from backend.config import settings
from backend.database import db_connection
from backend.services.inventory_service import ensure_fifo_allocations


PURCHASE_TEMPLATE = "《外贸企业出口退税进货明细申报表》导入模板-纳税号-批次号.xlsx"
EXPORT_TEMPLATE = "《外贸企业出口退税出口明细申报表》导入模板-纳税号-批次号.xlsx"
FOREX_TEMPLATE = "出口业务外汇情况表模版.xlsx"
ENTERPRISE_TEMPLATE = "企业收汇情况表模版.xlsx"


def decimal_value(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    return value if isinstance(value, Decimal) else Decimal(str(value))


def excel_number(value: Any) -> float | None:
    if value is None:
        return None
    return float(value) if isinstance(value, Decimal) else value


def normalized_rate(value: Any) -> Decimal | None:
    if value is None:
        return None
    rate = decimal_value(value)
    return rate / Decimal("100") if abs(rate) > 20 else rate


def multiply(amount: Any, rate: Decimal | None) -> float | None:
    if amount is None or rate is None:
        return None
    return float(decimal_value(amount) * rate)


def customs_key(value: Any) -> str:
    digits = "".join(re.findall(r"\d", str(value or "")))
    return digits[:18]


def source_batch(row: dict) -> str:
    value = str(row.get("declaration_batch") or "").strip()
    if value.isdigit():
        value = value.zfill(3)
    return value


def generated_batch_code(position: int, skip_002: bool = False) -> str:
    """供应商排序批次；仅申报年月202512按业务规则保留并跳过002。"""
    number = position if not skip_002 or position == 1 else position + 1
    return f"{number:03d}"


def generated_relation(row: dict, batch: str) -> str:
    correlation_no = str(row.get("correlation_no") or "").strip()
    return f"{correlation_no[:6]}{batch}{correlation_no[9:]}"


def validate_relation_parts(row: dict, source_name: str) -> None:
    correlation_no = str(row.get("correlation_no") or "").strip()
    declaration_month = str(row.get("declaration_month") or "").strip()
    declaration_batch = source_batch(row)
    if (
        len(correlation_no) != 17
        or correlation_no[:6] != declaration_month
        or correlation_no[6:9] != declaration_batch
    ):
        raise ValueError(
            f"{source_name}关联号与申报年月/批次不一致: "
            f"关联号={correlation_no}, 申报年月={declaration_month}, 申报批次={declaration_batch}"
        )


def copy_row_style(sheet, source_row: int, target_row: int, max_column: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def prepare_rows(sheet, start_row: int, count: int, max_column: int, style_row: int) -> int:
    last_row = start_row + count - 1
    clear_to = max(sheet.max_row, last_row)
    for row in range(start_row, clear_to + 1):
        for column in range(1, max_column + 1):
            sheet.cell(row, column).value = None
    for row in range(start_row, last_row + 1):
        copy_row_style(sheet, style_row, row, max_column)
    return last_row


def write_matrix(sheet, start_row: int, rows: list[list[Any]]) -> None:
    for row_offset, values in enumerate(rows):
        for column, value in enumerate(values, start=1):
            sheet.cell(start_row + row_offset, column).value = excel_number(value)


def set_date_formats(sheet, rows: range, columns: tuple[int, ...]) -> None:
    for row in rows:
        for column in columns:
            if sheet.cell(row, column).value:
                sheet.cell(row, column).number_format = "yyyy-mm-dd"


def set_text_formats(sheet, rows: range, columns: tuple[int, ...]) -> None:
    for row in rows:
        for column in columns:
            if sheet.cell(row, column).value is not None:
                sheet.cell(row, column).number_format = "@"


def rebuild_purchase_validations(sheet, last_row: int) -> None:
    sheet.data_validations.dataValidation = []
    tax_type = DataValidation(type="list", formula1="税种!$A$2:$A$200")
    voucher = DataValidation(
        type="list",
        formula1='INDIRECT(IF(LEFT(D9,1)="V","增值税凭证种类","消费税凭证种类"))',
    )
    sheet.add_data_validation(tax_type)
    sheet.add_data_validation(voucher)
    tax_type.add(f"D9:D{last_row}")
    voucher.add(f"E9:E{last_row}")


def rebuild_export_validations(sheet, last_row: int) -> None:
    sheet.data_validations.dataValidation = []
    for column, formula in (
        ("N", "'退（免）税业务类型'!$A$2:$A$200"),
        ("O", "'国内启运方式'!$A$2:$A$200"),
        ("P", "'海关总署认证企业类型'!$A$2:$A$200"),
    ):
        validation = DataValidation(type="list", formula1=formula)
        sheet.add_data_validation(validation)
        validation.add(f"{column}9:{column}{last_row}")


def rebuild_forex_validations(sheet, last_row: int) -> None:
    sheet.data_validations.dataValidation = []
    for column, formula in (
        ("G", "'币种代码'!$A$2:$A$200"),
        ("M", "'币种代码'!$A$2:$A$200"),
        ("T", "'视同收汇原因代码'!$A$2:$A$200"),
        ("U", "'视同收汇举证材料种类代码'!$A$2:$A$200"),
    ):
        validation = DataValidation(type="list", formula1=formula)
        sheet.add_data_validation(validation)
        validation.add(f"{column}9:{column}{last_row}")


def shift_column_reference_left(cell_reference: str) -> str | None:
    match = re.fullmatch(r"(\$?)([A-Z]+)(\$?\d+)", cell_reference)
    if not match:
        return cell_reference
    absolute_column, column_name, row_reference = match.groups()
    column_index = 0
    for character in column_name:
        column_index = column_index * 26 + ord(character) - ord("A") + 1
    if column_index <= 1:
        return None
    return f"{absolute_column}{get_column_letter(column_index - 1)}{row_reference}"


def shift_validation_ranges_left(sheet) -> None:
    retained = []
    for validation in sheet.data_validations.dataValidation:
        shifted_ranges = []
        for reference in str(validation.sqref).split():
            boundaries = reference.split(":", 1)
            shifted = [shift_column_reference_left(value) for value in boundaries]
            if all(shifted):
                shifted_ranges.append(":".join(shifted))
        if shifted_ranges:
            validation.sqref = " ".join(shifted_ranges)
            retained.append(validation)
    sheet.data_validations.dataValidation = retained


def remove_forex_column_a(sheet) -> None:
    """Match remove_col_a.py: shift row 8 onward left and move validations."""
    max_row = sheet.max_row
    max_column = sheet.max_column
    for row in range(8, max_row + 1):
        for column in range(2, max_column + 1):
            sheet.cell(row, column - 1).value = sheet.cell(row, column).value
            sheet.cell(row, column).value = None
    for row in range(1, max_row + 1):
        sheet.cell(row, max_column).value = None
    shift_validation_ranges_left(sheet)


def save_purchase_file(template: Path, output: Path, batch: str, rows: list[dict]) -> None:
    workbook = load_workbook(template)
    sheet = workbook["模板"]
    matrix = [
        [
            row["declaration_month"], batch, generated_relation(row, batch), row["tax_type"],
            "02|增值税专用发票", row["purchase_voucher_no"], row["invoice_date"],
            row["supplier_tax_id"], row["export_product_code"], row["export_product_name"],
            row["measurement_unit"], row["quantity"], row["taxable_amount"],
            row["levy_rate_percent"], row["refund_rate_percent"],
            row["refundable_tax_amount"], row["remark"],
        ]
        for row in rows
    ]
    last_row = prepare_rows(sheet, 9, len(matrix), 17, 9)
    write_matrix(sheet, 9, matrix)
    set_date_formats(sheet, range(9, last_row + 1), (7,))
    set_text_formats(sheet, range(9, last_row + 1), (1, 2, 3, 6, 8, 9))
    sheet.auto_filter.ref = f"A8:Q{last_row}"
    rebuild_purchase_validations(sheet, last_row)
    workbook.save(output)


def save_export_file(template: Path, output: Path, batch: str, rows: list[dict]) -> None:
    workbook = load_workbook(template)
    sheet = workbook["模板"]
    matrix = [
        [
            row["declaration_month"], batch, generated_relation(row, batch),
            row["customs_declaration_no"], row["agent_export_certificate_no"],
            row["export_invoice_no"], row["export_date"], row["export_product_code"],
            row["export_product_name"], row["measurement_unit"], row["export_quantity"],
            row["fob_value_usd"], row["declared_product_code"],
            row["tax_refund_business_type"], None, None, row["remark"],
        ]
        for row in rows
    ]
    last_row = prepare_rows(sheet, 9, len(matrix), 17, 9)
    write_matrix(sheet, 9, matrix)
    set_date_formats(sheet, range(9, last_row + 1), (7,))
    set_text_formats(sheet, range(9, last_row + 1), (1, 2, 3, 4, 5, 6, 8, 13))
    sheet.auto_filter.ref = f"A8:Q{last_row}"
    rebuild_export_validations(sheet, last_row)
    workbook.save(output)


def matched_receipts(export_row: dict, receipt_index: dict[str, list[dict]]) -> list[dict | None]:
    matches = receipt_index.get(customs_key(export_row["customs_declaration_no"]), [])
    return matches or [None]


def allocated_payment(
    export_row: dict,
    receipt: dict | None,
    fob_totals: dict[str, Decimal],
) -> Decimal | None:
    if not receipt:
        return None
    total_fob = fob_totals.get(customs_key(export_row["customs_declaration_no"]), Decimal("0"))
    if not total_fob:
        return None
    payment = decimal_value(
        receipt.get("payment_amount_usd") or receipt.get("foreign_exchange_received_usd")
    )
    return payment * decimal_value(export_row["fob_value_usd"]) / total_fob


def forex_values(
    supplier_tax_id: str,
    batch: str,
    export_row: dict,
    receipt: dict | None,
    fob_totals: dict[str, Decimal],
) -> list[Any]:
    monthly_rate = normalized_rate(receipt.get("monthly_exchange_rate")) if receipt else None
    actual_rate = normalized_rate(receipt.get("actual_exchange_rate")) if receipt else None
    # 凭证总金额直接取回款汇总表的“回单结汇金额（RMB）”。
    voucher_total = receipt.get("settlement_receipt_amount_rmb") if receipt else None
    # 已匹配回款时，出口货物收汇金额必须与本行出口货物销售额一致。
    # 两个折合人民币金额分别使用月度汇率和实际汇率，差异只来自汇率。
    received = decimal_value(export_row["fob_value_usd"]) if receipt else None
    return [
        supplier_tax_id,
        export_row["declaration_month"],
        batch,
        export_row["customs_declaration_no"],
        export_row["agent_export_certificate_no"],
        export_row["export_invoice_no"],
        "USD|美元",
        export_row["fob_value_usd"],
        monthly_rate,
        multiply(export_row["fob_value_usd"], monthly_rate),
        receipt.get("receipt_date") if receipt else None,
        receipt.get("bank_transaction_no") if receipt else None,
        "USD|美元",
        voucher_total,
        received,
        actual_rate,
        multiply(received, actual_rate),
        settings.payer_name,
        None, None, None, None, None,
        receipt.get("contract_no") if receipt else export_row.get("contract_no"),
        export_row.get("remark"),
    ]


def save_forex_file(
    template: Path,
    output: Path,
    supplier_tax_id: str,
    batch: str,
    export_rows: list[dict],
    receipt_index: dict[str, list[dict]],
    fob_totals: dict[str, Decimal],
) -> None:
    workbook = load_workbook(template)
    sheet = workbook["出口业务收汇情况表模板"]
    matrix = [
        forex_values(supplier_tax_id, batch, export_row, receipt, fob_totals)
        for export_row in export_rows
        for receipt in matched_receipts(export_row, receipt_index)
    ]
    last_row = prepare_rows(sheet, 9, len(matrix), 25, 9)
    write_matrix(sheet, 9, matrix)
    set_date_formats(sheet, range(9, last_row + 1), (11, 23))
    set_text_formats(sheet, range(9, last_row + 1), (1, 2, 3, 4, 5, 6, 7, 12, 13, 18, 20, 21, 24))
    sheet.column_dimensions["A"].width = 24
    rebuild_forex_validations(sheet, last_row)
    remove_forex_column_a(sheet)
    workbook.save(output)


def save_all_forex_file(
    template: Path,
    output: Path,
    supplier_exports: list[dict],
    supplier_codes: dict[str, str],
    receipt_index: dict[str, list[dict]],
    fob_totals: dict[str, Decimal],
) -> None:
    workbook = load_workbook(template)
    sheet = workbook["出口业务收汇情况表模板"]
    matrix = []
    for export_row in supplier_exports:
        supplier_tax_id = export_row["_supplier_tax_id"]
        batch = supplier_codes[supplier_tax_id]
        matrix.extend(
            forex_values(supplier_tax_id, batch, export_row, receipt, fob_totals)
            for receipt in matched_receipts(export_row, receipt_index)
        )
    last_row = prepare_rows(sheet, 9, len(matrix), 25, 9)
    write_matrix(sheet, 9, matrix)
    set_date_formats(sheet, range(9, last_row + 1), (11, 23))
    set_text_formats(sheet, range(9, last_row + 1), (1, 2, 3, 4, 5, 6, 7, 12, 13, 18, 20, 21, 24))
    sheet.column_dimensions["A"].width = 24
    rebuild_forex_validations(sheet, last_row)
    workbook.save(output)


def aggregate_receipts(
    export_row: dict,
    receipt_index: dict[str, list[dict]],
    fob_totals: dict[str, Decimal],
) -> dict[str, Any]:
    receipts = [row for row in matched_receipts(export_row, receipt_index) if row]
    monthly_rate = normalized_rate(receipts[0].get("monthly_exchange_rate")) if receipts else None
    fob_rmb = multiply(export_row["fob_value_usd"], monthly_rate)
    received_usd = sum(
        (allocated_payment(export_row, row, fob_totals) or Decimal("0") for row in receipts),
        Decimal("0"),
    )
    received_rmb = sum(
        (
            (allocated_payment(export_row, row, fob_totals) or Decimal("0"))
            * (normalized_rate(row.get("actual_exchange_rate")) or Decimal("0"))
            for row in receipts
        ),
        Decimal("0"),
    )
    voucher_total_rmb = sum(
        (
            decimal_value(row.get("foreign_exchange_received_usd"))
            * (normalized_rate(row.get("actual_exchange_rate")) or Decimal("0"))
            for row in receipts
        ),
        Decimal("0"),
    )
    contract_total_rmb = sum(
        (
            decimal_value(row.get("declared_contract_amount_usd"))
            * (normalized_rate(row.get("monthly_exchange_rate")) or Decimal("0"))
            for row in {row["id"]: row for row in receipts}.values()
        ),
        Decimal("0"),
    )
    return {
        "receipts": receipts,
        "fob_rmb": fob_rmb,
        "received_usd": received_usd,
        "received_rmb": received_rmb,
        "voucher_total_rmb": voucher_total_rmb,
        "contract_total_rmb": contract_total_rmb,
    }


def save_enterprise_file(
    template: Path,
    output: Path,
    supplier_exports: list[dict],
    supplier_codes: dict[str, str],
    receipt_index: dict[str, list[dict]],
    fob_totals: dict[str, Decimal],
) -> None:
    workbook = load_workbook(template)
    sheet = workbook["Sheet2"]
    matrix = []
    for sequence, export_row in enumerate(supplier_exports, start=1):
        batch = supplier_codes[export_row["_supplier_tax_id"]]
        aggregate = aggregate_receipts(export_row, receipt_index, fob_totals)
        receipts = aggregate["receipts"]
        fob_rmb = aggregate["fob_rmb"]
        dates = "\n".join(str(row["receipt_date"]) for row in receipts if row.get("receipt_date"))
        vouchers = "\n".join(dict.fromkeys(str(row["bank_transaction_no"]) for row in receipts if row.get("bank_transaction_no")))
        contracts = "\n".join(dict.fromkeys(str(row["contract_no"]) for row in receipts if row.get("contract_no")))
        matrix.append(
            [
                sequence, export_row["declaration_month"], batch,
                export_row["customs_declaration_no"], export_row["export_date"],
                fob_rmb, export_row["fob_value_usd"], fob_rmb,
                None, export_row["fob_value_usd"], dates, vouchers,
                aggregate["voucher_total_rmb"], None, None, None, contracts,
                aggregate["contract_total_rmb"], None if receipts else "未匹配回款",
            ]
        )

    last_data_row = prepare_rows(sheet, 4, len(matrix), 19, 4)
    write_matrix(sheet, 4, matrix)
    set_date_formats(sheet, range(4, last_data_row + 1), (5,))
    set_text_formats(sheet, range(4, last_data_row + 1), (2, 3, 4, 9, 11, 12, 15, 16, 17))
    for row in range(4, last_data_row + 1):
        for column in (6, 7, 8, 10, 13, 14, 18):
            sheet.cell(row, column).number_format = "#,##0.00"

    total_row = last_data_row + 1
    copy_row_style(sheet, 38, total_row, 19)
    sheet.cell(total_row, 1).value = "合计"
    for column in (6, 7, 8, 10):
        letter = get_column_letter(column)
        sheet.cell(total_row, column).value = f"=SUM({letter}4:{letter}{last_data_row})"
        sheet.cell(total_row, column).number_format = "#,##0.00"
    sheet.data_validations.dataValidation = []
    validation = DataValidation(type="list", formula1='"是,否"')
    sheet.add_data_validation(validation)
    validation.add(f"I4:I{last_data_row}")
    workbook.save(output)


def save_generation_error_file(
    output: Path, generation_errors: list[dict[str, Any]]
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "生成错误清单"
    sheet.append(
        [
            "序号", "报关单号", "合同协议号", "项号", "SKU",
            "商品名称", "需求数量", "单位", "错误原因",
        ]
    )
    for position, error in enumerate(generation_errors, start=1):
        sheet.append(
            [
                position,
                str(error.get("customs_declaration_no") or ""),
                str(error.get("contract_no") or ""),
                str(error.get("item_no") or ""),
                str(error.get("sku") or ""),
                str(error.get("product_name") or ""),
                str(error.get("quantity") or ""),
                str(error.get("unit") or ""),
                str(error.get("error") or ""),
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{max(sheet.max_row, 1)}"
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 24
    sheet.column_dimensions["F"].width = 24
    sheet.column_dimensions["G"].width = 14
    sheet.column_dimensions["H"].width = 10
    sheet.column_dimensions["I"].width = 100
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in sheet.iter_rows(min_row=2, min_col=9, max_col=9):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(output)


def load_export_data(cursor, generation_id: str) -> tuple[list[dict], list[dict], list[dict], dict]:
    cursor.execute("SELECT * FROM tax_refund_purchase_details ORDER BY correlation_no, allocation_sequence, id")
    purchases_before = cursor.fetchall()
    cursor.execute("SELECT * FROM tax_refund_export_details ORDER BY correlation_no")
    all_exports = cursor.fetchall()
    cursor.execute(
        "SELECT contract_no, item_no, specification FROM customs_declaration_items"
    )
    customs_skus = {
        (str(row["contract_no"] or "").strip(), str(int(row["item_no"])).zfill(3)):
            row["specification"]
        for row in cursor.fetchall()
        if row.get("contract_no") and str(row.get("item_no") or "").isdigit()
    }
    manual_relations = {
        row["correlation_no"]
        for row in purchases_before
        if row.get("source_type") == "MANUAL_EXCEL"
    }
    existing_fifo_relations = {
        row["correlation_no"]
        for row in purchases_before
        if row.get("source_type") == "FIFO_INVOICE"
    }
    fifo_exports = []
    for export in all_exports:
        if export["correlation_no"] in manual_relations:
            continue
        is_customs_generated = str(export.get("uploaded_file_name") or "").startswith(
            "报关资料生成:"
        )
        if export["correlation_no"] not in existing_fifo_relations and not is_customs_generated:
            continue
        full_customs_no = str(export.get("customs_declaration_no") or "")
        export["inventory_sku"] = customs_skus.get(
            (str(export.get("contract_no") or "").strip(), full_customs_no[-3:])
        )
        fifo_exports.append(export)

    allocation_stats = ensure_fifo_allocations(cursor, fifo_exports, generation_id)
    cursor.execute("SELECT * FROM tax_refund_purchase_details ORDER BY correlation_no, allocation_sequence, id")
    purchases = cursor.fetchall()
    purchase_relations = {row["correlation_no"] for row in purchases}
    exports = [row for row in all_exports if row["correlation_no"] in purchase_relations]
    cursor.execute("SELECT * FROM foreign_exchange_receipts ORDER BY id")
    receipts = cursor.fetchall()

    export_relations = {row["correlation_no"] for row in exports}
    missing = purchase_relations - export_relations
    if missing:
        sample = ", ".join(sorted(missing)[:5])
        raise ValueError(f"有{len(missing)}个进货关联号缺少出口数据，例如：{sample}")
    purchases_by_relation: dict[str, list[dict]] = defaultdict(list)
    for row in purchases:
        purchases_by_relation[row["correlation_no"]].append(row)
    for relation, rows in purchases_by_relation.items():
        source_types = {row.get("source_type") for row in rows}
        if "MANUAL_EXCEL" in source_types and "FIFO_INVOICE" in source_types:
            raise ValueError(
                f"关联号{relation}同时存在手工进货明细和FIFO库存明细，请先确认保留哪一种来源"
            )
    for purchase in purchases:
        validate_relation_parts(purchase, "进货明细")
    for export in exports:
        validate_relation_parts(export, "出口明细")
        for purchase in purchases_by_relation[export["correlation_no"]]:
            if source_batch(export) != source_batch(purchase):
                raise ValueError(
                    f"进出口申报批次不一致: 关联号={export['correlation_no']}, "
                    f"出口批次={source_batch(export)}, 进货批次={source_batch(purchase)}"
                )
    return purchases, exports, receipts, allocation_stats


def split_exports_by_supplier(
    purchases: list[dict],
    exports: list[dict],
    supplier_codes: dict[str, str],
) -> tuple[dict[str, list[dict]], list[dict]]:
    """按FIFO实际进货数量拆分出口数量和FOB，使每个供应商批次只申报自身份额。"""
    quantities: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: defaultdict(lambda: Decimal("0"))
    )
    for purchase in purchases:
        relation = purchase["correlation_no"]
        supplier = str(purchase.get("supplier_tax_id") or "").strip()
        if not supplier:
            raise ValueError(f"关联号{relation}的进货明细缺少供应商纳税号")
        quantities[relation][supplier] += decimal_value(purchase["quantity"])

    rows_by_supplier: dict[str, list[dict]] = defaultdict(list)
    all_supplier_rows: list[dict] = []
    for export in exports:
        relation = export["correlation_no"]
        supplier_quantities = quantities.get(relation)
        if not supplier_quantities:
            raise ValueError(f"关联号{relation}没有可用于拆分的进货数量")
        export_quantity = decimal_value(export["export_quantity"])
        if export_quantity <= 0:
            raise ValueError(f"关联号{relation}出口数量必须大于0，无法按供应商拆分")
        allocated_quantity = sum(supplier_quantities.values(), Decimal("0"))
        if allocated_quantity != export_quantity:
            raise ValueError(
                f"关联号{relation}进货数量合计{allocated_quantity}与出口数量"
                f"{export_quantity}不一致，无法按供应商拆分"
            )

        suppliers = sorted(
            supplier_quantities,
            key=lambda supplier: (supplier_codes[supplier], supplier),
        )
        original_fob = decimal_value(export["fob_value_usd"])
        remaining_fob = original_fob
        for position, supplier in enumerate(suppliers):
            quantity = supplier_quantities[supplier]
            if position == len(suppliers) - 1:
                supplier_fob = remaining_fob
            else:
                supplier_fob = (
                    original_fob * quantity / export_quantity
                ).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
                remaining_fob -= supplier_fob
            supplier_export = dict(export)
            supplier_export["export_quantity"] = quantity
            supplier_export["fob_value_usd"] = supplier_fob
            supplier_export["_supplier_tax_id"] = supplier
            rows_by_supplier[supplier].append(supplier_export)
            all_supplier_rows.append(supplier_export)
    return dict(rows_by_supplier), all_supplier_rows


def generate_final_package(
    generation_errors: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    generation_errors = list(generation_errors or [])
    template_dir = Path(settings.template_dir)
    templates = {
        "purchase": template_dir / PURCHASE_TEMPLATE,
        "export": template_dir / EXPORT_TEMPLATE,
        "forex": template_dir / FOREX_TEMPLATE,
        "enterprise": template_dir / ENTERPRISE_TEMPLATE,
    }
    missing_templates = [str(path) for path in templates.values() if not path.exists()]
    if missing_templates:
        raise ValueError(f"缺少模板文件: {', '.join(missing_templates)}")
    generation_id = str(uuid4())
    directory_suffix = generation_id.replace("-", "")[:8]
    output_root = Path(settings.export_output_dir)
    final_directory = output_root / "汇总"
    temp_directory = output_root / f".汇总_{directory_suffix}.tmp"
    backup_directory = output_root / f".汇总_{directory_suffix}.bak"
    generated_files = 0
    swapped = False
    had_previous_output = final_directory.exists()

    with db_connection() as connection:
        try:
            with connection.cursor() as cursor:
                purchases, exports, receipts, allocation_stats = load_export_data(
                    cursor, generation_id
                )

            supplier_totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
            purchases_by_supplier: dict[str, list[dict]] = defaultdict(list)
            for row in purchases:
                supplier = row["supplier_tax_id"]
                supplier_totals[supplier] += decimal_value(row["refundable_tax_amount"])
                purchases_by_supplier[supplier].append(row)
            ranked_suppliers = sorted(
                supplier_totals, key=lambda value: (-supplier_totals[value], value)
            )
            declaration_months = {
                str(row.get("declaration_month") or "").strip() for row in purchases
            }
            skip_002 = declaration_months == {"202512"}
            supplier_codes = {
                supplier: generated_batch_code(position, skip_002=skip_002)
                for position, supplier in enumerate(ranked_suppliers, start=1)
            }
            exports_by_supplier, supplier_exports = split_exports_by_supplier(
                purchases, exports, supplier_codes
            )
            generated_relations: dict[str, str] = {}
            for purchase in purchases:
                batch = supplier_codes[purchase["supplier_tax_id"]]
                output_relation = generated_relation(purchase, batch)
                previous = generated_relations.get(output_relation)
                if previous and previous != purchase["correlation_no"]:
                    raise ValueError(
                        f"生成关联号重复: {output_relation}，"
                        f"来源关联号={previous},{purchase['correlation_no']}"
                    )
                generated_relations[output_relation] = purchase["correlation_no"]
            receipt_index: dict[str, list[dict]] = defaultdict(list)
            for row in receipts:
                receipt_index[customs_key(row["customs_declaration_no"])].append(row)
            fob_totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
            for row in exports:
                fob_totals[customs_key(row["customs_declaration_no"])] += decimal_value(
                    row["fob_value_usd"]
                )

            output_root.mkdir(parents=True, exist_ok=True)
            temp_directory.mkdir(parents=True)
            for supplier in ranked_suppliers:
                batch = supplier_codes[supplier]
                supplier_dir = temp_directory / batch
                supplier_dir.mkdir()
                purchase_rows = purchases_by_supplier[supplier]
                export_rows = exports_by_supplier[supplier]
                save_purchase_file(
                    templates["purchase"],
                    supplier_dir
                    / f"《外贸企业出口退税进货明细申报表》导入模板-{supplier}-{batch}.xlsx",
                    batch,
                    purchase_rows,
                )
                save_export_file(
                    templates["export"],
                    supplier_dir
                    / f"《外贸企业出口退税出口明细申报表》导入模板-{supplier}-{batch}.xlsx",
                    batch,
                    export_rows,
                )
                save_forex_file(
                    templates["forex"],
                    supplier_dir / f"出口业务收汇情况表_{supplier}_{batch}.xlsx",
                    supplier,
                    batch,
                    export_rows,
                    receipt_index,
                    fob_totals,
                )
                generated_files += 3

            enterprise_path = temp_directory / "企业收汇情况表.xlsx"
            save_enterprise_file(
                templates["enterprise"], enterprise_path, supplier_exports,
                supplier_codes, receipt_index, fob_totals,
            )
            generated_files += 1
            save_all_forex_file(
                templates["forex"], temp_directory / "出口业务收汇情况表_全部汇总.xlsx",
                supplier_exports, supplier_codes, receipt_index, fob_totals,
            )
            generated_files += 1
            if generation_errors:
                save_generation_error_file(
                    temp_directory / "生成错误清单.xlsx", generation_errors
                )
                generated_files += 1

            if final_directory.exists():
                final_directory.rename(backup_directory)
            temp_directory.rename(final_directory)
            swapped = True
            connection.commit()
        except Exception:
            connection.rollback()
            shutil.rmtree(temp_directory, ignore_errors=True)
            if swapped and final_directory.exists():
                shutil.rmtree(final_directory, ignore_errors=True)
            if backup_directory.exists() and not final_directory.exists():
                backup_directory.rename(final_directory)
            elif swapped and not had_previous_output:
                shutil.rmtree(final_directory, ignore_errors=True)
            raise

    if backup_directory.exists():
        shutil.rmtree(backup_directory, ignore_errors=True)
    package_path = output_root / "外汇退税生成文件.zip"
    temporary_package = output_root / f".外汇退税生成文件_{directory_suffix}.tmp.zip"
    with zipfile.ZipFile(
        temporary_package, "w", compression=zipfile.ZIP_DEFLATED
    ) as archive:
        for path in sorted(final_directory.rglob("*.xlsx")):
            archive.write(
                path,
                arcname=path.relative_to(final_directory).as_posix(),
            )
    temporary_package.replace(package_path)
    return {
        "generation_id": generation_id,
        "output_dir": str(final_directory),
        "supplier_count": len(ranked_suppliers),
        "batch_folder_count": len(ranked_suppliers),
        "batch_package_count": len(ranked_suppliers),
        "generated_files": generated_files,
        "purchase_rows": len(purchases),
        "export_rows": len(exports),
        "supplier_export_rows": len(supplier_exports),
        "receipt_rows": len(receipts),
        "error_count": len(generation_errors),
        "errors": generation_errors,
        "download_package": {
            "name": package_path.name,
            "url": "/api/export/download-package",
        },
        **allocation_stats,
    }
