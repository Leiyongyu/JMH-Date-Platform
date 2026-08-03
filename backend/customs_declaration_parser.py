from __future__ import annotations

import hashlib
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# 商品编码必须直接位于单元格开头；编码后可以紧跟空格、逗号或用途描述。
# 源文件中的J、B、M等字母前缀视为录入错误，应在上传前修正。
PRODUCT_CODE_PATTERN = re.compile(
    r"^\s*(\d{8,13})\s*(.*)$",
    re.DOTALL,
)
ITEM_NO_PATTERN = re.compile(r"^\d+(?:\.0+)?$")
QUANTITY_UNIT_PATTERN = re.compile(r"^\s*([+-]?[\d,]+(?:\.\d+)?)\s*(.*?)\s*$")


def text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    value = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return value or None


def compact(value: Any) -> str:
    return re.sub(r"\s+", "", text(value) or "").upper()


def source_document_key(contract_no: str, file_name: str) -> str:
    """同一逻辑文件重复上传时整份覆盖；同合同不同文件必须分别保留。"""
    base_name = file_name.replace("\\", "/").rsplit("/", 1)[-1].strip().lower()
    identity = f"{contract_no.strip().upper()}|{base_name}"
    return hashlib.sha256(identity.encode("utf-8")).hexdigest()


def find_sheet(workbook, expected_name: str) -> Worksheet:
    expected = compact(expected_name)
    for sheet in workbook.worksheets:
        if compact(sheet.title) == expected:
            return sheet
    raise ValueError(f"Excel缺少“{expected_name}”工作表")


def value_right_of_label(sheet: Worksheet, label_pattern: re.Pattern[str]) -> str | None:
    for row in sheet.iter_rows():
        for cell in row:
            cell_text = text(cell.value)
            if not cell_text or not label_pattern.search(cell_text):
                continue
            for column in range(cell.column + 1, min(sheet.max_column, cell.column + 5) + 1):
                candidate = text(sheet.cell(cell.row, column).value)
                if candidate:
                    return candidate
    return None


def locate_header(sheet: Worksheet) -> tuple[int, dict[str, int]]:
    required = ("项号", "商品编码", "商品名称及规格型号", "数量及单位", "单价/总价/币制")
    for row_number in range(1, sheet.max_row + 1):
        columns: dict[str, int] = {}
        for column in range(1, sheet.max_column + 1):
            value = compact(sheet.cell(row_number, column).value)
            for header in required:
                if compact(header) == value:
                    columns[header] = column
        if all(header in columns for header in required):
            return row_number, columns
    raise ValueError("报关单页缺少商品表头：项号、商品编码、商品名称及规格型号、数量及单位、单价/总价/币制")


def split_code_and_use(raw_value: Any, specification: str | None) -> tuple[str, str | None]:
    raw_text = text(raw_value)
    if not raw_text:
        raise ValueError("商品编码为空")
    match = PRODUCT_CODE_PATTERN.match(raw_text)
    if not match:
        raise ValueError(f"无法从“{raw_text}”中识别商品编码")
    product_code = match.group(1)
    product_use = match.group(2).strip().lstrip(",，;；").strip()
    if specification:
        # 源文件常把规格同时附在商品编码描述末尾，拆分后避免用途字段重复保存规格。
        specification_variants = {specification.strip()}
        if specification.upper().startswith("J"):
            specification_variants.add(specification.strip()[1:])
        for variant in sorted(specification_variants, key=len, reverse=True):
            product_use = re.sub(
                rf"[\s;；,，]*{re.escape(variant)}\s*$",
                "",
                product_use,
                flags=re.IGNORECASE,
            ).strip()
    return product_code, product_use or None


def decimal_value(value: str | None) -> Decimal | None:
    if not value:
        return None
    try:
        return Decimal(value.replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def split_quantity_and_units(
    raw_value: Any,
) -> tuple[Decimal | None, str | None, Decimal | None, str | None]:
    """拆分“20个/21千克”为主数量、主单位、第二数量、第二单位。"""
    raw_text = text(raw_value)
    if not raw_text:
        return None, None, None, None
    parsed: list[tuple[Decimal | None, str | None]] = []
    for segment in raw_text.split("/")[:2]:
        match = QUANTITY_UNIT_PATTERN.match(segment)
        if not match:
            parsed.append((None, text(segment)))
            continue
        parsed.append((decimal_value(match.group(1)), text(match.group(2))))
    while len(parsed) < 2:
        parsed.append((None, None))
    return parsed[0][0], parsed[0][1], parsed[1][0], parsed[1][1]


def split_price_total_currency(
    raw_value: Any,
) -> tuple[Decimal | None, Decimal | None, str | None]:
    """拆分“43.03/215.15/USD”为单价、总价、币制。"""
    raw_text = text(raw_value)
    if not raw_text:
        return None, None, None
    parts = [part.strip() for part in raw_text.split("/")]
    unit_price = decimal_value(parts[0]) if parts else None
    total_price = decimal_value(parts[1]) if len(parts) > 1 else None
    currency = text(parts[2]) if len(parts) > 2 else None
    return unit_price, total_price, currency.upper() if currency else None


def parse_customs_declaration_workbook(
    content: bytes,
    file_name: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not file_name.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("上传文件必须是 .xlsx 或 .xlsm")

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    invoice_sheet = find_sheet(workbook, "INVOICE")
    declaration_sheet = find_sheet(workbook, "报关单")

    invoice_no = value_right_of_label(
        invoice_sheet, re.compile(r"\bINV\.?\s*NO\.?\s*:?", re.IGNORECASE)
    )
    contract_no = value_right_of_label(declaration_sheet, re.compile(r"合同协议号"))
    if not invoice_no:
        raise ValueError("INVOICE页未找到“INV. NO:”对应的发票号")
    if not contract_no:
        raise ValueError("报关单页未找到“合同协议号”")

    header_row, columns = locate_header(declaration_sheet)
    name_column = columns["商品名称及规格型号"]
    specification_column = name_column + 1
    source_hash = hashlib.sha256(content).hexdigest()
    document_key = source_document_key(contract_no, file_name)
    upload_batch_id = str(uuid4())
    records: list[dict[str, Any]] = []

    for row_number in range(header_row + 1, declaration_sheet.max_row + 1):
        raw_item_no = text(declaration_sheet.cell(row_number, columns["项号"]).value)
        if not raw_item_no or not ITEM_NO_PATTERN.match(raw_item_no):
            continue
        item_no = str(int(float(raw_item_no)))
        specification = text(declaration_sheet.cell(row_number, specification_column).value)
        quantity_and_unit = text(
            declaration_sheet.cell(row_number, columns["数量及单位"]).value
        )
        price_total_currency = text(
            declaration_sheet.cell(row_number, columns["单价/总价/币制"]).value
        )
        quantity_value, quantity_unit, second_quantity_value, second_quantity_unit = (
            split_quantity_and_units(quantity_and_unit)
        )
        unit_price, total_price, currency = split_price_total_currency(price_total_currency)
        try:
            product_code, product_use = split_code_and_use(
                declaration_sheet.cell(row_number, columns["商品编码"]).value,
                specification,
            )
        except ValueError as exc:
            raise ValueError(f"报关单页第{row_number}行：{exc}") from exc

        records.append(
            {
                "contract_no": contract_no,
                "invoice_no": invoice_no,
                "source_document_key": document_key,
                "document_total_usd": None,
                "customs_match_status": "UNMATCHED",
                "customs_declaration_no": None,
                "declaration_date": None,
                "declaration_month": None,
                "declaration_batch": None,
                "sequence_no": None,
                "correlation_no": None,
                "export_date": None,
                "item_no": item_no,
                "product_code": product_code,
                "product_use": product_use,
                "product_name": text(declaration_sheet.cell(row_number, name_column).value),
                "specification": specification,
                "quantity_and_unit": quantity_and_unit,
                "quantity_value": quantity_value,
                "quantity_unit": quantity_unit,
                "second_quantity_value": second_quantity_value,
                "second_quantity_unit": second_quantity_unit,
                "price_total_currency": price_total_currency,
                "unit_price": unit_price,
                "total_price": total_price,
                "currency": currency,
                "upload_batch_id": upload_batch_id,
                "uploaded_file_name": file_name,
                "source_hash": source_hash,
                "source_sheet": declaration_sheet.title,
                "source_row": row_number,
            }
        )

    workbook.close()
    if not records:
        raise ValueError("报关单页未解析到有效商品明细")
    document_total_usd = sum(
        (record["total_price"] or Decimal("0")) for record in records
    )
    for record in records:
        record["document_total_usd"] = document_total_usd
    return records, {
        "kind": "customs_declaration",
        "file_name": file_name,
        "source_hash": source_hash,
        "upload_batch_id": upload_batch_id,
        "contract_no": contract_no,
        "invoice_no": invoice_no,
        "source_document_key": document_key,
        "document_total_usd": document_total_usd,
        "rows": len(records),
    }
