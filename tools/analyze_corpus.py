"""Summarize workbook sheet layouts across the customs Excel corpus."""

from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


def normalize_sheet(name: str) -> str:
    return " ".join(name.strip().upper().split())


def main() -> None:
    root = Path(sys.argv[1])
    files = [
        path
        for path in root.rglob("*")
        if path.is_file()
        and path.suffix.lower() in {".xlsx", ".xlsm"}
        and not path.name.startswith("~$")
    ]
    combinations: Counter[tuple[str, ...]] = Counter()
    sheet_counts: Counter[str] = Counter()
    examples: dict[str, list[str]] = defaultdict(list)
    failures: list[dict[str, str]] = []
    workbook_rows: list[dict[str, object]] = []

    for path in files:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheets = tuple(normalize_sheet(name) for name in workbook.sheetnames)
            combinations[sheets] += 1
            for sheet in sheets:
                sheet_counts[sheet] += 1
                if len(examples[sheet]) < 3:
                    examples[sheet].append(str(path))
            workbook_rows.append(
                {
                    "path": str(path),
                    "folder_contract_no": path.parent.name,
                    "sheets": list(sheets),
                }
            )
            workbook.close()
        except Exception as exc:
            failures.append({"path": str(path), "error": str(exc)})

    result = {
        "file_count": len(files),
        "sheet_counts": dict(sheet_counts.most_common()),
        "sheet_examples": examples,
        "sheet_combinations": [
            {"count": count, "sheets": list(combo)}
            for combo, count in combinations.most_common()
        ],
        "failures": failures,
        "non_customs_candidates": [
            row
            for row in workbook_rows
            if not {"合同", "INVOICE", "PACKING LIST", "报关单"}.issubset(row["sheets"])
        ],
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
