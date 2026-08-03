import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from backend.database import db_connection
from backend.exporter import customs_key


root = Path(sys.argv[1])
batch_dirs = sorted(path for path in root.iterdir() if path.is_dir())
assert all(path.name.isdigit() and len(path.name) == 3 for path in batch_dirs)
expected_folder_names = ["001"] + [f"{number:03d}" for number in range(3, 292)]
assert [path.name for path in batch_dirs] == expected_folder_names

purchase_rows = 0
export_rows = 0
forex_rows = 0
forex_amount_rows = 0
validated_files = 0
batch_package_count = 0
export_batch_by_relation = {}
export_batch_counts = Counter()
suppliers = set()
target_file = None
sample_supplier = "91445300MA4ULHCGX2"
sample_supplier_counts = Counter()
sample_supplier_folder = None


def validation_range(column: str, last_row: int) -> str:
    return f"{column}9" if last_row == 9 else f"{column}9:{column}{last_row}"


for batch_dir in batch_dirs:
    batch = batch_dir.name
    files = [file for file in batch_dir.glob("*.xlsx") if not file.name.startswith("~$")]
    assert len(files) == 3, (batch_dir, len(files))
    relations = defaultdict(dict)
    file_types = defaultdict(set)
    for file in files:
        assert file.stem[-3:] == batch, file
        supplier = (
            file.stem.rsplit("_", 2)[-2]
            if file.name.startswith("出口业务收汇情况表_")
            else file.stem.rsplit("-", 2)[-2]
        )
        suppliers.add(supplier)
        if supplier == sample_supplier:
            sample_supplier_folder = batch
        workbook = load_workbook(file, read_only=False, data_only=False)
        if "进货明细申报表" in file.name:
            sheet = workbook["模板"]
            rows = [
                (
                    str(sheet.cell(row, 1).value),
                    str(sheet.cell(row, 2).value).zfill(3),
                    str(sheet.cell(row, 3).value),
                )
                for row in range(9, sheet.max_row + 1)
                if sheet.cell(row, 3).value
            ]
            assert all(month == relation[:6] and row_batch == batch == relation[6:9] for month, row_batch, relation in rows)
            assert all(not (month == "202512" and row_batch == "002") for month, row_batch, _ in rows)
            relations[supplier]["purchase"] = {relation for _, _, relation in rows}
            file_types[supplier].add("purchase")
            purchase_rows += len(rows)
            if supplier == sample_supplier:
                sample_supplier_counts["purchase"] += len(rows)
            assert sheet.auto_filter.ref.endswith(str(8 + len(rows)))
        elif "出口明细申报表" in file.name:
            sheet = workbook["模板"]
            rows = [
                (
                    str(sheet.cell(row, 1).value),
                    str(sheet.cell(row, 2).value).zfill(3),
                    str(sheet.cell(row, 3).value),
                )
                for row in range(9, sheet.max_row + 1)
                if sheet.cell(row, 3).value
            ]
            assert all(month == relation[:6] and row_batch == batch == relation[6:9] for month, row_batch, relation in rows)
            assert all(not (month == "202512" and row_batch == "002") for month, row_batch, _ in rows)
            relations[supplier]["export"] = {relation for _, _, relation in rows}
            file_types[supplier].add("export")
            export_batch_by_relation.update({relation: row_batch for _, row_batch, relation in rows})
            export_batch_counts.update(row_batch for _, row_batch, _ in rows)
            if any(relation == "20251200100001925" for _, _, relation in rows):
                target_file = file
            export_rows += len(rows)
            if supplier == sample_supplier:
                sample_supplier_counts["export"] += len(rows)
            assert sheet.auto_filter.ref.endswith(str(8 + len(rows)))
        else:
            sheet = workbook["出口业务收汇情况表模板"]
            forex_rows_data = [
                (str(sheet.cell(row, 1).value), str(sheet.cell(row, 2).value).zfill(3))
                for row in range(9, sheet.max_row + 1)
                if sheet.cell(row, 3).value is not None
            ]
            assert sheet.cell(8, 1).value == "*申报年月"
            assert sheet.cell(8, 25).value is None
            assert all(row_batch == batch for _, row_batch in forex_rows_data)
            assert all(not (month == "202512" and row_batch == "002") for month, row_batch in forex_rows_data)
            for row in range(9, 9 + len(forex_rows_data)):
                sales_amount = sheet.cell(row, 7).value
                received_amount = sheet.cell(row, 14).value
                if received_amount is not None:
                    assert sales_amount == received_amount, (file, row, sales_amount, received_amount)
                    forex_amount_rows += 1
            validation_ranges = {
                str(validation.sqref) for validation in sheet.data_validations.dataValidation
            }
            assert validation_ranges == {
                validation_range("F", 8 + len(forex_rows_data)),
                validation_range("L", 8 + len(forex_rows_data)),
                validation_range("S", 8 + len(forex_rows_data)),
                validation_range("T", 8 + len(forex_rows_data)),
            }, (file, validation_ranges)
            file_types[supplier].add("forex")
            forex_rows += len(forex_rows_data)
            if supplier == sample_supplier:
                sample_supplier_counts["forex"] += len(forex_rows_data)
        validated_files += 1
    for supplier in relations:
        assert relations[supplier]["purchase"] == relations[supplier]["export"], (batch_dir, supplier)
        assert file_types[supplier] == {"purchase", "export", "forex"}, (batch_dir, supplier)
    batch_package_count += len(relations)
    assert len(relations) == 1, batch_dir

target_relation = "20251200100001925"
assert export_batch_by_relation[target_relation] == "001"
assert target_file is not None and target_file.parent.name == "001"
assert sample_supplier_folder is not None
assert sample_supplier_counts == {"purchase": 18, "export": 18, "forex": 18}

enterprise = root / "企业收汇情况表.xlsx"
workbook = load_workbook(enterprise, read_only=False, data_only=False)
sheet = workbook["Sheet2"]
enterprise_rows = sum(sheet.cell(row, 1).value not in (None, "合计") for row in range(4, sheet.max_row + 1))
assert enterprise_rows == 2286, enterprise_rows
total_row = 4 + enterprise_rows
assert sheet.cell(total_row, 1).value == "合计"
assert all(sheet.cell(row, 9).value is None for row in range(4, total_row))
assert all(sheet.cell(row, 14).value is None for row in range(4, total_row))
assert all(sheet.cell(row, 10).value == sheet.cell(row, 7).value for row in range(4, total_row))
enterprise_batch_counts = Counter(str(sheet.cell(row, 3).value).zfill(3) for row in range(4, total_row))
assert enterprise_batch_counts == export_batch_counts
assert all(sheet.cell(total_row, column).data_type == "f" for column in (6, 7, 8, 10))
assert all(sheet.cell(total_row, column).value is None for column in (13, 14, 18))
validated_files += 1

all_forex = root / "出口业务收汇情况表_全部汇总.xlsx"
workbook = load_workbook(all_forex, read_only=False, data_only=False)
sheet = workbook["出口业务收汇情况表模板"]
assert sheet.cell(8, 1).value == "供货方纳税号"
assert sheet.cell(9, 1).value is not None
all_forex_rows = sum(sheet.cell(row, 4).value is not None for row in range(9, sheet.max_row + 1))
assert all_forex_rows == export_rows, (all_forex_rows, export_rows)
all_forex_batch_counts = Counter(
    str(sheet.cell(row, 3).value).zfill(3) for row in range(9, 9 + all_forex_rows)
)
assert all_forex_batch_counts == export_batch_counts
with db_connection() as connection:
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT customs_declaration_no, settlement_receipt_amount_rmb "
            "FROM foreign_exchange_receipts"
        )
        receipt_totals_by_customs = defaultdict(set)
        for receipt in cursor.fetchall():
            amount = receipt["settlement_receipt_amount_rmb"]
            receipt_totals_by_customs[customs_key(receipt["customs_declaration_no"])].add(
                None if amount is None else float(amount)
            )
for row in range(9, 9 + all_forex_rows):
    sales_amount = sheet.cell(row, 8).value
    voucher_total = sheet.cell(row, 14).value
    received_amount = sheet.cell(row, 15).value
    if received_amount is not None:
        assert sales_amount == received_amount, (row, sales_amount, received_amount)
        assert voucher_total in receipt_totals_by_customs[
            customs_key(sheet.cell(row, 4).value)
        ], (row, voucher_total, sheet.cell(row, 4).value)
validation_ranges = {str(validation.sqref) for validation in sheet.data_validations.dataValidation}
assert validation_ranges == {
    f"G9:G{8 + all_forex_rows}",
    f"M9:M{8 + all_forex_rows}",
    f"T9:T{8 + all_forex_rows}",
    f"U9:U{8 + all_forex_rows}",
}
validated_files += 1

print(
    {
        "batch_dirs": len(batch_dirs),
        "suppliers": len(suppliers),
        "batch_packages": batch_package_count,
        "validated_files": validated_files,
        "purchase_rows": purchase_rows,
        "export_rows": export_rows,
        "forex_rows": forex_rows,
        "forex_amount_rows": forex_amount_rows,
        "enterprise_rows": enterprise_rows,
        "all_forex_rows": all_forex_rows,
        "directory_xlsx_files": len(
            [file for file in root.rglob("*.xlsx") if not file.name.startswith("~$")]
        ),
    }
)
