"""Analyze Sheet1 data quality and types before database import."""

from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


sys.stdout.reconfigure(encoding="utf-8")
path = Path(sys.argv[1])
frame = pd.read_excel(path, sheet_name="Sheet1", header=0, dtype=object)
workbook = load_workbook(path, read_only=False, data_only=False)
worksheet = workbook["Sheet1"]
formula_cells = [cell.coordinate for row in worksheet.iter_rows() for cell in row if cell.data_type == "f"]

result = {
    "data_rows": len(frame),
    "columns": [str(column) for column in frame.columns],
    "kept_columns": [str(column) for column in frame.columns[:-1]],
    "ignored_last_column": str(frame.columns[-1]),
    "nonempty_by_column": {str(column): int(frame[column].notna().sum()) for column in frame.columns},
    "value_types": {
        str(column): dict(Counter(type(value).__name__ for value in frame[column].dropna()).most_common())
        for column in frame.columns
    },
    "duplicate_contract_rows": int(frame["合同编号"].duplicated(keep=False).sum()),
    "unique_contracts": int(frame["合同编号"].nunique(dropna=True)),
    "formula_count": len(formula_cells),
    "formula_examples": formula_cells[:20],
}
print(json.dumps(result, ensure_ascii=False, indent=2))
