"""Inspect customs workbook structure without modifying the source file."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def clean(value):
    if pd.isna(value):
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value).strip()


def main() -> None:
    path = Path(sys.argv[1])
    workbook = load_workbook(path, read_only=False, data_only=False)
    result = {"path": str(path), "sheets": []}
    for worksheet in workbook.worksheets:
        frame = pd.read_excel(path, sheet_name=worksheet.title, header=None, dtype=object)
        preview = []
        for row_index, row in frame.iloc[:40, :30].iterrows():
            values = [clean(value) for value in row.tolist()]
            if any(value not in (None, "") for value in values):
                preview.append({"excel_row": int(row_index) + 1, "values": values})
        result["sheets"].append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "merged_cells": [str(item) for item in worksheet.merged_cells.ranges],
                "preview": preview,
            }
        )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
