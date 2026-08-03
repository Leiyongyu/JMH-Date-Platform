"""Inspect Sheet1 layout for the foreign-exchange receipts workbook."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


sys.stdout.reconfigure(encoding="utf-8")
path = Path(sys.argv[1])
workbook = load_workbook(path, read_only=False, data_only=False)
worksheet = workbook["Sheet1"]
frame = pd.read_excel(path, sheet_name="Sheet1", header=None, dtype=object)


def value(item):
    if pd.isna(item):
        return None
    if hasattr(item, "isoformat"):
        return item.isoformat()
    return str(item).strip()


def rows(part):
    output = []
    for index, row in part.iterrows():
        values = [value(item) for item in row.tolist()]
        if any(item not in (None, "") for item in values):
            output.append({"excel_row": int(index) + 1, "values": values})
    return output


print(
    json.dumps(
        {
            "sheets": workbook.sheetnames,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "merged_cells": [str(cell) for cell in worksheet.merged_cells.ranges],
            "first_rows": rows(frame.iloc[:30]),
            "last_rows": rows(frame.iloc[-10:]),
        },
        ensure_ascii=False,
        indent=2,
    )
)
