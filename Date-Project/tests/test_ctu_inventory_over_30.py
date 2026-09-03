from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from unittest.mock import Mock

import pytest
from openpyxl import load_workbook

from backend.repositories import clearance_repository
from backend.services import clearance_export_service, clearance_service


class _FixedDateTime(datetime):
    @classmethod
    def now(cls, tz=None):
        value = cls(2026, 9, 3, 10, 0, 0)
        return value if tz is None else value.replace(tzinfo=tz)


class _FakeDomain:
    def __init__(self, response):
        self.response = response
        self.calls = []

    def request_endpoint(self, endpoint_key, body):
        self.calls.append((endpoint_key, body))
        return self.response


class _FakeConnection:
    def __init__(self):
        self.committed = False
        self.rolled_back = False

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def commit(self):
        self.committed = True

    def rollback(self):
        self.rolled_back = True


def test_ctu_age_buckets_use_numeric_start_and_decimal_quantity():
    quantity, unparsed = clearance_service._ctu_over_30_quantity(
        [
            {"name": "0-15天库龄", "qty": 20},
            {"name": "16-30天库龄", "qty": 4},
            {"name": "31-90天库龄", "qty": "3.5"},
            {"name": "91天以上库龄", "qty": 2},
            {"name": "新版档位", "qty": 99},
        ]
    )

    assert quantity == Decimal("5.5")
    assert unparsed == {"新版档位"}


def test_ctu_warehouse_mapping_includes_uk_and_all_ebay_warehouses():
    assert clearance_service.CTU_WAREHOUSES["19561"][0] == "EU"
    ebay_wids = {
        wid
        for wid, (group, _name) in clearance_service.CTU_WAREHOUSES.items()
        if group == "EBAY-1"
    }
    assert ebay_wids == {"18674", "18675", "18676", "19578", "19579"}


def test_sync_ctu_inventory_filters_unmapped_and_builds_five_groups(monkeypatch):
    domain = _FakeDomain(
        {
            "code": 0,
            "total": 2,
            "data": [
                {
                    "wid": 18678,
                    "product_id": 1001,
                    "sku": "SKU-1",
                    "product_total": 10,
                    "stock_price": "10.50",
                    "average_age": 40,
                    "stock_age_list": [
                        {"name": "0-15天库龄", "qty": 5},
                        {"name": "31-90天库龄", "qty": 3},
                        {"name": "91天以上库龄", "qty": 2},
                    ],
                },
                {
                    "wid": 99999,
                    "sku": "UNMAPPED",
                    "stock_age_list": [{"name": "31天以上", "qty": 7}],
                },
            ],
        }
    )
    connection = _FakeConnection()
    persisted = {}

    def replace(conn, month, rows, groups):
        persisted.update(month=month, rows=rows, groups=groups)
        return {"ctu_ods_inserted_rows": len(rows)}

    monkeypatch.setattr(clearance_service, "datetime", _FixedDateTime)
    monkeypatch.setattr(
        clearance_service, "LingXingInventoryDomain", lambda: domain
    )
    monkeypatch.setattr(
        clearance_service.repo,
        "ctu_warehouse_names",
        lambda wids: {"18678": "领星-成都US1中转仓"},
    )
    monkeypatch.setattr(clearance_service.repo, "connection", lambda: connection)
    monkeypatch.setattr(
        clearance_service.repo, "replace_ctu_inventory_month", replace
    )

    result = clearance_service.sync_ctu_inventory("2026-09")

    assert result["extract_rows"] == 2
    assert result["ods_rows"] == 1
    assert result["unmapped_wid_rows"] == 1
    assert result["unmapped_wids"] == ["99999"]
    assert connection.committed is True
    assert connection.rolled_back is False
    assert persisted["rows"][0]["warehouse_name"] == "领星-成都US1中转仓"
    assert persisted["rows"][0]["over_30_qty"] == Decimal("5")
    assert persisted["rows"][0]["over_30_cost"] == Decimal("52.50")
    assert len(persisted["groups"]) == 5
    us1 = next(row for row in persisted["groups"] if row["group_code"] == "US1")
    assert us1["over_30_qty"] == Decimal("5")
    assert domain.calls[0][1]["length"] == 800
    assert set(domain.calls[0][1]["wid"].split(",")) == set(
        clearance_service.CTU_WAREHOUSES
    )


def test_sync_ctu_empty_response_never_replaces_existing_rows(monkeypatch):
    domain = _FakeDomain({"code": 0, "total": 0, "data": []})
    replace = Mock()
    monkeypatch.setattr(clearance_service, "datetime", _FixedDateTime)
    monkeypatch.setattr(
        clearance_service, "LingXingInventoryDomain", lambda: domain
    )
    monkeypatch.setattr(
        clearance_service.repo, "replace_ctu_inventory_month", replace
    )

    with pytest.raises(RuntimeError, match="拒绝清空"):
        clearance_service.sync_ctu_inventory("2026-09")

    replace.assert_not_called()


def test_sync_ctu_history_is_rejected_before_domain_creation(monkeypatch):
    domain_factory = Mock()
    monkeypatch.setattr(clearance_service, "datetime", _FixedDateTime)
    monkeypatch.setattr(
        clearance_service, "LingXingInventoryDomain", domain_factory
    )

    with pytest.raises(ValueError, match="不支持补拉历史月份"):
        clearance_service.sync_ctu_inventory("2026-08")

    domain_factory.assert_not_called()


def test_clearance_summary_uses_direct_ctu_total_without_us3_duplication(
    monkeypatch,
):
    base = {
        "inventory_0_90_qty": Decimal("1"),
        "inventory_0_90_cost": Decimal("2"),
        "inventory_91_180_qty": Decimal("3"),
        "inventory_91_180_cost": Decimal("4"),
        "inventory_181_plus_qty": Decimal("5"),
        "inventory_181_plus_cost": Decimal("6"),
        "total_inventory_qty": Decimal("9"),
        "total_inventory_cost": Decimal("12"),
        "pulled_at": datetime(2026, 9, 3, 9, 0, 0),
    }
    monkeypatch.setattr(
        clearance_repository,
        "list_groups",
        lambda month: {
            "pull_month": "2026-09",
            "items": [{**base, "group_code": "US2-MJ"}, {**base, "group_code": "US1-ZXY"}],
            "total": 2,
        },
    )
    monkeypatch.setattr(
        clearance_repository,
        "ctu_inventory_age_summary",
        lambda month: {
            "ctu_over_30_qty": Decimal("7"),
            "ctu_over_30_cost": Decimal("70"),
        },
    )

    result = clearance_repository.summary("2026-09")

    assert result["ctu_over_30_qty"] == Decimal("7")
    assert result["ctu_over_30_cost"] == Decimal("70")


def test_export_contains_fba_and_ctu_detail_sheets(monkeypatch, tmp_path):
    monkeypatch.setattr(
        clearance_export_service.repo,
        "inventory_age_details",
        lambda month: {
            "pull_month": "2026-09",
            "items": [{"pull_month": "2026-09", "group_code": "EU"}],
        },
    )
    monkeypatch.setattr(
        clearance_export_service.repo,
        "ctu_inventory_age_details",
        lambda month: {
            "pull_month": "2026-09",
            "items": [
                {
                    "pull_month": "2026-09",
                    "group_code": "US1",
                    "wid": "18678",
                    "warehouse_name": "CTUAMZ-US1中转仓",
                    "sku": "SKU-1",
                    "over_30_qty": Decimal("5"),
                    "over_30_cost": Decimal("52.50"),
                }
            ],
        },
    )
    monkeypatch.setattr(
        clearance_export_service,
        "settings",
        type("Settings", (), {"export_output_dir": str(tmp_path)})(),
    )

    file_path, download_name = clearance_export_service.export_inventory_age_details(
        "2026-09"
    )
    workbook = load_workbook(file_path, read_only=True)

    assert workbook.sheetnames == ["库龄明细", "成都仓30天以上明细"]
    assert download_name.startswith("2026-09-库龄明细-")
    ctu_sheet = workbook["成都仓30天以上明细"]
    assert ctu_sheet.cell(2, 5).value == "SKU-1"
    assert ctu_sheet.cell(2, 10).value == 5
    assert ctu_sheet.cell(2, 11).value == 52.5