from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from backend.services import inventory_report_etl_service as etl
from backend.services import inventory_report_export_service as export_service


def test_detail_derived_fields_keep_existing_formulas_and_precision():
    row = {
        "department_code": "AMZ-EU",
        "local_end_in_transit_total_cost": Decimal("10"),
        "local_end_inventory_total_cost": Decimal("20"),
        "overseas_end_in_transit_qty": Decimal("10"),
        "fba_end_in_transit_qty": Decimal("4"),
        "overseas_end_in_transit_total_cost": Decimal("30"),
        "fba_end_in_transit_total_cost": Decimal("40"),
        "overseas_end_inventory_qty": Decimal("50"),
        "fba_end_inventory_qty": Decimal("30"),
        "overseas_end_inventory_total_cost": Decimal("50"),
        "fba_end_inventory_total_cost": Decimal("60"),
        "actual_achievement_amount": Decimal("100"),
        "next_month_opening_inventory_qty": Decimal("100"),
        "monthly_sales_qty": Decimal("20"),
    }

    etl._apply_detail_report_derived_fields(row)
    result = etl._report_json_ready(row)

    assert result["total_goods_value"] == "210.00"
    assert result["fba_transit_inventory_amount"] == "180"
    assert result["turnover_days_by_value"] == "15.098039"
    assert result["opening_inventory_sales_ratio"] == "5.700000"
    assert result["turnover_days_by_sku"] == "130.900000"


def test_total_derived_fields_require_complete_detail_metrics():
    details = [
        {
            "department_code": "AMZ-EU",
            "actual_achievement_amount": Decimal("100"),
            "next_month_opening_inventory_qty": Decimal("50"),
            "monthly_sales_qty": Decimal("10"),
            "overseas_end_inventory_total_cost": Decimal("20"),
        },
        {
            "department_code": "AMZ-US1",
            "actual_achievement_amount": None,
            "next_month_opening_inventory_qty": Decimal("40"),
            "monthly_sales_qty": None,
            "fba_end_inventory_total_cost": Decimal("30"),
        },
    ]
    total = {
        "is_total": 1,
        "overseas_end_inventory_total_cost": Decimal("20"),
        "fba_end_inventory_total_cost": Decimal("30"),
    }

    etl._apply_total_report_derived_fields(total, details)

    assert total["turnover_days_by_value"] is None
    assert total["monthly_sales_qty"] is None
    assert total["opening_inventory_sales_ratio"] is None
    assert total["turnover_days_by_sku"] is None


def test_dimension_summary_returns_backend_total_and_derived_fields(monkeypatch):
    monkeypatch.setattr(
        etl.repo,
        "dimension_summary",
        lambda _dimension, _month: {
            "stat_month": "2026-07",
            "items": [
                {
                    "platform_code": "AMZ",
                    "dimension_type": "STORE",
                    "dimension_value": "店铺A",
                    "department_code": "AMZ-EU",
                    "source_rows": 1,
                    "overseas_end_in_transit_qty": Decimal("10"),
                    "fba_end_in_transit_qty": Decimal("4"),
                    "overseas_end_in_transit_total_cost": Decimal("30"),
                    "fba_end_in_transit_total_cost": Decimal("40"),
                    "overseas_end_inventory_qty": Decimal("50"),
                    "fba_end_inventory_qty": Decimal("30"),
                    "overseas_end_inventory_total_cost": Decimal("50"),
                    "fba_end_inventory_total_cost": Decimal("60"),
                }
            ],
        },
    )
    monkeypatch.setattr(etl.repo, "usd_rate", lambda _month: Decimal("10"))
    monkeypatch.setattr(
        etl.repo,
        "amz_sales_amount_by_department",
        lambda _month: {"AMZ-EU": Decimal("100")},
    )
    monkeypatch.setattr(etl.repo, "ebay_sales_amount", lambda _month: None)
    monkeypatch.setattr(
        etl.repo,
        "amz_sales_amount_by_store",
        lambda _month: [
            {
                "department_code": "AMZ-EU",
                "store_name": "EU-店铺A-DE",
                "sales_amount": Decimal("100"),
            }
        ],
    )
    monkeypatch.setattr(
        etl.repo,
        "amz_sales_volume_by_store",
        lambda _month: [
            {
                "department_code": "AMZ-EU",
                "store_name": "EU-店铺A-DE",
                "sales_volume": Decimal("20"),
            }
        ],
    )
    monkeypatch.setattr(
        etl,
        "_inventory_health_maps",
        lambda _month: (
            {},
            {("AMZ-EU", "店铺A"): Decimal("2")},
            {},
            {"AMZ"},
        ),
    )

    result = etl.get_dimension_summary("STORE", "2026-07")
    row = result["items"][0]
    total = result["total"]

    assert row["total_goods_value"] == "180.00"
    assert row["next_month_opening_inventory_qty"] == "80"
    assert row["monthly_sales_qty"] == "20"
    assert row["opening_inventory_sales_ratio"] == "4.700000"
    assert row["turnover_days_by_sku"] == "116.900000"
    assert total["is_dimension_total"] == 1
    assert total["total_goods_value"] == "180.00"
    assert total["monthly_sales_qty"] == "20"
    assert total["opening_inventory_sales_ratio"] == "4.700000"

def test_group_export_uses_dynamic_month_headers_and_all_rows(monkeypatch):
    rows = [
        {
            "department_name": f"组别{index}",
            "total_goods_value": "100.00",
            "fba_transit_inventory_amount": "80.00",
            "turnover_days_by_value": "12.000000",
            "next_month_opening_inventory_qty": "10",
            "monthly_sales_qty": "2",
            "opening_inventory_sales_ratio": "5.000000",
            "turnover_days_by_sku": "70.000000",
        }
        for index in range(1, 8)
    ]
    monkeypatch.setattr(
        export_service,
        "get_department_summary",
        lambda _month: {
            "report_month": "2026-08",
            "items": rows,
        },
    )

    filename, content = export_service.export_monthly_inventory_report(
        "2026-07", "GROUP"
    )
    workbook = load_workbook(BytesIO(content), read_only=True)
    sheet = workbook["月度库存-组别"]
    values = list(sheet.iter_rows(values_only=True))
    headers = list(values[0])

    assert filename.startswith("2026-08-月度库存-组别-")
    assert len(values) == 8
    assert "8月初库存数量" in headers
    assert "8月销量" in headers
    assert "9月初库销比" in headers
    assert "9月周转天数（货值）" in headers


def test_dimension_export_prepends_backend_total(monkeypatch):
    monkeypatch.setattr(
        export_service,
        "get_dimension_summary",
        lambda _dimension, _month: {
            "report_month": "2026-08",
            "total": {
                "is_dimension_total": 1,
                "dimension_value": "合计",
                "total_goods_value": "300.00",
            },
            "items": [
                {"dimension_value": "店铺A", "platform_code": "AMZ"},
                {"dimension_value": "店铺B", "platform_code": "AMZ"},
            ],
        },
    )

    _filename, content = export_service.export_monthly_inventory_report(
        "2026-07", "STORE"
    )
    workbook = load_workbook(BytesIO(content), read_only=True)
    sheet = workbook["月度库存-店铺"]
    values = list(sheet.iter_rows(values_only=True))

    assert len(values) == 4
    assert values[1][0] == "合计（仅Amazon FBA）"
    assert values[2][0] == "店铺A"


def test_export_rejects_empty_month(monkeypatch):
    monkeypatch.setattr(
        export_service,
        "get_department_summary",
        lambda _month: {"report_month": "2026-08", "items": []},
    )

    with pytest.raises(ValueError, match="没有可导出"):
        export_service.export_monthly_inventory_report("2026-07", "GROUP")