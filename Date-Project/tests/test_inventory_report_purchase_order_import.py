from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook

from backend.parsers.inventory_report_purchase_order_parser import (
    parse_inventory_report_purchase_order_excel,
    purchase_warehouse_assignment,
)


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产品信息"
    sheet.append(
        [
            "采购单号",
            "采购仓库",
            "SKU",
            "店铺",
            "单价",
            "待到货量",
            "产品维度",
            "采购仓库（明细）",
        ]
    )
    sheet.append(
        ["PO-1", "CTUAMZ-US3中转仓", "SKU-1", "", 12.5, 3, "汽配", "CTUAMZ-US3中转仓"]
    )
    sheet.append([None, None, "SKU-2", "", 2, 4, "汽配", "CTUAMZ-US3中转仓"])
    sheet.merge_cells("A2:A3")
    sheet.merge_cells("B2:B3")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_purchase_order_parser_handles_merged_warehouse_and_line_cost():
    result = parse_inventory_report_purchase_order_excel(
        _workbook_bytes(), "采购单.xlsx", "2026-07", "tester"
    )

    assert result["source_rows"] == 2
    assert len(result["rows"]) == 2
    assert result["total_pending_arrival_qty"] == Decimal("7")
    assert result["total_pending_cost"] == Decimal("45.5")
    assert [row["purchase_order_no"] for row in result["rows"]] == ["PO-1", "PO-1"]
    assert [row["sku_pending_total_cost"] for row in result["rows"]] == [
        Decimal("37.5"), Decimal("8")
    ]
    assert all(row["department_code"] == "AMZ-US2-MJ" for row in result["rows"])


def test_purchase_warehouse_group_rules():
    assert purchase_warehouse_assignment("CTUebay-UK中转仓") == (
        "EBAY", "EBAY-1", "EBAY-1"
    )
    assert purchase_warehouse_assignment("CTUAMZ-UK中转仓") == (
        "AMZ", "EU", "AMZ-EU"
    )
    assert purchase_warehouse_assignment("CTUAMZ-EU中转仓") == (
        "AMZ", "EU", "AMZ-EU"
    )
    assert purchase_warehouse_assignment("CTUAMZ-US1中转仓") == (
        "AMZ", "US1", "AMZ-US1"
    )
    assert purchase_warehouse_assignment("CTUAMZ-US2中转仓") == (
        "AMZ", "US2", "AMZ-US2"
    )
    assert purchase_warehouse_assignment("CTUAMZ-US3中转仓") == (
        "AMZ", "US3", "AMZ-US2-MJ"
    )
