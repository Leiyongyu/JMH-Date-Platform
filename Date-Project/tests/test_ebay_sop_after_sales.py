from contextlib import nullcontext
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from backend.services import ebay_sop_after_sales_service as service


def _workbook_bytes(headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Worksheet"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _history_workbook_bytes(after_sales_rows, sales_rows=None):
    workbook = Workbook()
    after_sheet = workbook.active
    after_sheet.title = "售后数据"
    after_sheet.append([
        "订单号", "付款时间", "退款时间", "仓库商品标题", "售后数量", "产品SKU",
        "售后小类划分", "售后大类划分", "数据来源", "平台", "售后备注",
    ])
    for row in after_sales_rows:
        after_sheet.append(row)
    sales_sheet = workbook.create_sheet("销量")
    sales_sheet.append(["付款时间", "数据来源", "产品SKU", "销量"])
    for row in sales_rows or [["2026年7月", "eBay-US", "SKU-H1", 10]]:
        sales_sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_rg_sku_multiplier_is_explicit_and_does_not_use_normal_sku_tail():
    assert service._normalize_rg_sku("ABC+4RG") == ("ABC", Decimal("4"))
    assert service._normalize_rg_sku("ABC*3RG") == ("ABC", Decimal("3"))
    assert service._normalize_rg_sku("ABC-4+RG") == ("ABC-4", Decimal("4"))
    assert service._normalize_rg_sku("ZMY-DE-30477-0740+RG") == (
        "ZMY-DE-30477-0740", Decimal("1")
    )


def test_sales_cleaning_filters_voided_and_pre_cutoff_rows_and_maps_currency():
    rows = [
        {
            "source_key": "1",
            "platform_order_no": "ORDER-1",
            "payment_time": datetime(2026, 7, 1, 10),
            "package_status": "已发货",
            "inventory_sku": "SKU-1+4RG",
            "purchase_quantity": Decimal("2"),
            "receivable_goods": Decimal("20"),
            "currency_code": "USD",
        },
        {
            "source_key": "2",
            "platform_order_no": "ORDER-2",
            "payment_time": datetime(2026, 7, 1, 11),
            "package_status": "已作废",
            "inventory_sku": "SKU-2",
            "purchase_quantity": Decimal("1"),
            "receivable_goods": Decimal("8"),
            "currency_code": "GBP",
        },
        {
            "source_key": "3",
            "platform_order_no": "ORDER-3",
            "payment_time": datetime(2026, 6, 13, 23),
            "package_status": "已发货",
            "inventory_sku": "SKU-3",
            "purchase_quantity": Decimal("1"),
            "receivable_goods": Decimal("9"),
            "currency_code": "EUR",
        },
        {
            "source_key": "4",
            "platform_order_no": "ORDER-4",
            "payment_time": datetime(2026, 7, 1, 12),
            "shipping_status": "已作废",
            "package_status": "已发货",
            "inventory_sku": "SKU-4",
            "purchase_quantity": Decimal("1"),
            "receivable_goods": Decimal("9"),
            "currency_code": "USD",
        },
    ]
    result = service._build_sales_rows(rows, "batch")
    assert len(result) == 1
    assert result[0]["business_sku"] == "SKU-1"
    assert result[0]["data_source"] == "eBay-US"
    assert result[0]["sales_quantity"] == Decimal("8")
    assert result[0]["unit_price_total"] == Decimal("40")
    assert result[0]["order_count"] == 1


def test_history_workbook_keeps_existing_categories_and_source():
    content = _history_workbook_bytes([[
            "ORDER-H1", datetime(2026, 7, 1), datetime(2026, 7, 20), "商品A", 2,
            "SKU-H1", "产品不适配", "不适配", "ebay-us", "eBay", "车型不匹配",
        ]])
    parsed = service._parse_history_workbook(content, "history.xlsx", "batch")
    assert parsed["total_rows"] == 2
    assert parsed["skipped_rows"] == 0
    assert len(parsed["after_sales_raw_rows"]) == 1
    assert parsed["after_sales_dwd_rows"][0]["big_category"] == "不适配"
    assert parsed["after_sales_dwd_rows"][0]["small_category"] == "产品不适配"
    assert parsed["after_sales_dwd_rows"][0]["data_source"] == "eBay-US"
    assert parsed["after_sales_dwd_rows"][0]["classify_method"] == "history"
    assert parsed["sales_dwd_rows"][0]["sales_quantity"] == Decimal("10")


def test_history_raw_keeps_every_source_row_and_dwd_prefers_explicit_category():
    content = _history_workbook_bytes([
        [
            "ORDER-D1", datetime(2026, 6, 1), datetime(2026, 6, 10), "商品", 1,
            "SKU-D1", "其他", "其他", "eBay-DE", "eBay", "待确认",
        ],
        [
            "ORDER-D1", datetime(2026, 6, 1), datetime(2026, 6, 10), "商品", 1,
            "SKU-D1", "产品不适配", "不适配", "eBay-DE", "eBay", "车型不匹配",
        ],
        [
            "", datetime(2026, 6, 1), datetime(2026, 6, 10), "商品", 1,
            "", "其他", "其他", "eBay-DE", "eBay", "缺少业务键的原始行",
        ],
    ])
    parsed = service._parse_history_workbook(content, "history.xlsx", "batch")
    assert len(parsed["after_sales_raw_rows"]) == 3
    assert parsed["skipped_rows"] == 1
    assert len(parsed["after_sales_dwd_rows"]) == 1
    assert parsed["after_sales_dwd_rows"][0]["big_category"] == "不适配"
    assert parsed["after_sales_dwd_rows"][0]["small_category"] == "产品不适配"


def test_history_sales_sheet_aggregates_month_source_and_sku():
    content = _history_workbook_bytes(
        [[
            "ORDER-H2", datetime(2026, 7, 1), datetime(2026, 7, 20), "商品", 1,
            "SKU-H2", "其他", "其他", "eBay-DE", "eBay", "其他",
        ]],
        [
            ["2026年1月", "eBay-US", "SKU-1", 10],
            ["2026年1月", "eBay-US", "SKU-1", 2],
            ["2026年2月", "eBay-DE", "SKU-2", 7],
        ],
    )
    parsed = service._parse_history_workbook(content, "history.xlsx", "batch")
    assert len(parsed["sales_raw_rows"]) == 3
    assert len(parsed["sales_dwd_rows"]) == 2
    january = next(
        row for row in parsed["sales_dwd_rows"]
        if row["month_start"] == date(2026, 1, 1)
    )
    assert january["month_end"] == date(2026, 1, 31)
    assert january["sales_quantity"] == Decimal("12")


def test_future_after_sales_uses_classifier_and_currency_source(monkeypatch):
    def classify(rows, platform="AMZ"):
        assert platform == "EBAY"
        for row in rows:
            row["classification_hash"] = "classified"
        return {
            "classified": {
                "big_category": "其他",
                "small_category": "其他",
                "classify_method": "fallback",
                "confidence": Decimal("0"),
            }
        }

    monkeypatch.setattr(service, "classify_rows", classify)
    rows, skipped = service._build_future_after_sales_rows(
        [{
            "source_key": "future-1",
            "platform_order_no": "ORDER-F1",
            "shipping_status": "已发货,已退款",
            "package_status": "已发货",
            "payment_time": datetime(2026, 8, 1, 10),
            "marked_ship_time": datetime(2026, 8, 2, 10),
            "inventory_sku": "SKU-F1",
            "purchase_quantity": Decimal("1"),
            "currency_code": "GBP",
            "platform_name": "eBay",
            "exception_status": "",
            "sku_status": "",
            "logistics_channel": "",
        }, {
            "source_key": "future-2",
            "platform_order_no": "ORDER-F2",
            "shipping_status": "已作废",
            "package_status": "已作废",
            "payment_time": datetime(2026, 8, 1, 11),
            "marked_ship_time": datetime(2026, 8, 2, 11),
            "inventory_sku": "SKU-F2",
            "purchase_quantity": Decimal("1"),
            "currency_code": "USD",
            "platform_name": "eBay",
            "exception_status": "",
            "sku_status": "",
            "logistics_channel": "",
        }],
        "batch",
    )
    assert skipped == 0
    assert len(rows) == 2
    assert rows[0]["after_type"] == "退货"
    assert rows[0]["data_source"] == "eBay-UK"
    assert rows[0]["big_category"] == "其他"
    void_row = next(row for row in rows if row["order_no"] == "ORDER-F2")
    assert void_row["after_type"] == "退款"
    assert void_row["data_source"] == "eBay-US"


def test_product_summary_recalculates_selected_date_range(monkeypatch):
    captured = {}
    monkeypatch.setattr(service.repo, "coverage", lambda: (date(2026, 1, 1), date(2026, 8, 31)))
    monkeypatch.setattr(service.repo, "has_partial_monthly_history", lambda *_: False)
    monkeypatch.setattr(
        service.repo,
        "after_sales_rows",
        lambda start, end: [{
            "order_no": "ORDER-1", "business_sku": "SKU-1",
            "big_category": "不适配", "small_category": "产品不适配",
            "after_quantity": Decimal("2"), "data_source": "eBay-DE",
        }],
    )
    monkeypatch.setattr(
        service.repo,
        "sales_by_sku_source",
        lambda start, end: {("SKU-1", "eBay-DE"): Decimal("100")},
    )
    monkeypatch.setattr(
        service.repo,
        "replace_summary_period",
        lambda start, end, rows: captured.update(start=start, end=end, rows=rows),
    )
    monkeypatch.setattr(
        service.repo,
        "summary_filtered",
        lambda *args, **kwargs: captured["rows"],
    )
    result = service.list_product_summary(
        date(2026, 7, 1), date(2026, 7, 31), None, None, "SKU-1", 1, 20
    )
    assert captured["start"] == date(2026, 7, 1)
    assert captured["end"] == date(2026, 7, 31)
    assert result["items"][0]["after_quantity"] == Decimal("2")
    assert result["items"][0]["sales_volume"] == Decimal("100")
    assert result["items"][0]["after_sales_rate"] == Decimal("0.02")


def test_ebay_query_accepts_whole_month_ranges_and_latest_partial_month(monkeypatch):
    coverage_start = date(2026, 1, 1)
    coverage_end = date(2026, 8, 10)
    monkeypatch.setattr(service.repo, "has_partial_monthly_history", lambda *_: False)
    service._validate_range(
        date(2026, 7, 1), date(2026, 7, 31), coverage_start, coverage_end
    )
    service._validate_range(
        date(2026, 8, 1), date(2026, 8, 10), coverage_start, coverage_end
    )
    service._validate_range(
        date(2026, 5, 1), date(2026, 7, 31), coverage_start, coverage_end
    )
    try:
        service._validate_range(
            date(2026, 8, 2), date(2026, 8, 9), coverage_start, coverage_end
        )
        assert False, "custom day range must be rejected"
    except ValueError as exc:
        assert "完整自然月区间" in str(exc)


def _mock_import_batch(monkeypatch):
    monkeypatch.setattr(
        service,
        "_start_batch",
        lambda content, file_name, import_type, operator: {
            "batch_id": "batch",
            "import_type": import_type,
            "file_name": file_name,
            "file_sha256": "hash",
            "operator": operator,
        },
    )
    monkeypatch.setattr(service, "_finish_batch", lambda *args, **kwargs: None)
    monkeypatch.setattr(service, "_fail_batch", lambda *args, **kwargs: None)
    monkeypatch.setattr(service.repo, "import_lock", lambda *args: nullcontext())


def test_sales_import_skips_existing_order_and_keeps_all_skus_of_new_order(monkeypatch):
    _mock_import_batch(monkeypatch)
    headers = [
        "平台订单号", "发货状态", "包裹状态", "付款时间", "平台", "币种",
        "应收货款(订单级别)", "库存SKU", "SKU状态", "购买数量",
    ]
    content = _workbook_bytes(headers, [
        ["ORDER-OLD", "已发货", "已发货", datetime(2026, 8, 1), "eBay", "USD", 10, "SKU-OLD", "", 1],
        ["ORDER-NEW", "已发货", "已发货", datetime(2026, 8, 2), "eBay", "USD", 20, "SKU-A", "", 1],
        ["ORDER-NEW", "已发货", "已发货", datetime(2026, 8, 2), "eBay", "USD", 30, "SKU-B", "", 2],
    ])
    captured = {}
    monkeypatch.setattr(
        service.repo,
        "existing_order_numbers",
        lambda data_kind, order_numbers: {"ORDER-OLD"},
    )
    monkeypatch.setattr(
        service.repo,
        "append_sales_orders",
        lambda raw, dwd: (
            captured.update(raw=raw, dwd=dwd) or (len(raw), len(dwd))
        ),
    )

    result = service.import_ebay_sales(content, "sales.xlsx", "tester")

    assert {row["platform_order_no"] for row in captured["raw"]} == {"ORDER-NEW"}
    assert {row["business_sku"] for row in captured["dwd"]} == {"SKU-A", "SKU-B"}
    assert result["new_order_count"] == 1
    assert result["existing_order_count"] == 1
    assert result["skipped_existing_rows"] == 1


def test_after_sales_import_skips_existing_order_and_keeps_new_order_skus(monkeypatch):
    _mock_import_batch(monkeypatch)
    headers = [
        "平台订单号", "发货状态", "包裹状态", "付款时间", "平台", "币种",
        "应收货款(订单级别)", "库存SKU", "SKU状态", "购买数量",
    ]
    content = _workbook_bytes(headers, [
        ["ORDER-OLD", "已发货,已退款", "已发货", datetime(2026, 8, 1), "eBay", "GBP", 10, "SKU-OLD", "", 1],
        ["ORDER-NEW", "已发货,已退款", "已发货", datetime(2026, 8, 2), "eBay", "GBP", 20, "SKU-A", "", 1],
        ["ORDER-NEW", "已发货,已退款", "已发货", datetime(2026, 8, 2), "eBay", "GBP", 30, "SKU-B", "", 1],
    ])
    captured = {}

    def classify(rows, platform="AMZ"):
        result = {}
        for index, row in enumerate(rows):
            key = f"classification-{index}"
            row["classification_hash"] = key
            result[key] = {
                "big_category": "其他",
                "small_category": "其他",
                "classify_method": "fallback",
                "confidence": Decimal("0"),
            }
        return result

    monkeypatch.setattr(service, "classify_rows", classify)
    monkeypatch.setattr(
        service.repo,
        "existing_order_numbers",
        lambda data_kind, order_numbers: {"ORDER-OLD"},
    )
    monkeypatch.setattr(
        service.repo,
        "append_after_sales_orders",
        lambda raw, dwd: (
            captured.update(raw=raw, dwd=dwd) or (len(raw), len(dwd))
        ),
    )

    result = service.import_ebay_after_sales(content, "after-sales.xlsx", "tester")

    assert {row["platform_order_no"] for row in captured["raw"]} == {"ORDER-NEW"}
    assert {row["business_sku"] for row in captured["dwd"]} == {"SKU-A", "SKU-B"}
    assert result["new_order_count"] == 1
    assert result["existing_order_count"] == 1
    assert result["skipped_existing_rows"] == 1


def test_history_import_replaces_history_partition(monkeypatch):
    _mock_import_batch(monkeypatch)
    content = _history_workbook_bytes([[
        "ORDER-H1", datetime(2026, 7, 1), datetime(2026, 7, 20), "商品", 1,
        "SKU-H1", "其他", "其他", "eBay-DE", "eBay", "其他",
    ]])
    captured = {}

    def replace_history(after_raw, sales_raw, after_dwd, sales_dwd):
        captured.update(
            after_raw=after_raw,
            sales_raw=sales_raw,
            after_dwd=after_dwd,
            sales_dwd=sales_dwd,
        )
        return len(after_raw), len(sales_raw), len(after_dwd), len(sales_dwd)

    monkeypatch.setattr(service.repo, "replace_history_data", replace_history)

    result = service.import_ebay_history(content, "history.xlsx", "tester")

    assert len(captured["after_raw"]) == 1
    assert len(captured["sales_raw"]) == 1
    assert len(captured["after_dwd"]) == 1
    assert len(captured["sales_dwd"]) == 1
    assert result["dwd_rows"] == 2
