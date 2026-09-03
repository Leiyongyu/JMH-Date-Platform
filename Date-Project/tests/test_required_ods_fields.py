from datetime import datetime
from decimal import Decimal
from types import SimpleNamespace

from openpyxl import load_workbook

from backend.services.amazon_profit_sync_service import _transform_row
from backend.parsers.performance_owner_rule_parser import _raw
from backend.services.clearance_service import (
    REQUIRED_INVENTORY_FIELDS,
    _inventory_age_fields,
    _inventory_identity_fields,
)
from backend.services import clearance_export_service
from backend.repositories.clearance_repository import resolve_store_name


def test_amazon_profit_transform_keeps_only_required_product_and_amount_fields():
    item = {
        "currency_code": "CNY",
        "gross_profit": "12.34",
        "amount": "100.00",
        "refund_amount": "-5.50",
        "promotion_discount": "-2.50",
        "net_amount": "100.00",
        "price_list": [
            {
                "sid": 12576,
                "seller_sku": "MSKU-1",
                "local_sku": "SKU-1",
                "asin": "B000000001",
            }
        ],
        "seller_store_countries": [{"country": "DE"}],
        "unused_large_field": {"should_not": "be persisted"},
    }

    row, ods_row = _transform_row(
        item, "2026-07", "batch-1", datetime(2026, 8, 6, 20, 0, 0)
    )

    assert row is not None
    assert ods_row["sid"] == "12576"
    assert ods_row["seller_sku"] == "MSKU-1"
    assert ods_row["local_sku"] == "SKU-1"
    assert ods_row["asin"] == "B000000001"
    assert ods_row["gross_profit"] == Decimal("12.34")
    assert ods_row["amount"] == Decimal("100.00")
    assert ods_row["refund_amount"] == Decimal("-5.50")
    assert ods_row["promotion_discount"] == Decimal("-2.50")
    assert ods_row["net_sales_amount"] == Decimal("100.00")
    assert row["promotion_discount"] == Decimal("-2.50")
    assert row["net_sales_amount"] == Decimal("92.00")
    assert "raw_json" not in ods_row
    assert "unused_large_field" not in ods_row
    assert "raw_json" not in row


def test_amazon_profit_business_net_sales_accepts_positive_refund_history():
    row, ods_row = _transform_row(
        {
            "amount": "100.00",
            "refund_amount": "5.50",
            "promotion_discount": "2.50",
            "net_amount": "100.00",
            "price_list": [{"sid": 12576, "seller_sku": "MSKU-2"}],
        },
        "2026-07",
        "batch-2",
        datetime(2026, 8, 6, 20, 0, 0),
    )

    assert row is not None
    assert ods_row["net_sales_amount"] == Decimal("100.00")
    assert row["net_sales_amount"] == Decimal("92.00")


def test_amazon_profit_net_sales_matches_confirmed_lingxing_examples():
    examples = (
        ("JYKS-OTH-230395-0828", "24355.14", "-105.06", "-11321.73", "12928.35"),
        ("RQJ-US-70106-0264", "2195.16", "0", "-8027.82", "-5832.66"),
    )
    for sku, amount, discount, refund, expected in examples:
        row, _ = _transform_row(
            {
                "amount": amount,
                "promotion_discount": discount,
                "refund_amount": refund,
                "net_amount": amount,
                "price_list": [{"sid": 12576, "seller_sku": sku}],
            },
            "2026-08",
            "batch-examples",
            datetime(2026, 9, 2, 12, 0, 0),
        )
        assert row is not None
        assert row["net_sales_amount"] == Decimal(expected)


def test_clearance_inventory_transform_keeps_all_and_only_requested_age_fields():
    item = {
        field: str(index + 1)
        for index, field in enumerate(REQUIRED_INVENTORY_FIELDS)
    }
    item["unused_stock_field"] = "999"

    result = _inventory_age_fields(item)

    assert tuple(result) == REQUIRED_INVENTORY_FIELDS
    assert result["inv_age_0_to_30_days"] == Decimal("1")
    assert result["inv_age_365_plus_price"] == Decimal("20")
    assert "unused_stock_field" not in result


def test_clearance_inventory_keeps_shared_warehouse_identity_fields():
    result = _inventory_identity_fields(
        {
            "name": " AMZ-EU-智贸云欧洲仓 ",
            "seller_group_name": "EU-智贸云-FR,EU-智贸云-DE",
            "share_type": "2",
            "unused_field": "discarded",
        }
    )

    assert result == {
        "warehouse_name": "AMZ-EU-智贸云欧洲仓",
        "seller_group_name": "EU-智贸云-FR,EU-智贸云-DE",
        "share_type": 2,
    }


def test_owner_rule_raw_row_is_one_structured_rule_per_month():
    rule = {
        "platform": "amazon",
        "stat_month": "2026-07",
        "group_code": "US1",
        "rule_type": "STORE",
        "match_key": "赵昕怡",
        "principal_name": "赵昕怡",
        "source_file_name": "负责人.xlsx",
        "source_sheet": "US1",
        "source_row": 2,
        "import_batch_id": "owner-batch-1",
    }

    result = _raw(rule)

    assert result["stat_month"] == "2026-07"
    assert result["match_key"] == "赵昕怡"
    assert "raw_json" not in result


def test_inventory_age_detail_export_contains_structured_columns(monkeypatch, tmp_path):
    monkeypatch.setattr(
        clearance_export_service,
        "settings",
        SimpleNamespace(export_output_dir=str(tmp_path)),
    )
    monkeypatch.setattr(
        clearance_export_service.repo,
        "inventory_age_details",
        lambda month: {
            "pull_month": month,
            "items": [
                {
                    "pull_month": month,
                    "region_name": "欧洲组",
                    "group_code": "EU",
                    "sid": "12576",
                    "store_name": "EU-示例店铺-DE",
                    "shared_store_names": "",
                    "seller_sku": "MSKU-1",
                    "sku": "SKU-1",
                    **{field: Decimal("1.25") for field in REQUIRED_INVENTORY_FIELDS},
                    "pulled_at": datetime(2026, 8, 6, 20, 0, 0),
                    "sync_batch_id": "batch-1",
                }
            ],
        },
    )
    monkeypatch.setattr(
        clearance_export_service.repo,
        "ctu_inventory_age_details",
        lambda month: {"pull_month": month, "items": []},
    )

    file_path, download_name = clearance_export_service.export_inventory_age_details(
        "2026-08"
    )
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook["库龄明细"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    assert download_name.startswith("2026-08-库龄明细-")
    assert headers[:7] == [
        "快照月份", "区域", "组别", "店铺名称", "共享店铺列表", "MSKU", "SKU"
    ]
    assert "365天以上成本" in headers
    assert "raw_json" not in headers
    assert sheet["A2"].value == "2026-08"
    assert sheet["D2"].value == "EU-示例店铺-DE"
    assert sheet["E2"].value is None
    assert sheet["H2"].value == 1.25


def test_inventory_age_detail_export_contains_shared_warehouse(monkeypatch, tmp_path):
    monkeypatch.setattr(
        clearance_export_service,
        "settings",
        SimpleNamespace(export_output_dir=str(tmp_path)),
    )
    monkeypatch.setattr(
        clearance_export_service.repo,
        "inventory_age_details",
        lambda month: {
            "pull_month": month,
            "items": [
                {
                    "pull_month": month,
                    "region_name": "欧洲组",
                    "group_code": "EU",
                    "sid": "0",
                    "store_name": "AMZ-EU-智贸云欧洲仓",
                    "shared_store_names": "EU-智贸云-FR,EU-智贸云-DE",
                    "seller_sku": "49-02CK-R3OX",
                    "sku": "SKU-2",
                    **{field: Decimal("0") for field in REQUIRED_INVENTORY_FIELDS},
                    "pulled_at": datetime(2026, 8, 6, 20, 0, 0),
                    "sync_batch_id": "batch-shared",
                }
            ],
        },
    )
    monkeypatch.setattr(
        clearance_export_service.repo,
        "ctu_inventory_age_details",
        lambda month: {"pull_month": month, "items": []},
    )

    file_path, _ = clearance_export_service.export_inventory_age_details("2026-08")
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook["库龄明细"]

    assert sheet["D2"].value == "AMZ-EU-智贸云欧洲仓"
    assert sheet["E2"].value == "EU-智贸云-FR,EU-智贸云-DE"
    assert sheet["F2"].value == "49-02CK-R3OX"


def test_inventory_age_export_store_name_falls_back_to_zero():
    shops = {"12576": "EU-示例店铺-DE", "12577": "  "}

    assert resolve_store_name(shops, "12576", "普通仓") == "EU-示例店铺-DE"
    assert resolve_store_name(shops, "0", "AMZ-EU-共享仓") == "AMZ-EU-共享仓"
    assert resolve_store_name(shops, "12577", "普通仓") == "0"
    assert resolve_store_name(shops, "99999", "普通仓") == "0"
    assert resolve_store_name(shops, None, None) == "0"
