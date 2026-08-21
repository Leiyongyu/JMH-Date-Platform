from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from backend.repositories import amz_sop_repository as repo
from backend.services import amz_sop_after_sales_service as service
from backend.services.amz_sop_after_sales_service import (
    _aggregate_product_rows,
    _automatic_start_date,
    _build_summary,
    _call_lingxing_with_retry,
    _default_start_date,
    _is_rate_limit_error,
    _late_update_months,
    _merge_refund_return,
    _transform_after_sales,
    _transform_sales_period,
    data_source_for_shop,
    export_filtered_summary,
    request_range_summary,
)
from backend.services.amz_sop_classifier import _fallback


def test_weekly_start_refreshes_current_month_when_previous_is_finalized(monkeypatch):
    monkeypatch.setattr(
        service.state_repo, "previous_month_finalized", lambda *_: True
    )
    assert _automatic_start_date(date(2026, 9, 6), False) == (
        date(2026, 9, 1), False
    )


def test_first_weekly_run_in_new_month_forces_previous_full_month(monkeypatch):
    monkeypatch.setattr(
        service.state_repo, "previous_month_finalized", lambda *_: False
    )
    assert _automatic_start_date(date(2026, 9, 6), False) == (
        date(2026, 8, 1), True
    )


def _after_row(source_key, after_type, small_category, quantity=1):
    return {
        "source_key": source_key,
        "amazon_order_id": "ORDER-1",
        "sid": "12514",
        "business_sku": "SKU-1",
        "data_source": "AMZ-EU",
        "after_type": after_type,
        "small_category": small_category,
        "big_category": "其他" if small_category == "其他" else "不适配",
        "after_quantity": Decimal(str(quantity)),
        "after_reason": "",
        "return_status": "",
        "inventory_attributes": "",
        "buyers_note": "",
        "after_reason_zh": "",
        "return_status_zh": "",
        "inventory_attributes_zh": "",
        "buyers_note_zh": "",
        "after_sales_note": "其他",
        "classify_method": "fallback",
        "confidence": Decimal("0"),
        "merged_from_source_key": None,
        "coexist_flag": 0,
    }


def test_shop_name_maps_to_required_sources():
    assert data_source_for_shop("欧洲-伶斯勋旗舰店") == "AMZ-EU"
    assert data_source_for_shop("美国-恒游千") == "AMZ-US1"
    assert data_source_for_shop("US3-新志楠") == "AMZ-US2"
    assert data_source_for_shop("未配置店铺") == "AMZ-OTHER"


def test_summary_keeps_sales_sku_with_zero_after_sales():
    rows = _build_summary(
        [], {("SKU-ZERO", "AMZ-EU"): Decimal("12")},
        date(2026, 8, 1), date(2026, 8, 31), "batch",
    )
    assert len(rows) == 1
    assert rows[0]["business_sku"] == "SKU-ZERO"
    assert rows[0]["small_category"] == "无售后"
    assert rows[0]["after_sales_rate"] == Decimal("0")


def test_not_compatible_reason_is_deterministically_classified_as_product_mismatch():
    result = _fallback({
        "classification_hash": "test-hash",
        "after_reason": "NOT_COMPATIBLE",
        "return_status": "Unit returned to inventory",
        "inventory_attributes": "SELLABLE",
        "buyers_note": "Passt nicht",
    }, "2026-08-10")
    assert result["big_category"] == "不适配"
    assert result["small_category"] == "产品不适配"
    assert result["after_reason_zh"] == "商品不适配"
    assert result["buyers_note_zh"] == "不适配"
    assert result["classify_method"] == "rule"


def test_explicit_amazon_mismatch_reason_codes_do_not_depend_on_ai():
    cases = {
        "PART_NOT_COMPATIBLE": "产品不适配",
        "POOR_FIT": "产品不适配",
        "APPAREL_TOO_SMALL": "产品不适配",
        "APPAREL_TOO_LARGE": "产品不适配",
        "NOT_AS_DESCRIBED": "listing货描不符",
    }
    for reason, expected_small in cases.items():
        result = _fallback({
            "classification_hash": reason,
            "after_reason": reason,
            "return_status": "",
            "inventory_attributes": "",
            "buyers_note": "",
        }, "2026-08-10")
        assert result["big_category"] == "不适配"
        assert result["small_category"] == expected_small
        assert result["classify_method"] == "rule"
        assert result["confidence"] == 0.98


def test_clear_listing_mismatch_note_takes_priority_over_generic_fit_wording():
    result = _fallback({
        "classification_hash": "listing-note",
        "after_reason": "",
        "return_status": "",
        "inventory_attributes": "",
        "buyers_note": "Product does not fit and description doesn't match",
    }, "2026-08-10")
    assert result["small_category"] == "listing货描不符"


def test_default_range_is_current_natural_year_then_current_month():
    end_date = date(2026, 8, 9)
    assert _default_start_date(end_date, True) == date(2026, 1, 1)
    assert _default_start_date(end_date, False) == date(2026, 8, 1)


def test_late_updates_rebuild_complete_old_months_only_once():
    assert _late_update_months(
        {
            date(2026, 5, 9),
            date(2026, 5, 20),
            date(2026, 7, 31),
            date(2026, 8, 15),
        },
        date(2026, 8, 13),
        date(2026, 8, 19),
    ) == [
        (date(2026, 5, 1), date(2026, 5, 31)),
        (date(2026, 7, 1), date(2026, 7, 31)),
    ]


def test_sales_rows_keep_the_requested_period_as_one_snapshot():
    start_date = date(2025, 8, 8)
    end_date = date(2026, 8, 7)
    ods_rows, dwd_rows, skipped = _transform_sales_period(
        start_date,
        end_date,
        [{
            "price_list": [{
                "sid": "12514",
                "seller_sku": "MSKU-1",
                "local_sku": "SKU-1",
                "asin": "B000000001",
            }],
            "currency_code": "EUR",
            "volume": 20,
            "gross_profit": "10.5",
            "amount": "100",
            "refund_amount": "2",
        }],
        {"12514": "伶斯勋"},
        "batch",
    )
    assert skipped == 0
    assert len(ods_rows) == 1
    assert len(dwd_rows) == 1
    assert ods_rows[0]["period_start"] == start_date
    assert ods_rows[0]["period_end"] == end_date


def test_after_sales_keeps_all_items_of_the_same_order():
    rows = _transform_after_sales(
        [{
            "amazon_order_id": "ORDER-1",
            "sid": "12514",
            "item_list": [
                {
                    "item_identifier": "ITEM-1",
                    "local_sku": "SKU-1",
                    "after_type": "退款",
                    "after_quantity": 1,
                    "after_time": "2026-08-01 10:00:00",
                    "row_index": 1,
                },
                {
                    "item_identifier": "ITEM-2",
                    "local_sku": "SKU-2",
                    "after_type": "退货",
                    "after_quantity": 2,
                    "after_time": "2026-08-01 11:00:00",
                    "row_index": 2,
                },
            ],
        }],
        {"12514": "伶斯勋"},
        "batch",
    )
    assert len(rows) == 2
    assert {row["amazon_order_id"] for row in rows} == {"ORDER-1"}
    assert len({row["source_key"] for row in rows}) == 2


def test_filtered_export_uses_current_filters_and_page_columns(monkeypatch):
    captured = {}

    def filtered(
        start_date, end_date, big_category, small_category, sku,
        selected_ids, selected_skus
    ):
        captured["args"] = (
            start_date, end_date, big_category, small_category, sku,
            selected_ids, selected_skus
        )
        return [{
            "big_category": "不适配",
            "small_category": "产品不适配",
            "business_sku": "SKU-1",
            "after_quantity": Decimal("3"),
            "order_count": 2,
            "order_numbers": "ORDER-1、ORDER-2",
            "source_after_quantity_text": "AMZ-EU:3",
            "source_sales_volume_text": "AMZ-EU:200",
            "sales_volume": Decimal("200"),
            "after_sales_rate": Decimal("0.015"),
        }]

    monkeypatch.setattr(repo, "summary_filtered", filtered)
    monkeypatch.setattr(repo, "summary_period_exists", lambda *_: True)
    monkeypatch.setattr(
        repo, "sales_date_bounds",
        lambda: (date(2026, 1, 1), date(2026, 8, 10)),
    )
    content = export_filtered_summary(
        date(2026, 7, 1), date(2026, 7, 31), "不适配", "产品", "SKU-1",
        [7, 9], ["SKU-1"]
    )
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook["售后数据"]
    assert captured["args"] == (
        date(2026, 7, 1), date(2026, 7, 31), "不适配", "产品", "SKU-1",
        [7, 9], ["SKU-1"]
    )
    assert sheet.max_row == 2
    assert sheet["A2"].value == "不适配"
    assert sheet["C2"].value == "SKU-1"
    assert sheet["J2"].value == 0.015
    assert sheet["J2"].number_format == "0.00%"


def test_product_summary_uses_sales_once_and_expands_category_details():
    rows = [
        {
            "id": 1,
            "big_category": "供应商加强包装",
            "small_category": "到货破损",
            "business_sku": "DAS-10130-0040",
            "after_quantity": Decimal("119"),
            "order_count": 115,
            "order_numbers": "ORDER-1、ORDER-2",
            "source_after_quantity_text": "AMZ-EU:119",
            "source_sales_volume_text": "AMZ-EU:2000",
            "sales_volume": Decimal("2000"),
            "after_sales_rate": Decimal("0.0595"),
        },
        {
            "id": 2,
            "big_category": "其他",
            "small_category": "其他",
            "business_sku": "DAS-10130-0040",
            "after_quantity": Decimal("63"),
            "order_count": 63,
            "order_numbers": "ORDER-2、ORDER-3",
            "source_after_quantity_text": "AMZ-EU:63",
            "source_sales_volume_text": "AMZ-EU:2000",
            "sales_volume": Decimal("2000"),
            "after_sales_rate": Decimal("0.0315"),
        },
    ]
    result = _aggregate_product_rows(rows, 1, 20)
    assert result["total"] == 1
    product = result["items"][0]
    assert product["business_sku"] == "DAS-10130-0040"
    assert product["after_quantity"] == Decimal("182")
    assert product["sales_volume"] == Decimal("2000")
    assert product["order_count"] == 3
    assert product["after_sales_rate"] == Decimal("0.091")
    assert product["source_after_quantity_text"] == "AMZ-EU:182"
    assert len(product["children"]) == 2


def test_uncached_range_is_submitted_once_in_background(monkeypatch):
    start_date = date(2026, 7, 1)
    end_date = date(2026, 7, 31)
    key = (start_date, end_date)
    submitted = []
    service._RANGE_TASKS.pop(key, None)
    monkeypatch.setattr(repo, "summary_period_exists", lambda *_: False)
    monkeypatch.setattr(
        repo, "sales_date_bounds",
        lambda: (date(2026, 1, 1), date(2026, 8, 7)),
    )
    monkeypatch.setattr(
        service, "_enqueue_range_summary",
        lambda *args: submitted.append(args),
    )
    try:
        first = request_range_summary(start_date, end_date)
        second = request_range_summary(start_date, end_date)
        assert first["status"] == "building"
        assert second["status"] == "building"
        assert len(submitted) == 1
    finally:
        service._RANGE_TASKS.pop(key, None)


def test_amz_query_accepts_whole_month_ranges_and_latest_partial_month():
    coverage_start = date(2026, 1, 1)
    coverage_end = date(2026, 8, 7)
    service._validate_month_range(
        date(2026, 7, 1), date(2026, 7, 31), coverage_start, coverage_end
    )
    service._validate_month_range(
        date(2026, 8, 1), date(2026, 8, 7), coverage_start, coverage_end
    )
    service._validate_month_range(
        date(2026, 5, 1), date(2026, 7, 31), coverage_start, coverage_end
    )
    try:
        service._validate_month_range(
            date(2026, 7, 3), date(2026, 7, 20), coverage_start, coverage_end
        )
        assert False, "custom day range must be rejected"
    except ValueError as exc:
        assert "完整自然月区间" in str(exc)


def test_amz_periods_are_calendar_months_not_cached_arbitrary_ranges(monkeypatch):
    monkeypatch.setattr(repo.state_repo, "periods", lambda *_: [])
    monkeypatch.setattr(
        repo, "sales_date_bounds",
        lambda: (date(2026, 6, 14), date(2026, 8, 7)),
    )
    assert repo.periods(3) == [
        {"period_start": date(2026, 8, 1), "period_end": date(2026, 8, 7)},
        {"period_start": date(2026, 7, 1), "period_end": date(2026, 7, 31)},
        {"period_start": date(2026, 6, 14), "period_end": date(2026, 6, 30)},
    ]


def test_refund_other_uses_consistent_return_category_and_removes_return():
    refund = _after_row("refund", "退款", "其他")
    returned = _after_row("return", "退货", "产品不适配")
    returned["big_category"] = "不适配"
    returned["after_reason_zh"] = "型号不匹配"
    result = _merge_refund_return([refund, returned])
    assert len(result) == 1
    assert result[0]["source_key"] == "refund"
    assert result[0]["small_category"] == "产品不适配"
    assert result[0]["merged_from_source_key"] == "return"


def test_remaining_refund_and_return_are_marked_as_coexisting():
    refund = _after_row("refund", "退款", "客户主观退换货")
    returned = _after_row("return", "退货", "产品不适配")
    result = _merge_refund_return([refund, returned])
    assert len(result) == 2
    assert all(row["coexist_flag"] == 1 for row in result)


def test_summary_accumulates_orders_sources_and_sales():
    rows = [
        {
            **_after_row("a", "退款", "产品不适配", 2),
            "amazon_order_id": "ORDER-1",
            "big_category": "不适配",
            "data_source": "AMZ-EU",
        },
        {
            **_after_row("b", "退款", "产品不适配", 1),
            "amazon_order_id": "ORDER-2",
            "big_category": "不适配",
            "data_source": "AMZ-US1",
        },
    ]
    result = _build_summary(
        rows,
        {
            ("SKU-1", "AMZ-EU"): Decimal("119"),
            ("SKU-1", "AMZ-US1"): Decimal("81"),
        },
        date(2026, 6, 1),
        date(2026, 6, 30),
        "batch",
    )
    assert len(result) == 1
    assert result[0]["after_quantity"] == Decimal("3")
    assert result[0]["order_count"] == 2
    assert result[0]["source_sales_volume_text"] == "AMZ-EU:119；AMZ-US1:81"
    assert result[0]["sales_volume"] == Decimal("200")
    assert result[0]["after_sales_rate"] == Decimal("0.015")


def test_summary_uses_lingxing_quantity_and_deduplicates_order_shop_records():
    rows = [
        _after_row("a", "退款", "产品不适配", 2),
        _after_row("b", "退款", "产品不适配", 1),
        {**_after_row("c", "退款", "产品不适配", 1), "sid": "12515"},
    ]
    result = _build_summary(
        rows,
        {("SKU-1", "AMZ-EU"): Decimal("100")},
        date(2026, 1, 1),
        date(2026, 7, 31),
        "batch",
    )
    assert result[0]["after_quantity"] == Decimal("3")
    assert result[0]["order_count"] == 1
    assert result[0]["after_sales_rate"] == Decimal("0.03")
    assert result[0]["calculation_version"] == "QUANTITY-V3"


def test_rate_limit_error_is_recognized():
    assert _is_rate_limit_error(
        RuntimeError("领星接口返回失败：code=3001008，message=new requests too frequently")
    )
    assert not _is_rate_limit_error(RuntimeError("领星接口返回失败：code=500，message=系统错误"))


def test_lingxing_call_retries_then_succeeds_on_rate_limit(monkeypatch):
    calls = []

    def flaky():
        calls.append(1)
        if len(calls) < 3:
            raise RuntimeError("code=3001008 new requests too frequently")
        return "ok"

    monkeypatch.setattr(service.time, "sleep", lambda _: None)
    assert _call_lingxing_with_retry(flaky) == "ok"
    assert len(calls) == 3


def test_lingxing_call_rethrows_non_rate_limit_errors_immediately(monkeypatch):
    calls = []

    def broken():
        calls.append(1)
        raise RuntimeError("领星接口返回失败：code=400，message=参数错误")

    monkeypatch.setattr(service.time, "sleep", lambda _: None)
    try:
        _call_lingxing_with_retry(broken)
        assert False, "expected RuntimeError"
    except RuntimeError as exc:
        assert "参数错误" in str(exc)
    assert len(calls) == 1
