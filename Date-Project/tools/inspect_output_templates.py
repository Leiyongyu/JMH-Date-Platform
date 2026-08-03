from pathlib import Path

from openpyxl import load_workbook


TEMPLATES = [
    Path(r"D:\JMH\出口业务收汇情况表\外汇退税\数据源\模版\企业收汇情况表模版.xlsx"),
    Path(r"D:\JMH\出口业务收汇情况表\外汇退税\数据源\模版\《外贸企业出口退税出口明细申报表》导入模板-纳税号-批次号.xlsx"),
    Path(r"D:\JMH\出口业务收汇情况表\外汇退税\数据源\模版\《外贸企业出口退税进货明细申报表》导入模板-纳税号-批次号.xlsx"),
    Path(r"D:\JMH\出口业务收汇情况表\外汇退税\数据源\模版\出口业务外汇情况表模版.xlsx"),
]


for template in TEMPLATES:
    workbook = load_workbook(template, data_only=False)
    print(f"\n===== {template.name} =====")
    for sheet in workbook.worksheets:
        print(
            f"sheet={sheet.title!r} rows={sheet.max_row} cols={sheet.max_column} "
            f"freeze={sheet.freeze_panes} merged={[str(value) for value in sheet.merged_cells.ranges]}"
        )
        print(f"print_area={sheet.print_area} auto_filter={sheet.auto_filter.ref!r}")
        formulas = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formulas.append(f"{cell.coordinate}:{cell.value}")
        print(f"formulas={formulas[:30]} total={len(formulas)}")
        for row_no in range(1, min(sheet.max_row, 18) + 1):
            values = [sheet.cell(row_no, column).value for column in range(1, sheet.max_column + 1)]
            if any(value is not None for value in values):
                print(f"row {row_no}: {values}")
        widths = {
            key: dimension.width
            for key, dimension in sheet.column_dimensions.items()
            if dimension.width is not None
        }
        print(f"column_widths={widths}")
