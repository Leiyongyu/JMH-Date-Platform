"""Inspect candidate detail workbooks before database schema design."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


sys.stdout.reconfigure(encoding="utf-8")


def clean(value):
    if pd.isna(value):
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value).strip()


def summarize(path: Path):
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheets = []
    for worksheet in workbook.worksheets:
        frame = pd.read_excel(path, sheet_name=worksheet.title, header=None, dtype=object)
        nonempty = []
        for index, row in frame.iloc[:20].iterrows():
            values = [clean(value) for value in row.tolist()]
            while values and values[-1] is None:
                values.pop()
            if values:
                nonempty.append({"row": int(index) + 1, "values": values})
        tail = []
        for index, row in frame.iloc[-5:].iterrows():
            values = [clean(value) for value in row.tolist()]
            while values and values[-1] is None:
                values.pop()
            if values:
                tail.append({"row": int(index) + 1, "values": values})
        formulas = [cell.coordinate for row in worksheet.iter_rows() for cell in row if cell.data_type == "f"]
        sheets.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "merged_cells": [str(item) for item in worksheet.merged_cells.ranges],
                "formula_count": len(formulas),
                "formula_examples": formulas[:10],
                "first_rows": nonempty,
                "last_rows": tail,
            }
        )
    return {"path": str(path), "sheets": sheets}


print(json.dumps([summarize(Path(item)) for item in sys.argv[1:]], ensure_ascii=False, indent=2))
