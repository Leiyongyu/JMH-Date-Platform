from __future__ import annotations

import hashlib
import json
import time
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from queue import Queue
from threading import Lock, Thread
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from backend.integrations.lingxing.domains.after_sales import LingXingAfterSalesDomain
from backend.integrations.lingxing.domains.order_profit import LingXingOrderProfitDomain
from backend.repositories import amz_sop_repository as repo
from backend.repositories import after_sales_state_repository as state_repo
from backend.services.amz_sop_classifier import BIG_CATEGORIES, classify_rows


TASK_CODE = "amz_sop_after_sales_chain"
TASK_NAME = "AMZ-SOP售后链路"
SUMMARY_METRIC_VERSION = "QUANTITY-V3"
# 领星 OrderProfit / afterSaleList 令牌桶容量=1，严格限流（3001008）。
_LINGXING_RATE_LIMIT_RETRIES = 4
_LINGXING_RATE_LIMIT_BASE_WAIT_SEC = 10
_LINGXING_REQUEST_INTERVAL_SEC = 2.0
_RANGE_TASK_LOCK = Lock()
_RANGE_TASKS: dict[tuple[date, date], dict[str, str]] = {}
_RANGE_TASK_QUEUE: Queue[tuple[date, date]] = Queue()
_RANGE_WORKER_STARTED = False

EU_SHOPS = (
    "伶斯勋", "重庆宣依", "王耀伟", "斯露星", "帝蓝泰江", "信博嘉立", "智贸云",
    "寻梦云", "保加利亚本土号001", "钟燕本土",
)
US1_SHOPS = (
    "恒游千", "刘子洋", "邱存帅", "智链通", "水金余", "勇成励", "川服喜",
    "鼎晟华", "茂林怡然", "穆厚", "智鑫弘", "邓诗琪", "睿启君", "吉西瑞雅",
    "胡思妍", "重庆茁凯", "优贝诺",
)
US2_SHOPS = (
    "奥诺兰", "旌越问", "峻跃昆昇", "韵盛余", "拓维茂宏", "斯鑫雅", "新志楠", "富琳顿",
)


class AmzSopEtlError(RuntimeError):
    def __init__(self, stage: str, message: str, metrics: dict[str, Any] | None = None):
        super().__init__(message)
        self.stage = stage
        self.metrics = metrics or {}


def _is_rate_limit_error(exc: RuntimeError) -> bool:
    message = str(exc)
    return "too frequently" in message or "3001008" in message


def _call_lingxing_with_retry(fn, *args, **kwargs):
    """调用领星接口；限流(3001008)时退避重试，其它错误直接上抛。"""
    last: RuntimeError | None = None
    for attempt in range(_LINGXING_RATE_LIMIT_RETRIES):
        try:
            return fn(*args, **kwargs)
        except RuntimeError as exc:
            if not _is_rate_limit_error(exc):
                raise
            last = exc
            time.sleep(_LINGXING_RATE_LIMIT_BASE_WAIT_SEC * (attempt + 1))
    raise RuntimeError(f"领星接口限流重试失败：{last}")


def run_amz_sop_chain(
    start_date: date | None = None,
    end_date: date | None = None,
    request_id: str = "",
) -> dict[str, Any]:
    stage = "VALIDATE"
    batch_id = f"{SUMMARY_METRIC_VERSION}-CHAIN-{uuid4()}"
    metrics: dict[str, Any] = {
        "sync_batch_id": batch_id,
        "extract_rows": 0,
        "ods_rows": 0,
        "dwd_rows": 0,
        "skipped_rows": 0,
    }
    try:
        manual_range = start_date is not None or end_date is not None
        if manual_range and (start_date is None or end_date is None):
            raise ValueError("手工补跑必须同时提供 start_date 和 end_date")
        summary_end = end_date or date.today()
        source_status = repo.source_table_status()
        # 售后源表为空也可能代表该月真实没有售后，不能据此反复全量初始化。
        initial_load = not source_status["sales_ready"] and not state_repo.has_any_month("AMZ")
        forced_previous_month = False
        if start_date:
            summary_start = start_date
        else:
            summary_start, forced_previous_month = _automatic_start_date(
                summary_end, initial_load
            )
        if summary_start > summary_end:
            raise ValueError("start_date 不能晚于 end_date")
        if (summary_end - summary_start).days > 365:
            raise ValueError("单次链路日期范围不能超过366天")

        shops = repo.shop_map()
        sids = sorted(shops)
        if not sids:
            raise ValueError("shop_list 中没有可用的 Amazon 店铺 SID")

        stage = "ORDER_PROFIT_EXTRACT"
        profit_domain = LingXingOrderProfitDomain()
        sales_remote_rows = 0
        sales_ods_rows = 0
        sales_dwd_rows = 0
        sales_skipped = 0
        for month_start in repo.natural_month_starts(summary_start, summary_end):
            month_end = _month_end(month_start, summary_end)
            remote = _call_lingxing_with_retry(
                profit_domain.fetch_monthly_profit,
                sids, month_start, month_end, currency_code="原币种",
            )
            sales_remote_rows += len(remote)
            ods_rows, dwd_sales_rows, skipped = _transform_sales_period(
                month_start, month_end, remote, shops, batch_id
            )
            repo.replace_sales_period(month_start, month_end, ods_rows, dwd_sales_rows)
            sales_ods_rows += len(ods_rows)
            sales_dwd_rows += len(dwd_sales_rows)
            sales_skipped += skipped
            time.sleep(_LINGXING_REQUEST_INTERVAL_SEC)

        stage = "AFTER_SALES_EXTRACT"
        after_domain = LingXingAfterSalesDomain()
        after_pull_start = summary_start
        new_orders = _fetch_after_sales_chunked(
            after_domain, sids, after_pull_start, summary_end + timedelta(days=1), 1
        )
        updated_orders: list[dict[str, Any]] = []
        if not manual_range and not initial_load:
            updated_orders = _fetch_after_sales_chunked(
                after_domain,
                sids,
                summary_start,
                summary_end + timedelta(days=1),
                3,
            )
        normalized_after = _transform_after_sales(
            [*new_orders, *updated_orders], shops, batch_id
        )
        deduped_after = {row["source_key"]: row for row in normalized_after}
        updated_order_ids = [
            _text(order.get("amazon_order_id")) for order in updated_orders
        ]
        previous_updated_dates = repo.after_sales_dates_for_orders(updated_order_ids)
        returned_order_ids = [
            _text(order.get("amazon_order_id"))
            for order in [*new_orders, *updated_orders]
        ]
        after_ods_rows = repo.replace_after_sales_orders(
            list(deduped_after.values()), returned_order_ids
        )

        stage = "CLASSIFY_AND_MERGE"
        updated_ids = {value for value in updated_order_ids if value}
        current_updated_dates = {
            row["after_time"].date()
            for row in deduped_after.values()
            if row.get("amazon_order_id") in updated_ids
            and isinstance(row.get("after_time"), datetime)
        }
        late_update_months = _late_update_months(
            previous_updated_dates | current_updated_dates,
            summary_start,
            summary_end,
        )
        late_update_dwd_rows = 0
        for month_start, month_end in late_update_months:
            month_dwd_rows = _build_period_dwd(
                month_start, month_end, batch_id
            )
            repo.replace_dwd_period(month_start, month_end, month_dwd_rows)
            sales_count, after_count = state_repo.month_counts(
                "AMZ", month_start, month_end
            )
            state_repo.refresh_month(
                "AMZ", month_start, month_end, sales_count, after_count,
                month_end < date.today().replace(day=1), batch_id,
            )
            late_update_dwd_rows += len(month_dwd_rows)

        dwd_rows = _build_period_dwd(summary_start, summary_end, batch_id)

        stage = "DWS_SUMMARY"
        sales_volumes = repo.sales_volume_by_sku_source(
            summary_start, summary_end
        )
        summary_rows = _build_summary(
            dwd_rows,
            sales_volumes,
            summary_start,
            summary_end,
            batch_id,
        )
        repo.replace_dwd_and_summary(
            summary_start, summary_end, dwd_rows, summary_rows
        )
        today_month = date.today().replace(day=1)
        refreshed_months = []
        for month_start in repo.natural_month_starts(summary_start, summary_end):
            natural_end = _natural_month_end(month_start)
            pulled_end = min(natural_end, summary_end)
            sales_count, after_count = state_repo.month_counts(
                "AMZ", month_start, pulled_end
            )
            finalized = pulled_end == natural_end and month_start < today_month
            state_repo.refresh_month(
                "AMZ", month_start, pulled_end, sales_count, after_count,
                finalized, batch_id,
            )
            refreshed_months.append(month_start.strftime("%Y-%m"))
        state_repo.mark_range_ready(
            "AMZ", summary_start, summary_end,
            SUMMARY_METRIC_VERSION, len(summary_rows),
        )

        metrics.update({
            "extract_rows": sales_remote_rows + len(new_orders) + len(updated_orders),
            "ods_rows": sales_ods_rows + after_ods_rows,
            "dwd_rows": sales_dwd_rows + len(dwd_rows),
            "skipped_rows": sales_skipped + len(normalized_after) - len(deduped_after),
        })
        return {
            "task_name": TASK_NAME,
            "sync_batch_id": batch_id,
            "period_start": summary_start.isoformat(),
            "period_end": summary_end.isoformat(),
            "sales_pull_start": summary_start.isoformat(),
            "after_sales_pull_start": after_pull_start.isoformat(),
            "load_mode": (
                "manual_range" if manual_range
                else "initial_current_year" if initial_load
                else "new_month_previous_full_and_current"
                if forced_previous_month
                else "incremental_current_month"
            ),
            "forced_previous_month": forced_previous_month,
            "refreshed_months": refreshed_months,
            "source_table_status_before_run": source_status,
            "sid_count": len(sids),
            "sales_extract_rows": sales_remote_rows,
            "sales_ods_rows": sales_ods_rows,
            "sales_dwd_rows": sales_dwd_rows,
            "after_sales_order_rows": len(new_orders) + len(updated_orders),
            "after_sales_ods_rows": after_ods_rows,
            "after_sales_dwd_rows": len(dwd_rows),
            "late_update_rebuild_months": [
                month_start.strftime("%Y-%m")
                for month_start, _ in late_update_months
            ],
            "late_update_dwd_rows": late_update_dwd_rows,
            "summary_rows": len(summary_rows),
            "extract_rows": metrics["extract_rows"],
            "ods_rows": metrics["ods_rows"],
            "dwd_rows": metrics["dwd_rows"],
            "skipped_rows": metrics["skipped_rows"],
            "etl_stage": "COMPLETED",
            "request_id": request_id,
        }
    except Exception as exc:
        if isinstance(exc, AmzSopEtlError):
            raise
        raise AmzSopEtlError(stage, str(exc), metrics) from exc


def ensure_range_summary(start_date: date, end_date: date) -> bool:
    """Build and cache one whole-month reporting range when queried first."""
    coverage_start, coverage_end = repo.sales_date_bounds()
    _validate_month_range(start_date, end_date, coverage_start, coverage_end)
    if repo.summary_period_exists(start_date, end_date, SUMMARY_METRIC_VERSION):
        return False
    with state_repo.range_lock("AMZ", start_date, end_date):
        if repo.summary_period_exists(start_date, end_date, SUMMARY_METRIC_VERSION):
            return False
        coverage_start, coverage_end = repo.sales_date_bounds()
        _validate_month_range(start_date, end_date, coverage_start, coverage_end)

        batch_id = f"{SUMMARY_METRIC_VERSION}-RANGE-{uuid4()}"
        _backfill_missing_sales_months(start_date, end_date, batch_id)
        sales_volumes = repo.sales_volume_by_sku_source(start_date, end_date)
        after_rows = repo.processed_after_sales_rows(start_date, end_date)
        summary_rows = _build_summary(
            after_rows, sales_volumes, start_date, end_date, batch_id
        )
        repo.replace_summary_period(start_date, end_date, summary_rows)
        state_repo.mark_range_ready(
            "AMZ", start_date, end_date,
            SUMMARY_METRIC_VERSION, len(summary_rows),
        )
        return True


def _backfill_missing_sales_months(
    start_date: date, end_date: date, batch_id: str
) -> None:
    """Fetch and store sales snapshots for any natural month lacking one."""
    missing = repo.missing_sales_months(start_date, end_date)
    if not missing:
        return
    shops = repo.shop_map()
    sids = sorted(shops)
    if not sids:
        raise ValueError("shop_list 中没有可用的 Amazon 店铺 SID")
    profit_domain = LingXingOrderProfitDomain()
    for month_start in missing:
        month_end = _month_end(month_start, end_date)
        remote = _call_lingxing_with_retry(
            profit_domain.fetch_monthly_profit,
            sids, month_start, month_end, currency_code="原币种",
        )
        ods_rows, dwd_sales_rows, _ = _transform_sales_period(
            month_start, month_end, remote, shops, batch_id
        )
        repo.replace_sales_period(month_start, month_end, ods_rows, dwd_sales_rows)
        time.sleep(_LINGXING_REQUEST_INTERVAL_SEC)


def request_range_summary(start_date: date, end_date: date) -> dict[str, str]:
    """Return immediately and generate an uncached month range in one worker."""
    _validate_range_coverage(start_date, end_date)
    if repo.summary_period_exists(start_date, end_date, SUMMARY_METRIC_VERSION):
        return {"status": "ready", "message": "月份区间汇总已就绪"}
    key = (start_date, end_date)
    with _RANGE_TASK_LOCK:
        state = _RANGE_TASKS.get(key)
        if state and state["status"] == "ready":
            _RANGE_TASKS.pop(key, None)
            state = None
        if state and state["status"] == "error":
            _RANGE_TASKS.pop(key, None)
            raise ValueError(state["message"])
        if state:
            return dict(state)
        state = {
            "status": "building",
            "message": "首次查询该月份区间，正在后台生成销量快照和售后率",
        }
        _RANGE_TASKS[key] = state
        _enqueue_range_summary(start_date, end_date)
        return dict(state)


def _enqueue_range_summary(start_date: date, end_date: date) -> None:
    global _RANGE_WORKER_STARTED
    if not _RANGE_WORKER_STARTED:
        Thread(
            target=_range_summary_worker,
            name="amz-sop-range-worker",
            daemon=True,
        ).start()
        _RANGE_WORKER_STARTED = True
    _RANGE_TASK_QUEUE.put((start_date, end_date))


def _range_summary_worker() -> None:
    while True:
        start_date, end_date = _RANGE_TASK_QUEUE.get()
        try:
            _run_range_summary_task(start_date, end_date)
        finally:
            _RANGE_TASK_QUEUE.task_done()


def _run_range_summary_task(start_date: date, end_date: date) -> None:
    key = (start_date, end_date)
    try:
        ensure_range_summary(start_date, end_date)
        state = {"status": "ready", "message": "区间汇总生成完成"}
    except Exception as exc:
        state = {"status": "error", "message": f"区间汇总生成失败：{exc}"}
    with _RANGE_TASK_LOCK:
        _RANGE_TASKS[key] = state


def _validate_range_coverage(start_date: date, end_date: date) -> None:
    coverage_start, coverage_end = repo.sales_date_bounds()
    _validate_month_range(start_date, end_date, coverage_start, coverage_end)


def _validate_month_range(
    start_date: date,
    end_date: date,
    coverage_start: date | None,
    coverage_end: date | None,
) -> None:
    if start_date > end_date:
        raise ValueError("开始月份不能晚于结束月份")
    month_count = (
        (end_date.year - start_date.year) * 12
        + end_date.month - start_date.month + 1
    )
    if month_count > 12:
        raise ValueError("单次最多查询连续12个自然月")
    if not coverage_start or not coverage_end:
        raise ValueError("AMZ-SOP售后链路尚未生成可查询数据")
    if start_date < coverage_start or end_date > coverage_end:
        raise ValueError(
            f"当前售后数据覆盖范围为{coverage_start}至{coverage_end}，"
            "请先等待定时链路更新后再查询"
        )
    month_start = date(start_date.year, start_date.month, 1)
    next_month = date(
        end_date.year + (end_date.month == 12),
        end_date.month % 12 + 1,
        1,
    )
    month_end = date.fromordinal(next_month.toordinal() - 1)
    expected_start = max(coverage_start, month_start)
    expected_end = min(coverage_end, month_end)
    if start_date != expected_start or end_date != expected_end:
        raise ValueError(
            "请按完整自然月区间查询；当前选择应转换为"
            f"{expected_start}至{expected_end}"
        )


def list_product_summary(
    start_date: date | None,
    end_date: date | None,
    big_category: str | None,
    small_category: str | None,
    sku: str | None,
    page: int,
    page_size: int,
) -> dict[str, Any]:
    if not start_date or not end_date:
        latest_months = repo.periods(1)
        if not latest_months:
            return {
                "period_start": None,
                "period_end": None,
                "items": [],
                "total": 0,
                "summary": {},
                "range_generated": False,
            }
        start_date = latest_months[0]["period_start"]
        end_date = latest_months[0]["period_end"]
    range_state = request_range_summary(start_date, end_date)
    if range_state["status"] != "ready":
        return {
            "period_start": start_date,
            "period_end": end_date,
            "items": [],
            "total": 0,
            "summary": {},
            "range_generated": False,
            "range_status": range_state["status"],
            "range_message": range_state["message"],
        }
    range_generated = False
    rows = repo.summary_filtered(
        start_date, end_date, big_category, small_category, sku
    )
    result = _aggregate_product_rows(rows, page, page_size)
    all_rows = rows if not any((big_category, small_category, sku)) else repo.summary_filtered(
        start_date, end_date, None, None, None
    )
    range_summary = _aggregate_product_rows(all_rows, 1, 1)["summary"]
    range_sales_volume = repo.sales_total_for_period(start_date, end_date)
    result["summary"].update({
        "range_after_quantity": range_summary["after_quantity"],
        "range_sales_volume": range_sales_volume,
        "range_after_sales_rate": (
            range_summary["after_quantity"] / range_sales_volume
            if range_sales_volume else Decimal("0")
        ),
    })
    result.update({
        "period_start": start_date,
        "period_end": end_date,
        "range_generated": range_generated,
        "range_status": "ready",
        "range_message": "月份区间汇总已就绪",
    })
    return result


def _aggregate_product_rows(
    rows: list[dict[str, Any]], page: int, page_size: int
) -> dict[str, Any]:
    products: dict[str, dict[str, Any]] = {}
    all_orders: set[str] = set()
    all_categories: set[str] = set()
    for source_row in rows:
        row = dict(source_row)
        business_sku = str(row.get("business_sku") or "").strip()
        if not business_sku:
            continue
        product = products.setdefault(business_sku, {
            "id": business_sku,
            "business_sku": business_sku,
            "after_quantity": Decimal("0"),
            "sales_volume": Decimal("0"),
            "orders": set(),
            "source_after": defaultdict(Decimal),
            "source_sales_volume_text": "",
            "children": [],
        })
        quantity = _decimal(row.get("after_quantity"))
        product["after_quantity"] += quantity
        product["sales_volume"] = max(
            product["sales_volume"], _decimal(row.get("sales_volume"))
        )
        if row.get("source_sales_volume_text"):
            product["source_sales_volume_text"] = row["source_sales_volume_text"]
        for order_id in str(row.get("order_numbers") or "").split("、"):
            order_id = order_id.strip()
            if order_id:
                product["orders"].add(order_id)
                all_orders.add(order_id)
        for source, value in _parse_source_text(
            row.get("source_after_quantity_text")
        ).items():
            product["source_after"][source] += value
        all_categories.add(str(row.get("big_category") or ""))
        row["detail_key"] = str(row.get("id") or "")
        product["children"].append(row)

    product_rows: list[dict[str, Any]] = []
    for product in products.values():
        product["children"].sort(
            key=lambda item: (
                -_decimal(item.get("after_quantity")),
                str(item.get("big_category") or ""),
                str(item.get("small_category") or ""),
            )
        )
        sales_volume = product["sales_volume"]
        after_quantity = product["after_quantity"]
        product["order_count"] = len(product.pop("orders"))
        product["order_numbers"] = "、".join(sorted({
            order_id
            for child in product["children"]
            for order_id in str(child.get("order_numbers") or "").split("、")
            if order_id
        }))
        product["source_after_quantity_text"] = _source_text(
            product.pop("source_after")
        )
        product["after_sales_rate"] = (
            after_quantity / sales_volume if sales_volume else Decimal("0")
        )
        product["detail_count"] = len(product["children"])
        product_rows.append(product)
    product_rows.sort(
        key=lambda item: (-item["after_quantity"], item["business_sku"])
    )

    safe_page = max(page, 1)
    safe_size = max(1, min(page_size, 200))
    offset = (safe_page - 1) * safe_size
    total_after = sum(
        (item["after_quantity"] for item in product_rows), Decimal("0")
    )
    total_sales = sum(
        (item["sales_volume"] for item in product_rows), Decimal("0")
    )
    return {
        "items": product_rows[offset:offset + safe_size],
        "total": len(product_rows),
        "summary": {
            "sku_count": len(product_rows),
            "after_quantity": total_after,
            "order_count": len(all_orders),
            "category_count": len({value for value in all_categories if value}),
            "sales_volume": total_sales,
            "after_sales_rate": (
                total_after / total_sales if total_sales else Decimal("0")
            ),
        },
    }


def _parse_source_text(value: Any) -> dict[str, Decimal]:
    result: dict[str, Decimal] = defaultdict(Decimal)
    for part in str(value or "").replace(";", "；").split("；"):
        if ":" not in part:
            continue
        source, raw_value = part.rsplit(":", 1)
        source = source.strip()
        if source:
            result[source] += _decimal(raw_value)
    return result


def export_summary(start_date: date, end_date: date) -> bytes:
    ensure_range_summary(start_date, end_date)
    rows = repo.summary_all(start_date, end_date)
    workbook = Workbook()
    workbook.remove(workbook.active)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[row["big_category"]].append(row)
    for big_category in BIG_CATEGORIES:
        _append_summary_sheet(workbook, big_category, grouped.get(big_category, []))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_filtered_summary(
    start_date: date,
    end_date: date,
    big_category: str | None = None,
    small_category: str | None = None,
    sku: str | None = None,
    selected_ids: list[int] | None = None,
    selected_skus: list[str] | None = None,
) -> bytes:
    ensure_range_summary(start_date, end_date)
    rows = repo.summary_filtered(
        start_date, end_date, big_category, small_category, sku,
        selected_ids, selected_skus,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "售后数据"
    _write_summary_sheet(sheet, rows)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _append_summary_sheet(
    workbook: Workbook, title: str, rows: list[dict[str, Any]]
) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    _write_summary_sheet(sheet, rows)


def _write_summary_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    headers = (
        "售后原因（大）", "售后原因（小）", "SKU", "售后数量", "订单数量",
        "订单号", "数据来源及售后数量", "数据来源及销量", "销量", "售后率",
    )
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append((
            row["big_category"], row["small_category"], row["business_sku"],
            float(row["after_quantity"] or 0), int(row["order_count"] or 0),
            row.get("order_numbers") or "", row.get("source_after_quantity_text") or "",
            row.get("source_sales_volume_text") or "", float(row["sales_volume"] or 0),
            float(row["after_sales_rate"] or 0),
        ))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = (18, 22, 24, 12, 12, 48, 28, 28, 12, 12)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for cell in sheet["J"][1:]:
        cell.number_format = "0.00%"


def _default_start_date(end_date: date, initial_load: bool) -> date:
    """Initial load covers the whole natural year; later runs only the current month."""
    if initial_load:
        return date(end_date.year, 1, 1)
    return date(end_date.year, end_date.month, 1)


def _automatic_start_date(end_date: date, initial_load: bool) -> tuple[date, bool]:
    if initial_load:
        return _default_start_date(end_date, True), False
    current_month_start = date(end_date.year, end_date.month, 1)
    previous_month_end = current_month_start - timedelta(days=1)
    previous_month_start = date(
        previous_month_end.year, previous_month_end.month, 1
    )
    force_previous = not state_repo.previous_month_finalized(
        "AMZ", previous_month_start
    )
    return (
        previous_month_start if force_previous else current_month_start,
        force_previous,
    )


def _natural_month_end(month_start: date) -> date:
    if month_start.month == 12:
        next_month = date(month_start.year + 1, 1, 1)
    else:
        next_month = date(month_start.year, month_start.month + 1, 1)
    return next_month - timedelta(days=1)


def _month_end(month_start: date, cap: date) -> date:
    """Return the last day of month_start's month, capped at cap."""
    if month_start.month == 12:
        next_month = date(month_start.year + 1, 1, 1)
    else:
        next_month = date(month_start.year, month_start.month + 1, 1)
    month_end = date.fromordinal(next_month.toordinal() - 1)
    return min(month_end, cap)


def _fetch_after_sales_chunked(
    domain: LingXingAfterSalesDomain,
    sids: list[str],
    start_date: date,
    end_date_exclusive: date,
    date_type: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index in range(0, len(sids), 50):
        rows.extend(_call_lingxing_with_retry(
            domain.fetch,
            sids[index:index + 50], start_date, end_date_exclusive, date_type=date_type,
        ))
        time.sleep(_LINGXING_REQUEST_INTERVAL_SEC)
    return rows


def _build_period_dwd(
    start_date: date,
    end_date: date,
    batch_id: str,
) -> list[dict[str, Any]]:
    period_ods = repo.after_sales_rows(start_date, end_date)
    period_ods = [
        row for row in period_ods
        if not _contains_star(_business_sku(row))
    ]
    classifications = classify_rows(period_ods)
    return _merge_refund_return(
        _build_dwd_rows(period_ods, classifications, batch_id)
    )


def _late_update_months(
    affected_dates: set[date],
    current_start: date,
    current_end: date,
) -> list[tuple[date, date]]:
    """Return complete months containing updates outside the current rebuild window."""
    month_starts = {
        date(value.year, value.month, 1)
        for value in affected_dates
        if value < current_start or value > current_end
    }
    result: list[tuple[date, date]] = []
    for month_start in sorted(month_starts):
        next_month = date(
            month_start.year + (month_start.month == 12),
            month_start.month % 12 + 1,
            1,
        )
        result.append((
            month_start,
            date.fromordinal(next_month.toordinal() - 1),
        ))
    return result


def _transform_sales_period(
    period_start: date,
    period_end: date,
    remote_rows: list[dict[str, Any]],
    shops: dict[str, str],
    batch_id: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    ods_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    dwd_by_key: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    skipped = 0
    pulled_at = datetime.now()
    for item in remote_rows:
        price = _first_dict(item.get("price_list") or item.get("priceList") or item.get("price"))
        local = _first_dict(item.get("local_infos") or item.get("localInfos") or item.get("local_info"))
        sid = _text(price.get("sid") or item.get("sid") or _first_value(item.get("sids")))
        seller_sku = _text(
            price.get("seller_sku") or price.get("sellerSku") or price.get("msku")
            or item.get("seller_sku") or item.get("sellerSku") or item.get("msku")
        )
        local_sku = _text(
            price.get("local_sku") or price.get("localSku")
            or local.get("local_sku") or local.get("localSku") or item.get("local_sku")
        )
        if not sid or not seller_sku:
            skipped += 1
            continue
        store_name = shops.get(sid, "")
        currency = _text(item.get("currency_code") or item.get("currencyCode")) or "UNKNOWN"
        ods = {
            "period_start": period_start,
            "period_end": period_end,
            "sid": sid,
            "store_name": store_name,
            "data_source": data_source_for_shop(store_name),
            "seller_sku": seller_sku,
            "local_sku": local_sku or None,
            "asin": _text(price.get("asin") or item.get("asin")) or None,
            "currency_code": currency,
            "volume": _decimal(item.get("volume")),
            "gross_profit": _decimal(item.get("gross_profit") or item.get("grossProfit")),
            "amount": _decimal(item.get("amount")),
            "refund_amount": _decimal(item.get("refund_amount") or item.get("refundAmount")),
            "sync_batch_id": batch_id,
            "pulled_at": pulled_at,
        }
        ods["source_hash"] = _sha({
            key: ods[key]
            for key in (
                "period_start", "period_end", "sid", "seller_sku", "local_sku", "asin",
                "currency_code", "volume", "gross_profit", "amount", "refund_amount",
            )
        })
        ods_by_key[(sid, seller_sku, currency)] = ods
        if not local_sku or _contains_star_or_reb(local_sku) or _contains_star_or_reb(seller_sku):
            skipped += 1
            continue
        if ods["volume"] == 0:
            skipped += 1
            continue
        key = (sid, local_sku, seller_sku, currency)
        dwd = dwd_by_key.get(key)
        if dwd is None:
            dwd = {
                **{key: ods[key] for key in (
                    "period_start", "period_end", "sid", "store_name", "data_source", "seller_sku", "asin",
                    "currency_code", "volume", "gross_profit", "amount", "refund_amount",
                    "sync_batch_id",
                )},
                "business_sku": local_sku,
            }
            dwd_by_key[key] = dwd
        else:
            for field in ("volume", "gross_profit", "amount", "refund_amount"):
                dwd[field] += ods[field]
    return list(ods_by_key.values()), list(dwd_by_key.values()), skipped


def _transform_after_sales(
    orders: list[dict[str, Any]], shops: dict[str, str], batch_id: str
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    pulled_at = datetime.now()
    for order in orders:
        order_id = _text(order.get("amazon_order_id"))
        sid = _text(order.get("sid"))
        store_name = shops.get(sid, "")
        items = order.get("item_list") if isinstance(order.get("item_list"), list) else []
        for row_index, item in enumerate(items, 1):
            if not isinstance(item, dict):
                continue
            local_sku = _text(item.get("local_sku") or order.get("local_sku"))
            msku = _text(item.get("msku") or order.get("msku"))
            after_time = _datetime(item.get("after_time") or order.get("deal_time"))
            update_time = _datetime(item.get("data_update_time") or order.get("gmt_modified"))
            native_item_key = _text(item.get("item_identifier") or item.get("md5_v2"))
            source_key = hashlib.sha256(
                "|".join((
                    order_id, native_item_key, sid, local_sku, msku,
                    _text(item.get("after_type")), str(after_time or ""),
                    _text(item.get("row_index") or row_index),
                )).encode("utf-8")
            ).hexdigest()
            selected = {
                "source_key": source_key,
                "amazon_order_id": order_id,
                "sid": sid,
                "store_name": store_name,
                "data_source": data_source_for_shop(store_name),
                "local_sku": local_sku or None,
                "msku": msku or None,
                "asin": _text(item.get("asin") or order.get("asin")) or None,
                "after_type": _after_type(item.get("after_type")),
                "after_quantity": _decimal(item.get("after_quantity")),
                "after_reason": _text(item.get("after_reason")) or None,
                "return_status": _text(item.get("return_status")) or None,
                "inventory_attributes": _text(item.get("inventory_attributes")) or None,
                "buyers_note": _text(item.get("buyers_note")) or None,
                "after_time": after_time,
                "data_update_time": update_time,
                "item_identifier": _text(item.get("item_identifier")) or None,
                "md5_v2": _text(item.get("md5_v2")) or None,
                "sync_batch_id": batch_id,
                "pulled_at": pulled_at,
            }
            selected["source_hash"] = _sha({
                key: selected[key]
                for key in (
                    "source_key", "amazon_order_id", "sid", "local_sku", "msku",
                    "asin", "after_type", "after_quantity", "after_reason",
                    "return_status", "inventory_attributes", "buyers_note",
                    "after_time", "data_update_time", "item_identifier", "md5_v2",
                )
            })
            result.append(selected)
    return result


def _build_dwd_rows(
    rows: list[dict[str, Any]],
    classifications: dict[str, dict[str, Any]],
    batch_id: str,
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows:
        classification = classifications[row["classification_hash"]]
        values = {
            "source_key": row["source_key"],
            "amazon_order_id": row["amazon_order_id"],
            "sid": row["sid"],
            "store_name": row.get("store_name"),
            "data_source": row.get("data_source") or "AMZ-OTHER",
            "business_sku": _business_sku(row),
            "msku": row.get("msku"),
            "asin": row.get("asin"),
            "after_type": _after_type(row.get("after_type")),
            "coexist_flag": 0,
            "after_quantity": _decimal(row.get("after_quantity")),
            "after_reason": row.get("after_reason"),
            "return_status": row.get("return_status"),
            "inventory_attributes": row.get("inventory_attributes"),
            "buyers_note": row.get("buyers_note"),
            "after_reason_zh": classification.get("after_reason_zh"),
            "return_status_zh": classification.get("return_status_zh"),
            "inventory_attributes_zh": classification.get("inventory_attributes_zh"),
            "buyers_note_zh": classification.get("buyers_note_zh"),
            "big_category": classification.get("big_category") or "其他",
            "small_category": classification.get("small_category") or "其他",
            "classify_method": classification.get("classify_method") or "fallback",
            "confidence": _decimal(classification.get("confidence")),
            "merged_from_source_key": None,
            "after_time": row.get("after_time"),
            "data_update_time": row.get("data_update_time"),
            "sync_batch_id": batch_id,
        }
        values["after_sales_note"] = _after_sales_note(values)
        result.append(values)
    return result


def _merge_refund_return(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(row["amazon_order_id"], row["business_sku"])].append(row)
    deleted: set[str] = set()
    fields = (
        "after_reason", "return_status", "inventory_attributes", "buyers_note",
        "after_reason_zh", "return_status_zh", "inventory_attributes_zh", "buyers_note_zh",
        "big_category", "small_category", "classify_method", "confidence",
    )
    for group in grouped.values():
        refunds = [row for row in group if row["after_type"] == "退款"]
        returns = [row for row in group if row["after_type"] == "退货"]
        if not refunds or not returns:
            continue
        return_categories = {row["small_category"] for row in returns}
        fallback_refunds = [row for row in refunds if row["small_category"] == "其他"]
        if fallback_refunds and len(return_categories) == 1:
            source = returns[0]
            for refund in fallback_refunds:
                for field in fields:
                    refund[field] = source.get(field)
                refund["merged_from_source_key"] = source["source_key"]
                refund["after_sales_note"] = _after_sales_note(refund)
            deleted.update(row["source_key"] for row in returns)
        else:
            for row in [*refunds, *returns]:
                row["coexist_flag"] = 1
    return [row for row in rows if row["source_key"] not in deleted]


def _build_summary(
    rows: list[dict[str, Any]],
    sales_volumes: dict[tuple[str, str], Decimal],
    start_date: date,
    end_date: date,
    batch_id: str,
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    sales_sources_by_sku: dict[str, dict[str, Decimal]] = defaultdict(dict)
    for (sku, source), volume in sales_volumes.items():
        sales_sources_by_sku[sku][source] = volume
    for row in rows:
        key = (row["big_category"], row["small_category"], row["business_sku"])
        target = grouped.setdefault(key, {
            "orders": set(),
            "records": {},
            "source_records": defaultdict(dict),
        })
        order_id = _text(row.get("amazon_order_id"))
        source = _text(row.get("data_source")) or "AMZ-OTHER"
        sid = _text(row.get("sid")) or source
        record_key = (order_id, sid)
        quantity = max(_decimal(row.get("after_quantity")), Decimal("0"))
        target["orders"].add(order_id)
        target["records"][record_key] = max(
            target["records"].get(record_key, Decimal("0")), quantity
        )
        target["source_records"][source][record_key] = max(
            target["source_records"][source].get(record_key, Decimal("0")),
            quantity,
        )
    after_sales_skus = {key[2] for key in grouped}
    for sku in sales_sources_by_sku.keys() - after_sales_skus:
        grouped.setdefault(("其他", "无售后", sku), {
            "orders": set(), "records": {},
            "source_records": defaultdict(dict),
        })
    generated_at = datetime.now()
    result: list[dict[str, Any]] = []
    for (big, small, sku), values in grouped.items():
        source_sales = sales_sources_by_sku.get(sku, {})
        sales_volume = sum(source_sales.values(), Decimal("0"))
        after_quantity = sum(values["records"].values(), Decimal("0"))
        source_after = {
            source: sum(record_quantities.values(), Decimal("0"))
            for source, record_quantities in values["source_records"].items()
        }
        rate = after_quantity / sales_volume if sales_volume else Decimal("0")
        result.append({
            "period_start": start_date,
            "period_end": end_date,
            "big_category": big,
            "small_category": small,
            "business_sku": sku,
            "order_count": len(values["orders"]),
            "order_numbers": "、".join(sorted(values["orders"])),
            "source_after_quantity_text": _source_text(source_after),
            "source_sales_volume_text": _source_text(source_sales),
            "after_quantity": after_quantity,
            "sales_volume": sales_volume,
            "after_sales_rate": rate,
            "calculation_version": SUMMARY_METRIC_VERSION,
            "sync_batch_id": batch_id,
            "generated_at": generated_at,
        })
    return sorted(result, key=lambda row: (-row["after_quantity"], row["business_sku"]))


def data_source_for_shop(store_name: str) -> str:
    name = _text(store_name)
    if any(keyword in name for keyword in EU_SHOPS):
        return "AMZ-EU"
    if any(keyword in name for keyword in US1_SHOPS):
        return "AMZ-US1"
    if any(keyword in name for keyword in US2_SHOPS):
        return "AMZ-US2"
    return "AMZ-OTHER"


def _business_sku(row: dict[str, Any]) -> str:
    return _text(row.get("local_sku") or row.get("business_sku") or row.get("msku")) or "UNKNOWN"


def _after_sales_note(row: dict[str, Any]) -> str:
    parts = []
    labels = (
        ("售后原因", "after_reason_zh"),
        ("退货状态", "return_status_zh"),
        ("库存属性", "inventory_attributes_zh"),
        ("买家备注", "buyers_note_zh"),
    )
    for label, key in labels:
        value = _text(row.get(key))
        if value:
            parts.append(f"{label}：{value}")
    return "；".join(parts) or "其他"


def _source_text(values: dict[str, Decimal]) -> str:
    order = {"AMZ-EU": 0, "AMZ-US1": 1, "AMZ-US2": 2, "AMZ-OTHER": 3}
    return "；".join(
        f"{source}:{_decimal_text(quantity)}"
        for source, quantity in sorted(values.items(), key=lambda item: (order.get(item[0], 99), item[0]))
    )


def _decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def _contains_star(value: str) -> bool:
    return "STAR" in _text(value).upper()


def _contains_star_or_reb(value: str) -> bool:
    upper = _text(value).upper()
    return "STAR" in upper or "REB" in upper


def _after_type(value: Any) -> str:
    text = _text(value)
    return {"1": "退款", "2": "退货", "3": "换货"}.get(text, text or "其他")


def _first_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, list) and value and isinstance(value[0], dict):
        return value[0]
    return value if isinstance(value, dict) else {}


def _first_value(value: Any) -> Any:
    if isinstance(value, list) and value:
        first = value[0]
        if isinstance(first, dict):
            return first.get("sid") or first.get("id")
        return first
    return None


def _decimal(value: Any) -> Decimal:
    try:
        text = str(value if value is not None else "0").strip().replace(",", "")
        return Decimal(text or "0")
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = _text(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def _sha(value: Any) -> str:
    return hashlib.sha256(
        json.dumps(value, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()


def _text(value: Any) -> str:
    return str(value or "").strip()
