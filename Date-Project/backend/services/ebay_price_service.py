from __future__ import annotations

import re
import tempfile
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.request import urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

from backend.config import settings
from backend.integrations.ebay.client import EbayBrowseClient, format_item
from backend.repositories import ebay_price_repository as repo


_client = EbayBrowseClient()


def import_sku_oe_mapping(content: bytes, file_name: str) -> dict[str, Any]:
    sku_to_oes, total_rows, skipped_rows = _parse_mapping_workbook(content)
    if not sku_to_oes:
        raise ValueError("未读取到有效的 sku、oe 对照数据")
    result = repo.replace_sku_mappings(sku_to_oes, file_name)
    return {
        "totalRows": total_rows,
        "affectedSkus": result["affected_skus"],
        "createdSkus": result["created_skus"],
        "updatedSkus": result["updated_skus"],
        "insertedMappings": result["inserted_mappings"],
        "skippedRows": skipped_rows,
    }


def search_prices(keywords: list[str], site: str, input_type: str) -> dict[str, Any]:
    normalized_keywords = _normalize_keywords(keywords)
    if not normalized_keywords:
        raise ValueError("请输入至少一个 SKU 或 OE 号")
    if len(normalized_keywords) > settings.ebay_search_max_keywords:
        raise ValueError(f"单次最多查询 {settings.ebay_search_max_keywords} 个关键词")

    sku_mapping: dict[str, list[str]] = {}
    not_found_skus: list[str] = []
    oe_list: list[str] = []

    if input_type == "oe":
        oe_list = normalized_keywords
    else:
        mappings = repo.get_oes_by_skus(normalized_keywords)
        for keyword in normalized_keywords:
            oes = mappings.get(keyword)
            if oes:
                sku_mapping[keyword] = oes
                oe_list.extend(oes)
            elif input_type == "sku":
                not_found_skus.append(keyword)
            else:
                oe_list.append(keyword)
        if input_type == "sku" and not oe_list:
            raise ValueError("未找到任何 SKU 对应的 OE 号")

    oe_list = _unique(oe_list)
    if not oe_list:
        raise ValueError("未解析到可查询的 OE 号")

    results = _search_many_oes(oe_list, site)
    return {
        "site": site,
        "inputType": input_type,
        "oeList": oe_list,
        "skuMapping": sku_mapping,
        "notFoundSkus": not_found_skus,
        "results": results,
    }


def export_search_results(items: list[dict[str, Any]]) -> tuple[str, str]:
    if not items:
        raise ValueError("没有可导出的商品")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "eBay Results"
    headers = ["OE", "Price", "Title", "Seller", "Rate", "Shipping", "Link", "Images"]
    widths = [18, 14, 60, 18, 10, 14, 45, 18]
    for column, (header, width) in enumerate(zip(headers, widths), 1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = cell.font.copy(bold=True)
        sheet.column_dimensions[get_column_letter(column)].width = width

    for row_index, item in enumerate(items, 2):
        sheet.cell(row=row_index, column=1, value=str(item.get("oe") or ""))
        sheet.cell(row=row_index, column=2, value=str(item.get("price") or ""))
        sheet.cell(row=row_index, column=3, value=str(item.get("title") or ""))
        sheet.cell(row=row_index, column=4, value=str(item.get("seller") or ""))
        feedback = str(item.get("sellerFeedback") or "")
        sheet.cell(row=row_index, column=5, value=f"{feedback}%" if feedback else "")
        sheet.cell(row=row_index, column=6, value=str(item.get("shipping") or ""))
        link = str(item.get("link") or "")
        if link:
            link_cell = sheet.cell(row=row_index, column=7, value=link)
            link_cell.hyperlink = link
            link_cell.font = link_cell.font.copy(color="0563C1", underline="single")
        _add_first_image(sheet, row_index, item.get("images") or [])
        sheet.row_dimensions[row_index].height = 68

    export_dir = Path(settings.export_output_dir)
    export_dir.mkdir(parents=True, exist_ok=True)
    temp = tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        prefix="ebay_price_export_",
        dir=str(export_dir),
        delete=False,
    )
    temp.close()
    workbook.save(temp.name)
    return temp.name, f"ebay_price_export.xlsx"


def _search_many_oes(oe_list: list[str], site: str) -> list[dict[str, Any]]:
    workers = max(1, settings.ebay_search_max_workers)
    with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="ebay-search") as executor:
        return list(executor.map(lambda oe: _search_one_oe(oe, site), oe_list))


def _search_one_oe(oe: str, site: str) -> dict[str, Any]:
    data = _client.search_items(oe, site, limit=settings.ebay_search_limit)
    raw_items = data.get("itemSummaries") or []
    if not raw_items:
        return {"oe": oe, "count": 0, "items": [], "warning": f"OE '{oe}' 未找到结果"}
    items = []
    for raw_item in raw_items:
        item = format_item(raw_item)
        item["oe"] = oe
        items.append(item)
    items.sort(key=lambda row: row["pf"])
    items = items[: settings.ebay_search_top_n]
    return {"oe": oe, "count": len(items), "items": items, "warning": None}


def _parse_mapping_workbook(content: bytes) -> tuple[dict[str, list[str]], int, int]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            raise ValueError("Excel 文件为空") from None
        sku_index, oe_index = _mapping_header_indexes(headers)
        total_rows = 0
        skipped_rows = 0
        sku_to_oes: OrderedDict[str, list[str]] = OrderedDict()
        for row in rows:
            total_rows += 1
            sku = _cell_text(row[sku_index] if sku_index < len(row) else None)
            oe_text = _cell_text(row[oe_index] if oe_index < len(row) else None)
            oes = _split_oes(oe_text)
            if not sku or not oes:
                skipped_rows += 1
                continue
            sku_to_oes.setdefault(sku, [])
            sku_to_oes[sku] = _unique([*sku_to_oes[sku], *oes])
        return dict(sku_to_oes), total_rows, skipped_rows
    finally:
        workbook.close()


def _mapping_header_indexes(headers: tuple[Any, ...]) -> tuple[int, int]:
    normalized = {_normalize_header(value): index for index, value in enumerate(headers)}
    sku_index = normalized.get("sku")
    oe_index = normalized.get("oe")
    if sku_index is None or oe_index is None:
        raise ValueError("Excel 表头必须包含 sku 和 oe 两列")
    return sku_index, oe_index


def _normalize_header(value: Any) -> str:
    return _cell_text(value).lower().replace(" ", "").replace("_", "")


def _normalize_keywords(keywords: list[str]) -> list[str]:
    result = []
    for keyword in keywords:
        result.extend(_split_keywords(keyword))
    return _unique(result)


def _split_keywords(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[\n,，]+", str(value or "")) if part.strip()]


def _split_oes(value: str) -> list[str]:
    return _unique(_split_keywords(value))


def _unique(values: list[str]) -> list[str]:
    result = []
    seen = set()
    for value in values:
        normalized = value.strip()
        key = normalized.upper()
        if normalized and key not in seen:
            seen.add(key)
            result.append(normalized)
    return result


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _add_first_image(sheet, row_index: int, images: list[Any]) -> None:
    if not images:
        return
    image_url = str(images[0] or "")
    if not image_url.startswith(("http://", "https://")):
        return
    try:
        with urlopen(image_url, timeout=10) as response:
            data = BytesIO(response.read())
        image = XlImage(data)
        image.height = 80
        image.width = 80
        sheet.add_image(image, f"H{row_index}")
    except Exception:
        return
