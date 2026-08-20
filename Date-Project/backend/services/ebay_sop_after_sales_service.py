from __future__ import annotations

import hashlib
import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import from_excel

from backend.repositories import ebay_sop_repository as repo
from backend.services.amz_sop_classifier import BIG_CATEGORIES, classify_rows


DIGITAL_REQUIRED_HEADERS = {
    "平台订单号", "发货状态", "包裹状态", "付款时间", "平台", "币种",
    "应收货款(订单级别)", "库存SKU", "SKU状态", "购买数量",
}

HISTORY_REQUIRED_HEADERS = {
    "订单号", "付款时间", "退款时间", "仓库商品标题", "售后数量", "产品SKU",
    "售后小类划分", "售后大类划分", "数据来源", "平台", "售后备注",
}

HISTORY_SALES_REQUIRED_HEADERS = {"付款时间", "数据来源", "产品SKU", "销量"}

DIGITAL_COLUMN_MAP = {
    "平台订单号": "platform_order_no",
    "发货状态": "shipping_status",
    "包裹状态": "package_status",
    "异常状态": "exception_status",
    "物流渠道": "logistics_channel",
    "货运单号": "tracking_no",
    "付款时间": "payment_time",
    "标发时间": "marked_ship_time",
    "平台": "platform_name",
    "币种": "currency_code",
    "应收货款(订单级别)": "receivable_goods",
    "应收运费": "receivable_shipping",
    "平台费用": "platform_fee",
    "保险金额": "insurance_amount",
    "转账费": "transfer_fee",
    "收款Paypal": "paypal_receipt",
    "客户ID": "customer_id",
    "收件人": "recipient",
    "客户邮箱": "customer_email",
    "收件人电话1": "recipient_phone1",
    "收件人电话2": "recipient_phone2",
    "收件人国家": "recipient_country",
    "收件人所在省/州": "recipient_state",
    "收件人所在市": "recipient_city",
    "收件人所在地邮政编码": "recipient_postal_code",
    "收件地址": "recipient_address1",
    "收件地址2": "recipient_address2",
    "收件地址3": "recipient_address3",
    "收件人备用地址": "recipient_alt_address",
    "收件人门牌号": "recipient_house_no",
    "库存SKU": "inventory_sku",
    "SKU状态": "sku_status",
    "购买数量": "purchase_quantity",
    "申请到的库存数量": "allocated_quantity",
    "PayPal交易号": "paypal_transaction_no",
    "销售员": "salesperson",
    "汇率": "exchange_rate",
    "原币收入": "original_currency_income",
    "国家中文": "country_cn",
    "国家英文": "country_en",
}

DECIMAL_RAW_FIELDS = {
    "receivable_goods", "receivable_shipping", "platform_fee", "insurance_amount",
    "transfer_fee", "purchase_quantity", "allocated_quantity", "exchange_rate",
    "original_currency_income",
}

DATETIME_RAW_FIELDS = {"payment_time", "marked_ship_time"}

SALES_MIN_DATE = date(2026, 6, 14)
SUMMARY_METRIC_VERSION = "EBAY-QUANTITY-V1"


def import_ebay_sales(
    content: bytes, file_name: str, operator: str | None = None
) -> dict[str, Any]:
    return _import_digital(content, file_name, "SALES", operator)


def import_ebay_after_sales(
    content: bytes, file_name: str, operator: str | None = None
) -> dict[str, Any]:
    return _import_digital(content, file_name, "AFTER_SALES", operator)


def import_ebay_history(
    content: bytes, file_name: str, operator: str | None = None
) -> dict[str, Any]:
    batch = _start_batch(content, file_name, "HISTORY", operator)
    metrics = {"total_rows": 0, "raw_rows": 0, "dwd_rows": 0, "skipped_rows": 0}
    try:
        parsed = _parse_history_workbook(content, file_name, batch["batch_id"])
        metrics["total_rows"] = parsed["total_rows"]
        metrics["skipped_rows"] = parsed["skipped_rows"]
        if not parsed["after_sales_raw_rows"]:
            raise ValueError("历史文件没有可导入的售后数据")
        if not parsed["sales_raw_rows"]:
            raise ValueError("历史文件没有可导入的销量数据")
        with repo.import_lock("HISTORY"):
            (
                after_raw_count,
                sales_raw_count,
                after_dwd_count,
                sales_dwd_count,
            ) = repo.replace_history_data(
                parsed["after_sales_raw_rows"],
                parsed["sales_raw_rows"],
                parsed["after_sales_dwd_rows"],
                parsed["sales_dwd_rows"],
            )
        metrics["raw_rows"] = after_raw_count + sales_raw_count
        metrics["dwd_rows"] = after_dwd_count + sales_dwd_count
        replaced_months = sorted({
            row["month_start"].strftime("%Y-%m")
            for row in parsed["sales_dwd_rows"]
        } | {
            row["after_time"].strftime("%Y-%m")
            for row in parsed["after_sales_dwd_rows"]
        })
        _finish_batch(batch["batch_id"], metrics)
        return {
            **batch,
            **metrics,
            "after_sales_raw_rows": after_raw_count,
            "after_sales_dwd_rows": after_dwd_count,
            "sales_raw_rows": sales_raw_count,
            "sales_dwd_rows": sales_dwd_count,
            "replaced_months": replaced_months,
            "message": (
                "eBay标准售后及销量数据按月份覆盖完成："
                + "、".join(replaced_months)
            ),
        }
    except Exception as exc:
        _fail_batch(batch["batch_id"], metrics, exc)
        raise


def list_periods(limit: int = 24) -> list[dict[str, Any]]:
    return repo.periods(limit)


def _import_digital(
    content: bytes,
    file_name: str,
    data_kind: str,
    operator: str | None,
) -> dict[str, Any]:
    batch = _start_batch(content, file_name, data_kind, operator)
    metrics = {"total_rows": 0, "raw_rows": 0, "dwd_rows": 0, "skipped_rows": 0}
    try:
        parsed = _parse_digital_workbook(
            content, file_name, data_kind, batch["batch_id"]
        )
        metrics["total_rows"] = parsed["total_rows"]
        metrics["skipped_rows"] = parsed["skipped_rows"]
        if not parsed["raw_rows"]:
            raise ValueError("数字酋长文件没有可导入的数据")

        order_numbers = {
            _text(row.get("platform_order_no"), 128)
            for row in parsed["raw_rows"]
            if _text(row.get("platform_order_no"), 128)
        }
        with repo.import_lock(data_kind):
            existing_orders = repo.existing_order_numbers(data_kind, order_numbers)
            new_raw_rows = [
                row for row in parsed["raw_rows"]
                if _text(row.get("platform_order_no"), 128) not in existing_orders
            ]
            skipped_existing_rows = len(parsed["raw_rows"]) - len(new_raw_rows)

            if data_kind == "SALES":
                dwd_rows = _build_sales_rows(new_raw_rows, batch["batch_id"])
                new_order_count = len({
                    _text(row.get("platform_order_no"), 128)
                    for row in new_raw_rows if _is_valid_sales_row(row)
                })
                invalid_rows = sum(
                    1 for row in new_raw_rows if not _is_valid_sales_row(row)
                )
                if not dwd_rows and not existing_orders:
                    raise ValueError(
                        "文件中没有2026-06-14之后且发货/包裹状态非已作废的有效销量"
                    )
                metrics["raw_rows"], metrics["dwd_rows"] = (
                    repo.append_sales_orders(new_raw_rows, dwd_rows)
                    if dwd_rows else (0, 0)
                )
                result_dates = ({
                    "period_start": min(row["sale_date"] for row in dwd_rows),
                    "period_end": max(row["sale_date"] for row in dwd_rows),
                } if dwd_rows else {})
            else:
                dwd_rows, invalid_rows = _build_future_after_sales_rows(
                    new_raw_rows, batch["batch_id"]
                )
                new_order_count = len({row["order_no"] for row in dwd_rows})
                if not dwd_rows and not existing_orders:
                    raise ValueError("文件中没有包含已退款或已作废状态的有效售后订单")
                metrics["raw_rows"], metrics["dwd_rows"] = (
                    repo.append_after_sales_orders(new_raw_rows, dwd_rows)
                    if dwd_rows else (0, 0)
                )
                result_dates = ({
                    "period_start": min(row["after_time"].date() for row in dwd_rows),
                    "period_end": max(row["after_time"].date() for row in dwd_rows),
                } if dwd_rows else {})

        metrics["skipped_rows"] = invalid_rows + skipped_existing_rows
        existing_order_count = len(existing_orders)
        _finish_batch(batch["batch_id"], metrics)
        label = "销量" if data_kind == "SALES" else "售后"
        return {
            **batch, **metrics, **result_dates,
            "new_order_count": new_order_count,
            "existing_order_count": existing_order_count,
            "skipped_existing_rows": skipped_existing_rows,
            "message": (
                f"eBay{label}增量导入完成：新增{new_order_count}个订单，"
                f"忽略{existing_order_count}个已存在订单"
            ),
        }
    except Exception as exc:
        _fail_batch(batch["batch_id"], metrics, exc)
        raise


def list_product_summary(
    start_date: date | None,
    end_date: date | None,
    big_category: str | None,
    small_category: str | None,
    sku: str | None,
    page: int,
    page_size: int,
) -> dict[str, Any]:
    coverage_start, coverage_end = repo.coverage()
    if not coverage_start or not coverage_end or coverage_start > coverage_end:
        return {
            "period_start": None, "period_end": None, "items": [], "total": 0,
            "summary": {}, "range_status": "ready", "range_generated": False,
        }
    if not start_date or not end_date:
        latest_months = repo.periods(1)
        if not latest_months:
            return {
                "period_start": None, "period_end": None, "items": [], "total": 0,
                "summary": {}, "range_status": "ready", "range_generated": False,
            }
        start_date = latest_months[0]["period_start"]
        end_date = latest_months[0]["period_end"]
    _validate_range(start_date, end_date, coverage_start, coverage_end)
    _rebuild_summary(start_date, end_date)
    rows = repo.summary_filtered(
        start_date, end_date, big_category, small_category, sku
    )
    result = _aggregate_product_rows(rows, page, page_size)
    all_rows = rows if not any((big_category, small_category, sku)) else repo.summary_filtered(
        start_date, end_date, None, None, None
    )
    range_summary = _aggregate_product_rows(all_rows, 1, 1)["summary"]
    range_sales_volume = repo.sales_total(start_date, end_date)
    result["summary"].update({
        "range_after_quantity": range_summary["after_quantity"],
        "range_sales_volume": range_sales_volume,
        "range_after_sales_rate": (
            range_summary["after_quantity"] / range_sales_volume
            if range_sales_volume else Decimal("0")
        ),
    })
    result.update({
        "period_start": start_date,
        "period_end": end_date,
        "range_generated": True,
        "range_status": "ready",
        "range_message": "eBay月份区间售后率已实时计算",
    })
    return result


def export_summary(start_date: date, end_date: date) -> bytes:
    coverage_start, coverage_end = repo.coverage()
    _validate_range(start_date, end_date, coverage_start, coverage_end)
    _rebuild_summary(start_date, end_date)
    rows = repo.summary_filtered(start_date, end_date)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get("big_category") or "其他")].append(row)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for category in BIG_CATEGORIES:
        sheet = workbook.create_sheet(category[:31])
        _write_summary_sheet(sheet, grouped.get(category, []))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_filtered_summary(
    start_date: date,
    end_date: date,
    big_category: str | None = None,
    small_category: str | None = None,
    sku: str | None = None,
    selected_skus: list[str] | None = None,
) -> bytes:
    coverage_start, coverage_end = repo.coverage()
    _validate_range(start_date, end_date, coverage_start, coverage_end)
    _rebuild_summary(start_date, end_date)
    rows = repo.summary_filtered(
        start_date, end_date, big_category, small_category, sku, selected_skus
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "eBay售后数据"
    _write_summary_sheet(sheet, rows)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _parse_digital_workbook(
    content: bytes, file_name: str, data_kind: str, batch_id: str
) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    raw_rows: list[dict[str, Any]] = []
    total_rows = 0
    skipped_rows = 0
    valid_sheet_found = False
    try:
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            header_values = next(iterator, None)
            if not header_values or not any(_text(value) for value in header_values):
                continue
            headers = [_text(value) for value in header_values]
            missing = sorted(DIGITAL_REQUIRED_HEADERS - set(headers))
            if missing:
                raise ValueError(
                    f"工作表“{sheet.title}”不是数字酋长订单模板，缺少表头：{'、'.join(missing)}"
                )
            valid_sheet_found = True
            positions = {header: index for index, header in enumerate(headers) if header}
            for row_no, values in enumerate(iterator, 2):
                if not values or not any(value not in (None, "") for value in values):
                    continue
                total_rows += 1
                source = {
                    header: values[index] if index < len(values) else None
                    for header, index in positions.items()
                }
                raw = _digital_raw_row(
                    source, file_name, sheet.title, row_no, data_kind, batch_id
                )
                raw_rows.append(raw)
                if not raw["platform_order_no"] or not raw["inventory_sku"]:
                    skipped_rows += 1
    finally:
        workbook.close()
    if not valid_sheet_found:
        raise ValueError("Excel中没有可识别的数字酋长订单工作表")
    return {
        "total_rows": total_rows,
        "raw_rows": raw_rows,
        "skipped_rows": skipped_rows,
    }


def _parse_history_workbook(
    content: bytes, file_name: str, batch_id: str
) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    after_sales_raw_rows: list[dict[str, Any]] = []
    after_sales_dwd_by_key: dict[str, dict[str, Any]] = {}
    dwd_scores: dict[str, tuple[int, int, int, int]] = {}
    sales_raw_rows: list[dict[str, Any]] = []
    sales_dwd_by_key: dict[tuple[date, date, str, str], dict[str, Any]] = {}
    total_rows = 0
    skipped_rows = 0
    after_sales_sheet_found = False
    sales_sheet_found = False
    try:
        for sheet in workbook.worksheets:
            iterator = sheet.iter_rows(values_only=True)
            header_values = next(iterator, None)
            if not header_values or not any(_text(value) for value in header_values):
                continue
            headers = [_text(value) for value in header_values]
            positions = {header: index for index, header in enumerate(headers) if header}
            is_after_sales_sheet = HISTORY_REQUIRED_HEADERS.issubset(set(headers))
            is_sales_sheet = HISTORY_SALES_REQUIRED_HEADERS.issubset(set(headers))
            if not is_after_sales_sheet and not is_sales_sheet:
                if sheet.max_row <= 1:
                    continue
                raise ValueError(
                    f"工作表“{sheet.title}”无法识别：应为“售后数据”或“销量”模板"
                )
            after_sales_sheet_found = after_sales_sheet_found or is_after_sales_sheet
            sales_sheet_found = sales_sheet_found or is_sales_sheet
            for row_no, values in enumerate(iterator, 2):
                if not values or not any(value not in (None, "") for value in values):
                    continue
                total_rows += 1
                source = {
                    header: values[index] if index < len(values) else None
                    for header, index in positions.items()
                }
                if is_after_sales_sheet:
                    raw, candidate, score = _history_after_sales_row(
                        source, file_name, sheet.title, row_no, batch_id
                    )
                    after_sales_raw_rows.append(raw)
                    if candidate is None:
                        skipped_rows += 1
                        continue
                    source_key = candidate["source_key"]
                    if score > dwd_scores.get(source_key, (-1, -1, -1, -1)):
                        after_sales_dwd_by_key[source_key] = candidate
                        dwd_scores[source_key] = score
                    continue

                raw, candidate = _history_sales_row(
                    source, file_name, sheet.title, row_no, batch_id
                )
                sales_raw_rows.append(raw)
                if candidate is None:
                    skipped_rows += 1
                    continue
                key = (
                    candidate["month_start"], candidate["month_end"],
                    candidate["data_source"], candidate["business_sku"],
                )
                target = sales_dwd_by_key.setdefault(key, candidate)
                if target is not candidate:
                    target["sales_quantity"] += candidate["sales_quantity"]
    finally:
        workbook.close()
    if not after_sales_sheet_found:
        raise ValueError("历史文件缺少“售后数据”工作表")
    if not sales_sheet_found:
        raise ValueError("历史文件缺少“销量”工作表")
    return {
        "total_rows": total_rows,
        "after_sales_raw_rows": after_sales_raw_rows,
        "after_sales_dwd_rows": list(after_sales_dwd_by_key.values()),
        "sales_raw_rows": sales_raw_rows,
        "sales_dwd_rows": list(sales_dwd_by_key.values()),
        "skipped_rows": skipped_rows,
    }


def _history_after_sales_row(
    source: dict[str, Any],
    file_name: str,
    sheet_name: str,
    row_no: int,
    batch_id: str,
) -> tuple[dict[str, Any], dict[str, Any] | None, tuple[int, int, int, int]]:
    order_no = _text(source.get("订单号"), 128)
    sku = _text(source.get("产品SKU"), 255)
    refund_time = _datetime(source.get("退款时间"))
    quantity = max(_decimal(source.get("售后数量")), Decimal("0"))
    raw = {
        "source_key": _sha("HISTORY_RAW", file_name, sheet_name, row_no),
        "order_no": order_no or None,
        "payment_time": _datetime(source.get("付款时间")),
        "refund_time": refund_time,
        "product_title": _text(source.get("仓库商品标题")) or None,
        "after_quantity": quantity,
        "product_sku": sku or None,
        "small_category": _text(source.get("售后小类划分"), 100) or None,
        "big_category": _text(source.get("售后大类划分"), 50) or None,
        "data_source": _text(source.get("数据来源"), 50) or None,
        "platform_name": _text(source.get("平台"), 100) or None,
        "after_sales_note": _text(source.get("售后备注")) or None,
        "source_file": file_name,
        "source_sheet": sheet_name,
        "source_row_no": row_no,
        "import_batch_id": batch_id,
    }
    if not order_no or not sku or not refund_time or quantity <= 0:
        return raw, None, (-1, -1, -1, -1)
    source_key = _sha("HISTORY", order_no, sku, refund_time.isoformat())
    big, small = _valid_category(
        source.get("售后大类划分"), source.get("售后小类划分")
    )
    candidate = {
        "source_key": source_key,
        "source_kind": "HISTORY",
        "order_no": order_no,
        "payment_time": raw["payment_time"],
        "after_time": refund_time,
        "product_title": raw["product_title"],
        "after_quantity": quantity,
        "business_sku": sku,
        "after_type": "退款",
        "big_category": big,
        "small_category": small,
        "data_source": _normalize_data_source(source.get("数据来源")),
        "platform_name": raw["platform_name"],
        "after_sales_note": raw["after_sales_note"] or "其他",
        "classify_method": "history",
        "confidence": Decimal("1"),
        "import_batch_id": batch_id,
    }
    score = (
        int(big != "其他" and small != "其他"),
        int(bool(raw["after_sales_note"])),
        len(raw["after_sales_note"] or ""),
        row_no,
    )
    return raw, candidate, score


def _history_sales_row(
    source: dict[str, Any],
    file_name: str,
    sheet_name: str,
    row_no: int,
    batch_id: str,
) -> tuple[dict[str, Any], dict[str, Any] | None]:
    month_bounds = _month_bounds(source.get("付款时间"))
    sku = _text(source.get("产品SKU"), 255)
    quantity = max(_decimal(source.get("销量")), Decimal("0"))
    data_source = _normalize_data_source(source.get("数据来源"))
    raw = {
        "source_key": _sha("HISTORY_SALES_RAW", file_name, sheet_name, row_no),
        "sale_month": month_bounds[0].strftime("%Y-%m") if month_bounds else None,
        "data_source": _text(source.get("数据来源"), 50) or None,
        "product_sku": sku or None,
        "sales_quantity": quantity,
        "source_file": file_name,
        "source_sheet": sheet_name,
        "source_row_no": row_no,
        "import_batch_id": batch_id,
    }
    if not month_bounds or not sku or quantity <= 0:
        return raw, None
    business_sku, factor = _normalize_rg_sku(sku)
    return raw, {
        "month_start": month_bounds[0],
        "month_end": month_bounds[1],
        "data_source": data_source,
        "business_sku": business_sku,
        "sales_quantity": quantity * factor,
        "import_batch_id": batch_id,
    }


def _digital_raw_row(
    source: dict[str, Any],
    file_name: str,
    sheet_name: str,
    row_no: int,
    data_kind: str,
    batch_id: str,
) -> dict[str, Any]:
    raw: dict[str, Any] = {
        field: Decimal("0") if field in DECIMAL_RAW_FIELDS else None
        for field in DIGITAL_COLUMN_MAP.values()
    }
    for header, field in DIGITAL_COLUMN_MAP.items():
        value = source.get(header)
        if field in DECIMAL_RAW_FIELDS:
            raw[field] = _decimal(value)
        elif field in DATETIME_RAW_FIELDS:
            raw[field] = _datetime(value)
        else:
            raw[field] = _text(value) or None
    raw["platform_order_no"] = _text(raw.get("platform_order_no"), 128)
    raw["inventory_sku"] = _text(raw.get("inventory_sku"), 255)
    raw.update({
        "data_kind": data_kind,
        "source_file": file_name,
        "source_sheet": sheet_name,
        "source_row_no": row_no,
        "import_batch_id": batch_id,
    })
    raw["source_key"] = _sha(
        f"{data_kind}_RAW", file_name, sheet_name, row_no
    )
    return raw


def _build_sales_rows(
    raw_rows: list[dict[str, Any]], batch_id: str
) -> list[dict[str, Any]]:
    grouped: dict[tuple[date, str, str, str], dict[str, Any]] = {}
    for raw in raw_rows:
        if not _is_valid_sales_row(raw):
            continue
        payment_time = raw.get("payment_time")
        sku = _text(raw.get("inventory_sku"), 255)
        quantity = _decimal(raw.get("purchase_quantity"))
        amount = _decimal(raw.get("receivable_goods"))
        unit_price = amount / quantity if quantity else Decimal("0")
        business_sku, factor = _normalize_rg_sku(sku)
        adjusted_quantity = quantity * factor
        adjusted_unit_price = unit_price * factor
        currency = _text(raw.get("currency_code"), 20).upper() or "UNKNOWN"
        data_source = _data_source_for_currency(currency)
        key = (payment_time.date(), data_source, business_sku, currency)
        target = grouped.setdefault(key, {
            "sale_date": payment_time.date(),
            "data_source": data_source,
            "business_sku": business_sku,
            "currency_code": currency,
            "sales_quantity": Decimal("0"),
            "unit_price_total": Decimal("0"),
            "sales_amount": Decimal("0"),
            "orders": set(),
            "import_batch_id": batch_id,
        })
        target["sales_quantity"] += adjusted_quantity
        target["unit_price_total"] += adjusted_unit_price
        target["sales_amount"] += amount
        target["orders"].add(raw["platform_order_no"])
    result = []
    for target in grouped.values():
        target["order_count"] = len(target.pop("orders"))
        result.append(target)
    return result


def _is_valid_sales_row(raw: dict[str, Any]) -> bool:
    payment_time = raw.get("payment_time")
    return bool(
        raw.get("platform_order_no")
        and _text(raw.get("inventory_sku"))
        and payment_time
        and payment_time.date() >= SALES_MIN_DATE
        and "已作废" not in _text(raw.get("shipping_status"))
        and "已作废" not in _text(raw.get("package_status"))
        and _decimal(raw.get("purchase_quantity")) > 0
    )


def _build_future_after_sales_rows(
    raw_rows: list[dict[str, Any]], batch_id: str
) -> tuple[list[dict[str, Any]], int]:
    candidates: list[dict[str, Any]] = []
    skipped = 0
    for raw in raw_rows:
        shipping_status = _text(raw.get("shipping_status"))
        package_status = _text(raw.get("package_status"))
        order_no = _text(raw.get("platform_order_no"), 128)
        sku = _text(raw.get("inventory_sku"), 255)
        payment_time = raw.get("payment_time")
        after_time = raw.get("marked_ship_time") or payment_time
        quantity = _decimal(raw.get("purchase_quantity"))
        if not order_no or not sku or not after_time or quantity <= 0:
            skipped += 1
            continue
        business_sku, factor = _normalize_rg_sku(sku)
        note_parts = [
            f"发货状态：{shipping_status}" if shipping_status else "",
            f"包裹状态：{package_status}" if package_status else "",
            f"异常状态：{_text(raw.get('exception_status'))}" if raw.get("exception_status") else "",
            f"SKU状态：{_text(raw.get('sku_status'))}" if raw.get("sku_status") else "",
            f"物流渠道：{_text(raw.get('logistics_channel'))}" if raw.get("logistics_channel") else "",
        ]
        candidate = {
            "raw": raw,
            "source_key": raw["source_key"],
            "order_no": order_no,
            "payment_time": payment_time,
            "after_time": after_time,
            "after_quantity": quantity * factor,
            "business_sku": business_sku,
            "after_type": "退货" if "已发货" in shipping_status else "退款",
            "data_source": _data_source_for_currency(raw.get("currency_code")),
            "platform_name": _text(raw.get("platform_name"), 100) or "eBay",
            "after_sales_note": "；".join(part for part in note_parts if part) or "其他",
            "after_reason": _text(raw.get("exception_status")),
            "return_status": package_status,
            "inventory_attributes": _text(raw.get("sku_status")),
            "buyers_note": "；".join(part for part in note_parts if part),
        }
        candidates.append(candidate)
    classifications = classify_rows(candidates, platform="EBAY")
    result = []
    for row in candidates:
        classified = classifications.get(row.get("classification_hash"), {})
        result.append({
            "source_key": _sha(
                "AFTER_SALES",
                row["order_no"],
                row["business_sku"],
                row["payment_time"],
                row["after_time"],
                row["raw"].get("paypal_transaction_no"),
            ),
            "source_kind": "AFTER_SALES",
            "order_no": row["order_no"],
            "payment_time": row["payment_time"],
            "after_time": row["after_time"],
            "product_title": None,
            "after_quantity": row["after_quantity"],
            "business_sku": row["business_sku"],
            "after_type": row["after_type"],
            "big_category": classified.get("big_category") or "其他",
            "small_category": classified.get("small_category") or "其他",
            "data_source": row["data_source"],
            "platform_name": row["platform_name"],
            "after_sales_note": row["after_sales_note"],
            "classify_method": classified.get("classify_method") or "fallback",
            "confidence": _decimal(classified.get("confidence")),
            "import_batch_id": batch_id,
        })
    return result, skipped


def _rebuild_summary(start_date: date, end_date: date) -> None:
    after_rows = repo.after_sales_rows(start_date, end_date)
    sales = repo.sales_by_sku_source(start_date, end_date)
    sales_by_sku: dict[str, dict[str, Decimal]] = defaultdict(dict)
    for (sku, source), quantity in sales.items():
        sales_by_sku[sku][source] = quantity
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in after_rows:
        sku = _text(row.get("business_sku"))
        if not sku:
            continue
        key = (
            _text(row.get("big_category")) or "其他",
            _text(row.get("small_category")) or "其他",
            sku,
        )
        target = grouped.setdefault(key, {
            "orders": set(), "source_after": defaultdict(Decimal),
            "after_quantity": Decimal("0"),
        })
        quantity = _decimal(row.get("after_quantity"))
        source = _normalize_data_source(row.get("data_source"))
        target["orders"].add(_text(row.get("order_no")))
        target["source_after"][source] += quantity
        target["after_quantity"] += quantity
    generated_at = datetime.now()
    summary_rows = []
    for (big, small, sku), values in grouped.items():
        source_sales = sales_by_sku.get(sku, {})
        sales_volume = sum(source_sales.values(), Decimal("0"))
        after_quantity = values["after_quantity"]
        summary_rows.append({
            "period_start": start_date,
            "period_end": end_date,
            "big_category": big,
            "small_category": small,
            "business_sku": sku,
            "order_count": len(values["orders"]),
            "order_numbers": "、".join(sorted(order for order in values["orders"] if order)),
            "source_after_quantity_text": _source_text(values["source_after"]),
            "source_sales_volume_text": _source_text(source_sales),
            "after_quantity": after_quantity,
            "sales_volume": sales_volume,
            "after_sales_rate": after_quantity / sales_volume if sales_volume else Decimal("0"),
            "calculation_version": SUMMARY_METRIC_VERSION,
            "generated_at": generated_at,
        })
    repo.replace_summary_period(start_date, end_date, summary_rows)


def _aggregate_product_rows(
    rows: list[dict[str, Any]], page: int, page_size: int
) -> dict[str, Any]:
    products: dict[str, dict[str, Any]] = {}
    all_orders: set[str] = set()
    all_categories: set[str] = set()
    for source_row in rows:
        row = dict(source_row)
        sku = _text(row.get("business_sku"))
        if not sku:
            continue
        product = products.setdefault(sku, {
            "id": sku, "business_sku": sku, "after_quantity": Decimal("0"),
            "sales_volume": Decimal("0"), "orders": set(),
            "source_after": defaultdict(Decimal), "source_sales_volume_text": "",
            "children": [],
        })
        product["after_quantity"] += _decimal(row.get("after_quantity"))
        product["sales_volume"] = max(product["sales_volume"], _decimal(row.get("sales_volume")))
        if row.get("source_sales_volume_text"):
            product["source_sales_volume_text"] = row["source_sales_volume_text"]
        for order in _text(row.get("order_numbers")).split("、"):
            if order:
                product["orders"].add(order)
                all_orders.add(order)
        for source, quantity in _parse_source_text(row.get("source_after_quantity_text")).items():
            product["source_after"][source] += quantity
        all_categories.add(_text(row.get("big_category")))
        product["children"].append(row)
    result = []
    for product in products.values():
        product["children"].sort(
            key=lambda item: (-_decimal(item.get("after_quantity")), _text(item.get("small_category")))
        )
        orders = product.pop("orders")
        product["order_count"] = len(orders)
        product["order_numbers"] = "、".join(sorted(orders))
        product["source_after_quantity_text"] = _source_text(product.pop("source_after"))
        product["after_sales_rate"] = (
            product["after_quantity"] / product["sales_volume"]
            if product["sales_volume"] else Decimal("0")
        )
        product["detail_count"] = len(product["children"])
        result.append(product)
    result.sort(key=lambda item: (-item["after_quantity"], item["business_sku"]))
    safe_page = max(page, 1)
    safe_size = max(1, min(page_size, 200))
    offset = (safe_page - 1) * safe_size
    total_after = sum((item["after_quantity"] for item in result), Decimal("0"))
    total_sales = sum((item["sales_volume"] for item in result), Decimal("0"))
    return {
        "items": result[offset:offset + safe_size],
        "total": len(result),
        "summary": {
            "sku_count": len(result),
            "after_quantity": total_after,
            "order_count": len(all_orders),
            "category_count": len({item for item in all_categories if item}),
            "sales_volume": total_sales,
            "after_sales_rate": total_after / total_sales if total_sales else Decimal("0"),
        },
    }


def _write_summary_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    headers = (
        "售后原因（大）", "售后原因（小）", "SKU", "售后数量", "订单数量",
        "订单号", "数据来源及售后数量", "数据来源及销量", "销量", "售后率",
    )
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append((
            row["big_category"], row["small_category"], row["business_sku"],
            float(row["after_quantity"] or 0), int(row["order_count"] or 0),
            row.get("order_numbers") or "", row.get("source_after_quantity_text") or "",
            row.get("source_sales_volume_text") or "", float(row["sales_volume"] or 0),
            float(row["after_sales_rate"] or 0),
        ))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = (18, 22, 24, 12, 12, 48, 28, 28, 12, 12)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for cell in sheet["J"][1:]:
        cell.number_format = "0.00%"


def _start_batch(
    content: bytes, file_name: str, import_type: str, operator: str | None
) -> dict[str, Any]:
    batch_id = f"EBAY-SOP-{import_type}-{datetime.now():%Y%m%d%H%M%S}-{uuid4().hex[:8]}"
    file_sha256 = hashlib.sha256(content).hexdigest()
    row = {
        "batch_id": batch_id,
        "import_type": import_type,
        "file_name": file_name,
        "file_sha256": file_sha256,
        "operator": operator,
    }
    repo.start_import_batch(row)
    return row


def _finish_batch(batch_id: str, metrics: dict[str, int]) -> None:
    repo.finish_import_batch(batch_id, "SUCCESS", **metrics)


def _fail_batch(batch_id: str, metrics: dict[str, int], exc: Exception) -> None:
    repo.finish_import_batch(
        batch_id, "FAILED", **metrics, error_message=str(exc)[:4000]
    )


def _validate_range(
    start_date: date,
    end_date: date,
    coverage_start: date | None,
    coverage_end: date | None,
) -> None:
    if start_date > end_date:
        raise ValueError("开始月份不能晚于结束月份")
    month_count = (
        (end_date.year - start_date.year) * 12
        + end_date.month - start_date.month + 1
    )
    if month_count > 12:
        raise ValueError("单次最多查询连续12个自然月")
    if not coverage_start or not coverage_end:
        raise ValueError("eBay销量或售后数据尚未上传完整")
    if start_date < coverage_start or end_date > coverage_end:
        raise ValueError(
            f"eBay销量与售后共同覆盖范围为{coverage_start}至{coverage_end}"
        )
    month_start = date(start_date.year, start_date.month, 1)
    next_month = date(
        end_date.year + (end_date.month == 12),
        end_date.month % 12 + 1,
        1,
    )
    month_end = date.fromordinal(next_month.toordinal() - 1)
    expected_start = max(coverage_start, month_start)
    expected_end = min(coverage_end, month_end)
    if start_date != expected_start or end_date != expected_end:
        raise ValueError(
            "请按完整自然月区间查询；当前选择应转换为"
            f"{expected_start}至{expected_end}"
        )
    if repo.has_partial_monthly_history(start_date, end_date):
        raise ValueError(
            "所选日期包含仅有月汇总销量的历史月份，请从月初选到月末；"
            "2026年8月及后续日明细不受此限制"
        )


def _normalize_rg_sku(value: str) -> tuple[str, Decimal]:
    sku = _text(value, 255)
    marker = re.search(r"(?i)(?:\+|\*)(\d{1,3})RG$", sku)
    if marker:
        factor = Decimal(marker.group(1))
        if factor > 0:
            return sku[:marker.start()].rstrip("+*"), factor
    suffix = re.search(r"(?i)\+RG$", sku)
    if suffix:
        base = sku[:suffix.start()]
        # 数字以独立短尾码出现在 +RG 前时作为倍数；普通料号长尾码
        # （例如 -0740+RG）不能误判成 740 倍。
        factor_match = re.search(r"(?:^|[-_*])(\d{1,2})$", base)
        factor = Decimal(factor_match.group(1)) if factor_match else Decimal("1")
        return base, factor if factor > 0 else Decimal("1")
    return sku, Decimal("1")


def _data_source_for_currency(value: Any) -> str:
    return {
        "USD": "eBay-US",
        "GBP": "eBay-UK",
        "EUR": "eBay-DE",
    }.get(_text(value).upper(), "eBay-OTHER")


def _normalize_data_source(value: Any) -> str:
    normalized = re.sub(r"[^a-z]", "", _text(value).lower())
    if normalized in {"ebayus", "ebayusa"}:
        return "eBay-US"
    if normalized in {"ebayuk", "ebaygb"}:
        return "eBay-UK"
    if normalized in {"ebayde", "ebaygermany"}:
        return "eBay-DE"
    return _text(value, 50) or "eBay-OTHER"


def _valid_category(big_value: Any, small_value: Any) -> tuple[str, str]:
    big = _text(big_value, 50)
    small = _text(small_value, 100)
    if big not in BIG_CATEGORIES or not small:
        return "其他", "其他"
    return big, small


def _source_text(values: dict[str, Decimal]) -> str:
    order = {"eBay-US": 0, "eBay-UK": 1, "eBay-DE": 2, "eBay-OTHER": 3}
    return "；".join(
        f"{source}:{_decimal_text(quantity)}"
        for source, quantity in sorted(values.items(), key=lambda item: (order.get(item[0], 99), item[0]))
    )


def _parse_source_text(value: Any) -> dict[str, Decimal]:
    result: dict[str, Decimal] = defaultdict(Decimal)
    for part in _text(value).replace(";", "；").split("；"):
        if ":" not in part:
            continue
        source, raw_quantity = part.rsplit(":", 1)
        if source.strip():
            result[source.strip()] += _decimal(raw_quantity)
    return result


def _decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def _month_bounds(value: Any) -> tuple[date, date] | None:
    if isinstance(value, datetime):
        year, month = value.year, value.month
    elif isinstance(value, date):
        year, month = value.year, value.month
    else:
        text = _text(value)
        match = re.search(r"(20\d{2})\s*(?:年|[-/.])\s*(\d{1,2})", text)
        if not match:
            return None
        year, month = int(match.group(1)), int(match.group(2))
    if not 1 <= month <= 12:
        return None
    month_start = date(year, month, 1)
    next_month = date(year + (month == 12), month % 12 + 1, 1)
    return month_start, date.fromordinal(next_month.toordinal() - 1)


def _datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            return parsed if isinstance(parsed, datetime) else datetime.combine(parsed, datetime.min.time())
        except Exception:
            return None
    text = _text(value)
    if not text:
        return None
    for pattern in (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def _decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value if value is not None else "0").replace(",", "").strip() or "0")
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _sha(*values: Any) -> str:
    payload = "|".join(str(value or "").strip() for value in values)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _text(value: Any, limit: int | None = None) -> str:
    result = str(value or "").strip()
    return result[:limit] if limit else result
