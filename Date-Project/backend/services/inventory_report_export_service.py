from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.services.inventory_report_etl_service import (
    get_department_summary,
    get_dimension_summary,
)


EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
DIMENSION_LABELS = {"GROUP": "组别", "STORE": "店铺", "OWNER": "负责人"}
Header = tuple[str, Callable[[dict[str, Any]], Any], str]


def export_monthly_inventory_report(
    stat_month: str | None,
    dimension_type: str,
) -> tuple[str, bytes]:
    """按维度导出月度库存报表；页面与Excel共用后端计算结果。"""
    dimension = str(dimension_type or "").strip().upper()
    if dimension not in DIMENSION_LABELS:
        raise ValueError("dimension_type必须是GROUP、STORE或OWNER")
    if dimension == "GROUP":
        data = get_department_summary(stat_month)
        rows = list(data.get("items") or [])
        headers = _group_headers(data.get("report_month"))
    else:
        data = get_dimension_summary(dimension, stat_month)
        detail_rows = list(data.get("items") or [])
        total = data.get("total")
        rows = ([total] if total else []) + detail_rows
        headers = _dimension_headers(dimension)
    if not rows:
        report_month = data.get("report_month") or stat_month or "当前月份"
        raise ValueError(f"{report_month} 没有可导出的月度库存数据")

    workbook = Workbook(write_only=True)
    _append_sheet(
        workbook,
        f"月度库存-{DIMENSION_LABELS[dimension]}",
        headers,
        rows,
    )
    output = BytesIO()
    workbook.save(output)
    report_month = data.get("report_month") or stat_month or "当前月份"
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = (
        f"{report_month}-月度库存-{DIMENSION_LABELS[dimension]}-"
        f"{timestamp}.xlsx"
    )
    return filename, output.getvalue()


def _group_headers(report_month: str | None) -> list[Header]:
    business_label = _month_label(report_month, "当月")
    next_label = _month_label(_next_month(report_month), "次月")
    return [
        ("组别", _field("department_name"), "text"),
        ("总货值", _field("total_goods_value"), "money"),
        ("本地仓-期末在途数量", _field("local_end_in_transit_qty"), "qty"),
        ("本地仓-期末在途总成本", _field("local_end_in_transit_total_cost"), "money"),
        ("本地仓-期末库存数量", _field("local_end_inventory_qty"), "qty"),
        ("本地仓-期末库存总成本", _field("local_end_inventory_total_cost"), "money"),
        ("海外仓/FBA仓-期末在途数量", _combined("overseas_end_in_transit_qty", "fba_end_in_transit_qty"), "qty"),
        ("海外仓/FBA仓-期末在途总成本", _combined("overseas_end_in_transit_total_cost", "fba_end_in_transit_total_cost"), "money"),
        ("海外仓/FBA仓-期末库存数量", _combined("overseas_end_inventory_qty", "fba_end_inventory_qty"), "qty"),
        ("海外仓/FBA仓-期末库存总成本", _combined("overseas_end_inventory_total_cost", "fba_end_inventory_total_cost"), "money"),
        ("FBA在途金额+FBA在库金额", _field("fba_transit_inventory_amount"), "money"),
        ("库存健康度", _field("inventory_health_rate"), "percent"),
        ("销售目标（USD）", _field("sales_target_usd"), "money"),
        ("实际达成（USD）", _field("actual_achievement_amount_usd"), "money"),
        ("目标达成率", _field("target_achievement_rate"), "percent"),
        (f"{next_label}周转天数（货值）", _field("turnover_days_by_value"), "decimal"),
        (f"{business_label}初库存数量", _field("next_month_opening_inventory_qty"), "qty"),
        (f"{business_label}销量", _field("monthly_sales_qty"), "qty"),
        (f"{next_label}初库销比", _field("opening_inventory_sales_ratio"), "decimal"),
        (f"{business_label}周转天数（SKU）", _field("turnover_days_by_sku"), "decimal"),
        ("成都仓30天以上货值", _field("ctu_over_30_cost"), "money"),
        ("90-180库龄成本", _field("inventory_age_90_180_cost"), "money"),
        ("180+库龄成本", _field("inventory_age_180_plus_cost"), "money"),
    ]


def _dimension_headers(dimension: str) -> list[Header]:
    first_title = "店铺" if dimension == "STORE" else "负责人"
    return [
        (first_title, _dimension_name(dimension), "text"),
        ("平台", _platform_name, "text"),
        ("组别", _field("department_code"), "text"),
        ("总货值", _field("total_goods_value"), "money"),
        ("海外仓/FBA仓-期末在途数量", _combined("overseas_end_in_transit_qty", "fba_end_in_transit_qty"), "qty"),
        ("海外仓/FBA仓-期末在途总成本", _combined("overseas_end_in_transit_total_cost", "fba_end_in_transit_total_cost"), "money"),
        ("海外仓/FBA仓-期末库存数量", _combined("overseas_end_inventory_qty", "fba_end_inventory_qty"), "qty"),
        ("海外仓/FBA仓-期末库存总成本", _combined("overseas_end_inventory_total_cost", "fba_end_inventory_total_cost"), "money"),
        ("FBA在途金额+FBA在库金额", _field("fba_transit_inventory_amount"), "money"),
        ("库存健康度", _field("inventory_health_rate"), "percent"),
        ("销售目标（USD）", _field("sales_target_usd"), "money"),
        ("实际达成（USD）", _field("actual_achievement_amount_usd"), "money"),
        ("目标达成率", _field("target_achievement_rate"), "percent"),
    ]


def _append_sheet(
    workbook: Workbook,
    title: str,
    headers: list[Header],
    rows: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True)
    header_cells = []
    for column_index, (column_title, _accessor, _value_type) in enumerate(
        headers, start=1
    ):
        sheet.column_dimensions[get_column_letter(column_index)].width = max(
            14, min(32, len(column_title) * 2 + 4)
        )
        cell = WriteOnlyCell(sheet, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cells.append(cell)
    sheet.append(header_cells)

    for row in rows:
        cells = []
        for _column_title, accessor, value_type in headers:
            value = accessor(row)
            cell = WriteOnlyCell(
                sheet,
                value=_excel_value(value, value_type),
            )
            if value_type == "qty":
                cell.number_format = "#,##0.######"
            elif value_type == "money":
                cell.number_format = "#,##0.00"
            elif value_type == "percent":
                cell.number_format = "0.00%"
            elif value_type == "decimal":
                cell.number_format = "#,##0.00"
            cells.append(cell)
        sheet.append(cells)
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    )


def _field(key: str) -> Callable[[dict[str, Any]], Any]:
    return lambda row: row.get(key)


def _combined(
    left_key: str,
    right_key: str,
) -> Callable[[dict[str, Any]], Any]:
    return lambda row: _decimal(row.get(left_key)) + _decimal(
        row.get(right_key)
    )


def _dimension_name(
    dimension: str,
) -> Callable[[dict[str, Any]], str]:
    def value(row: dict[str, Any]) -> str:
        if int(row.get("is_dimension_total") or 0) == 1:
            return "合计（仅Amazon FBA）" if dimension == "STORE" else "合计"
        return str(row.get("dimension_value") or "")

    return value


def _platform_name(row: dict[str, Any]) -> str:
    platform = str(row.get("platform_code") or "").upper()
    if platform == "EBAY":
        return "eBay"
    return "Amazon" if platform == "AMZ" else ""


def _excel_value(value: Any, value_type: str) -> Any:
    if value is None or value == "":
        return None
    if value_type == "text":
        return str(value)
    return float(_decimal(value))


def _decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _next_month(month: str | None) -> str | None:
    if not month or len(month) != 7:
        return None
    year, month_number = (int(part) for part in month.split("-", 1))
    if month_number == 12:
        return f"{year + 1:04d}-01"
    return f"{year:04d}-{month_number + 1:02d}"


def _month_label(month: str | None, fallback: str) -> str:
    if not month or len(month) != 7:
        return fallback
    return f"{int(month[5:7])}月"