from __future__ import annotations

import tempfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.config import settings
from backend.repositories import clearance_repository as repo


HEADERS = [
    ("快照月份", "pull_month", "text"),
    ("区域", "region_name", "text"),
    ("组别", "group_code", "text"),
    ("店铺名称", "store_name", "text"),
    ("共享店铺列表", "shared_store_names", "text"),
    ("MSKU", "seller_sku", "text"),
    ("SKU", "sku", "text"),
    ("0-30天数量", "inv_age_0_to_30_days", "qty"),
    ("0-30天成本", "inv_age_0_to_30_price", "money"),
    ("31-60天数量", "inv_age_31_to_60_days", "qty"),
    ("31-60天成本", "inv_age_31_to_60_price", "money"),
    ("61-90天数量", "inv_age_61_to_90_days", "qty"),
    ("61-90天成本", "inv_age_61_to_90_price", "money"),
    ("0-90天数量", "inv_age_0_to_90_days", "qty"),
    ("0-90天成本", "inv_age_0_to_90_price", "money"),
    ("91-180天数量", "inv_age_91_to_180_days", "qty"),
    ("91-180天成本", "inv_age_91_to_180_price", "money"),
    ("181-270天数量", "inv_age_181_to_270_days", "qty"),
    ("181-270天成本", "inv_age_181_to_270_price", "money"),
    ("271-330天数量", "inv_age_271_to_330_days", "qty"),
    ("271-330天成本", "inv_age_271_to_330_price", "money"),
    ("271-365天数量", "inv_age_271_to_365_days", "qty"),
    ("271-365天成本", "inv_age_271_to_365_price", "money"),
    ("331-365天数量", "inv_age_331_to_365_days", "qty"),
    ("331-365天成本", "inv_age_331_to_365_price", "money"),
    ("365天以上数量", "inv_age_365_plus_days", "qty"),
    ("365天以上成本", "inv_age_365_plus_price", "money"),
    ("拉取时间", "pulled_at", "datetime"),
    ("同步批次ID", "sync_batch_id", "text"),
]


def export_inventory_age_details(pull_month: str | None) -> tuple[str, str]:
    data = repo.inventory_age_details(pull_month)
    month = data["pull_month"]
    rows = data["items"]
    if not month or not rows:
        raise ValueError(f"{pull_month or '当前月份'} 没有可导出的库龄明细")

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title="库龄明细")
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    widths = [12, 12, 12, 28, 62, 28, 24] + [20] * 20 + [20, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True)
    header_cells = []
    for title, _, _ in HEADERS:
        cell = WriteOnlyCell(sheet, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cells.append(cell)
    sheet.append(header_cells)

    for row in rows:
        cells = []
        for _, key, value_type in HEADERS:
            value = row.get(key)
            if key == "store_name" and not str(value or "").strip():
                value = "0"
            cell = WriteOnlyCell(sheet, value=_excel_value(value))
            if value_type == "qty":
                cell.number_format = "#,##0"
            elif value_type == "money":
                cell.number_format = "#,##0.00"
            elif value_type == "datetime":
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            cells.append(cell)
        sheet.append(cells)

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"
    export_dir = Path(settings.export_output_dir)
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    download_name = f"{month}-库龄明细-{timestamp}.xlsx"
    temp = tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        prefix=f"inventory_age_detail_{month}_",
        dir=str(export_dir),
        delete=False,
    )
    temp.close()
    workbook.save(temp.name)
    return temp.name, download_name


def _excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value
    return value
