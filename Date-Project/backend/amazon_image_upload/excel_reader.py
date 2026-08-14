"""Excel SKU 列表读取。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def read_skus(excel_path: str, sku_column: str = "SKU") -> list[str]:
    """读取 Excel 中指定列的 SKU，去重去空，保留出现顺序。

    Args:
        excel_path: .xlsx 文件路径
        sku_column: 列名（首行表头匹配），默认 "SKU"

    Returns:
        SKU 字符串列表
    """
    p = Path(excel_path)
    if not p.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        wb.close()
        return []

    # 找到目标列索引
    col_idx = None
    for i, cell in enumerate(header):
        if cell is not None and str(cell).strip().lower() == sku_column.strip().lower():
            col_idx = i
            break
    if col_idx is None:
        wb.close()
        raise ValueError(f"Excel 中找不到列名: {sku_column}")

    skus: list[str] = []
    seen: set[str] = set()
    for row in rows:
        if not row or col_idx >= len(row):
            continue
        val = row[col_idx]
        if val is None:
            continue
        sku = str(val).strip()
        if not sku or sku in seen:
            continue
        seen.add(sku)
        skus.append(sku)

    wb.close()
    return skus
