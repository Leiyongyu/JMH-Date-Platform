"""
eBay 价格抓取工具 — FastAPI 子应用
==================================
从 Flask 版 (app_api_v3.py) 转换而来，保持所有业务逻辑不变。

策略：两遍搜索 + 类目锁定（category lock）+ 可选 getItem 核验

依赖：
  pip install requests openpyxl fastapi python-multipart
"""

from __future__ import annotations

import os
import re
import tempfile
import threading
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path

import requests as _requests
import urllib3
from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ═══════════════════════ 凭据（环境变量）═══════════════════════

CLIENT_ID = os.getenv("EBAY_CLIENT_ID", "").strip()
CLIENT_SECRET = os.getenv("EBAY_CLIENT_SECRET", "").strip()

# ═══════════════════════ v3 策略配置 ═══════════════════════

CATEGORY_MIN_SHARE = 0.40
VERIFY_ENABLED = True
VERIFY_TOP_N = 5

# ═══════════════════════ SKU-OE 对应关系配置 ═══════════════════════

SKU_OE_FILE = str(Path(__file__).resolve().parent / "SKU主OE对照表.xlsx")
_sku_oe_map: dict[str, str] = {}
_middle_code_oe_map: dict[str, str] = {}
_custom_middle_codes: dict[str, str] = {}
_sku_oe_lock = threading.Lock()


def _extract_middle_code(sku: str) -> str:
    """从 SKU 中提取中间码：去掉第一个 '-' 前的前缀，如 HME-220292 → 220292"""
    if not sku:
        return ""
    idx = sku.find("-")
    if 0 < idx < len(sku) - 1:
        return sku[idx + 1:]
    return ""


def load_sku_oe_map() -> None:
    """从 Excel 文件加载 SKU-OE 对应关系（含中间码）"""
    global _sku_oe_map, _middle_code_oe_map, _custom_middle_codes
    with _sku_oe_lock:
        try:
            if not os.path.exists(SKU_OE_FILE):
                _sku_oe_map = {}
                _middle_code_oe_map = {}
                _custom_middle_codes = {}
                return

            wb = load_workbook(SKU_OE_FILE, read_only=True)
            ws = wb.active
            new_map: dict[str, str] = {}
            mc_map: dict[str, str] = {}
            custom_mc: dict[str, str] = {}

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            has_mc_col = (
                header_row
                and len(header_row) > 2
                and header_row[2]
                and "中间码" in str(header_row[2])
            )

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    sku = str(row[0]).strip()
                    oe = str(row[1]).strip()
                    if sku and oe:
                        new_map[sku] = oe
                        auto_mc = _extract_middle_code(sku)
                        if has_mc_col and len(row) > 2 and row[2]:
                            mc = str(row[2]).strip()
                            if mc and mc.upper() != (auto_mc or "").upper():
                                custom_mc[sku] = mc
                        else:
                            mc = auto_mc
                        if mc:
                            mc_map[mc.upper()] = oe

            wb.close()
            _sku_oe_map = new_map
            _middle_code_oe_map = mc_map
            _custom_middle_codes = custom_mc
        except Exception as e:
            print(f"加载 SKU-OE 对应表失败: {e}")


def save_sku_oe_map() -> bool:
    """保存 SKU-OE 对应关系到 Excel 文件（含中间码列）"""
    global _sku_oe_map
    with _sku_oe_lock:
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(["SKU", "主OE", "中间码"])
            for sku, oe in sorted(_sku_oe_map.items()):
                mc = _custom_middle_codes.get(sku) or _extract_middle_code(sku)
                ws.append([sku, oe, mc])
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 15
            wb.save(SKU_OE_FILE)
            wb.close()
            _reload_middle_code_map()
            return True
        except Exception as e:
            print(f"保存 SKU-OE 对应表失败: {e}")
            return False


def _reload_middle_code_map() -> None:
    """重建中间码映射。调用方需已持有 _sku_oe_lock"""
    global _middle_code_oe_map
    mc_map: dict[str, str] = {}
    for sku, oe in _sku_oe_map.items():
        mc = _custom_middle_codes.get(sku) or _extract_middle_code(sku)
        if mc:
            mc_map[mc.upper()] = oe
    _middle_code_oe_map = mc_map


def get_oe_by_sku(sku: str) -> str | None:
    """通过 SKU 或中间码查找 OE"""
    if not sku:
        return None
    key = sku.strip()
    oe = _sku_oe_map.get(key)
    if oe:
        return oe
    return _middle_code_oe_map.get(key.upper())


load_sku_oe_map()

# ── eBay 站点配置 ──
SITES: dict[str, dict] = {
    "de": {
        "domain": "ebay.de",
        "locale": "de-DE",
        "marketplace": "EBAY_DE",
        "currency": "EUR",
        "global_id": "EBAY-GERMANY",
    },
    "uk": {
        "domain": "ebay.co.uk",
        "locale": "en-GB",
        "marketplace": "EBAY_GB",
        "currency": "GBP",
        "global_id": "EBAY-GB",
    },
    "us": {
        "domain": "ebay.com",
        "locale": "en-US",
        "marketplace": "EBAY_US",
        "currency": "USD",
        "global_id": "EBAY-US",
    },
}

# ── OAuth2 Token 缓存 ──
_token_cache: dict = {"token": None, "expires": 0}
_token_lock = threading.Lock()

# ── API 端点 ──
OAUTH_URL = "https://api.ebay.com/identity/v1/oauth2/token"
BROWSE_URL = "https://api.ebay.com/buy/browse/v1/item_summary/search"
ITEM_URL = "https://api.ebay.com/buy/browse/v1/item/"


def get_access_token() -> str:
    """获取或刷新 App 级别的 OAuth2 Access Token"""
    with _token_lock:
        if _token_cache["token"] and time.time() < _token_cache["expires"]:
            return _token_cache["token"]

        resp = _requests.post(
            OAUTH_URL,
            data={
                "grant_type": "client_credentials",
                "scope": "https://api.ebay.com/oauth/api_scope",
            },
            auth=(CLIENT_ID, CLIENT_SECRET),
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=15,
            verify=False,
        )

        if resp.status_code != 200:
            raise RuntimeError(
                f"获取 eBay Token 失败 (HTTP {resp.status_code})：{resp.text}\n"
                f"请检查环境变量 EBAY_CLIENT_ID 和 EBAY_CLIENT_SECRET 是否正确。"
            )

        data = resp.json()
        _token_cache["token"] = data["access_token"]
        _token_cache["expires"] = time.time() + data.get("expires_in", 7200) - 60
        return _token_cache["token"]


def _api_get(
    url: str, params: dict, marketplace: str, retry_on_401: bool = True
) -> tuple[dict | None, str | None]:
    """带鉴权的 GET 请求，401 时自动刷新 token 重试一次"""
    token = get_access_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "X-EBAY-C-MARKETPLACE-ID": marketplace,
        "X-EBAY-C-ENDUSERCTX": (
            "affiliateCampaignId=<ePNCampaignId>,"
            "affiliateReferenceId=<referenceId>"
        ),
    }
    resp = _requests.get(
        url, params=params, headers=headers, timeout=30, verify=False
    )

    if resp.status_code == 200:
        return resp.json(), None

    if resp.status_code == 401 and retry_on_401:
        _token_cache["token"] = None
        token = get_access_token()
        headers["Authorization"] = f"Bearer {token}"
        resp = _requests.get(
            url, params=params, headers=headers, timeout=30, verify=False
        )
        if resp.status_code == 200:
            return resp.json(), None

    return None, (
        f"eBay API 返回错误 (HTTP {resp.status_code}): {resp.text[:300]}"
    )


def search_ebay(
    keyword: str,
    marketplace: str,
    limit: int = 100,
    offset: int = 0,
    sort: str = "",
    category_id: str = "",
    fixed_price_only: bool = False,
    fieldgroups: str = "",
    free_shipping: bool = False,
) -> tuple[dict | None, str | None]:
    """
    调用 eBay Browse API 搜索商品
    sort:             "" 默认(Best Match)，"price" 价格升序
    category_id:      限定叶子类目（v3 类目锁定用）
    fixed_price_only: 只保留一口价（排除拍卖，v3 第二遍搜索用）
    fieldgroups:      传 "MATCHING_ITEMS,CATEGORY_REFINEMENTS" 可拿 refinement 类目分布
    free_shipping:    仅返回免运费商品
    """
    filters = ["conditions:{NEW}"]
    if fixed_price_only:
        filters.append("buyingOptions:{FIXED_PRICE}")
    if free_shipping:
        filters.append("deliveryOptions:{FREE_SHIPPING}")

    params: dict = {
        "q": keyword,
        "limit": min(limit, 200),
        "offset": offset,
        "filter": ",".join(filters),
    }
    if sort:
        params["sort"] = sort
    if category_id:
        params["category_ids"] = category_id
    if fieldgroups:
        params["fieldgroups"] = fieldgroups

    return _api_get(BROWSE_URL, params, marketplace)


def get_item_product_type(item_id: str, marketplace: str) -> str:
    """
    调 getItem 接口，读取 localizedAspects 中的产品类型字段
    用于 v3 可选核验：确认最低价候选确实是整件而非零件
    """
    if not item_id:
        return ""
    data, error = _api_get(ITEM_URL + item_id, {}, marketplace)
    if error or not data:
        return ""
    for aspect in data.get("localizedAspects", []):
        name = (aspect.get("name") or "").strip().lower()
        if name in ("produktart", "product type", "type"):
            return (aspect.get("value") or "").strip()
    return ""


def _clean_item_url(url: str) -> str:
    """清理 eBay 商品链接，去掉追踪参数"""
    if not url:
        return ""
    idx = url.find("/itm/")
    if idx < 0:
        return url
    after = url[idx + 5:]
    m = re.search(r"[?_#]", after)
    return url[: idx + 5] + (after[: m.start()] if m else after)


def _upgrade_img(url: str) -> str:
    """将图片 URL 的尺寸升级到 s-l500"""
    return re.sub(r"s-l\d+", "s-l500", url) if url else url


def format_item(item: dict) -> dict:
    """将 Browse API 返回的单条 item 标准化"""
    title = item.get("title", "")
    price_info = item.get("price", {})
    price_value = float(price_info.get("value", 0))
    currency = price_info.get("currency", "EUR")
    price_str = f"{price_value:.2f} {currency}"

    condition = item.get("condition", "")
    condition_id = item.get("conditionId", "")

    images: list[str] = []
    seen_urls: set[str] = set()
    for src_list in [
        item.get("thumbnailImages", []),
        item.get("additionalImages", []),
    ]:
        for img in src_list:
            u = img.get("imageUrl", "")
            if u and u not in seen_urls:
                seen_urls.add(u)
                images.append(_upgrade_img(u))
    if not images:
        primary = item.get("image", {})
        if primary and primary.get("imageUrl"):
            images.append(_upgrade_img(primary["imageUrl"]))
    images = images[:5]

    item_web_url = _clean_item_url(item.get("itemWebUrl", ""))
    item_id = item.get("itemId", "")

    seller = item.get("seller", {})
    seller_username = seller.get("username", "")
    seller_feedback = seller.get("feedbackPercentage", "")

    buying_options = item.get("buyingOptions", [])
    shipping_options = item.get("shippingOptions", [])
    shipping_cost = ""
    if shipping_options:
        ship_price = shipping_options[0].get("shippingCost", {})
        if ship_price:
            ship_val = float(ship_price.get("value", 0))
            ship_curr = ship_price.get("currency", currency)
            shipping_cost = (
                f"{ship_val:.2f} {ship_curr}" if ship_val > 0 else "免运费"
            )

    # 从 estimatedAvailabilities 提取预计已售数量
    estimated_sold = None
    for avail in item.get("estimatedAvailabilities") or []:
        qty = avail.get("estimatedSoldQuantity")
        if qty is not None:
            try:
                parsed = max(int(qty), 0)
                estimated_sold = max(parsed, estimated_sold or 0)
            except (ValueError, TypeError):
                pass

    return {
        "title": title,
        "price": price_str,
        "pf": price_value,
        "currency": currency,
        "condition": condition,
        "conditionId": condition_id,
        "categoryId": item.get("categoryId", ""),
        "images": images,
        "link": item_web_url,
        "itemId": item_id,
        "seller": seller_username,
        "sellerFeedback": seller_feedback,
        "shipping": shipping_cost,
        "buyingOptions": buying_options,
        "estimatedSold": estimated_sold,
    }


def parse_oe_list(text: str) -> list[str]:
    """解析用户输入的 OE 号列表"""
    lines: list[str] = []
    for line in text.replace(",", "\n").replace("，", "\n").split("\n"):
        oe = line.strip()
        if oe:
            lines.append(oe)
    seen: set[str] = set()
    result: list[str] = []
    for oe in lines:
        if oe.upper() not in seen:
            seen.add(oe.upper())
            result.append(oe)
    return result


def _pick_dominant_category(
    data: dict, items_raw: list[dict]
) -> tuple[str, str, float, list[dict]]:
    """
    从搜索响应中选出主导类目
    优先读 refinement.categoryDistributions（eBay 官方类目分布）；
    若 API 未返回 refinement 数据，则从 items 自身的 categoryId 投票统计。
    返回 (category_id, category_name, share, dist_top)
    dist_top 为类目分布前 5 名 [{id, name, count}]，用于前端诊断展示；
    类目不可信时 category_id 为 ""
    """
    refinement = data.get("refinement", {}) or {}
    dists = refinement.get("categoryDistributions", []) or []

    # ── 路径 A：API 返回了 refinement 数据 ──
    if dists:
        dist_top = [
            {
                "id": d.get("categoryId", ""),
                "name": d.get("categoryName", ""),
                "count": d.get("matchCount", 0),
            }
            for d in sorted(
                dists, key=lambda x: x.get("matchCount", 0), reverse=True
            )[:5]
        ]
        total = sum(d.get("matchCount", 0) for d in dists)
        top = max(dists, key=lambda d: d.get("matchCount", 0))
        top_count = top.get("matchCount", 0)
        share = (top_count / total) if total > 0 else 0.0
        if top.get("categoryId") and share >= CATEGORY_MIN_SHARE:
            return top["categoryId"], top.get("categoryName", ""), share, dist_top
        return "", "", share, dist_top

    # 部分响应没有 categoryDistributions，但有 dominantCategoryId
    dom_id = refinement.get("dominantCategoryId", "")
    if dom_id:
        return dom_id, "", 1.0, []

    # ── 路径 B：refinement 完全缺失，从 items 自身统计类目 ──
    cat_counter: Counter[str] = Counter()
    cat_name_map: dict[str, str] = {}
    for item in items_raw or []:
        cid = ""
        cname = ""
        cats = item.get("categories", []) or []
        if cats:
            leaf = cats[-1]
            cid = leaf.get("categoryId", "")
            cname = leaf.get("categoryName", "")
        if not cid:
            leaf_ids = item.get("leafCategoryIds", []) or []
            if leaf_ids:
                cid = leaf_ids[0]
        if cid:
            cat_counter[cid] += 1
            if cname and cid not in cat_name_map:
                cat_name_map[cid] = cname

    if not cat_counter:
        return "", "", 0.0, []

    dist_top = [
        {"id": cid, "name": cat_name_map.get(cid, ""), "count": cnt}
        for cid, cnt in cat_counter.most_common(5)
    ]

    top_cid, top_count = cat_counter.most_common(1)[0]
    total = sum(cat_counter.values())
    share = top_count / total if total > 0 else 0.0

    if share >= CATEGORY_MIN_SHARE:
        return top_cid, cat_name_map.get(top_cid, ""), share, dist_top
    return "", "", share, dist_top


def _verify_top_items(items: list[dict], marketplace: str) -> list[dict]:
    """
    可选核验：对最便宜的前 VERIFY_TOP_N 条调 getItem 取 Produktart，
    与多数派类型不符的标记 suspect 并排到队尾（不删除，保留人工判断空间）
    """
    if not VERIFY_ENABLED or not items:
        return items

    check_n = min(VERIFY_TOP_N, len(items))
    head = items[:check_n]
    tail = items[check_n:]

    types: list[str] = []
    for it in head:
        ptype = get_item_product_type(it["itemId"], marketplace)
        it["productType"] = ptype
        types.append(ptype.lower() if ptype else "")

    known = [t for t in types if t]
    if len(known) < 2:
        return items

    modal_type = max(set(known), key=known.count)
    if known.count(modal_type) < 2:
        return items

    normal: list[dict] = []
    suspect: list[dict] = []
    for it, t in zip(head, types):
        if t and t != modal_type:
            it["suspect"] = True
            suspect.append(it)
        else:
            normal.append(it)

    return normal + suspect + tail


def _finalize_oe(
    items_raw: list[dict],
    oe: str,
    strategy: str,
    mp: str,
    cat_info: dict | None,
    verify: bool = False,
) -> tuple[list[dict], None, str, dict | None, list[float]]:
    """公共收尾：format → 设 oe → 排序 → 可选核验 → 截断前 20"""
    result = [format_item(it) for it in items_raw]
    for it in result:
        it["oe"] = oe
    result.sort(key=lambda x: x["pf"])
    all_prices = [it["pf"] for it in result]
    if verify:
        result = _verify_top_items(result, mp)
    return result[:20], None, strategy, cat_info, all_prices


def _raw_cat_id(item: dict) -> str:
    """从原始 item 取叶子类目 ID（与 _pick_dominant_category 路径 B 逻辑一致）"""
    cats = item.get("categories") or []
    if cats:
        return cats[-1].get("categoryId", "")
    leaf_ids = item.get("leafCategoryIds") or []
    return leaf_ids[0] if leaf_ids else ""


def scrape_one_oe(
    oe: str,
    site_key: str,
    free_shipping: bool = False,
    search_strategy: str = "category_lock",
) -> tuple[list[dict], str | None, int, str, dict | None, list[float]]:
    """
    搜索单个 OE 号
    search_strategy:
      best_match     — 单遍搜索，Best Match 排序，本地按价格取前 20（v1 策略，相关性优先）
      category_lock  — 两遍搜索 + 类目锁定 + getItem 核验（v3 策略，价格精确优先）
    返回 (items, warning, sold_total, strategy, cat_info, all_prices)
    """
    cfg = SITES[site_key]
    mp = cfg["marketplace"]

    # ═══ Best Match 策略：单遍搜索，相关性优先 ═══
    if search_strategy == "best_match":
        data, error = search_ebay(oe, mp, limit=50, free_shipping=free_shipping)
        if error:
            return [], error, 0, "error", None, []

        items_raw = data.get("itemSummaries", [])
        if not items_raw:
            return (
                [],
                f"OE '{oe}' 在 {cfg['domain']} 上未找到结果",
                0,
                "error",
                None,
                [],
            )

        sold_total = sum(item.get("soldQuantity", 0) or 0 for item in items_raw)
        items, _, _, _, all_prices = _finalize_oe(
            items_raw, oe, "best_match", mp, None
        )
        return items, None, sold_total, "best_match", None, all_prices

    # ═══ 类目锁定策略：两遍搜索 + 类目锁定 ═══
    data1, error = search_ebay(
        oe,
        mp,
        limit=100,
        fieldgroups="MATCHING_ITEMS,CATEGORY_REFINEMENTS",
        free_shipping=free_shipping,
    )
    if error:
        return [], error, 0, "error", None, []

    items_raw1 = data1.get("itemSummaries", [])
    if not items_raw1:
        return (
            [],
            f"OE '{oe}' 在 {cfg['domain']} 上未找到结果",
            0,
            "error",
            None,
            [],
        )

    sold_total = sum(item.get("soldQuantity", 0) or 0 for item in items_raw1)

    cat_id, cat_name, share, dist_top = _pick_dominant_category(data1, items_raw1)
    cat_info = {
        "id": cat_id,
        "name": cat_name,
        "share": round(share, 3),
        "top": dist_top,
    }

    # ═══ 类目不可信 → 退回全局 Best Match（只取相关性最高的前 40 条） ═══
    if not cat_id:
        items, _, strategy, ci, all_prices = _finalize_oe(
            items_raw1[:40], oe, "best_match_fallback", mp, cat_info
        )
        return items, None, sold_total, strategy, ci, all_prices

    # ═══ 第二遍：锁类目 + 价格升序 + 仅一口价 ═══
    data2, error2 = search_ebay(
        oe,
        mp,
        limit=50,
        sort="price",
        category_id=cat_id,
        fixed_price_only=True,
        free_shipping=free_shipping,
    )
    items_raw2 = (
        (data2 or {}).get("itemSummaries", []) if not error2 else []
    )

    if items_raw2:
        items, _, strategy, ci, all_prices = _finalize_oe(
            items_raw2, oe, "category_locked", mp, cat_info, verify=True
        )
        return items, None, sold_total, strategy, ci, all_prices

    # ═══ 第二遍为空 → 退回第一遍中属于整件类目的商品 ═══
    pool = [it for it in items_raw1 if _raw_cat_id(it) == cat_id]
    if pool:
        items, _, strategy, ci, all_prices = _finalize_oe(
            pool, oe, "category_pool_fallback", mp, cat_info, verify=True
        )
        return items, None, sold_total, strategy, ci, all_prices

    # 类目 id 与第一遍摘要对不上（少见），退回全局 Best Match
    items, _, strategy, ci, all_prices = _finalize_oe(
        items_raw1[:40], oe, "best_match_fallback", mp, cat_info
    )
    return items, None, sold_total, strategy, ci, all_prices


# ═══════════════════════ 批量搜索（异步任务） ═══════════════════════

tasks: dict[str, dict] = {}


def search_batch(
    oe_list: list[str],
    site_key: str,
    task_id: str,
    free_shipping: bool = False,
    search_strategy: str = "category_lock",
) -> None:
    """后台线程：3 并发搜索 OE 号"""
    total = len(oe_list)
    all_results: list[dict | None] = []
    completed = 0
    lock = threading.Lock()

    def do_one(oe: str) -> dict:
        items, warning, sold, strategy, cat_info, all_prices = scrape_one_oe(
            oe, site_key, free_shipping=free_shipping, search_strategy=search_strategy
        )
        return {
            "oe": oe,
            "count": len(items),
            "items": items,
            "warning": warning if not items else None,
            "soldTotal": sold,
            "strategy": strategy,
            "categoryInfo": cat_info,
            "allPrices": all_prices,
        }

    try:
        with ThreadPoolExecutor(max_workers=3) as pool:
            future_map = {pool.submit(do_one, oe): oe for oe in oe_list}
            oe_index = {oe: i for i, oe in enumerate(oe_list)}
            all_results = [None] * total

            for fut in as_completed(future_map):
                oe = future_map[fut]
                result = fut.result()
                all_results[oe_index[oe]] = result
                with lock:
                    completed += 1
                tasks[task_id] = {
                    "status": "running",
                    "progress": {
                        "current": completed,
                        "total": total,
                        "oe": oe,
                    },
                }

        tasks[task_id] = {
            "status": "done",
            "total_oe": total,
            "results": all_results,
        }

    except Exception as e:
        tasks[task_id] = {"status": "error", "msg": str(e)}


def cleanup_tasks() -> None:
    """清理已完成的旧任务（保留最近 20 条）"""
    terminal = [
        tid
        for tid, t in tasks.items()
        if t.get("status") in ("done", "error")
    ]
    if len(terminal) > 20:
        for tid in terminal[:-20]:
            del tasks[tid]


# ═══════════════════════ 路由 ═══════════════════════

router = APIRouter()


def _start_search_task(
    oe_list: list[str],
    site: str,
    free_shipping: bool,
    search_strategy: str,
    sku_mapping: dict | None = None,
    not_found_skus: list[str] | None = None,
    file_info: dict | None = None,
) -> dict:
    """公共逻辑：启动后台搜索任务并返回响应"""
    oe_list = list(dict.fromkeys(oe_list))
    if not oe_list:
        raise HTTPException(status_code=400, detail="未能识别有效的产品编号")

    task_id = str(int(time.time() * 1000))[-8:]
    tasks[task_id] = {
        "status": "running",
        "progress": {"current": 0, "total": len(oe_list), "oe": "启动中..."},
    }

    t = threading.Thread(
        target=search_batch,
        args=(oe_list, site, task_id, free_shipping, search_strategy),
    )
    t.daemon = True
    t.start()

    resp: dict = {
        "task_id": task_id,
        "oe_list": oe_list,
        "sku_mapping": sku_mapping or {},
        "not_found_skus": not_found_skus or [],
    }
    if file_info:
        resp["file_info"] = file_info
    return resp


def _resolve_oe_list(
    input_list: list[str], input_type: str
) -> tuple[list[str], dict[str, str], list[str]]:
    """根据输入类型解析 OE 列表，返回 (oe_list, sku_mapping, not_found_skus)"""
    oe_list: list[str] = []
    sku_mapping: dict[str, str] = {}
    not_found_skus: list[str] = []

    if input_type == "oe":
        oe_list = input_list
    elif input_type == "sku":
        with _sku_oe_lock:
            for item in input_list:
                oe = _sku_oe_map.get(item) or _middle_code_oe_map.get(
                    item.upper()
                )
                if oe:
                    oe_list.append(oe)
                    sku_mapping[item] = oe
                else:
                    not_found_skus.append(item)
    else:
        # auto：先查 SKU，再查中间码，都没有则当作 OE 直接搜
        with _sku_oe_lock:
            for item in input_list:
                oe = _sku_oe_map.get(item) or _middle_code_oe_map.get(
                    item.upper()
                )
                if oe:
                    oe_list.append(oe)
                    sku_mapping[item] = oe
                else:
                    oe_list.append(item)

    return oe_list, sku_mapping, not_found_skus


def _parse_excel_keywords_from_bytes(
    content: bytes, filename: str
) -> tuple[list[str], str, int, int, str]:
    """
    解析上传的 Excel 文件内容，提取 SKU / OE 列表。
    支持两种格式：
      一列表：A1 为 OE 相关表头，A2 起每行一个 OE。
      两列表：A1 为 SKU、B1 为 OE，数据从第 2 行开始。
    返回 (keywords, input_type, total_rows, skipped_rows, file_name)
    """
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Excel 文件没有工作表")

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Excel 文件为空")

    header = rows[0]
    if not header:
        raise ValueError("Excel 文件表头为空")

    def normalize_header(val: object) -> str:
        return re.sub(r"[\s_]", "", str(val or "")).lower()

    col_a = normalize_header(header[0]) if len(header) > 0 and header[0] else ""
    col_b = normalize_header(header[1]) if len(header) > 1 and header[1] else ""

    has_sku_col = col_a in ("sku", "skuid", "skuno")
    has_oe_col = col_b in ("oe", "oeno", "oenumber", "oenr", "partnumber", "partno")

    keywords: list[str] = []
    total_rows = 0
    skipped_rows = 0

    if has_sku_col and has_oe_col:
        for row in rows[1:]:
            total_rows += 1
            sku = str(row[0]).strip() if len(row) > 0 and row[0] else ""
            oe = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if oe:
                keywords.append(oe)
            elif sku:
                keywords.append(sku)
            else:
                skipped_rows += 1
        input_type = "auto"
    else:
        for row in rows[1:]:
            total_rows += 1
            val = str(row[0]).strip() if len(row) > 0 and row[0] else ""
            for part in re.split(r"[\r\n,，;；]+", val):
                part = part.strip()
                if part:
                    keywords.append(part)
            if not val:
                skipped_rows += 1
        input_type = "oe"

    # 去重保序
    seen: set[str] = set()
    unique_keywords: list[str] = []
    for kw in keywords:
        key_upper = kw.upper()
        if key_upper not in seen:
            seen.add(key_upper)
            unique_keywords.append(kw)

    return unique_keywords, input_type, total_rows, skipped_rows, filename


# ═══════════════════════ 搜索 API ═══════════════════════


@router.post("/api/scrape")
async def api_scrape(
    nkw: str = Form(""),
    site: str = Form("de"),
    input_type: str = Form("auto"),
    free_shipping: str = Form("false"),
    search_strategy: str = Form("category_lock"),
):
    cleanup_tasks()

    raw_text = nkw.strip()
    fs = free_shipping.strip().lower() == "true"
    ss = search_strategy.strip()
    if ss not in ("best_match", "category_lock"):
        ss = "category_lock"

    if site not in SITES:
        raise HTTPException(status_code=400, detail="不支持的站点")
    if not raw_text:
        raise HTTPException(status_code=400, detail="请输入至少一个产品编号")

    input_list = parse_oe_list(raw_text)
    if not input_list:
        raise HTTPException(status_code=400, detail="未能识别有效的产品编号")

    oe_list, sku_mapping, not_found_skus = _resolve_oe_list(
        input_list, input_type.strip()
    )
    return _start_search_task(
        oe_list, site, fs, ss, sku_mapping, not_found_skus
    )


@router.post("/api/scrape-file")
async def api_scrape_file(
    site: str = Form("de"),
    free_shipping: str = Form("false"),
    search_strategy: str = Form("category_lock"),
    file: UploadFile = File(...),
):
    """通过上传 Excel 文件批量查询"""
    cleanup_tasks()

    fs = free_shipping.strip().lower() == "true"
    ss = search_strategy.strip()
    if ss not in ("best_match", "category_lock"):
        ss = "category_lock"

    if site not in SITES:
        raise HTTPException(status_code=400, detail="不支持的站点")

    if not file.filename:
        raise HTTPException(status_code=400, detail="请选择要上传的 Excel 文件")

    fname_lower = file.filename.lower()
    if not (
        fname_lower.endswith(".xlsx")
        or fname_lower.endswith(".xlsm")
        or fname_lower.endswith(".xls")
    ):
        raise HTTPException(
            status_code=400, detail="仅支持 .xlsx、.xlsm、.xls 文件"
        )

    try:
        content = await file.read()
        keywords, input_type_val, total_rows, skipped_rows, file_name = (
            _parse_excel_keywords_from_bytes(content, file.filename)
        )
    except Exception as e:
        raise HTTPException(
            status_code=400, detail=f"Excel 解析失败: {str(e)}"
        )

    if not keywords:
        raise HTTPException(
            status_code=400, detail="Excel 中未读取到有效的 SKU 或 OE 数据"
        )

    oe_list, sku_mapping, not_found_skus = _resolve_oe_list(
        keywords, input_type_val
    )

    file_info = {
        "fileName": file_name,
        "totalRows": total_rows,
        "skippedRows": skipped_rows,
        "inputType": input_type_val,
    }
    return _start_search_task(
        oe_list, site, fs, ss, sku_mapping, not_found_skus, file_info
    )


@router.get("/api/status/{task_id}")
async def api_status(task_id: str):
    task = tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务未找到")
    if task["status"] == "running":
        return {"status": "running", "progress": task.get("progress")}
    if task["status"] == "error":
        return {"status": "error", "msg": task["msg"]}
    return {
        "status": "done",
        "total_oe": task["total_oe"],
        "results": task["results"],
    }


@router.get("/api/check")
async def api_check():
    configured = bool(
        CLIENT_ID
        and CLIENT_SECRET
        and CLIENT_ID != "YOUR_APP_ID_HERE"
        and CLIENT_SECRET != "YOUR_CERT_ID_HERE"
    )
    return {"configured": configured}


# ═══════════════════════ SKU-OE 管理 API ═══════════════════════


@router.get("/api/sku/list")
async def api_sku_list():
    with _sku_oe_lock:
        items = [
            {
                "sku": sku,
                "oe": oe,
                "middleCode": _custom_middle_codes.get(sku)
                or _extract_middle_code(sku),
            }
            for sku, oe in sorted(_sku_oe_map.items())
        ]
    return {"items": items, "total": len(items)}


@router.post("/api/sku/add")
async def api_sku_add(request: Request):
    data = await request.json()
    if not data:
        raise HTTPException(status_code=400, detail="缺少数据")

    sku = data.get("sku", "").strip()
    oe = data.get("oe", "").strip()
    middle_code = data.get("middleCode", "").strip()

    if not sku or not oe:
        raise HTTPException(status_code=400, detail="SKU 和 OE 都不能为空")

    with _sku_oe_lock:
        if sku in _sku_oe_map:
            raise HTTPException(
                status_code=400, detail=f"SKU '{sku}' 已存在"
            )
        _sku_oe_map[sku] = oe
        if middle_code:
            _custom_middle_codes[sku] = middle_code

    if save_sku_oe_map():
        return {"success": True, "sku": sku, "oe": oe, "middleCode": middle_code}
    else:
        with _sku_oe_lock:
            _sku_oe_map.pop(sku, None)
            _custom_middle_codes.pop(sku, None)
        raise HTTPException(status_code=500, detail="保存到 Excel 文件失败")


@router.post("/api/sku/update")
async def api_sku_update(request: Request):
    data = await request.json()
    if not data:
        raise HTTPException(status_code=400, detail="缺少数据")

    sku = data.get("sku", "").strip()
    oe = data.get("oe", "").strip()
    middle_code = data.get("middleCode", "").strip()

    if not sku or not oe:
        raise HTTPException(status_code=400, detail="SKU 和 OE 都不能为空")

    with _sku_oe_lock:
        if sku not in _sku_oe_map:
            raise HTTPException(
                status_code=400, detail=f"SKU '{sku}' 不存在"
            )
        old_oe = _sku_oe_map[sku]
        old_mc = _custom_middle_codes.get(sku)
        _sku_oe_map[sku] = oe
        if middle_code:
            _custom_middle_codes[sku] = middle_code
        else:
            _custom_middle_codes.pop(sku, None)

    if save_sku_oe_map():
        return {"success": True, "sku": sku, "oe": oe, "middleCode": middle_code}
    else:
        with _sku_oe_lock:
            _sku_oe_map[sku] = old_oe
            if old_mc is not None:
                _custom_middle_codes[sku] = old_mc
            else:
                _custom_middle_codes.pop(sku, None)
        raise HTTPException(status_code=500, detail="保存到 Excel 文件失败")


@router.post("/api/sku/delete")
async def api_sku_delete(request: Request):
    data = await request.json()
    if not data:
        raise HTTPException(status_code=400, detail="缺少数据")

    sku = data.get("sku", "").strip()
    if not sku:
        raise HTTPException(status_code=400, detail="缺少 SKU")

    with _sku_oe_lock:
        if sku not in _sku_oe_map:
            raise HTTPException(
                status_code=400, detail=f"SKU '{sku}' 不存在"
            )
        old_oe = _sku_oe_map.pop(sku)

    if save_sku_oe_map():
        return {"success": True, "sku": sku}
    else:
        with _sku_oe_lock:
            _sku_oe_map[sku] = old_oe
        raise HTTPException(status_code=500, detail="保存到 Excel 文件失败")


@router.post("/api/sku/lookup")
async def api_sku_lookup(request: Request):
    data = await request.json()
    if not data or "skus" not in data:
        raise HTTPException(status_code=400, detail="缺少 SKU 列表")

    skus = data["skus"]
    results: list[dict] = []
    not_found: list[str] = []

    with _sku_oe_lock:
        for sku in skus:
            sku = sku.strip()
            if sku:
                oe = _sku_oe_map.get(sku) or _middle_code_oe_map.get(
                    sku.upper()
                )
                if oe:
                    results.append({"sku": sku, "oe": oe})
                else:
                    not_found.append(sku)

    return {
        "found": results,
        "not_found": not_found,
        "total": len(results),
    }


@router.post("/api/sku/refresh")
async def api_sku_refresh():
    load_sku_oe_map()
    return {"success": True, "total": len(_sku_oe_map)}


@router.post("/api/sku/import")
async def api_sku_import(file: UploadFile = File(...)):
    """从 Excel 文件批量导入 SKU-OE 映射（覆盖式）"""
    global _sku_oe_map

    if not file.filename:
        raise HTTPException(
            status_code=400, detail="请选择要导入的 Excel 文件"
        )

    fname_lower = file.filename.lower()
    if not (
        fname_lower.endswith(".xlsx")
        or fname_lower.endswith(".xlsm")
        or fname_lower.endswith(".xls")
    ):
        raise HTTPException(
            status_code=400, detail="仅支持 .xlsx、.xlsm、.xls 文件"
        )

    try:
        content = await file.read()
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            raise HTTPException(
                status_code=400, detail="Excel 文件没有工作表"
            )

        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()

        if len(rows) < 2:
            raise HTTPException(
                status_code=400,
                detail="Excel 文件至少需要表头+1行数据",
            )

        header = rows[0]

        def norm(v: object) -> str:
            return re.sub(r"[\s_]", "", str(v or "")).lower()

        col_sku = col_oe = -1
        for i, h in enumerate(header):
            n = norm(h)
            if n in ("sku", "skuid", "skuno"):
                col_sku = i
            elif n in (
                "oe",
                "主oe",
                "主oeno",
                "oeno",
                "oenumber",
                "oenr",
                "partnumber",
                "partno",
            ):
                col_oe = i

        if col_sku < 0:
            col_sku = 0
        if col_oe < 0:
            col_oe = 1

        new_map: dict[str, str] = {}
        total_rows = 0
        skipped_rows = 0
        for row in rows[1:]:
            total_rows += 1
            sku = (
                str(row[col_sku]).strip()
                if len(row) > col_sku and row[col_sku]
                else ""
            )
            oe = (
                str(row[col_oe]).strip()
                if len(row) > col_oe and row[col_oe]
                else ""
            )
            if not sku or not oe:
                skipped_rows += 1
                continue
            new_map[sku] = oe

        if not new_map:
            raise HTTPException(
                status_code=400,
                detail="Excel 中未读取到有效的 SKU-OE 数据",
            )

        # 增量合并：没有的追加，有的覆盖
        with _sku_oe_lock:
            old_skus = set(_sku_oe_map.keys())
            new_skus = set(new_map.keys())
            created = len(new_skus - old_skus)
            updated = len(new_skus & old_skus)
            _sku_oe_map.update(new_map)

        save_sku_oe_map()

        return {
            "success": True,
            "totalRows": total_rows,
            "skippedRows": skipped_rows,
            "imported": len(new_map),
            "created": created,
            "updated": updated,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=400, detail=f"Excel 解析失败: {str(e)}"
        )


@router.post("/api/export")
async def api_export(request: Request, background_tasks: BackgroundTasks):
    """服务端生成带图片的 Excel 文件"""
    data = await request.json()
    if not data or "items" not in data:
        raise HTTPException(status_code=400, detail="缺少导出数据")

    items = data["items"]
    if not items:
        raise HTTPException(status_code=400, detail="没有可导出的商品")

    wb = Workbook()
    ws = wb.active
    ws.title = "eBay Results"

    headers = [
        "OE",
        "Price",
        "Title",
        "Seller",
        "Rate",
        "Shipping",
        "Link",
        "Images",
    ]
    col_widths = [16, 12, 50, 14, 8, 12, 40, 18]
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = cell.font.copy(bold=True)
        ws.column_dimensions[get_column_letter(c)].width = w

    img_height = 80
    row_height_px = img_height + 10

    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item.get("oe", ""))
        ws.cell(row=i, column=2, value=item.get("price", ""))
        ws.cell(row=i, column=3, value=item.get("title", ""))
        ws.cell(row=i, column=4, value=item.get("seller", ""))
        fb = item.get("sellerFeedback", "")
        ws.cell(row=i, column=5, value=f"{fb}%" if fb else "")
        ws.cell(row=i, column=6, value=item.get("shipping", ""))
        link = item.get("link", "")
        if link:
            cell = ws.cell(row=i, column=7, value=link)
            cell.hyperlink = link
            cell.font = cell.font.copy(color="0563C1", underline="single")

        img_url = (item.get("images") or [None])[0]
        if img_url:
            try:
                resp = _requests.get(img_url, timeout=10, verify=False)
                if resp.status_code == 200:
                    img_data = BytesIO(resp.content)
                    xl_img = XlImage(img_data)
                    xl_img.height = img_height
                    xl_img.width = img_height
                    anchor = f"H{i}"
                    ws.add_image(xl_img, anchor)
            except Exception:
                pass

        ws.row_dimensions[i].height = row_height_px * 0.75

    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, prefix="ebay_export_"
    )
    wb.save(tmp.name)
    tmp.close()
    tmp_path = tmp.name

    background_tasks.add_task(_safe_remove, tmp_path)

    return FileResponse(
        tmp_path,
        filename=f"ebay_export_{int(time.time())}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _safe_remove(path: str) -> None:
    """安全删除临时文件"""
    try:
        os.remove(path)
    except OSError:
        pass
