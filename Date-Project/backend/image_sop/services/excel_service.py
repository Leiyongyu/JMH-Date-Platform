from __future__ import annotations

import logging
import mimetypes
import os
from datetime import datetime
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.units import pixels_to_EMU

from backend.image_sop.models import ImageRequirement

logger = logging.getLogger(__name__)

# openpyxl 实际能稳定嵌入的图片格式（webp 排除，因其需要系统 libwebp 且 MIME 注册不稳定）
SUPPORTED_IMG_EXT = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".tif", ".wmf", ".emf"}

# 参考图在 Excel 中的显示尺寸（200*200，小尺寸节省空间）
GRID_COLS = 1          # 单列排版，每张图独占全宽
GRID_SINGLE_W = 200    # 单张图宽度（像素）
GRID_SINGLE_H = 200    # 单张图高度（像素）
GRID_GAP = 4           # 图之间垂直间距（像素）
COMPOSITE_W = 200      # 合成图在 Excel 中显示宽度
COMPOSITE_H = 200      # 合成图在 Excel 中显示高度
SECTION_GAP_EMU = 8 * 9525  # 不同来源区域之间的间距
EMU_PER_PIXEL = 9525

# 合成图内部生成参数：场景 60% / 产品 40%，上下拼接
COMPOSITE_TOTAL_W = 200   # 合成图总宽度
COMPOSITE_TOTAL_H = 200   # 合成图总高度
COMPOSITE_GAP = 4         # 上下图间隔
COMPOSITE_SCENE_H = int((COMPOSITE_TOTAL_H - COMPOSITE_GAP) * 0.60)  # 场景占上半部 60%
COMPOSITE_PRODUCT_H = COMPOSITE_TOTAL_H - COMPOSITE_SCENE_H - COMPOSITE_GAP  # 产品占下半部 40%
COMPOSITE_PANEL_W = COMPOSITE_TOTAL_W  # 面板宽度=全宽
COMPOSITE_TEMP_DIR_NAME = "_composites"

# ---- 猴子补丁：防止 openpyxl 在 _register_mimetypes 中因 .webp 导致 KeyError ----
import openpyxl.packaging.manifest as _manifest


def _safe_register_mimetypes(self, filenames):
    """注册 MIME 类型，跳过 mimetypes 中不存在的扩展名（如 .webp），避免 KeyError。"""
    for fn in filenames:
        ext = os.path.splitext(fn)[-1]  # 保留大小写以匹配 mimetypes.types_map
        if not ext:
            continue
        try:
            mime = mimetypes.types_map[True][ext]
        except KeyError:
            # .webp 等未注册扩展名直接跳过
            continue
        fe = _manifest.FileExtension(ext[1:], mime)
        self.Default.append(fe)


_manifest.Manifest._register_mimetypes = _safe_register_mimetypes
# ---- 猴子补丁结束 ----



class ExcelService:
    def __init__(self, export_dir: Path) -> None:
        self.export_dir = export_dir
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def _make_reference_composite(
        self,
        scene_path: Path,
        product_path: Path,
        prefix: str = "",
    ) -> Path | None:
        """将场景图（上 60%）和产品细节图（下 40%）拼接为一张参考合成图。

        美工拿到这张合成图后可以直接照此构图设计，无需额外揣测。
        合成图尺寸匹配设计稿 1600*1600，上下拼接提供足够信息量。
        返回合成图的 Path，失败返回 None。
        """
        try:
            from PIL import Image as PilImage

            scene = PilImage.open(scene_path).convert("RGB")
            product = PilImage.open(product_path).convert("RGB")

            composite = PilImage.new("RGB", (COMPOSITE_TOTAL_W, COMPOSITE_TOTAL_H), (255, 255, 255))

            # ── 上半部：场景图（占 60%）──
            scene_thumb = self._cover_crop(scene, COMPOSITE_PANEL_W, COMPOSITE_SCENE_H)
            composite.paste(scene_thumb, (0, 0))

            # ── 灰色分隔线 ──
            gap_y = COMPOSITE_SCENE_H
            for yy in range(gap_y, gap_y + COMPOSITE_GAP):
                for xx in range(COMPOSITE_TOTAL_W):
                    composite.putpixel((xx, yy), (200, 200, 200))

            # ── 下半部：产品细节图（占 40%）──
            prod_thumb = self._cover_crop(product, COMPOSITE_PANEL_W, COMPOSITE_PRODUCT_H)
            bottom_y = COMPOSITE_SCENE_H + COMPOSITE_GAP
            composite.paste(prod_thumb, (0, bottom_y))

            # ── 标签 ──
            try:
                from PIL import ImageDraw, ImageFont
                draw = ImageDraw.Draw(composite)
                # 左上角标签
                draw.rectangle([(0, 0), (COMPOSITE_PANEL_W, 18)], fill=(50, 50, 50, 180))
                draw.rectangle([(0, bottom_y), (COMPOSITE_PANEL_W, bottom_y + 18)], fill=(50, 50, 50, 180))
                try:
                    font = ImageFont.load_default()
                    draw.text((6, 1), "SCENE", fill=(255, 255, 255), font=font)
                    draw.text((6, bottom_y + 1), "PRODUCT", fill=(255, 255, 255), font=font)
                except Exception:
                    pass
            except Exception:
                pass

            # 保存
            temp_dir = self.export_dir / COMPOSITE_TEMP_DIR_NAME
            temp_dir.mkdir(parents=True, exist_ok=True)
            save_name = f"{prefix}composite_{scene_path.stem}_{product_path.stem}.jpg".replace(" ", "_")[:120]
            output = temp_dir / save_name
            composite.save(output, "JPEG", quality=88)
            return output
        except Exception as exc:
            logger.warning(f"合成参考图失败: {exc}")
            return None

    @staticmethod
    def _cover_crop(img, target_w: int, target_h: int):
        """覆盖式裁剪：等比缩放填满目标尺寸，居中裁剪多余部分。"""
        from PIL import Image as PilImage
        img_w, img_h = img.size
        target_ratio = target_w / target_h
        img_ratio = img_w / img_h

        if img_ratio > target_ratio:
            # 图片更宽：按高度等比缩到 target_h，裁剪左右
            new_h = target_h
            new_w = int(target_h * img_ratio)
        else:
            # 图片更高：按宽度等比缩到 target_w，裁剪上下
            new_w = target_w
            new_h = int(target_w / img_ratio)

        img = img.resize((new_w, new_h), PilImage.Resampling.LANCZOS)
        left = (new_w - target_w) // 2
        top = (new_h - target_h) // 2
        return img.crop((left, top, left + target_w, top + target_h))

    @staticmethod
    def _col_row_from_anchor(anchor: str) -> tuple[int, int]:
        """把 'F3' 这样的 anchor 转成 0-based (col, row)。"""
        from openpyxl.utils import coordinate_to_tuple
        row, col = coordinate_to_tuple(anchor)
        return col - 1, row - 1

    @staticmethod
    def _place_image_grid(
        sheet,
        anchor: str,
        paths: list[Path],
        y_start_emu: int,
        single_w: int,
        single_h: int,
        cols: int = 2,
        gap: int = 4,
    ) -> int:
        """把多张图片按网格平铺到同一锚点单元格内，返回下一个可用 top 偏移（EMU）。"""
        base_col, base_row = ExcelService._col_row_from_anchor(anchor)
        for idx, img_path in enumerate(paths):
            row = idx // cols
            col = idx % cols
            x_offset = col * (single_w + gap) * EMU_PER_PIXEL
            y_offset = y_start_emu + row * (single_h + gap) * EMU_PER_PIXEL
            try:
                img = Image(str(img_path))
                img.width = single_w
                img.height = single_h
                img.anchor = OneCellAnchor(
                    _from=AnchorMarker(
                        col=base_col,
                        row=base_row,
                        colOff=x_offset,
                        rowOff=y_offset,
                    ),
                    ext=XDRPositiveSize2D(
                        cx=pixels_to_EMU(single_w),
                        cy=pixels_to_EMU(single_h),
                    ),
                )
                sheet.add_image(img)
            except Exception:
                pass
        rows = (len(paths) + cols - 1) // cols
        return y_start_emu + max(0, rows) * (single_h + gap) * EMU_PER_PIXEL

    def _embed_image_in_column(
        self,
        sheet,
        img_path: Path,
        col_0: int,
        row_0: int,
        row_span: int,
        img_w: int,
        img_h: int,
    ) -> bool:
        """把一张图锚到独立列，用 TwoCellAnchor 填满该列合并格，避免 WPS 丢图。"""
        if img_path.suffix.lower() not in SUPPORTED_IMG_EXT:
            return False
        try:
            img = Image(str(img_path))
            img.width = img_w
            img.height = img_h
            img.anchor = TwoCellAnchor(
                _from=AnchorMarker(col=col_0, row=row_0, colOff=0, rowOff=0),
                to=AnchorMarker(
                    col=col_0 + 1,
                    row=row_0 + max(1, row_span),
                    colOff=0,
                    rowOff=0,
                ),
                editAs="oneCell",
            )
            sheet.add_image(img)
            return True
        except Exception as exc:
            logger.warning("嵌入图片失败 %s: %s", img_path.name, exc)
            return False

    @staticmethod
    def _nas_cache_name(nas_path: str) -> str:
        return nas_path.replace("/", "_").lstrip("_")

    def _resolve_upload_path_from_url(
        self,
        candidate: str,
        upload_files: dict[str, Path],
        web_ref_files: dict[str, Path],
    ) -> Path | None:
        """从 reference URL 解析本地文件路径（含 NAS 代理 URL）。"""
        if not candidate:
            return None

        name_aliases: list[str] = []
        if "/api/nas/" in candidate:
            parsed = urlparse(candidate)
            nas_path = unquote(parse_qs(parsed.query).get("p", [""])[0])
            if nas_path:
                cache_name = self._nas_cache_name(nas_path)
                name_aliases.append(cache_name)
                # NAS 下载会把 png/webp 转成 jpg，upload_files 的 key 已是 .jpg
                stem = Path(cache_name).stem
                for ext in (".jpg", ".jpeg", ".png", ".webp"):
                    name_aliases.append(stem + ext)

        filename = ""
        if "?" in candidate and "/api/nas/" not in candidate:
            filename = candidate.rsplit("/", 1)[-1].split("?")[0]
        elif "/api/nas/" not in candidate:
            filename = candidate.rsplit("/", 1)[-1]
        if filename:
            name_aliases.append(filename)
            stem = Path(filename).stem
            for ext in (".jpg", ".jpeg", ".png", ".webp"):
                name_aliases.append(stem + ext)

        seen_names: set[str] = set()
        unique_aliases: list[str] = []
        for name in name_aliases:
            key = name.lower()
            if not name or key in seen_names:
                continue
            seen_names.add(key)
            unique_aliases.append(name)

        for name in unique_aliases:
            if name in upload_files:
                return upload_files[name]
            if name in web_ref_files:
                return web_ref_files[name]

        alias_stems = {Path(name).stem.lower() for name in unique_aliases}
        for up_name, up_path in upload_files.items():
            if Path(up_name).stem.lower() in alias_stems or up_path.stem.lower() in alias_stems:
                return up_path
        for web_name, web_path in web_ref_files.items():
            if Path(web_name).stem.lower() in alias_stems or web_path.stem.lower() in alias_stems:
                return web_path
        return None

    def _make_detail_local_grid(self, paths: list[Path], prefix: str = "") -> Path | None:
        """把 2~4 张局部图拼成一张网格图，避免 WPS 同一单元格多图只显示第一张。"""
        valid = [p for p in paths if p.exists()][:4]
        if not valid:
            return None
        if len(valid) == 1:
            return valid[0]
        try:
            from PIL import Image as PilImage

            cell = 160
            gap = 8
            count = len(valid)
            if count == 2:
                cols, rows = 2, 1
            elif count == 3:
                cols, rows = 3, 1
            else:
                cols, rows = 2, 2
            width = cols * cell + (cols + 1) * gap
            height = rows * cell + (rows + 1) * gap
            canvas = PilImage.new("RGB", (width, height), (247, 247, 249))
            for idx, img_path in enumerate(valid):
                with PilImage.open(img_path) as src:
                    thumb = self._cover_crop(src.convert("RGB"), cell, cell)
                row, col = divmod(idx, cols)
                x = gap + col * (cell + gap)
                y = gap + row * (cell + gap)
                canvas.paste(thumb, (x, y))
            temp_dir = self.export_dir / COMPOSITE_TEMP_DIR_NAME
            temp_dir.mkdir(parents=True, exist_ok=True)
            save_name = f"{prefix}detail_local_grid_{count}.jpg".replace(" ", "_")[:120]
            output = temp_dir / save_name
            canvas.save(output, "JPEG", quality=90)
            logger.info("局部图网格合成: %s 张 -> %s", count, output.name)
            return output
        except Exception as exc:
            logger.warning("局部图网格合成失败: %s", exc)
            return None

    def _ensure_jpg(self, path: Path | None) -> Path | None:
        """将图片统一转换为 JPEG。无法转换时保留原文件并返回 None。"""
        if not path:
            return None
        if not path.exists():
            logger.warning("图片不存在，无法转为 JPEG: %s", path)
            return None
        try:
            from PIL import Image

            jpg_path = path.with_suffix(".jpg")
            # 如果已经存在有效的同名 jpg 文件且大于 2KB，直接复用
            if (
                jpg_path.exists()
                and jpg_path.stat().st_size > 2 * 1024
                and jpg_path != path
            ):
                return jpg_path

            # 无论扩展名是什么，都用 PIL 打开并重新保存为 JPEG。
            # 这能防止 "jpg 扩展名但内容是 webp" 的情况导致 openpyxl 内部打包成 webp。
            with Image.open(path) as img:
                if img.mode in ("RGBA", "P", "LA"):
                    img = img.convert("RGB")
                img.save(jpg_path, "JPEG", quality=92)
            if jpg_path.exists() and jpg_path.stat().st_size > 0:
                # 删除原始文件（如果不同）
                if jpg_path != path and "nas_cache" not in str(path).replace("\\", "/"):
                    try:
                        path.unlink(missing_ok=True)
                    except Exception:
                        pass
                return jpg_path
        except Exception as exc:
            logger.warning("图片转换为 JPEG 失败 %s: %s", path.name, exc)
            return None
        return None

    def _convert_file_map_or_raise(
        self, file_map: dict[str, Path], label: str
    ) -> dict[str, Path]:
        converted_map: dict[str, Path] = {}
        failed: list[str] = []
        for name, path in file_map.items():
            converted = self._ensure_jpg(path)
            if converted is None:
                failed.append(name)
            else:
                converted_map[name] = converted
        if failed:
            raise ValueError(
                f"{label}无法转为 JPEG（{len(failed)} 张）："
                + "、".join(failed[:8])
                + ("…" if len(failed) > 8 else "")
                + "。请更换图片后重新生成。"
            )
        return converted_map

    def _require_converted_path(self, path: Path | None, label: str) -> Path | None:
        if path is None:
            return None
        converted = self._ensure_jpg(path)
        if converted is None:
            raise ValueError(f"{label}无法转为 JPEG：{path.name}。请更换图片后重新生成。")
        return converted

    def _assert_supported_workbook_images(self, workbook) -> None:
        removed: list[str] = []
        for ws in workbook.worksheets:
            safe_images = []
            for img in ws._images:
                img_path = getattr(img, "path", None)
                if img_path and Path(str(img_path)).suffix.lower() not in SUPPORTED_IMG_EXT:
                    removed.append(Path(str(img_path)).name)
                    continue
                safe_images.append(img)
            ws._images = safe_images
        if removed:
            raise ValueError(
                "Excel 含无法嵌入的图片格式，已中止导出："
                + "、".join(removed[:8])
            )


    def export_sop_premium(
        self,
        sku: str,
        requirements: list[ImageRequirement],
        main_image_path: Path | None = None,
        upload_files: dict[str, Path] | None = None,
        web_ref_files: dict[str, Path] | None = None,
    ) -> Path:
        """导出高级A+图 Excel（9列模板：编号/尺寸/文字/需求/参考1~5）。"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        file_map = self._convert_file_map_or_raise(upload_files or {}, "参考图")
        web_ref_files = self._convert_file_map_or_raise(web_ref_files or {}, "场景搜图")
        main_image_path = self._require_converted_path(main_image_path, "主图")

        # ── 表头行 ──
        headers = ["编号", "尺寸", "文字", "需求", "参考1", "参考2", "参考3", "参考4", "参考5"]
        header_row = 1
        thin_border = Border(
            left=Side(style="thin", color="FF9CA3AF"),
            right=Side(style="thin", color="FF9CA3AF"),
            top=Side(style="thin", color="FF9CA3AF"),
            bottom=Side(style="thin", color="FF9CA3AF"),
        )
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FF111827")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="FFD1D5DB")
            cell.border = thin_border

        # ── 品牌/风格提示行 ──
        note_row = 2
        sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
        note_cell = sheet.cell(row=note_row, column=1, value=f"高级A+套图 SOP - {sku}")
        note_cell.font = Font(bold=True, size=12, color="FF1E293B")
        note_cell.alignment = Alignment(horizontal="left", vertical="center")
        note_cell.fill = PatternFill("solid", fgColor="FFFEF3C7")
        sheet.row_dimensions[note_row].height = 28
        for col_idx in range(1, 10):
            sheet.cell(row=note_row, column=col_idx).border = thin_border

        # ── 数据行 ──
        data_palette = ["FFE5F0FA", "FFF8E9E1", "FFE5F0FA", "FFF8E9E1", "FFE5F0FA"]
        current_row = 3  # 数据从第3行开始
        # 高级A+每个需求占3行（图多行高更大）
        block_rows = 3

        for item in requirements:
            block_start = current_row
            block_end = current_row + block_rows - 1
            fill_color = data_palette[(item.index - 1) % len(data_palette)]

            # 合并单元格 (除参考图列外)
            for col in range(1, 5):  # A-D
                sheet.merge_cells(
                    start_row=block_start, start_column=col,
                    end_row=block_end, end_column=col,
                )
            # 参考图列也合并
            for col in range(5, 10):  # E-I
                sheet.merge_cells(
                    start_row=block_start, start_column=col,
                    end_row=block_end, end_column=col,
                )

            # 填充数据
            values = {
                1: item.index,           # A: 编号
                2: item.size,            # B: 尺寸
                3: item.copy_text,       # C: 文字
                4: item.design_request,  # D: 需求
            }
            for col, val in values.items():
                cell = sheet.cell(row=block_start, column=col, value=val)
                cell.alignment = Alignment(
                    horizontal="center" if col <= 2 else "left",
                    vertical="top" if col >= 3 else "center",
                    wrap_text=True,
                )
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = thin_border

            # 图片单元格
            ref_cell = sheet.cell(row=block_start, column=5, value="")
            ref_cell.fill = PatternFill("solid", fgColor=fill_color)
            ref_cell.border = thin_border
            ref_cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(6, 10):
                c = sheet.cell(row=block_start, column=col)
                c.fill = PatternFill("solid", fgColor=fill_color)
                c.border = thin_border

            # 行高
            content_len = len(item.copy_text or "") + len(item.design_request or "")
            row_height = max(120, min(260, 80 + content_len // 5))
            for row_idx in range(block_start, block_end + 1):
                sheet.row_dimensions[row_idx].height = row_height / block_rows

            # ── 插入参考图 ──
            ref_paths = self._resolve_reference_paths(
                item.reference_image, item.reference_images or [],
                main_image_path, file_map, web_ref_files or {},
            )
            detail_paths = self._resolve_reference_paths(
                item.detail_image, [], main_image_path, file_map, web_ref_files or {},
            )
            scene_paths = self._resolve_reference_paths(
                item.scene_image, [], main_image_path, file_map, web_ref_files or {},
            )
            ref_source = getattr(item, "reference_source", "") or ""

            # 收集要展示的图片（按 reference_source 决定顺序）
            display_paths: list[Path] = []
            labels: list[str] = []

            def _append_path(path: Path | None, label: str = "") -> None:
                if not path or path.suffix.lower() not in SUPPORTED_IMG_EXT:
                    return
                if path not in display_paths:
                    display_paths.append(path)
                    if label:
                        labels.append(label)

            if ref_source == "detail":
                for dp in detail_paths[:1]:
                    _append_path(dp, "细节图")
            elif ref_source == "scene":
                for sp in scene_paths[:1]:
                    _append_path(sp, "场景图")
            elif ref_source == "composite":
                for sp in scene_paths[:1]:
                    _append_path(sp, "场景图")
                for dp in detail_paths[:1]:
                    _append_path(dp, "细节图")
            elif ref_source == "main_scene":
                for rp in ref_paths[:1]:
                    _append_path(rp, "主图")
                for sp in scene_paths[:1]:
                    _append_path(sp, "场景图")
                if len(display_paths) < 2 and len(ref_paths) >= 2:
                    display_paths = [ref_paths[0], ref_paths[1]]
                    labels = ["主图", "场景图"]
            elif ref_source == "detail_local":
                for rp in ref_paths[:4]:
                    _append_path(rp, "局部图")
            else:
                for sp in scene_paths[:2]:
                    _append_path(sp, "场景图")
                for dp in detail_paths[:2]:
                    _append_path(dp, "细节图")

            # 兜底
            if not display_paths:
                for rp in ref_paths[:5]:
                    _append_path(rp)

            # 在参考1~5列中嵌入图片
            if display_paths:
                base_col = 4  # E列 (0-based)
                base_row_idx = block_start - 1  # 0-based
                is_detail_local = ref_source == "detail_local"
                img_w = 140 if is_detail_local else 120
                img_h = 140 if is_detail_local else 120
                for i, img_path in enumerate(display_paths[:5]):
                    ok = self._embed_image_in_column(
                        sheet,
                        img_path,
                        col_0=base_col + i,
                        row_0=base_row_idx,
                        row_span=block_rows,
                        img_w=img_w,
                        img_h=img_h,
                    )
                    if not ok:
                        raise ValueError(
                            f"高级A+ 参考图嵌入失败：{img_path.name}。请更换图片后重新导出。"
                        )

                # 根据图片撑大行高
                image_h_px = img_h + GRID_GAP
                for row_idx in range(block_start, block_end + 1):
                    current_h = sheet.row_dimensions[row_idx].height or 20
                    sheet.row_dimensions[row_idx].height = max(current_h, image_h_px / block_rows)

                # 参考图标签
                if is_detail_local:
                    for i, _path in enumerate(display_paths[:4]):
                        cell = sheet.cell(row=block_start, column=5 + i)
                        cell.value = f"局部图{i + 1}"
                        cell.font = Font(italic=True, color="FF2563EB", size=9)
                elif labels:
                    sheet.cell(row=block_start, column=5).value = " / ".join(labels)
                    sheet.cell(row=block_start, column=5).font = Font(italic=True, color="FF2563EB", size=9)

            else:
                sheet.cell(row=block_start, column=5).value = "（暂无参考图）"
                sheet.cell(row=block_start, column=5).font = Font(italic=True, color="FF6B7280", size=9)

            current_row += block_rows

        # ── 列宽 ──
        widths = {1: 8, 2: 14, 3: 60, 4: 60, 5: 24, 6: 24, 7: 24, 8: 24, 9: 24}
        for col_idx, width in widths.items():
            col_letter = chr(64 + col_idx) if col_idx <= 26 else ""
            if col_letter:
                sheet.column_dimensions[col_letter].width = width

        # ── 补边框 ──
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, 10):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if not cell.border.left.style:
                    cell.border = thin_border

        # ── 保存 ──
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"SOP_{sku}_Premium_{timestamp}.xlsx"
        output_file = self.export_dir / filename

        self._assert_supported_workbook_images(workbook)
        workbook.save(output_file)
        return output_file

    def export_sop(
        self,
        sku: str,
        requirements: list[ImageRequirement],
        main_image_path: Path | None = None,
        upload_files: dict[str, Path] | None = None,
        web_ref_files: dict[str, Path] | None = None,
    ) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        file_map = self._convert_file_map_or_raise(upload_files or {}, "参考图")
        web_ref_files = self._convert_file_map_or_raise(web_ref_files or {}, "场景搜图")
        main_image_path = self._require_converted_path(main_image_path, "主图")

        sheet.merge_cells("A1:I1")
        title_cell = sheet["A1"]
        title_cell.value = f"图片SOP - {sku}"
        title_cell.font = Font(bold=True, size=14, color="FFFFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill("solid", fgColor="1E293B")
        sheet.row_dimensions[1].height = 28

        headers = ["编号", "主题", "尺寸", "文案", "需求", "参考图1", "参考图2", "参考图3", "参考图4"]
        header_row = 2
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FF111827")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="FFD1D5DB")

        thin_border = Border(
            left=Side(style="thin", color="FF9CA3AF"),
            right=Side(style="thin", color="FF9CA3AF"),
            top=Side(style="thin", color="FF9CA3AF"),
            bottom=Side(style="thin", color="FF9CA3AF"),
        )

        data_palette = ["FFE5F0FA", "FFF8E9E1"]
        current_row = 3
        for item in requirements:
            ref_source = getattr(item, "reference_source", "") or ""
            is_detail_local = ref_source == "detail_local"
            block_rows = 2
            block_start = current_row
            block_end = current_row + block_rows - 1
            for col_idx in range(1, (9 if is_detail_local else 6) + 1):
                col_letter = chr(64 + col_idx)
                sheet.merge_cells(f"{col_letter}{block_start}:{col_letter}{block_end}")

            fill_color = data_palette[(item.index - 1) % len(data_palette)]
            values = {
                "A": item.index,
                "B": item.theme or f"图片{item.index}",
                "C": item.size,
                "D": item.copy_text,
                "E": item.design_request,
                "F": "",
            }
            for col in ("A", "B", "C", "D", "E", "F"):
                cell = sheet[f"{col}{block_start}"]
                cell.value = values[col]
                cell.alignment = Alignment(
                    horizontal="center" if col in ("A", "B", "C", "F") else "left",
                    vertical="top" if col in ("D", "E") else "center",
                    wrap_text=True,
                )
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = thin_border
            if is_detail_local:
                for col_idx in range(7, 10):
                    cell = sheet.cell(row=block_start, column=col_idx, value="")
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            content_len = len(item.copy_text or "") + len(item.design_request or "")
            text_row_height = max(100, min(220, 60 + content_len // 6))
            img_w = 140 if is_detail_local else GRID_SINGLE_W
            img_h = 140 if is_detail_local else GRID_SINGLE_H
            row_height = text_row_height
            if is_detail_local:
                row_height = max(text_row_height, img_h + 24)

            for row_idx in range(block_start, block_end + 1):
                sheet.row_dimensions[row_idx].height = row_height / block_rows

            ref_paths = self._resolve_reference_paths(
                item.reference_image,
                item.reference_images,
                main_image_path,
                file_map,
                web_ref_files or {},
            )
            # 额外解析 detail_image 和 scene_image（分离的细节图/场景图 URL）
            detail_paths = self._resolve_reference_paths(
                item.detail_image,
                [],
                main_image_path,
                file_map,
                web_ref_files or {},
            )
            scene_paths = self._resolve_reference_paths(
                item.scene_image,
                [],
                main_image_path,
                file_map,
                web_ref_files or {},
            )
            ref_source = getattr(item, "reference_source", "") or ""

            if ref_paths or detail_paths or scene_paths:
                # 根据 reference_source 确定要显示的图片及其顺序
                display_images: list[tuple[Path, str]] = []  # (path, label)

                if ref_source == "detail":
                    # 只显示细节图
                    for dp in detail_paths[:1]:
                        display_images.append((dp, "细节图"))
                elif ref_source == "scene":
                    # 只显示场景图
                    for sp in scene_paths[:1]:
                        display_images.append((sp, "场景图"))
                elif ref_source == "composite":
                    # 场景图在上，细节图在下
                    for sp in scene_paths[:1]:
                        display_images.append((sp, "场景图"))
                    for dp in detail_paths[:1]:
                        display_images.append((dp, "细节图"))
                elif ref_source == "main_scene":
                    # eBay 第1张：主图 + 第一张场景物品图
                    for rp in ref_paths[:1]:
                        display_images.append((rp, "主图"))
                    for sp in scene_paths[:1]:
                        display_images.append((sp, "场景图"))
                    if len(display_images) < 2 and len(ref_paths) >= 2:
                        display_images = [
                            (ref_paths[0], "主图"),
                            (ref_paths[1], "场景图"),
                        ]
                elif ref_source == "detail_local":
                    for rp in ref_paths[:4]:
                        display_images.append((rp, "局部图"))
                else:
                    # fallback：沿用旧的拆分逻辑
                    operator_paths, web_paths = self._split_reference_sources(
                        ref_paths, file_map, web_ref_files or {}
                    )
                    if web_paths:
                        display_images.append((web_paths[0], ""))
                    if operator_paths:
                        display_images.append((operator_paths[0], ""))

                if not display_images:
                    # 从 ref_paths 中兜底取
                    for rp in ref_paths[:2]:
                        display_images.append((rp, ""))

                anchor = f"F{block_start}"
                has_ref_image = False
                max_display = 4 if is_detail_local else 2
                place_w = img_w
                place_h = img_h
                grid_paths = [
                    p for p, _ in display_images[:max_display]
                    if p.suffix.lower() in SUPPORTED_IMG_EXT
                ]

                if is_detail_local and grid_paths:
                    base_col, base_row = self._col_row_from_anchor(anchor)
                    placed = 0
                    for i, img_path in enumerate(grid_paths[:4]):
                        ok = self._embed_image_in_column(
                            sheet,
                            img_path,
                            col_0=base_col + i,
                            row_0=base_row,
                            row_span=block_rows,
                            img_w=place_w,
                            img_h=place_h,
                        )
                        if not ok:
                            raise ValueError(
                                f"局部图嵌入失败：{img_path.name}。"
                                f"已放入 {placed}/{len(grid_paths[:4])} 张，请更换图片后重新导出。"
                            )
                        placed += 1
                        has_ref_image = True
                        label_cell = sheet.cell(row=block_start, column=6 + i)
                        label_cell.value = f"局部图{i + 1}"
                        label_cell.font = Font(italic=True, color="FF2563EB", size=9)
                        label_cell.alignment = Alignment(
                            horizontal="center", vertical="top", wrap_text=True
                        )
                    logger.info(
                        "局部图横向排放: sku=%s resolved=%d placed=%d cols=F-I paths=%s",
                        sku,
                        len(grid_paths[:4]),
                        placed,
                        [p.name for p in grid_paths[:4]],
                    )
                else:
                    base_col, base_row = self._col_row_from_anchor(anchor)
                    y_offset = 0
                    embed_failed: list[str] = []
                    for img_path, _img_label in display_images[:max_display]:
                        if img_path.suffix.lower() not in SUPPORTED_IMG_EXT:
                            embed_failed.append(img_path.name)
                            continue
                        try:
                            img = Image(str(img_path))
                            img.width = img_w
                            img.height = img_h
                            img.anchor = OneCellAnchor(
                                _from=AnchorMarker(
                                    col=base_col,
                                    row=base_row,
                                    colOff=0,
                                    rowOff=y_offset,
                                ),
                                ext=XDRPositiveSize2D(
                                    cx=pixels_to_EMU(img_w),
                                    cy=pixels_to_EMU(img_h),
                                ),
                            )
                            sheet.add_image(img)
                            y_offset += (img_h + GRID_GAP) * EMU_PER_PIXEL
                            has_ref_image = True
                        except Exception as exc:
                            logger.warning("嵌入图片失败 %s: %s", img_path.name, exc)
                            embed_failed.append(img_path.name)
                    if embed_failed:
                        raise ValueError(
                            "参考图嵌入失败："
                            + "、".join(embed_failed[:8])
                            + "。请更换图片后重新导出。"
                        )
                    if has_ref_image and not is_detail_local:
                        image_area_px = (y_offset / EMU_PER_PIXEL) + 10
                        adjusted = max(row_height, int(image_area_px))
                        for row_idx in range(block_start, block_end + 1):
                            sheet.row_dimensions[row_idx].height = min(
                                adjusted / block_rows, 409
                            )

                # 添加参考图来源标签（局部图已在各列单独标注）
                if not is_detail_local:
                    if has_ref_image:
                        source_labels = [lbl for _, lbl in display_images[:max_display] if lbl]
                        sheet[f"F{block_start}"].value = " / ".join(source_labels) if source_labels else ""
                        if source_labels:
                            sheet[f"F{block_start}"].font = Font(italic=True, color="FF2563EB", size=9)
                        else:
                            sheet[f"F{block_start}"].value = ""
                    else:
                        sheet[f"F{block_start}"].value = "（暂无参考图）"
                        sheet[f"F{block_start}"].font = Font(italic=True, color="FF6B7280", size=9)
                elif not has_ref_image:
                    sheet[f"F{block_start}"].value = "（暂无参考图）"
                    sheet[f"F{block_start}"].font = Font(italic=True, color="FF6B7280", size=9)

            current_row += block_rows

        widths = {1: 8, 2: 18, 3: 14, 4: 52, 5: 72, 6: 22, 7: 22, 8: 22, 9: 22}
        for col_idx, width in widths.items():
            sheet.column_dimensions[chr(64 + col_idx)].width = width

        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, 10):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if not cell.border.left.style:
                    cell.border = thin_border

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"SOP_{sku}_{timestamp}.xlsx"
        output_file = self.export_dir / filename

        self._assert_supported_workbook_images(workbook)
        workbook.save(output_file)
        return output_file

    def _resolve_reference_paths(
        self,
        reference_image: str,
        reference_images: list[str],
        main_image_path: Path | None,
        upload_files: dict[str, Path],
        web_ref_files: dict[str, Path] | None = None,
    ) -> list[Path]:
        """解析单个需求的所有参考图路径。
        优先级：运营参考图 > Web 搜索图 > 主图兜底。
        运营参考图不会被 Web 搜索图挤出。
        """
        web_ref_files = web_ref_files or {}
        candidates: list[str] = []
        if reference_images:
            candidates.extend(reference_images)
        if reference_image and reference_image not in candidates:
            candidates.append(reference_image)

        # 分两个通道：运营图（upload_files 中的）和 Web 图（web_ref_files 中的）
        operator_paths: list[Path] = []
        web_paths: list[Path] = []
        seen: set[str] = set()

        for candidate in candidates:
            if not candidate:
                continue

            dedup_key = candidate if "/api/nas/" in candidate else ""
            if not dedup_key:
                if "?" in candidate:
                    dedup_key = candidate.rsplit("/", 1)[-1].split("?")[0]
                else:
                    dedup_key = candidate.rsplit("/", 1)[-1]
            if not dedup_key or dedup_key in seen:
                continue
            seen.add(dedup_key)

            resolved = self._resolve_upload_path_from_url(
                candidate, upload_files, web_ref_files
            )
            if resolved and str(resolved) in seen:
                resolved = None
            if resolved and resolved.exists():
                if resolved.suffix.lower() not in SUPPORTED_IMG_EXT:
                    continue
                seen.add(str(resolved))
                if str(resolved) in {str(p) for p in web_ref_files.values()}:
                    web_paths.append(resolved)
                else:
                    operator_paths.append(resolved)
                continue

            filename = dedup_key
            if filename in web_ref_files:
                resolved = web_ref_files[filename]
                if resolved and resolved.suffix.lower() not in SUPPORTED_IMG_EXT:
                    continue
                if resolved and resolved.exists() and str(resolved) not in seen:
                    seen.add(str(resolved))
                    web_paths.append(resolved)

        paths = operator_paths + web_paths
        paths = [p for p in paths if p.suffix.lower() in SUPPORTED_IMG_EXT]
        # 主图只能出现在主图行，其他行没有参考图时留空，不兜底主图
        logger.info(
            "resolve refs: candidates=%d operator=%d web=%d paths=%d",
            len(candidates),
            len(operator_paths),
            len(web_paths),
            len(paths),
        )
        return paths

    def _split_reference_sources(
        self,
        ref_paths: list[Path],
        upload_files: dict[str, Path],
        web_ref_files: dict[str, Path],
    ) -> tuple[list[Path], list[Path]]:
        """把参考图路径拆分为运营图（产品）和 Web 搜索图（场景/物品）。"""
        upload_set = {str(p) for p in upload_files.values()}
        web_set = {str(p) for p in web_ref_files.values()}
        operator_paths: list[Path] = []
        web_paths: list[Path] = []
        for p in ref_paths:
            p_str = str(p)
            if p_str in web_set:
                web_paths.append(p)
            elif p_str in upload_set:
                operator_paths.append(p)
            # main_image 兜底既不归于运营也不归于网络，不单独显示
        return operator_paths, web_paths

    def _build_reference_composites(
        self,
        ref_paths: list[Path],
        upload_files: dict[str, Path],
        web_ref_files: dict[str, Path],
        prefix: str = "",
    ) -> list[Path]:
        """将运营参考图（产品细节）和 Web 搜索图（场景/车型）配对拼接为左右合成图。

        拼法：左侧放 Web 搜索到的场景图，右侧放运营提供的产品细节图。
        美工拿到后可以直接参照构图设计，无需额外揣测。
        """
        # 区分两类图
        upload_set = {str(p) for p in upload_files.values()}
        web_set = {str(p) for p in web_ref_files.values()}
        operator_paths: list[Path] = []
        web_paths: list[Path] = []
        for p in ref_paths:
            p_str = str(p)
            if p_str in upload_set:
                operator_paths.append(p)
            elif p_str in web_set:
                web_paths.append(p)

        # 没有运营图 → 不做拼接
        if not operator_paths:
            return []

        # 只保留 1 张合成图：用第一个运营产品图 + 第一个 Web 场景图。
        # 避免多个运营产品图重复拼出多张产品细节图，造成美工困惑。
        op_path = operator_paths[0]
        web = web_paths[0] if web_paths else None
        if web is None:
            return []
        comp = self._make_reference_composite(
            scene_path=web,
            product_path=op_path,
            prefix=prefix,
        )
        return [comp] if comp else []
