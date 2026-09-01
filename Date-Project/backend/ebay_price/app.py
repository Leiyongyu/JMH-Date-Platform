"""
eBay 价格查询工具 v3 — FastAPI 版
==================================
策略：两遍搜索 + 类目锁定（category lock）+ 可选 getItem 核验
从原 Flask 独立服务改写，挂载到 Date-Project 后端。
"""

from __future__ import annotations

import os
import re
import tempfile
import threading
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path

import requests as http_requests
import urllib3
from fastapi import FastAPI, File, Form, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

from backend.config import settings
from backend.ebay_price import repository as repo

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = FastAPI(title="eBay Price Tool v3")

# ═══════════════════════ 配置 ═══════════════════════

CATEGORY_MIN_SHARE = 0.40
VERIFY_ENABLED = True
VERIFY_TOP_N = 5

SITES = {
    "de": {"domain": "ebay.de", "locale": "de-DE", "marketplace": "EBAY_DE",
           "currency": "EUR", "global_id": "EBAY-GERMANY"},
    "uk": {"domain": "ebay.co.uk", "locale": "en-GB", "marketplace": "EBAY_GB",
           "currency": "GBP", "global_id": "EBAY-GB"},
    "us": {"domain": "ebay.com", "locale": "en-US", "marketplace": "EBAY_US",
           "currency": "USD", "global_id": "EBAY-US"},
}

OAUTH_URL = "https://api.ebay.com/identity/v1/oauth2/token"
BROWSE_URL = "https://api.ebay.com/buy/browse/v1/item_summary/search"
ITEM_URL = "https://api.ebay.com/buy/browse/v1/item/"

# ═══════════════════════ SKU-OE 缓存 ═══════════════════════

_sku_oe_map: dict[str, str] = {}
_middle_code_oe_map: dict[str, str] = {}
_custom_middle_codes: dict[str, str] = {}
_sku_oe_lock = threading.Lock()


def _extract_middle_code(sku: str) -> str:
    if not sku:
        return ""
    idx = sku.find("-")
    if 0 <= idx < len(sku) - 1:
        return sku[idx + 1:]
    return ""


def load_sku_oe_map():
    global _sku_oe_map, _middle_code_oe_map, _custom_middle_codes
    with _sku_oe_lock:
        try:
            rows = repo.list_all()
            new_map: dict[str, str] = {}
            mc_map: dict[str, str] = {}
            custom_mc: dict[str, str] = {}
            for row in rows:
                sku = (row.get("sku") or "").strip()
                oe = (row.get("oe") or "").strip()
                mc_raw = (row.get("middle_code") or "").strip()
                if sku and oe:
                    new_map[sku] = oe
                    auto_mc = _extract_middle_code(sku)
                    if mc_raw and mc_raw.upper() != (auto_mc or "").upper():
                        custom_mc[sku] = mc_raw
                    mc = mc_raw if mc_raw else auto_mc
                    if mc:
                        mc_map[mc.upper()] = oe
            _sku_oe_map = new_map
            _middle_code_oe_map = mc_map
            _custom_middle_codes = custom_mc
        except Exception as e:
            print(f"加载 SKU-OE 映射失败: {e}")


def _reload_middle_code_map():
    global _middle_code_oe_map
    mc_map: dict[str, str] = {}
    for sku, oe in _sku_oe_map.items():
        mc = _custom_middle_codes.get(sku) or _extract_middle_code(sku)
        if mc:
            mc_map[mc.upper()] = oe
    _middle_code_oe_map = mc_map


def get_oe_by_sku(sku: str) -> str | None:
    if not sku:
        return None
    key = sku.strip()
    oe = _sku_oe_map.get(key)
    if oe:
        return oe
    return _middle_code_oe_map.get(key.upper())


# ═══════════════════════ OAuth2 Token ═══════════════════════

_token_cache = {"token": None, "expires": 0.0}
_token_lock = threading.Lock()


def _get_ebay_credentials() -> tuple[str, str]:
    return (
        os.getenv("EBAY_CLIENT_ID", ""),
        os.getenv("EBAY_CLIENT_SECRET", ""),
    )


def get_access_token() -> str:
    with _token_lock:
        if _token_cache["token"] and time.time() < _token_cache["expires"]:
            return _token_cache["token"]

        client_id, client_secret = _get_ebay_credentials()
        if not client_id or not client_secret:
            raise RuntimeError("eBay API 凭据未配置，请在 .env 中设置 EBAY_CLIENT_ID 和 EBAY_CLIENT_SECRET")

        resp = http_requests.post(
            OAUTH_URL,
            data={"grant_type": "client_credentials",
                  "scope": "https://api.ebay.com/oauth/api_scope"},
            auth=(client_id, client_secret),
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=15,
            verify=False,
        )
        if resp.status_code != 200:
            raise RuntimeError(f"获取 eBay Token 失败 (HTTP {resp.status_code}): {resp.text[:200]}")

        data = resp.json()
        _token_cache["token"] = data["access_token"]
        _token_cache["expires"] = time.time() + data.get("expires_in", 7200) - 60
        return _token_cache["token"]


# ═══════════════════════ eBay API 调用 ═══════════════════════

def _api_get(url: str, params: dict, marketplace: str, retry_on_401: bool = True):
    token = get_access_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "X-EBAY-C-MARKETPLACE-ID": marketplace,
        "X-EBAY-C-ENDUSERCTX": "affiliateCampaignId=<ePNCampaignId>,affiliateReferenceId=<referenceId>",
    }
    resp = http_requests.get(url, params=params, headers=headers, timeout=30, verify=False)
    if resp.status_code == 200:
        return resp.json(), None
    if resp.status_code == 401 and retry_on_401:
        _token_cache["token"] = None
        token = get_access_token()
        headers["Authorization"] = f"Bearer {token}"
        resp = http_requests.get(url, params=params, headers=headers, timeout=30, verify=False)
        if resp.status_code == 200:
            return resp.json(), None
    return None, f"eBay API 错误 (HTTP {resp.status_code}): {resp.text[:300]}"


def search_ebay(keyword: str, marketplace: str, limit: int = 100, offset: int = 0,
                sort: str = "", category_id: str = "", fixed_price_only: bool = False,
                fieldgroups: str = "", free_shipping: bool = False):
    filters = ["conditions:{NEW}"]
    if fixed_price_only:
        filters.append("buyingOptions:{FIXED_PRICE}")
    if free_shipping:
        filters.append("deliveryOptions:{FREE_SHIPPING}")
    params = {
        "q": keyword,
        "limit": min(limit, 200),
        "offset": offset,
        "filter": ",".join(filters),
    }
    if sort:
        params["sort"] = sort
    if category_id:
        params["category_ids"] = category_id
    if fieldgroups:
        params["fieldgroups"] = fieldgroups
    return _api_get(BROWSE_URL, params, marketplace)


def get_item_product_type(item_id: str, marketplace: str) -> str:
    if not item_id:
        return ""
    data, error = _api_get(ITEM_URL + item_id, {}, marketplace)
    if error or not data:
        return ""
    for aspect in data.get("localizedAspects", []):
        name = (aspect.get("name") or "").strip().lower()
        if name in ("produktart", "product type", "type"):
            return (aspect.get("value") or "").strip()
    return ""


# ═══════════════════════ 工具函数 ═══════════════════════

def _clean_item_url(url: str) -> str:
    if not url:
        return ""
    idx = url.find("/itm/")
    if idx < 0:
        return url
    after = url[idx + 5:]
    m = re.search(r'[?_#]', after)
    return url[:idx + 5] + (after[:m.start()] if m else after)


def _upgrade_img(url: str) -> str:
    return re.sub(r's-l\d+', 's-l500', url) if url else url


def format_item(item: dict) -> dict:
    title = item.get("title", "")
    price_info = item.get("price", {})
    price_value = float(price_info.get("value", 0))
    currency = price_info.get("currency", "EUR")
    condition = item.get("condition", "")
    condition_id = item.get("conditionId", "")

    images: list[str] = []
    seen_urls: set[str] = set()
    for src_list in [item.get("thumbnailImages", []), item.get("additionalImages", [])]:
        for img in src_list:
            u = img.get("imageUrl", "")
            if u and u not in seen_urls:
                seen_urls.add(u)
                images.append(_upgrade_img(u))
    if not images:
        primary = item.get("image", {})
        if primary and primary.get("imageUrl"):
            images.append(_upgrade_img(primary["imageUrl"]))
    images = images[:5]

    seller = item.get("seller", {})
    shipping_options = item.get("shippingOptions", [])
    shipping_cost = ""
    if shipping_options:
        ship_price = shipping_options[0].get("shippingCost", {})
        if ship_price:
            ship_val = float(ship_price.get("value", 0))
            ship_curr = ship_price.get("currency", currency)
            shipping_cost = f"{ship_val:.2f} {ship_curr}" if ship_val > 0 else "免运费"

    estimated_sold = None
    for avail in (item.get("estimatedAvailabilities") or []):
        qty = avail.get("estimatedSoldQuantity")
        if qty is not None:
            try:
                parsed = max(int(qty), 0)
                estimated_sold = max(parsed, estimated_sold or 0)
            except (ValueError, TypeError):
                pass

    return {
        "title": title,
        "price": f"{price_value:.2f} {currency}",
        "pf": price_value,
        "currency": currency,
        "condition": condition,
        "conditionId": condition_id,
        "categoryId": item.get("categoryId", ""),
        "images": images,
        "link": _clean_item_url(item.get("itemWebUrl", "")),
        "itemId": item.get("itemId", ""),
        "seller": seller.get("username", ""),
        "sellerFeedback": seller.get("feedbackPercentage", ""),
        "shipping": shipping_cost,
        "buyingOptions": item.get("buyingOptions", []),
        "estimatedSold": estimated_sold,
    }


def parse_oe_list(text: str) -> list[str]:
    lines = []
    for line in text.replace(",", "\n").replace("，", "\n").split("\n"):
        oe = line.strip()
        if oe:
            lines.append(oe)
    seen: set[str] = set()
    result: list[str] = []
    for oe in lines:
        if oe.upper() not in seen:
            seen.add(oe.upper())
            result.append(oe)
    return result


# ═══════════════════════ v3 搜索策略 ═══════════════════════

def _pick_dominant_category(data: dict, items_raw: list) -> tuple:
    refinement = data.get("refinement", {}) or {}
    dists = refinement.get("categoryDistributions", []) or []

    if dists:
        dist_top = [
            {"id": d.get("categoryId", ""), "name": d.get("categoryName", ""),
             "count": d.get("matchCount", 0)}
            for d in sorted(dists, key=lambda x: x.get("matchCount", 0), reverse=True)[:5]
        ]
        total = sum(d.get("matchCount", 0) for d in dists)
        top = max(dists, key=lambda d: d.get("matchCount", 0))
        top_count = top.get("matchCount", 0)
        share = (top_count / total) if total > 0 else 0.0
        if top.get("categoryId") and share >= CATEGORY_MIN_SHARE:
            return top["categoryId"], top.get("categoryName", ""), share, dist_top
        return "", "", share, dist_top

    dom_id = refinement.get("dominantCategoryId", "")
    if dom_id:
        return dom_id, "", 1.0, []

    cat_counter: Counter = Counter()
    cat_name_map: dict[str, str] = {}
    for item in (items_raw or []):
        cid = cname = ""
        cats = item.get("categories", []) or []
        if cats:
            leaf = cats[-1]
            cid = leaf.get("categoryId", "")
            cname = leaf.get("categoryName", "")
        if not cid:
            leaf_ids = item.get("leafCategoryIds", []) or []
            if leaf_ids:
                cid = leaf_ids[0]
        if cid:
            cat_counter[cid] += 1
            if cname and cid not in cat_name_map:
                cat_name_map[cid] = cname
    if not cat_counter:
        return "", "", 0.0, []

    dist_top = [
        {"id": cid, "name": cat_name_map.get(cid, ""), "count": cnt}
        for cid, cnt in cat_counter.most_common(5)
    ]
    top_cid, top_count = cat_counter.most_common(1)[0]
    total = sum(cat_counter.values())
    share = top_count / total if total > 0 else 0.0
    if share >= CATEGORY_MIN_SHARE:
        return top_cid, cat_name_map.get(top_cid, ""), share, dist_top
    return "", "", share, dist_top


def _verify_top_items(items: list[dict], marketplace: str) -> list[dict]:
    if not VERIFY_ENABLED or not items:
        return items
    check_n = min(VERIFY_TOP_N, len(items))
    head = items[:check_n]
    tail = items[check_n:]
    types = []
    for it in head:
        ptype = get_item_product_type(it["itemId"], marketplace)
        it["productType"] = ptype
        types.append(ptype.lower() if ptype else "")
    known = [t for t in types if t]
    if len(known) < 2:
        return items
    modal_type = max(set(known), key=known.count)
    if known.count(modal_type) < 2:
        return items
    normal, suspect = [], []
    for it, t in zip(head, types):
        if t and t != modal_type:
            it["suspect"] = True
            suspect.append(it)
        else:
            normal.append(it)
    return normal + suspect + tail


def _finalize_oe(items_raw: list, oe: str, strategy: str, mp: str,
                 cat_info: dict | None, verify: bool = False):
    result = [format_item(it) for it in items_raw]
    for it in result:
        it["oe"] = oe
    result.sort(key=lambda x: x["pf"])
    all_prices = [it["pf"] for it in result]
    if verify:
        result = _verify_top_items(result, mp)
    return result[:20], None, strategy, cat_info, all_prices


def _raw_cat_id(item: dict) -> str:
    cats = item.get("categories") or []
    if cats:
        return cats[-1].get("categoryId", "")
    leaf_ids = item.get("leafCategoryIds") or []
    return leaf_ids[0] if leaf_ids else ""


def scrape_one_oe(oe: str, site_key: str, free_shipping: bool = False,
                  search_strategy: str = "category_lock") -> tuple:
    cfg = SITES[site_key]
    mp = cfg["marketplace"]

    if search_strategy == "best_match":
        data, error = search_ebay(oe, mp, limit=50, free_shipping=free_shipping)
        if error:
            return [], error, 0, "error", None, []
        items_raw = data.get("itemSummaries", [])
        if not items_raw:
            return [], f"OE '{oe}' 在 {cfg['domain']} 上未找到结果", 0, "error", None, []
        sold_total = sum(item.get("soldQuantity", 0) or 0 for item in items_raw)
        items, _, _, _, all_prices = _finalize_oe(items_raw, oe, "best_match", mp, None)
        return items, None, sold_total, "best_match", None, all_prices

    data1, error = search_ebay(
        oe, mp, limit=100,
        fieldgroups="MATCHING_ITEMS,CATEGORY_REFINEMENTS",
        free_shipping=free_shipping,
    )
    if error:
        return [], error, 0, "error", None, []
    items_raw1 = data1.get("itemSummaries", [])
    if not items_raw1:
        return [], f"OE '{oe}' 在 {cfg['domain']} 上未找到结果", 0, "error", None, []

    sold_total = sum(item.get("soldQuantity", 0) or 0 for item in items_raw1)
    cat_id, cat_name, share, dist_top = _pick_dominant_category(data1, items_raw1)
    cat_info = {"id": cat_id, "name": cat_name, "share": round(share, 3), "top": dist_top}

    if not cat_id:
        items, _, strategy, ci, all_prices = _finalize_oe(
            items_raw1[:40], oe, "best_match_fallback", mp, cat_info)
        return items, None, sold_total, strategy, ci, all_prices

    data2, error2 = search_ebay(
        oe, mp, limit=50, sort="price",
        category_id=cat_id, fixed_price_only=True,
        free_shipping=free_shipping,
    )
    items_raw2 = (data2 or {}).get("itemSummaries", []) if not error2 else []

    if items_raw2:
        items, _, strategy, ci, all_prices = _finalize_oe(
            items_raw2, oe, "category_locked", mp, cat_info, verify=True)
        return items, None, sold_total, strategy, ci, all_prices

    pool = [it for it in items_raw1 if _raw_cat_id(it) == cat_id]
    if pool:
        items, _, strategy, ci, all_prices = _finalize_oe(
            pool, oe, "category_pool_fallback", mp, cat_info, verify=True)
        return items, None, sold_total, strategy, ci, all_prices

    items, _, strategy, ci, all_prices = _finalize_oe(
        items_raw1[:40], oe, "best_match_fallback", mp, cat_info)
    return items, None, sold_total, strategy, ci, all_prices


# ═══════════════════════ 批量搜索任务 ═══════════════════════

_tasks: dict[str, dict] = {}


def _search_batch(oe_list: list[str], site_key: str, task_id: str,
                  free_shipping: bool = False, search_strategy: str = "category_lock"):
    total = len(oe_list)
    all_results: list[dict | None] = [None] * total
    completed = 0
    lock = threading.Lock()

    def do_one(oe: str) -> dict:
        items, warning, sold, strategy, cat_info, all_prices = scrape_one_oe(
            oe, site_key, free_shipping=free_shipping, search_strategy=search_strategy)
        return {
            "oe": oe, "count": len(items), "items": items,
            "warning": warning if not items else None,
            "soldTotal": sold, "strategy": strategy,
            "categoryInfo": cat_info, "allPrices": all_prices,
        }

    try:
        with ThreadPoolExecutor(max_workers=3) as pool:
            future_map = {pool.submit(do_one, oe): oe for oe in oe_list}
            oe_index = {oe: i for i, oe in enumerate(oe_list)}
            for fut in as_completed(future_map):
                oe = future_map[fut]
                result = fut.result()
                all_results[oe_index[oe]] = result
                with lock:
                    completed += 1
                _tasks[task_id] = {
                    "status": "running",
                    "progress": {"current": completed, "total": total, "oe": oe},
                }
        _tasks[task_id] = {"status": "done", "total_oe": total, "results": all_results}
    except Exception as e:
        _tasks[task_id] = {"status": "error", "msg": str(e)}


def _cleanup_tasks():
    terminal = [tid for tid, t in _tasks.items() if t.get("status") in ("done", "error")]
    if len(terminal) > 20:
        for tid in terminal[:-20]:
            del _tasks[tid]


def _resolve_oe_list(input_list: list[str], input_type: str) -> tuple:
    oe_list: list[str] = []
    sku_mapping: dict[str, str] = {}
    not_found_skus: list[str] = []

    if input_type == "oe":
        oe_list = input_list
    elif input_type == "sku":
        with _sku_oe_lock:
            for item in input_list:
                oe = _sku_oe_map.get(item) or _middle_code_oe_map.get(item.upper())
                if oe:
                    oe_list.append(oe)
                    sku_mapping[item] = oe
                else:
                    not_found_skus.append(item)
    else:
        with _sku_oe_lock:
            for item in input_list:
                oe = _sku_oe_map.get(item) or _middle_code_oe_map.get(item.upper())
                if oe:
                    oe_list.append(oe)
                    sku_mapping[item] = oe
                else:
                    oe_list.append(item)

    return oe_list, sku_mapping, not_found_skus


def _parse_excel_keywords(file_bytes: bytes, filename: str) -> tuple:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Excel 文件没有工作表")

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        raise ValueError("Excel 文件为空")

    header = rows[0]
    if not header:
        raise ValueError("Excel 文件表头为空")

    def normalize_header(val):
        return re.sub(r'[\s_]', '', str(val or '')).lower()

    col_a = normalize_header(header[0]) if len(header) > 0 and header[0] else ''
    col_b = normalize_header(header[1]) if len(header) > 1 and header[1] else ''

    has_sku_col = col_a in ('sku', 'skuid', 'skuno')
    has_oe_col = col_b in ('oe', 'oeno', 'oenumber', 'oenr', 'partnumber', 'partno')

    keywords: list[str] = []
    total_rows = 0
    skipped_rows = 0

    if has_sku_col and has_oe_col:
        for row in rows[1:]:
            total_rows += 1
            sku = str(row[0]).strip() if len(row) > 0 and row[0] else ''
            oe = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if oe:
                keywords.append(oe)
            elif sku:
                keywords.append(sku)
            else:
                skipped_rows += 1
        input_type = 'auto'
    else:
        for row in rows[1:]:
            total_rows += 1
            val = str(row[0]).strip() if len(row) > 0 and row[0] else ''
            for part in re.split(r'[\r\n,，;；]+', val):
                part = part.strip()
                if part:
                    keywords.append(part)
            if not val:
                skipped_rows += 1
        input_type = 'oe'

    seen: set[str] = set()
    unique_keywords: list[str] = []
    for kw in keywords:
        key_upper = kw.upper()
        if key_upper not in seen:
            seen.add(key_upper)
            unique_keywords.append(kw)

    return unique_keywords, input_type, total_rows, skipped_rows, filename


# ═══════════════════════ FastAPI 路由 ═══════════════════════

@app.on_event("startup")
def startup():
    load_sku_oe_map()


@app.get("/")
def index():
    html_path = Path(__file__).parent / "static" / "ebay_tool_v3.html"
    if html_path.exists():
        return HTMLResponse(html_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>前端页面未找到</h1>", status_code=404)


@app.get("/api/check")
def api_check():
    client_id, client_secret = _get_ebay_credentials()
    configured = bool(client_id and client_secret)
    return {"configured": configured}


@app.post("/api/scrape")
async def api_scrape(
    nkw: str = Form(""),
    site: str = Form("de"),
    input_type: str = Form("auto"),
    free_shipping: str = Form("false"),
    search_strategy: str = Form("category_lock"),
):
    _cleanup_tasks()
    site = site.strip()
    input_type = input_type.strip()
    free_shipping_bool = free_shipping.strip().lower() == "true"
    search_strategy = search_strategy.strip()
    if search_strategy not in ("best_match", "category_lock"):
        search_strategy = "category_lock"

    if site not in SITES:
        return JSONResponse({"error": "不支持的站点"}, status_code=400)
    if not nkw.strip():
        return JSONResponse({"error": "请输入至少一个产品编号"}, status_code=400)

    input_list = parse_oe_list(nkw)
    if not input_list:
        return JSONResponse({"error": "未能识别有效的产品编号"}, status_code=400)

    oe_list, sku_mapping, not_found_skus = _resolve_oe_list(input_list, input_type)
    return _start_search_task(oe_list, site, free_shipping_bool, search_strategy,
                              sku_mapping, not_found_skus)


@app.post("/api/scrape-file")
async def api_scrape_file(
    file: UploadFile = File(...),
    site: str = Form("de"),
    free_shipping: str = Form("false"),
    search_strategy: str = Form("category_lock"),
):
    _cleanup_tasks()
    site = site.strip()
    free_shipping_bool = free_shipping.strip().lower() == "true"
    search_strategy = search_strategy.strip()
    if search_strategy not in ("best_match", "category_lock"):
        search_strategy = "category_lock"

    if site not in SITES:
        return JSONResponse({"error": "不支持的站点"}, status_code=400)

    fname_lower = (file.filename or "").lower()
    if not (fname_lower.endswith('.xlsx') or fname_lower.endswith('.xlsm') or fname_lower.endswith('.xls')):
        return JSONResponse({"error": "仅支持 .xlsx、.xlsm、.xls 文件"}, status_code=400)

    try:
        file_bytes = await file.read()
        keywords, input_type, total_rows, skipped_rows, file_name = _parse_excel_keywords(
            file_bytes, file.filename or "")
    except Exception as e:
        return JSONResponse({"error": f"Excel 解析失败: {e}"}, status_code=400)

    if not keywords:
        return JSONResponse({"error": "Excel 中未读取到有效的 SKU 或 OE 数据"}, status_code=400)

    oe_list, sku_mapping, not_found_skus = _resolve_oe_list(keywords, input_type)
    file_info = {
        "fileName": file_name, "totalRows": total_rows,
        "skippedRows": skipped_rows, "inputType": input_type,
    }
    return _start_search_task(oe_list, site, free_shipping_bool, search_strategy,
                              sku_mapping, not_found_skus, file_info)


def _start_search_task(oe_list: list[str], site: str, free_shipping: bool,
                       search_strategy: str, sku_mapping: dict | None = None,
                       not_found_skus: list | None = None, file_info: dict | None = None):
    oe_list = list(dict.fromkeys(oe_list))
    if not oe_list:
        return JSONResponse({"error": "未能识别有效的产品编号"}, status_code=400)

    task_id = str(int(time.time() * 1000))[-8:]
    _tasks[task_id] = {
        "status": "running",
        "progress": {"current": 0, "total": len(oe_list), "oe": "启动中..."},
    }

    t = threading.Thread(target=_search_batch, args=(oe_list, site, task_id, free_shipping, search_strategy))
    t.daemon = True
    t.start()

    resp: dict = {
        "task_id": task_id, "oe_list": oe_list,
        "sku_mapping": sku_mapping or {},
        "not_found_skus": not_found_skus or [],
    }
    if file_info:
        resp["file_info"] = file_info
    return resp


@app.get("/api/status/{task_id}")
def api_status(task_id: str):
    task = _tasks.get(task_id)
    if not task:
        return JSONResponse({"status": "not_found"}, status_code=404)
    if task["status"] == "running":
        return {"status": "running", "progress": task.get("progress")}
    if task["status"] == "error":
        return {"status": "error", "msg": task["msg"]}
    return {"status": "done", "total_oe": task["total_oe"], "results": task["results"]}


# ═══════════════════════ SKU-OE 管理 API ═══════════════════════

@app.get("/api/sku/list")
def api_sku_list():
    with _sku_oe_lock:
        items = [
            {"sku": sku, "oe": oe,
             "middleCode": _custom_middle_codes.get(sku) or _extract_middle_code(sku)}
            for sku, oe in sorted(_sku_oe_map.items())
        ]
    return {"items": items, "total": len(items)}


@app.post("/api/sku/add")
def api_sku_add(data: dict):
    sku = (data.get("sku") or "").strip()
    oe = (data.get("oe") or "").strip()
    middle_code = (data.get("middleCode") or "").strip()

    if not sku or not oe:
        return JSONResponse({"error": "SKU 和 OE 都不能为空"}, status_code=400)

    try:
        result = repo.add(sku, oe, middle_code)
    except ValueError as e:
        return JSONResponse({"error": str(e)}, status_code=400)

    with _sku_oe_lock:
        _sku_oe_map[sku] = oe
        if middle_code:
            _custom_middle_codes[sku] = middle_code
    _reload_middle_code_map()
    return {"success": True, **result}


@app.post("/api/sku/update")
def api_sku_update(data: dict):
    sku = (data.get("sku") or "").strip()
    oe = (data.get("oe") or "").strip()
    middle_code = (data.get("middleCode") or "").strip()

    if not sku or not oe:
        return JSONResponse({"error": "SKU 和 OE 都不能为空"}, status_code=400)

    try:
        result = repo.update(sku, oe, middle_code)
    except ValueError as e:
        return JSONResponse({"error": str(e)}, status_code=400)

    with _sku_oe_lock:
        _sku_oe_map[sku] = oe
        if middle_code:
            _custom_middle_codes[sku] = middle_code
        else:
            _custom_middle_codes.pop(sku, None)
    _reload_middle_code_map()
    return {"success": True, **result}


@app.post("/api/sku/delete")
def api_sku_delete(data: dict):
    sku = (data.get("sku") or "").strip()
    if not sku:
        return JSONResponse({"error": "缺少 SKU"}, status_code=400)

    try:
        repo.delete(sku)
    except ValueError as e:
        return JSONResponse({"error": str(e)}, status_code=400)

    with _sku_oe_lock:
        _sku_oe_map.pop(sku, None)
        _custom_middle_codes.pop(sku, None)
    _reload_middle_code_map()
    return {"success": True, "sku": sku}


@app.post("/api/sku/lookup")
def api_sku_lookup(data: dict):
    skus = data.get("skus", [])
    if not skus:
        return JSONResponse({"error": "缺少 SKU 列表"}, status_code=400)

    results = []
    not_found = []
    with _sku_oe_lock:
        for sku in skus:
            sku = sku.strip()
            if sku:
                oe = _sku_oe_map.get(sku) or _middle_code_oe_map.get(sku.upper())
                if oe:
                    results.append({"sku": sku, "oe": oe})
                else:
                    not_found.append(sku)

    return {"found": results, "not_found": not_found, "total": len(results)}


@app.post("/api/sku/refresh")
def api_sku_refresh():
    load_sku_oe_map()
    return {"success": True, "total": len(_sku_oe_map)}


@app.post("/api/sku/import")
async def api_sku_import(file: UploadFile = File(...)):
    fname_lower = (file.filename or "").lower()
    if not (fname_lower.endswith('.xlsx') or fname_lower.endswith('.xlsm') or fname_lower.endswith('.xls')):
        return JSONResponse({"error": "仅支持 .xlsx、.xlsm、.xls 文件"}, status_code=400)

    try:
        file_bytes = await file.read()
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return JSONResponse({"error": "Excel 文件没有工作表"}, status_code=400)

        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
        if len(rows) < 2:
            return JSONResponse({"error": "Excel 文件至少需要表头+1行数据"}, status_code=400)

        header = rows[0]

        def norm(v):
            return re.sub(r'[\s_]', '', str(v or '')).lower()

        col_sku = col_oe = -1
        for i, h in enumerate(header):
            n = norm(h)
            if n in ('sku', 'skuid', 'skuno'):
                col_sku = i
            elif n in ('oe', '主oe', '主oeno', 'oeno', 'oenumber', 'oenr', 'partnumber', 'partno'):
                col_oe = i
        if col_sku < 0:
            col_sku = 0
        if col_oe < 0:
            col_oe = 1

        import_rows: list[dict] = []
        total_rows = 0
        skipped_rows = 0
        for row in rows[1:]:
            total_rows += 1
            sku = str(row[col_sku]).strip() if len(row) > col_sku and row[col_sku] else ''
            oe = str(row[col_oe]).strip() if len(row) > col_oe and row[col_oe] else ''
            if not sku or not oe:
                skipped_rows += 1
                continue
            import_rows.append({"sku": sku, "oe": oe, "middle_code": None})

        if not import_rows:
            return JSONResponse({"error": "Excel 中未读取到有效的 SKU-OE 数据"}, status_code=400)

        stats = repo.batch_upsert(import_rows, file.filename or "")
        load_sku_oe_map()

        return {
            "success": True,
            "totalRows": total_rows,
            "skippedRows": skipped_rows,
            "imported": stats["imported"],
            "created": stats["created"],
            "updated": stats["updated"],
        }
    except Exception as e:
        return JSONResponse({"error": f"Excel 解析失败: {e}"}, status_code=400)


@app.post("/api/export")
def api_export(data: dict):
    items = data.get("items", [])
    if not items:
        return JSONResponse({"error": "没有可导出的商品"}, status_code=400)

    wb = Workbook()
    ws = wb.active
    ws.title = "eBay Results"

    headers = ["OE", "Price", "Title", "Seller", "Rate", "Shipping", "Link", "Images"]
    col_widths = [16, 12, 50, 14, 8, 12, 40, 18]
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = cell.font.copy(bold=True)
        ws.column_dimensions[get_column_letter(c)].width = w

    img_height = 80
    row_height_px = img_height + 10

    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item.get("oe", ""))
        ws.cell(row=i, column=2, value=item.get("price", ""))
        ws.cell(row=i, column=3, value=item.get("title", ""))
        ws.cell(row=i, column=4, value=item.get("seller", ""))
        fb = item.get("sellerFeedback", "")
        ws.cell(row=i, column=5, value=f"{fb}%" if fb else "")
        ws.cell(row=i, column=6, value=item.get("shipping", ""))
        link = item.get("link", "")
        if link:
            cell = ws.cell(row=i, column=7, value=link)
            cell.hyperlink = link
            cell.font = cell.font.copy(color="0563C1", underline="single")

        img_url = (item.get("images") or [None])[0]
        if img_url:
            try:
                resp = http_requests.get(img_url, timeout=10, verify=False)
                if resp.status_code == 200:
                    img_data = BytesIO(resp.content)
                    xl_img = XlImage(img_data)
                    xl_img.height = img_height
                    xl_img.width = img_height
                    ws.add_image(xl_img, f"H{i}")
            except Exception:
                pass

        ws.row_dimensions[i].height = row_height_px * 0.75

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, prefix="ebay_export_")
    wb.save(tmp.name)
    tmp.close()

    from starlette.background import BackgroundTask

    return FileResponse(
        tmp.name,
        filename=f"ebay_export_{int(time.time())}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(os.unlink, tmp.name),
    )
